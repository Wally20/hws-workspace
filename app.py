import os
import json
import calendar
import copy
import re
import sqlite3
import shutil
import subprocess
import tempfile
import threading
import time
import secrets
import hashlib
import hmac
import html
import base64
import mimetypes
import unicodedata
import zipfile
from io import BytesIO
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from decimal import Decimal, InvalidOperation
from datetime import date, datetime, time as dt_time, timedelta
from math import atan2, ceil, cos, sin
from typing import Any, Callable, Dict, List, Optional, Set, Tuple
from xml.etree import ElementTree as XmlElementTree
from zoneinfo import ZoneInfo

import requests
from django.conf import settings
from django.core.mail import EmailMessage
from flask import Flask, jsonify, redirect, render_template, request, send_from_directory, session, url_for
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from werkzeug.exceptions import HTTPException
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.security import check_password_hash, generate_password_hash

try:
    from pywebpush import WebPushException, webpush
except ImportError:  # pragma: no cover - optional in local development until requirements are installed.
    WebPushException = None
    webpush = None


ECWID_API_BASE = "https://app.ecwid.com/api/v3"
MONEYBIRD_API_BASE = "https://moneybird.com/api/v2"
VAT_SAVINGS_RATE = Decimal("0.09")
RIJKSOVERHEID_SCHOOL_HOLIDAYS_API_BASE = "https://opendata.rijksoverheid.nl/v1/infotypes/schoolholidays"
NAGER_PUBLIC_HOLIDAYS_API_BASE = "https://date.nager.at/api/v3/PublicHolidays"
BUNDLED_DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
DATA_DIR = os.getenv("DATA_DIR", BUNDLED_DATA_DIR)
DATABASE_PATH = os.path.join(DATA_DIR, "app.db")
DASHBOARD_EVENTS_PATH = os.path.join(DATA_DIR, "dashboard_events.json")
AGENDA_TRAININGS_PATH = os.path.join(DATA_DIR, "agenda_trainings.json")
CONTRACT_TEMPLATE_PATH = os.path.join(
    os.path.dirname(__file__),
    "static",
    "assets",
    "contracts",
    "HWS_Standaard_Overeenkomst.docx",
)
CONTRACT_WATERMARK_PATH = os.path.join(
    os.path.dirname(__file__),
    "static",
    "assets",
    "contracts",
    "HWS_watermark.png",
)
CONTRACT_WATERMARK_REL_ID = "rId999"
PPTX_XML_NAMESPACES = {
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "p": "http://schemas.openxmlformats.org/presentationml/2006/main",
}
DOCX_XML_NAMESPACES = {
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
    "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
    "wp": "http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing",
}
PPTX_SLIDE_WIDTH = 12192000
PPTX_SLIDE_HEIGHT = 6858000
FOOTBALL_DAYS_PDF_WIDTH = 960
FOOTBALL_DAYS_PDF_HEIGHT = 540
FOOTBALL_PLAYBOOK_CONTEXTS = {
    "voetbaldagen": {
        "playbookType": "voetbaldagen",
        "pageKey": "voetbaldagen",
        "pageTitle": "Voetbaldagen",
        "overviewDescription": "Open een opgeslagen draaiboek of maak een nieuwe voetbaldag aan.",
        "formDescription": "Bouw het draaiboek op uit gegevens, medewerkers, programma en onvoorziene omstandigheden.",
        "locationLabel": "Locatie",
        "locationPlaceholder": "Sportpark, veld of club",
        "showEcwidProduct": True,
        "showContingencies": True,
        "supportsSectionToggles": False,
        "supportsFieldTrainings": False,
        "supportsCycleDates": False,
        "supportsStaffSetupTasks": False,
        "defaultTitle": "Draaiboek Voetbaldagen",
        "titlePlaceholder": "Draaiboek Voetbaldag HWS",
        "saveAllLabel": "Alles van de voetbaldag opslaan",
        "pdfCoverTitle": "HWS VOETBALDAG",
        "introSubject": "voetbaldag",
        "overviewPath": "/voetbaldagen",
        "newPath": "/voetbaldagen/nieuw",
        "editPathPrefix": "/voetbaldagen",
        "registrationCountsApi": "/api/voetbaldagen/registration-counts",
        "exportPdfApi": "/api/voetbaldagen/export-pdf",
        "fallbackPdfFilename": "voetbaldag-draaiboek.pdf",
    },
    "samenwerkende-amateurclubs": {
        "playbookType": "samenwerkende-amateurclubs",
        "pageKey": "samenwerkende-amateurclubs",
        "pageTitle": "Samenwerkende Amateurclubs",
        "overviewDescription": "Open een opgeslagen draaiboek of maak een nieuw draaiboek voor een samenwerkende amateurclub aan.",
        "formDescription": "Bouw het draaiboek op uit clubgegevens, medewerkers, programma en trainingen.",
        "locationLabel": "Club",
        "locationPlaceholder": "Naam van de samenwerkende amateurclub",
        "showEcwidProduct": False,
        "showContingencies": False,
        "supportsSectionToggles": True,
        "supportsFieldTrainings": True,
        "supportsCycleDates": True,
        "supportsStaffSetupTasks": True,
        "defaultTitle": "Draaiboek Samenwerkende Amateurclubs",
        "titlePlaceholder": "Draaiboek Samenwerkende Amateurclub HWS",
        "saveAllLabel": "Alles van de samenwerkende amateurclub opslaan",
        "pdfCoverTitle": "SAMENWERKENDE AMATEURCLUBS",
        "introSubject": "samenwerkende amateurclub",
        "overviewPath": "/samenwerkende-amateurclubs",
        "newPath": "/samenwerkende-amateurclubs/nieuw",
        "editPathPrefix": "/samenwerkende-amateurclubs",
        "registrationCountsApi": "/api/samenwerkende-amateurclubs/registration-counts",
        "exportPdfApi": "/api/samenwerkende-amateurclubs/export-pdf",
        "fallbackPdfFilename": "samenwerkende-amateurclubs-draaiboek.pdf",
    },
}
EXERCISE_FIELD_MIN_X = 350000
EXERCISE_FIELD_MAX_X = 4700000
EXERCISE_FIELD_MIN_Y = 1600000
EXERCISE_FIELD_MAX_Y = 5200000
EXERCISE_TEXT_LABELS = (
    "OEFENING:",
    "TRAININGSOEFENING:",
    "DUUR:",
    "OMSCHRIJVING OEFENING:",
    "MATERIALEN:",
    "AFMETINGEN:",
    "COACHING:",
    "VARIATIE MAKKELIJKER MAKEN:",
    "VARIATIE MOEILIJKER MAKEN:",
)
EXERCISE_CATEGORY_OPTIONS = (
    "Dribbelvormen",
    "Pass-trapvormen",
    "Omschakelvormen",
    "1v1 vormen",
    "Afwerkvormen",
    "Partijvormen",
    "Fungames",
)
EXERCISE_AGE_GROUP_OPTIONS = ("O8", "O9", "O10", "O11", "O12", "O13", "O14", "O15")
EXERCISE_CATEGORY_ALIASES = {
    "DRIBBELVORMEN": "Dribbelvormen",
    "PASS-TRAPVORMEN": "Pass-trapvormen",
    "PASSTRAPVORMEN": "Pass-trapvormen",
    "PASS TRAPVORMEN": "Pass-trapvormen",
    "OMSCHAKELVORMEN": "Omschakelvormen",
    "1V1 VORMEN": "1v1 vormen",
    "1V1VORMEN": "1v1 vormen",
    "1 TEGEN 1 VORMEN": "1v1 vormen",
    "AFWERKVORMEN": "Afwerkvormen",
    "PARTIJVORMEN": "Partijvormen",
    "SPELVORMEN": "Fungames",
    "FUNGAMES": "Fungames",
}
EXERCISE_IMPORT_PREVIEW_DIR = os.path.join(DATA_DIR, "exercise_import_previews")
try:
    EXERCISE_FIELD_IMAGE_MAX_UPLOAD_MB = max(1, int(os.getenv("EXERCISE_FIELD_IMAGE_MAX_UPLOAD_MB", "5") or "5"))
except ValueError:
    EXERCISE_FIELD_IMAGE_MAX_UPLOAD_MB = 5
try:
    EXERCISE_VIDEO_MAX_UPLOAD_MB = max(1, int(os.getenv("BUNNY_VIDEO_MAX_UPLOAD_MB", "5000") or "5000"))
except ValueError:
    EXERCISE_VIDEO_MAX_UPLOAD_MB = 5000
AGENDA_DAY_PLAN_OPTIONS = (
    "Geen activiteit",
    "Voetbaldag",
    "Samenwerkende amateurclubs",
    "Techniektrainingen",
)
AGENDA_AMATEUR_CLUB_OPTIONS = (
    "WWNA",
    "ABS",
    "VV Oeken",
    "VV Gorssel",
)
AGENDA_TECHNIQUE_CLUB_OPTIONS = (
    "VV Diepenveen",
    "ABS",
    "Apeldoornse Boys",
)
AGENDA_CLUB_OPTIONS = tuple(dict.fromkeys((*AGENDA_AMATEUR_CLUB_OPTIONS, *AGENDA_TECHNIQUE_CLUB_OPTIONS)))
AGENDA_CLUB_CLASS_NAMES = {
    "WWNA": "agenda-event-club-wwna",
    "ABS": "agenda-event-club-abs",
    "VV Oeken": "agenda-event-club-vv-oeken",
    "VV Gorssel": "agenda-event-club-vv-gorssel",
    "VV Diepenveen": "agenda-event-club-vv-diepenveen",
    "Apeldoornse Boys": "agenda-event-club-apeldoornse-boys",
}
AGENDA_CLUB_OPTIONS_BY_TRAINING_TYPE = {
    "samenwerkende_amateurclub": AGENDA_AMATEUR_CLUB_OPTIONS,
    "techniektraining": AGENDA_TECHNIQUE_CLUB_OPTIONS,
}
TRAINER_FEE_ALL_CLUBS_VALUE = "Alle clubs"
TRAINER_FEE_ALL_ACTIVITIES_VALUE = "Alle activiteiten"
AGENDA_TRAINING_TYPE_OPTIONS = (
    {
        "value": "voetbaldag",
        "label": "Voetbaldag",
        "className": "agenda-event-type-voetbaldag",
    },
    {
        "value": "summercamp",
        "label": "SummerCamp",
        "className": "agenda-event-type-summercamp",
    },
    {
        "value": "samenwerkende_amateurclub",
        "label": "Samenwerkende amateurclub",
        "className": "agenda-event-type-samenwerkende-amateurclub",
    },
    {
        "value": "techniektraining",
        "label": "Techniektraining",
        "className": "agenda-event-type-techniektraining",
    },
    {
        "value": "clinic",
        "label": "Clinic",
        "className": "agenda-event-type-clinic",
    },
)
AGENDA_TRAINING_STATUS_OPTIONS = (
    {
        "value": "gepland",
        "label": "Gepland",
        "className": "agenda-event-status-gepland",
    },
    {
        "value": "gegeven",
        "label": "Gegeven",
        "className": "agenda-event-status-gegeven",
    },
    {
        "value": "geannuleerd",
        "label": "Geannuleerd",
        "className": "agenda-event-status-geannuleerd",
    },
)
TRAINER_FEE_ACTIVITY_OPTIONS = tuple(
    {
        "value": str(option["value"]),
        "label": str(option["label"]),
    }
    for option in AGENDA_TRAINING_TYPE_OPTIONS
)
TRAINER_FEE_TYPE_OPTIONS = (
    {
        "value": "samenwerkende_amateurclub",
        "label": "Samenwerkende amateurclub",
    },
    {
        "value": "techniektraining",
        "label": "Techniektraining",
    },
    {
        "value": "voetbaldag_summercamp",
        "label": "Voetbaldag/SummerCamp",
    },
)
TRAINER_FEE_CLUB_OPTIONS_BY_TYPE = {
    "samenwerkende_amateurclub": AGENDA_AMATEUR_CLUB_OPTIONS,
    "techniektraining": AGENDA_TECHNIQUE_CLUB_OPTIONS,
    "voetbaldag_summercamp": (TRAINER_FEE_ALL_CLUBS_VALUE,),
}
AGENDA_NO_ACTIVITY_COPY_REASONS = {
    "2026-10-19": "herfstvakantie",
    "2026-10-21": "herfstvakantie",
    "2026-12-21": "winterstop",
    "2026-12-23": "winterstop",
    "2026-12-28": "winterstop",
    "2026-12-30": "winterstop",
    "2027-01-04": "winterstop",
    "2027-01-06": "winterstop",
    "2027-02-22": "voorjaarsvakantie",
    "2027-02-24": "voorjaarsvakantie",
    "2027-03-29": "Eerste Paasdag",
    "2027-04-26": "meivakantie",
    "2027-04-28": "meivakantie",
    "2027-05-03": "meivakantie",
    "2027-05-05": "meivakantie",
    "2027-05-17": "Tweede Pinksterdag",
}
PROPOSAL_TYPE_OPTIONS = (
    {
        "value": "amateurclub",
        "label": "Samenwerkende amateurclub",
        "agenda_plan_type": "Samenwerkende amateurclubs",
    },
    {
        "value": "techniektrainingen",
        "label": "Club voor techniektrainingen",
        "agenda_plan_type": "Techniektrainingen",
    },
)
PROPOSAL_WEEKDAY_OPTIONS = (
    {"value": "monday", "label": "Maandag", "python_weekday": 0},
    {"value": "tuesday", "label": "Dinsdag", "python_weekday": 1},
    {"value": "wednesday", "label": "Woensdag", "python_weekday": 2},
    {"value": "thursday", "label": "Donderdag", "python_weekday": 3},
    {"value": "friday", "label": "Vrijdag", "python_weekday": 4},
    {"value": "saturday", "label": "Zaterdag", "python_weekday": 5},
    {"value": "sunday", "label": "Zondag", "python_weekday": 6},
)
PROPOSAL_TRAINING_KIND_OPTIONS = (
    {"value": "teamtraining", "label": "Teamtraining"},
    {"value": "circuittraining", "label": "Circuittraining"},
)
AGENDA_SUMMARY_FILTER_OPTIONS = (
    {
        "key": "total",
        "label": "Totaal",
        "description": "Alle ingevoerde dagen",
    },
    {
        "key": "season_2026_2027",
        "label": "Seizoen 2026/2027",
        "description": "Maandag 24 augustus 2026 t/m zondag 13 juni 2027",
        "start": date(2026, 8, 24),
        "end": date(2027, 6, 13),
    },
)
DUTCH_MONTH_NAMES = ["jan", "feb", "mrt", "apr", "mei", "jun", "jul", "aug", "sep", "okt", "nov", "dec"]
DUTCH_FULL_MONTH_NAMES = [
    "januari",
    "februari",
    "maart",
    "april",
    "mei",
    "juni",
    "juli",
    "augustus",
    "september",
    "oktober",
    "november",
    "december",
]
DUTCH_WEEKDAY_NAMES = [
    "Maandag",
    "Dinsdag",
    "Woensdag",
    "Donderdag",
    "Vrijdag",
    "Zaterdag",
    "Zondag",
]
ECWID_RESPONSE_FIELDS = (
    "total,count,offset,limit,"
    "items(id,orderNumber,createDate,status,paymentStatus,fulfillmentStatus,total,email,"
    "paymentMethod,shippingOption,items(productId,name,quantity,price,sku),"
    "shippingPerson(name,street,city,postalCode,phone),billingPerson(name,street,city,postalCode,phone),"
    "extraFields,orderExtraFields(id,title,value))"
)
ECWID_PROCESSING_FULFILLMENT_STATUS = "PROCESSING"
ECWID_DELIVERED_FULFILLMENT_STATUS = "DELIVERED"
ECWID_RETURNED_FULFILLMENT_STATUS = "RETURNED"

app = Flask(__name__)
app.config["SEND_FILE_MAX_AGE_DEFAULT"] = 31536000
ASSET_VERSION = str(int(time.time()))

CACHE_TTL_SECONDS = max(60, int(os.getenv("EXTERNAL_DATA_CACHE_TTL_SECONDS", "1800") or "1800"))
LOCAL_DATA_CACHE_TTL_SECONDS = max(1, int(os.getenv("LOCAL_DATA_CACHE_TTL_SECONDS", "30") or "30"))
AGENDA_EXTERNAL_CACHE_TTL_SECONDS = 43200
orders_cache: Dict[str, Any] = {
    "payload": None,
    "cached_at": 0.0,
}
cache_lock = threading.Lock()
refresh_in_progress = False
ecwid_orders_cache: Dict[str, Any] = {
    "payload": None,
    "cached_at": 0.0,
}
ecwid_orders_cache_lock = threading.Lock()
ecwid_refresh_in_progress = False
moneybird_cache: Dict[str, Any] = {
    "payload": None,
    "cached_at": 0.0,
}
moneybird_cache_lock = threading.Lock()
moneybird_refresh_in_progress = False
catalog_products_cache: Dict[str, Any] = {
    "payload": None,
    "cached_at": 0.0,
}
catalog_products_cache_lock = threading.Lock()
agenda_school_holidays_cache: Dict[str, Any] = {}
agenda_school_holidays_cache_lock = threading.Lock()
agenda_public_holidays_cache: Dict[str, Any] = {}
agenda_public_holidays_cache_lock = threading.Lock()
local_data_cache: Dict[Tuple[Any, ...], Dict[str, Any]] = {}
local_data_cache_lock = threading.Lock()
content_album_lock = threading.Lock()


def get_file_cache_fingerprint(path: str) -> Tuple[int, int]:
    try:
        stat_result = os.stat(path)
    except OSError:
        return (0, 0)
    return (stat_result.st_mtime_ns, stat_result.st_size)


def get_cached_local_data(cache_name: str, cache_args: Tuple[Any, ...], loader: Callable[[], Any]) -> Any:
    now = time.time()
    fingerprint = get_file_cache_fingerprint(DATABASE_PATH)
    cache_key = (cache_name, *cache_args)

    with local_data_cache_lock:
        cached_payload = local_data_cache.get(cache_key)
        if (
            cached_payload is not None
            and cached_payload.get("fingerprint") == fingerprint
            and now - float(cached_payload.get("cached_at") or 0.0) < LOCAL_DATA_CACHE_TTL_SECONDS
        ):
            return copy.deepcopy(cached_payload["payload"])

    payload = loader()
    with local_data_cache_lock:
        local_data_cache[cache_key] = {
            "payload": copy.deepcopy(payload),
            "fingerprint": fingerprint,
            "cached_at": now,
        }
    return payload


def clear_local_data_cache() -> None:
    with local_data_cache_lock:
        local_data_cache.clear()

DEFAULT_PASSWORD_HASH_METHOD = "scrypt" if hasattr(hashlib, "scrypt") else "pbkdf2:sha256"
PASSWORD_HASH_METHOD = os.getenv("PASSWORD_HASH_METHOD", "").strip() or DEFAULT_PASSWORD_HASH_METHOD
SESSION_PERSISTENT_SECONDS = max(86400, int(os.getenv("SESSION_PERSISTENT_SECONDS", "34560000") or "34560000"))
SESSION_IDLE_TIMEOUT_SECONDS = max(300, int(os.getenv("SESSION_IDLE_TIMEOUT_SECONDS", "3600") or "3600"))
SESSION_ABSOLUTE_TIMEOUT_SECONDS = max(
    SESSION_IDLE_TIMEOUT_SECONDS,
    int(os.getenv("SESSION_ABSOLUTE_TIMEOUT_SECONDS", str(SESSION_PERSISTENT_SECONDS)) or str(SESSION_PERSISTENT_SECONDS)),
)
CSRF_TOKEN_LENGTH = 48
GENERIC_AUTH_ERROR_MESSAGE = "De combinatie van inloggegevens is ongeldig of de actie kon niet worden voltooid."
ALLOWED_IMAGE_EXTENSIONS = {
    "image/jpeg": {".jpg", ".jpeg"},
    "image/png": {".png"},
    "image/webp": {".webp"},
    "image/avif": {".avif"},
}
ALLOWED_VIDEO_EXTENSIONS = {
    "video/mp4": {".mp4", ".m4v"},
    "video/webm": {".webm"},
    "video/quicktime": {".mov"},
}
SECURITY_HEADERS = {
    "Content-Security-Policy": (
        "default-src 'self'; "
        "base-uri 'self'; "
        "form-action 'self'; "
        "frame-ancestors 'none'; "
        "object-src 'none'; "
        "img-src 'self' data: https:; "
        "media-src 'self' https: blob:; "
        "script-src 'self' 'unsafe-inline'; "
        "style-src 'self' 'unsafe-inline'; "
        "font-src 'self' data:; "
        "worker-src 'self'; "
        "manifest-src 'self'; "
        "connect-src 'self' https://opendata.rijksoverheid.nl https://date.nager.at; "
        "upgrade-insecure-requests"
    ),
    "Referrer-Policy": "strict-origin-when-cross-origin",
    "X-Content-Type-Options": "nosniff",
    "X-Frame-Options": "DENY",
    "Permissions-Policy": "camera=(), geolocation=(), microphone=()",
    "Cross-Origin-Opener-Policy": "same-origin",
}
RATE_LIMIT_RULES = (
    (re.compile(r"^/login$"), 5, 300, "login"),
    (re.compile(r"^/uitnodiging/[^/]+$"), 5, 600, "invite"),
    (re.compile(r"^/api/v1/agenda/(?:events|calendar\.ics)$"), 120, 300, "agenda-api"),
    (re.compile(r"^/api/dashboard-events$"), 20, 300, "dashboard-events"),
    (re.compile(r"^/content(?:/\d+)?$"), 20, 300, "content"),
    (re.compile(r"^/trainers$"), 20, 300, "trainers"),
)
PROPOSAL_MIN_SEASON_START_YEAR = 2026

WORKSPACE_SEARCH_PAGES = (
    {
        "key": "dashboard",
        "title": "Dashboard",
        "path": "/",
        "section": "Algemeen",
        "description": "Dagoverzicht, planning, omzet en snelle signalen.",
        "keywords": ("home", "start", "overzicht", "vandaag", "planning"),
    },
    {
        "key": "agenda",
        "title": "Agenda",
        "path": "/agenda",
        "section": "Algemeen",
        "description": "Trainingsdagen, vakanties, activiteiten en planning.",
        "keywords": ("kalender", "planning", "training", "voetbaldag", "vakantie"),
    },
    {
        "key": "draaiboeken",
        "title": "Draaiboeken",
        "path": "/draaiboeken",
        "section": "Draaiboeken",
        "description": "Startpunt voor alle draaiboeken.",
        "keywords": ("playbook", "programma", "organisatie"),
    },
    {
        "key": "voetbaldagen",
        "title": "Voetbaldagen",
        "path": "/voetbaldagen",
        "section": "Draaiboeken",
        "description": "Draaiboeken voor voetbaldagen bekijken en beheren.",
        "keywords": ("draaiboek", "voetbaldag", "inschrijvingen", "programma"),
    },
    {
        "key": "voetbaldagen",
        "title": "Nieuwe voetbaldag",
        "path": "/voetbaldagen/nieuw",
        "section": "Draaiboeken",
        "description": "Maak direct een nieuw draaiboek voor een voetbaldag.",
        "keywords": ("nieuw", "maken", "draaiboek", "voetbaldag"),
    },
    {
        "key": "samenwerkende-amateurclubs",
        "title": "Samenwerkende amateurclubs",
        "path": "/samenwerkende-amateurclubs",
        "section": "Draaiboeken",
        "description": "Draaiboeken voor samenwerkende amateurclubs.",
        "keywords": ("club", "amateurclub", "samenwerking", "trainingen"),
    },
    {
        "key": "samenwerkende-amateurclubs",
        "title": "Nieuwe samenwerkende amateurclub",
        "path": "/samenwerkende-amateurclubs/nieuw",
        "section": "Draaiboeken",
        "description": "Maak direct een nieuw clubdraaiboek.",
        "keywords": ("nieuw", "maken", "club", "amateurclub", "draaiboek"),
    },
    {
        "key": "management",
        "title": "Management",
        "path": "/management",
        "section": "Management",
        "description": "Startpunt voor voorstellen, overeenkomsten, aanmeldingen en team.",
        "keywords": ("beheer", "admin", "team", "orders"),
    },
    {
        "key": "planning",
        "title": "Planning",
        "path": "/planning",
        "section": "Management",
        "description": "Programma's als planningstabel maken, bewaren en exporteren.",
        "keywords": ("planning", "programma", "tabel", "pdf"),
    },
    {
        "key": "api",
        "title": "API",
        "path": "/management/api",
        "section": "Management",
        "description": "Beheer de beveiligde koppeling met de HWS-agenda.",
        "keywords": ("api", "agenda", "koppeling", "token", "ics", "integratie"),
    },
    {
        "key": "begroting",
        "title": "Begroting",
        "path": "/begroting",
        "section": "Management",
        "description": "Inkomsten, trainerbudget en resultaat per seizoen begroten.",
        "keywords": ("budget", "kosten", "inkomsten", "seizoen"),
    },
    {
        "key": "materialen",
        "title": "Materialen",
        "path": "/materialen",
        "section": "Management",
        "description": "Materiaalvoorraad per club en opslag beheren.",
        "keywords": ("materiaal", "materialen", "voorraad", "clubs", "opslag"),
    },
    {
        "key": "orders",
        "title": "Aanmeldingen",
        "path": "/aanmeldingen",
        "section": "Management",
        "description": "Inschrijvingen en details per product bekijken.",
        "keywords": ("orders", "inschrijvingen", "producten", "deelnemers"),
    },
    {
        "key": "orders",
        "title": "Bestellingen",
        "path": "/bestellingen",
        "section": "Management",
        "description": "Bestellingen, betalingen en statussen bekijken.",
        "keywords": ("orders", "bestellingen", "betalingen", "ecwid"),
    },
    {
        "key": "leads",
        "title": "Leads",
        "path": "/leads",
        "section": "Management",
        "description": "Leads en geblokkeerde e-mails beheren.",
        "keywords": ("contacten", "emails", "marketing", "aanvragen"),
    },
    {
        "key": "trainers",
        "title": "Team",
        "path": "/trainers",
        "section": "Management",
        "description": "Teamleden, rollen en uitnodigingen beheren.",
        "keywords": ("trainers", "medewerkers", "rollen", "uitnodigingen"),
    },
    {
        "key": "voorstellen-maker",
        "title": "Voorstellen maker",
        "path": "/voorstellen-maker",
        "section": "Management",
        "description": "Voorstellen voor samenwerkingen en trainingen maken.",
        "keywords": ("voorstel", "offerte", "club", "techniektrainingen"),
    },
    {
        "key": "overeenkomsten",
        "title": "Overeenkomsten",
        "path": "/overeenkomsten",
        "section": "Management",
        "description": "Afspraken vastleggen en documenten exporteren.",
        "keywords": ("contract", "afspraak", "pdf", "docx"),
    },
    {
        "key": "overeenkomsten",
        "title": "Nieuwe overeenkomst",
        "path": "/overeenkomsten/nieuw",
        "section": "Management",
        "description": "Maak direct een nieuwe overeenkomst.",
        "keywords": ("nieuw", "contract", "afspraak", "document"),
    },
    {
        "key": "financien",
        "title": "Financiën",
        "path": "/financien",
        "section": "Financiën",
        "description": "Startpunt voor omzet, winst en spaarpot.",
        "keywords": ("geld", "omzet", "winst", "spaarpot"),
    },
    {
        "key": "financien",
        "title": "Automatische facturen",
        "path": "/financien/automatisch-facturen",
        "section": "Financiën",
        "description": "Automatische conceptfacturen instellen en verwerken.",
        "keywords": ("moneybird", "facturen", "automatisch", "clubs"),
    },
    {
        "key": "revenue",
        "title": "Omzet",
        "path": "/omzet",
        "section": "Financiën",
        "description": "Omzetoverzichten openen.",
        "keywords": ("revenue", "inkomsten", "totaal", "maand", "seizoen"),
    },
    {
        "key": "revenue",
        "title": "Omzet totaal",
        "path": "/omzet/totaal",
        "section": "Financiën",
        "description": "Totale omzet bekijken.",
        "keywords": ("revenue", "inkomsten", "totaal"),
    },
    {
        "key": "revenue",
        "title": "Omzet per maand",
        "path": "/omzet/per-maand",
        "section": "Financiën",
        "description": "Maandomzet bekijken.",
        "keywords": ("revenue", "inkomsten", "maand", "grafiek"),
    },
    {
        "key": "revenue",
        "title": "Winst",
        "path": "/omzet/winst",
        "section": "Financiën",
        "description": "Winstoverzicht bekijken.",
        "keywords": ("profit", "winst", "kosten"),
    },
    {
        "key": "revenue",
        "title": "Omzet per seizoen",
        "path": "/omzet/per-seizoen",
        "section": "Financiën",
        "description": "Seizoensomzet bekijken.",
        "keywords": ("revenue", "inkomsten", "seizoen"),
    },
    {
        "key": "spaarpot",
        "title": "Spaarpot",
        "path": "/spaarpot",
        "section": "Financiën",
        "description": "Spaarpot en reserveringen beheren.",
        "keywords": ("sparen", "btw", "reservering", "geld"),
    },
    {
        "key": "trainer-fees",
        "title": "Trainersvergoedingen",
        "path": "/trainersvergoedingen",
        "section": "Financiën",
        "description": "Trainersvergoedingen beheren.",
        "keywords": ("vergoeding", "trainers", "betaling"),
    },
    {
        "key": "oefenstof",
        "title": "Oefenstof",
        "path": "/oefenstof",
        "section": "Oefenstof",
        "description": "Startpunt voor oefeningen, trainingen en videos.",
        "keywords": ("training", "oefeningen", "bibliotheek", "videos"),
    },
    {
        "key": "oefeningen-bibliotheek",
        "title": "Oefeningen bibliotheek",
        "path": "/oefeningen-bibliotheek",
        "section": "Oefenstof",
        "description": "Oefeningen bekijken, importeren en bewerken.",
        "keywords": ("oefening", "bibliotheek", "import", "categorie", "leeftijd"),
    },
    {
        "key": "trainingen",
        "title": "Trainingen",
        "path": "/trainingen",
        "section": "Oefenstof",
        "description": "Trainingen maken en openen.",
        "keywords": ("training", "maker", "opgeslagen", "sessie"),
    },
    {
        "key": "trainingen",
        "title": "Opgeslagen trainingen",
        "path": "/trainingen/opgeslagen",
        "section": "Oefenstof",
        "description": "Eerder gemaakte trainingen openen.",
        "keywords": ("training", "opgeslagen", "archief"),
    },
    {
        "key": "trainingen",
        "title": "Training maker",
        "path": "/trainingen/maker",
        "section": "Oefenstof",
        "description": "Direct een training samenstellen.",
        "keywords": ("training", "maken", "oefeningen"),
    },
    {
        "key": "exercise-videos",
        "title": "Oefeningen videos",
        "path": "/oefeningen-videos",
        "section": "Oefenstof",
        "description": "Videos uploaden en koppelen aan oefeningen.",
        "keywords": ("video", "upload", "oefening", "koppelen"),
    },
    {
        "key": "marketing",
        "title": "Marketing",
        "path": "/marketing",
        "section": "Marketing",
        "description": "Startpunt voor social media, content en leads.",
        "keywords": ("content", "social", "media", "leads"),
    },
    {
        "key": "social-media",
        "title": "Social media",
        "path": "/social-media",
        "section": "Marketing",
        "description": "Social posts en ideeen beheren.",
        "keywords": ("instagram", "facebook", "post", "planning"),
    },
    {
        "key": "content",
        "title": "Content",
        "path": "/content",
        "section": "Marketing",
        "description": "Content en albums beheren.",
        "keywords": ("foto", "album", "media", "upload"),
    },
    {
        "key": "profile",
        "title": "Profiel",
        "path": "/profiel",
        "section": "Algemeen",
        "description": "Eigen profiel en account bekijken.",
        "keywords": ("account", "profiel", "persoonlijk", "gebruiker"),
    },
)


def get_workspace_search_pages_for_user(user: Optional[Dict[str, Any]]) -> List[Dict[str, Any]]:
    visible_pages = get_visible_pages_for_user(user)
    return [
        {
            "title": page["title"],
            "path": page["path"],
            "section": page["section"],
            "description": page["description"],
            "keywords": list(page["keywords"]),
        }
        for page in WORKSPACE_SEARCH_PAGES
        if page["key"] in visible_pages
    ]


def get_current_workspace_navigation_path(user: Optional[Dict[str, Any]], current_path: str) -> str:
    """Return the most specific menu path for overview and detail pages."""
    available_paths = [page["path"] for page in get_workspace_search_pages_for_user(user)]
    exact_matches = [path for path in available_paths if path == current_path]
    if exact_matches:
        return exact_matches[0]

    prefix_matches = [
        path
        for path in available_paths
        if path != "/" and current_path.startswith(f"{path}/")
    ]
    return max(prefix_matches, key=len, default="")


def get_asset_version() -> str:
    latest_mtime = 0
    static_root = os.path.join(os.path.dirname(__file__), "static")

    for root, _, filenames in os.walk(static_root):
        for filename in filenames:
            file_path = os.path.join(root, filename)
            try:
                latest_mtime = max(latest_mtime, int(os.path.getmtime(file_path)))
            except OSError:
                continue

    if latest_mtime:
        return str(latest_mtime)
    return ASSET_VERSION


@app.context_processor
def inject_asset_version():
    return {
        "asset_version": get_asset_version(),
        "legacy_csrf_token": ensure_csrf_token(),
    }


def load_dotenv() -> None:
    env_path = os.path.join(os.path.dirname(__file__), ".env")
    if not os.path.exists(env_path):
        return

    with open(env_path, "r", encoding="utf-8") as env_file:
        for raw_line in env_file:
            line = raw_line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue

            key, value = line.split("=", 1)
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = value


load_dotenv()


def get_env(name: str) -> str:
    return os.getenv(name, "").strip()


def get_env_int(name: str, default: int) -> int:
    value = get_env(name)
    if not value:
        return default
    try:
        return int(value)
    except ValueError:
        return default


PASSWORD_HASH_METHOD = get_env("PASSWORD_HASH_METHOD") or PASSWORD_HASH_METHOD
SESSION_PERSISTENT_SECONDS = max(86400, get_env_int("SESSION_PERSISTENT_SECONDS", SESSION_PERSISTENT_SECONDS))
SESSION_IDLE_TIMEOUT_SECONDS = max(300, get_env_int("SESSION_IDLE_TIMEOUT_SECONDS", SESSION_IDLE_TIMEOUT_SECONDS))
SESSION_ABSOLUTE_TIMEOUT_SECONDS = max(
    SESSION_IDLE_TIMEOUT_SECONDS,
    get_env_int("SESSION_ABSOLUTE_TIMEOUT_SECONDS", max(SESSION_ABSOLUTE_TIMEOUT_SECONDS, SESSION_PERSISTENT_SECONDS)),
)
AGENDA_SCHOOL_REGION = (get_env("AGENDA_SCHOOL_REGION") or "midden").strip().lower() or "midden"


def is_placeholder_value(value: str) -> bool:
    normalized = (value or "").strip()
    if not normalized:
        return True

    lowered = normalized.lower()
    compact = re.sub(r"[^a-z0-9]+", "", lowered)
    placeholder_fragments = (
        "hierjouw",
        "replacewith",
        "placeholder",
        "example",
        "yourstore",
        "yoursecret",
        "ecwidstoreid",
        "ecwidsecrettoken",
        "moneybirdapitoken",
        "moneybirdadministrationid",
    )
    if compact == "..." or any(fragment in compact for fragment in placeholder_fragments):
        return True

    return lowered.startswith("hier_jouw_") or lowered.startswith("replace-with-") or lowered in {
        "xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx",
        "secret_xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx",
        "12345678",
        "123456789012345678",
        "...",
    }


def get_secret_env(name: str) -> str:
    value = get_env(name)
    if value and not is_placeholder_value(value) and len(value) >= 32:
        return value
    return ""


def trusted_hosts_are_local(hosts: List[str]) -> bool:
    local_hosts = {"127.0.0.1", "localhost", "testserver"}
    normalized_hosts = {
        host.split(":", 1)[0].strip().lower()
        for host in hosts
        if host.strip()
    }
    return not normalized_hosts or normalized_hosts.issubset(local_hosts)


def should_require_configured_secret(trusted_hosts: List[str]) -> bool:
    return not app.debug and not trusted_hosts_are_local(trusted_hosts)


def get_env_bool(name: str, default: bool = False) -> bool:
    value = get_env(name)
    if not value:
        return default
    return value.lower() in {"1", "true", "yes", "on"}


def get_env_int(name: str, default: int) -> int:
    value = get_env(name)
    if not value:
        return default
    try:
        return int(value)
    except ValueError:
        return default


def configure_app() -> None:
    trusted_hosts = [item.strip() for item in get_env("TRUSTED_HOSTS").split(",") if item.strip()]
    session_cookie_secure_default = get_env_bool("SESSION_COOKIE_SECURE", default=not app.debug)
    configured_secret = get_secret_env("FLASK_SECRET_KEY")

    if not configured_secret and should_require_configured_secret(trusted_hosts):
        raise RuntimeError("FLASK_SECRET_KEY ontbreekt of is te zwak. Gebruik een random secret van minimaal 32 tekens.")

    app.config.update(
        SECRET_KEY=configured_secret or secrets.token_urlsafe(48),
        SESSION_COOKIE_NAME=get_env("SESSION_COOKIE_NAME") or "overzicht_session",
        SESSION_COOKIE_HTTPONLY=True,
        SESSION_COOKIE_SECURE=session_cookie_secure_default,
        SESSION_COOKIE_SAMESITE=get_env("SESSION_COOKIE_SAMESITE") or "Lax",
        PREFERRED_URL_SCHEME=get_env("PREFERRED_URL_SCHEME") or "https",
        PERMANENT_SESSION_LIFETIME=timedelta(seconds=SESSION_PERSISTENT_SECONDS),
    )

    if trusted_hosts:
        app.config["TRUSTED_HOSTS"] = trusted_hosts

    proxy_hops = max(0, get_env_int("REVERSE_PROXY_HOPS", 1))
    if proxy_hops:
        app.wsgi_app = ProxyFix(
            app.wsgi_app,
            x_for=proxy_hops,
            x_proto=proxy_hops,
            x_host=proxy_hops,
            x_port=proxy_hops,
            x_prefix=proxy_hops,
        )


configure_app()


def is_request_secure() -> bool:
    forwarded_proto = str(request.headers.get("X-Forwarded-Proto", "") or "").split(",")[0].strip().lower()
    if forwarded_proto:
        return forwarded_proto == "https"
    return bool(getattr(request, "is_secure", False))


def should_enforce_https() -> bool:
    return get_env_bool("FORCE_HTTPS", default=app.config.get("SESSION_COOKIE_SECURE", False))


def should_skip_https_redirect() -> bool:
    host = str(request.headers.get("Host", "") or "").split(":")[0].strip().lower()
    return host in {"localhost", "127.0.0.1", "testserver"}


def get_client_ip() -> str:
    forwarded_for = str(request.headers.get("X-Forwarded-For", "") or "").strip()
    if forwarded_for:
        return forwarded_for.split(",")[0].strip()
    for header_name in ("X-Real-IP", "CF-Connecting-IP"):
        header_value = str(request.headers.get(header_name, "") or "").strip()
        if header_value:
            return header_value
    return str(getattr(request, "remote_addr", "") or "unknown").strip() or "unknown"


def hash_password(password: str) -> str:
    return generate_password_hash(password.strip(), method=PASSWORD_HASH_METHOD, salt_length=16)


def password_needs_rehash(password_hash: str) -> bool:
    normalized_hash = str(password_hash or "").strip()
    return bool(normalized_hash) and not normalized_hash.startswith(f"{PASSWORD_HASH_METHOD}:")


def update_user_password_hash(profile_id: str, password: str) -> None:
    with get_db_connection() as connection:
        connection.execute(
            "UPDATE trainer_profiles SET password_hash = ? WHERE id = ?",
            (hash_password(password), profile_id.strip()),
        )
    clear_local_data_cache()


def ensure_csrf_token() -> str:
    token = str(session.get("csrf_token", "") or "").strip()
    if len(token) < CSRF_TOKEN_LENGTH:
        token = secrets.token_urlsafe(36)
        session["csrf_token"] = token
    return token


def is_safe_redirect_target(target: str) -> bool:
    normalized = str(target or "").strip()
    return normalized.startswith("/") and not normalized.startswith("//")


def get_request_csrf_token() -> str:
    header_token = str(
        request.headers.get("X-CSRF-Token", "") or request.headers.get("X-CSRFToken", "") or ""
    ).strip()
    if header_token:
        return header_token
    environ = getattr(request, "environ", {}) or getattr(request, "META", {})
    if isinstance(environ, dict):
        header_token = str(environ.get("HTTP_X_CSRF_TOKEN", "") or environ.get("HTTP_X_CSRFTOKEN", "")).strip()
        if header_token:
            return header_token
    return str(request.form.get("csrf_token", "") or "").strip()


def csrf_error_response() -> Any:
    if request.path.startswith("/api/") or request_prefers_json():
        return jsonify({"error": "Ongeldig of ontbrekend CSRF-token."}), 403
    return "Ongeldig of ontbrekend CSRF-token.", 403


def validate_csrf_token() -> Optional[Any]:
    if request.method in {"GET", "HEAD", "OPTIONS"}:
        return None
    session_token = ensure_csrf_token()
    request_token = get_request_csrf_token()
    if request_token and hmac.compare_digest(session_token, request_token):
        return None
    return csrf_error_response()


def rotate_authenticated_session(user_id: str) -> None:
    session.clear()
    session.permanent = True
    session["user_id"] = user_id
    session["csrf_token"] = secrets.token_urlsafe(36)
    now = int(time.time())
    session["session_started_at"] = now
    session["session_last_seen_at"] = now


def handle_session_timeout() -> Optional[Any]:
    user_id = str(session.get("user_id", "") or "").strip()
    if not user_id:
        ensure_csrf_token()
        return None

    now = int(time.time())
    if not session.get("session_started_at"):
        session["session_started_at"] = now
    session["session_last_seen_at"] = now
    session.permanent = True
    ensure_csrf_token()
    return None


def get_rate_limit_rule(path: str) -> Optional[Tuple[int, int, str]]:
    for pattern, max_attempts, window_seconds, scope in RATE_LIMIT_RULES:
        if pattern.match(path):
            return max_attempts, window_seconds, scope
    return None


def apply_rate_limit(max_attempts: int, window_seconds: int, scope: str) -> Optional[int]:
    user = get_current_user()
    identity = str(user["id"]) if user is not None else ""
    request_key = f"{scope}:{get_client_ip()}:{identity}:{request.path}"
    if scope == "login":
        request_key = f"{request_key}:{request.form.get('email', '').strip().lower()[:120]}"
    if scope == "invite":
        request_key = f"{request_key}:{request.path.rsplit('/', 1)[-1]}"

    now = time.time()
    window_start = now - window_seconds

    with get_db_connection() as connection:
        connection.execute("BEGIN IMMEDIATE")
        connection.execute(
            "DELETE FROM rate_limit_attempts WHERE created_at < ?",
            (window_start,),
        )
        row = connection.execute(
            """
            SELECT COUNT(*) AS total, MIN(created_at) AS first_attempt
            FROM rate_limit_attempts
            WHERE request_key = ? AND created_at >= ?
            """,
            (request_key, window_start),
        ).fetchone()
        total_attempts = int(row["total"] or 0)
        first_attempt = float(row["first_attempt"] or now)
        if total_attempts >= max_attempts:
            retry_after = max(1, int(window_seconds - (now - first_attempt)))
            return retry_after

        connection.execute(
            "INSERT INTO rate_limit_attempts (request_key, created_at) VALUES (?, ?)",
            (request_key, now),
        )
    return None


def rate_limit_error_response(retry_after: int) -> Any:
    message = "Te veel verzoeken. Probeer het over enkele minuten opnieuw."
    if request.path.startswith("/api/") or request_prefers_json():
        return jsonify({"error": message}), 429, {"Retry-After": str(retry_after)}
    return message, 429, {"Retry-After": str(retry_after)}


def validate_image_signature(content_type: str, file_bytes: bytes) -> bool:
    if content_type == "image/jpeg":
        return file_bytes.startswith(b"\xff\xd8\xff")
    if content_type == "image/png":
        return file_bytes.startswith(b"\x89PNG\r\n\x1a\n")
    if content_type == "image/webp":
        return len(file_bytes) > 12 and file_bytes.startswith(b"RIFF") and file_bytes[8:12] == b"WEBP"
    if content_type == "image/avif":
        return len(file_bytes) > 12 and file_bytes[4:12] in {b"ftypavif", b"ftypavis"}
    return False


def validate_video_signature(content_type: str, file_bytes: bytes) -> bool:
    if content_type in {"video/mp4", "video/quicktime"}:
        return len(file_bytes) > 12 and file_bytes[4:8] == b"ftyp"
    if content_type == "video/webm":
        return file_bytes.startswith(b"\x1a\x45\xdf\xa3")
    return False


def apply_security_headers(response: Any) -> Any:
    for header_name, header_value in SECURITY_HEADERS.items():
        if header_name == "Content-Security-Policy" and not is_request_secure():
            header_value = header_value.replace("upgrade-insecure-requests", "").strip()
            header_value = re.sub(r";\s*;", ";", header_value).rstrip("; ")
        response.headers.setdefault(header_name, header_value)
    if is_request_secure():
        response.headers.setdefault("Strict-Transport-Security", "max-age=31536000; includeSubDomains")
    return response


def get_config() -> Dict[str, str]:
    store_id = get_env("ECWID_STORE_ID")
    secret_token = get_env("ECWID_SECRET_TOKEN")
    moneybird_token = get_env("MONEYBIRD_API_TOKEN")
    moneybird_administration_id = get_env("MONEYBIRD_ADMINISTRATION_ID")

    return {
        "store_id": "" if is_placeholder_value(store_id) else store_id,
        "secret_token": "" if is_placeholder_value(secret_token) else secret_token,
        "moneybird_token": "" if is_placeholder_value(moneybird_token) else moneybird_token,
        "moneybird_administration_id": (
            "" if is_placeholder_value(moneybird_administration_id) else moneybird_administration_id
        ),
    }


def get_external_cache_fingerprint(include_moneybird: bool = False) -> Tuple[str, ...]:
    config = get_config()
    values = [config["store_id"], config["secret_token"]]
    if include_moneybird:
        values.extend([config["moneybird_token"], config["moneybird_administration_id"]])
    return tuple(hashlib.sha256(value.encode("utf-8")).hexdigest() if value else "" for value in values)


def get_content_storage_config() -> Dict[str, Any]:
    allowed_types_raw = get_env("BUNNY_IMAGE_ALLOWED_TYPES")
    allowed_types = [
        item.strip()
        for item in allowed_types_raw.split(",")
        if item.strip()
    ]
    if not allowed_types:
        allowed_types = ["image/jpeg", "image/png", "image/webp", "image/avif"]

    max_upload_mb = max(1, get_env_int("BUNNY_IMAGE_MAX_UPLOAD_MB", 15))
    max_request_mb = max(max_upload_mb, get_env_int("CONTENT_UPLOAD_MAX_REQUEST_MB", 250))
    max_upload_files = max(1, get_env_int("CONTENT_UPLOAD_MAX_FILES", 500))

    region = get_env("BUNNY_STORAGE_REGION") or "storage"
    zone = get_env("BUNNY_STORAGE_ZONE")
    access_key = get_env("BUNNY_STORAGE_ACCESS_KEY")
    api_access_key = get_env("BUNNY_API_ACCESS_KEY")
    public_base = get_env("BUNNY_IMAGE_PUBLIC_BASE").rstrip("/")
    base_path = get_env("BUNNY_IMAGE_BASE_PATH").strip().strip("/")
    if not base_path:
        base_path = "content"

    missing_config = [
        key
        for key, value in (
            ("BUNNY_STORAGE_ZONE", zone),
            ("BUNNY_STORAGE_ACCESS_KEY", access_key),
            ("BUNNY_IMAGE_PUBLIC_BASE", public_base),
        )
        if not value
    ]

    local_upload_root = get_env("LOCAL_UPLOAD_ROOT")
    if not local_upload_root:
        local_upload_root = os.path.join(os.path.dirname(__file__), "static", "uploads")

    return {
        "region": region,
        "zone": zone,
        "access_key": access_key,
        "api_access_key": api_access_key,
        "public_base": public_base,
        "base_path": base_path,
        "max_upload_mb": max_upload_mb,
        "max_request_mb": max_request_mb,
        "max_upload_files": max_upload_files,
        "allowed_types": allowed_types,
        "missing_config": missing_config,
        "bunny_enabled": not missing_config,
        "local_upload_root": local_upload_root,
    }


def get_exercise_video_storage_config() -> Dict[str, Any]:
    content_config = get_content_storage_config()
    allowed_types_raw = get_env("BUNNY_VIDEO_ALLOWED_TYPES")
    allowed_types = [
        item.strip()
        for item in allowed_types_raw.split(",")
        if item.strip()
    ]
    if not allowed_types:
        allowed_types = ["video/mp4", "video/webm", "video/quicktime"]

    public_base = (get_env("BUNNY_VIDEO_PUBLIC_BASE") or content_config["public_base"]).rstrip("/")
    base_path = get_env("BUNNY_VIDEO_BASE_PATH").strip().strip("/")
    if not base_path:
        base_path = "exercise-videos"

    missing_config = [
        key
        for key, value in (
            ("BUNNY_STORAGE_ZONE", content_config["zone"]),
            ("BUNNY_STORAGE_ACCESS_KEY", content_config["access_key"]),
            ("BUNNY_VIDEO_PUBLIC_BASE/BUNNY_IMAGE_PUBLIC_BASE", public_base),
        )
        if not value
    ]

    return {
        "region": content_config["region"],
        "zone": content_config["zone"],
        "access_key": content_config["access_key"],
        "api_access_key": content_config["api_access_key"],
        "public_base": public_base,
        "base_path": base_path,
        "max_upload_mb": EXERCISE_VIDEO_MAX_UPLOAD_MB,
        "max_request_mb": max(EXERCISE_VIDEO_MAX_UPLOAD_MB, content_config["max_request_mb"]),
        "allowed_types": allowed_types,
        "missing_config": missing_config,
        "bunny_enabled": not missing_config,
        "local_upload_root": content_config["local_upload_root"],
    }


app.config["MAX_CONTENT_LENGTH"] = max(
    get_content_storage_config()["max_request_mb"],
    get_exercise_video_storage_config()["max_request_mb"],
) * 1024 * 1024


def is_external_agenda_api_path(path: str) -> bool:
    return path in {"/api/v1/agenda/events", "/api/v1/agenda/calendar.ics"}


def is_public_path(path: str) -> bool:
    return (
        path.startswith("/static/")
        or path in {"/login", "/manifest.webmanifest", "/service-worker.js"}
        or path.startswith("/uitnodiging/")
        or is_external_agenda_api_path(path)
    )


def get_current_user() -> Optional[Dict[str, Any]]:
    user_id = str(session.get("user_id", "")).strip()
    if user_id:
        user = get_user_by_id(user_id)
        if user is not None:
            return user

    username = str(session.get("username", "")).strip()
    if username:
        user = get_user_by_username(username)
        if user is not None:
            session["user_id"] = user["id"]
            return user

    return None


@app.context_processor
def inject_current_user():
    return {"current_user": get_current_user()}


@app.context_processor
def inject_navigation_permissions():
    user = get_current_user()
    return {
        "visible_pages": get_visible_pages_for_user(user),
        "workspace_search_pages": get_workspace_search_pages_for_user(user),
        "current_workspace_navigation_path": get_current_workspace_navigation_path(user, request.path),
        "can_view_revenue": bool(user and user.get("isAdmin")),
    }


@app.before_request
def require_login():
    if should_enforce_https() and not is_request_secure() and not should_skip_https_redirect():
        secure_url = request.url.replace("http://", "https://", 1)
        return redirect(secure_url, code=301)

    if is_external_agenda_api_path(request.path):
        rate_limit_rule = get_rate_limit_rule(request.path)
        if rate_limit_rule is not None:
            max_attempts, window_seconds, scope = rate_limit_rule
            retry_after = apply_rate_limit(max_attempts, window_seconds, scope)
            if retry_after is not None:
                return rate_limit_error_response(retry_after)
        return None

    session_response = handle_session_timeout()
    if session_response is not None:
        return session_response

    if request.method not in {"GET", "HEAD", "OPTIONS"}:
        rate_limit_rule = get_rate_limit_rule(request.path)
        if rate_limit_rule is not None:
            max_attempts, window_seconds, scope = rate_limit_rule
            retry_after = apply_rate_limit(max_attempts, window_seconds, scope)
            if retry_after is not None:
                return rate_limit_error_response(retry_after)

        csrf_response = validate_csrf_token()
        if csrf_response is not None:
            return csrf_response

    if is_public_path(request.path):
        return None

    user = get_current_user()
    if user is None:
        return redirect(url_for("login_page", next=request.path))

    session["user_id"] = user["id"]
    return None


def format_ecwid_date(value: str) -> str:
    if not value:
        return ""

    for pattern in ("%Y-%m-%d %H:%M:%S %z", "%Y-%m-%d %H:%M:%S"):
        try:
            parsed = datetime.strptime(value, pattern)
            return parsed.isoformat()
        except ValueError:
            continue

    return value


def split_full_name(full_name: str) -> Tuple[str, str]:
    normalized_name = str(full_name or "").strip()
    if not normalized_name:
        return "", ""
    name_parts = [part for part in normalized_name.split() if part]
    if not name_parts:
        return "", ""
    if len(name_parts) == 1:
        return name_parts[0], ""
    return name_parts[0], " ".join(name_parts[1:])


def normalize_order_extra_fields(order: Dict[str, Any]) -> Dict[str, str]:
    normalized_fields: Dict[str, str] = {}

    for field in order.get("orderExtraFields", []) or []:
        if not isinstance(field, dict):
            continue
        key_candidates = [field.get("title"), field.get("id")]
        value = str(field.get("value", "") or "").strip()
        if not value:
            continue
        for key_candidate in key_candidates:
            normalized_key = normalize_match_text(str(key_candidate or ""))
            if normalized_key and normalized_key not in normalized_fields:
                normalized_fields[normalized_key] = value

    for raw_key, raw_value in (order.get("extraFields") or {}).items():
        value = str(raw_value or "").strip()
        normalized_key = normalize_match_text(str(raw_key or ""))
        if normalized_key and value and normalized_key not in normalized_fields:
            normalized_fields[normalized_key] = value

    return normalized_fields


def find_order_field_value(extra_fields: Dict[str, str], *field_names: str) -> str:
    for field_name in field_names:
        normalized_name = normalize_match_text(field_name)
        if not normalized_name:
            continue
        for key, value in extra_fields.items():
            key_tokens = set(key.split())
            name_tokens = set(normalized_name.split())
            if key == normalized_name or (name_tokens and name_tokens.issubset(key_tokens)):
                return value
    return ""


def extract_registration_details(order: Dict[str, Any], customer_name: str = "") -> Dict[str, str]:
    existing_details = order.get("registrationDetails")
    if isinstance(existing_details, dict) and existing_details:
        return {
            "firstName": str(existing_details.get("firstName", "") or "").strip(),
            "lastName": str(existing_details.get("lastName", "") or "").strip(),
            "birthDate": str(existing_details.get("birthDate", "") or "").strip(),
            "gender": str(existing_details.get("gender", "") or "").strip(),
            "address": str(existing_details.get("address", "") or "").strip(),
            "postalCode": str(existing_details.get("postalCode", "") or "").strip(),
            "city": str(existing_details.get("city", "") or "").strip(),
            "clubTeam": str(existing_details.get("clubTeam", "") or "").strip(),
            "phone": str(existing_details.get("phone", "") or "").strip(),
            "dietaryWishes": str(existing_details.get("dietaryWishes", "") or "").strip(),
            "comments": str(existing_details.get("comments", "") or "").strip(),
        }

    resolved_customer_name = (
        str(customer_name or "").strip()
        or str(order.get("customerName", "") or "").strip()
        or str(order.get("shippingPerson", {}).get("name", "") or "").strip()
        or str(order.get("billingPerson", {}).get("name", "") or "").strip()
    )
    extra_fields = normalize_order_extra_fields(order)
    fallback_first_name, fallback_last_name = split_full_name(resolved_customer_name)
    shipping_person = order.get("shippingPerson", {}) if isinstance(order.get("shippingPerson"), dict) else {}
    billing_person = order.get("billingPerson", {}) if isinstance(order.get("billingPerson"), dict) else {}

    return {
        "firstName": find_order_field_value(extra_fields, "voornaam", "first name", "firstname") or fallback_first_name,
        "lastName": find_order_field_value(extra_fields, "achternaam", "last name", "lastname") or fallback_last_name,
        "birthDate": find_order_field_value(
            extra_fields,
            "geboortedatum",
            "geboorte datum",
            "birth date",
            "date of birth",
            "birthday",
            "dob",
        ),
        "gender": find_order_field_value(extra_fields, "geslacht", "gender"),
        "address": (
            find_order_field_value(extra_fields, "adres", "straat", "street", "address")
            or str(shipping_person.get("street", "") or "").strip()
            or str(billing_person.get("street", "") or "").strip()
        ),
        "postalCode": (
            find_order_field_value(extra_fields, "postcode", "postal code", "postalcode", "zip")
            or str(shipping_person.get("postalCode", "") or "").strip()
            or str(billing_person.get("postalCode", "") or "").strip()
        ),
        "city": (
            find_order_field_value(extra_fields, "plaats", "woonplaats", "stad", "city")
            or str(shipping_person.get("city", "") or "").strip()
            or str(billing_person.get("city", "") or "").strip()
        ),
        "clubTeam": find_order_field_value(extra_fields, "club/team", "club team", "club", "team"),
        "phone": (
            find_order_field_value(
                extra_fields,
                "06-nummer",
                "06 nummer",
                "telefoonnummer",
                "telefoon",
                "mobiel",
                "phone",
                "mobile",
            )
            or str(shipping_person.get("phone", "") or "").strip()
            or str(billing_person.get("phone", "") or "").strip()
        ),
        "dietaryWishes": find_order_field_value(extra_fields, "dieetwensen", "dieet wensen", "allergieen", "allergieën", "dietary wishes"),
        "comments": find_order_field_value(extra_fields, "opmerkingen", "opmerking", "comments", "commentaar"),
    }


def normalize_order(order: Dict[str, Any]) -> Dict[str, Any]:
    customer_name = (
        order.get("shippingPerson", {}).get("name")
        or order.get("billingPerson", {}).get("name")
        or "Onbekende klant"
    )
    products = order.get("items", [])
    registration_details = extract_registration_details(order, customer_name)

    return {
        "id": order.get("id", ""),
        "createdAt": format_ecwid_date(order.get("createDate", "")),
        "orderNumber": order.get("orderNumber") or order.get("id", ""),
        "status": order.get("status", "UNKNOWN"),
        "paymentStatus": order.get("paymentStatus", "UNKNOWN"),
        "fulfillmentStatus": order.get("fulfillmentStatus", "UNKNOWN"),
        "total": order.get("total", 0),
        "email": order.get("email", ""),
        "customerName": customer_name,
        "paymentMethod": order.get("paymentMethod", "Onbekend"),
        "shippingMethod": order.get("shippingOption", "Niet opgegeven"),
        "itemCount": sum(item.get("quantity", 0) for item in products),
        "isRefunded": order.get("paymentStatus") == "REFUNDED",
        "registrationDetails": registration_details,
        "items": [
            {
                "productId": item.get("productId"),
                "name": item.get("name", "Naamloos product"),
                "quantity": item.get("quantity", 0),
                "price": item.get("price", 0),
                "sku": item.get("sku", ""),
            }
            for item in products
        ],
    }


def sort_orders_desc(orders: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    def sort_key(order: Dict[str, Any]) -> datetime:
        created_at = order.get("createdAt", "")
        if not created_at:
            return datetime.min
        try:
            return datetime.fromisoformat(created_at)
        except ValueError:
            return datetime.min

    return sorted(orders, key=sort_key, reverse=True)


def decorate_orders_for_list(orders: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    decorated_orders = []
    for order in orders:
        created_at = order.get("createdAt", "")
        display_date = "-"
        display_time = "-"

        if created_at:
            try:
                parsed = datetime.fromisoformat(created_at)
                display_date = parsed.strftime("%d-%m-%Y")
                display_time = parsed.strftime("%H:%M")
            except ValueError:
                pass

        item_names = ", ".join(item.get("name", "Naamloos product") for item in order.get("items", []))
        decorated_order = dict(order)
        decorated_order["displayDate"] = display_date
        decorated_order["displayTime"] = display_time
        decorated_order["itemNames"] = item_names or "-"
        decorated_orders.append(decorated_order)

    return decorated_orders


def normalize_product(product: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "id": None if product.get("id") is None else str(product.get("id")),
        "name": str(product.get("name", "") or "Naamloos product"),
        "sku": str(product.get("sku", "") or ""),
        "price": float(product.get("price") or 0),
        "enabled": bool(product.get("enabled", True)),
    }


def build_mock_catalog_products(orders: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    products_by_key: Dict[str, Dict[str, Any]] = {}
    for order in orders:
        for item in order.get("items", []):
            product_key = build_order_item_product_key(item)
            if product_key in products_by_key:
                continue
            products_by_key[product_key] = {
                "id": None if item.get("productId") is None else str(item.get("productId")),
                "name": str(item.get("name", "") or "Naamloos product"),
                "sku": str(item.get("sku", "") or ""),
                "price": float(item.get("price") or 0),
                "enabled": True,
            }
    return sorted(products_by_key.values(), key=lambda product: product["name"].lower())


def fetch_catalog_products_payload() -> Dict[str, Any]:
    config = get_config()
    if not config["store_id"] or not config["secret_token"]:
        mock_products = build_mock_catalog_products(mock_orders())
        return {
            "source": "mock",
            "items": mock_products,
            "message": (
                "Live Ecwid-koppeling staat nog niet aan. "
                "Voeg ECWID_STORE_ID en ECWID_SECRET_TOKEN toe."
            ),
        }

    all_products: List[Dict[str, Any]] = []
    offset = 0
    limit = 100
    total = 0

    try:
        while True:
            response = requests.get(
                f"{ECWID_API_BASE}/{config['store_id']}/products",
                headers={"Authorization": f"Bearer {config['secret_token']}"},
                params={
                    "offset": offset,
                    "limit": limit,
                    "responseFields": "total,count,items(id,name,sku,price,enabled)",
                },
                timeout=20,
            )
            response.raise_for_status()
            payload = response.json()
            items = payload.get("items", [])
            total = payload.get("total", total)
            all_products.extend(normalize_product(item) for item in items)

            if not items or len(items) < limit or len(all_products) >= total:
                break

            offset += limit
    except requests.RequestException:
        mock_products = build_mock_catalog_products(mock_orders())
        return {
            "source": "mock",
            "items": mock_products,
            "message": (
                "Ecwid-producten konden nu niet worden geladen. "
                "Tijdelijke voorbeelddata wordt getoond."
            ),
        }

    return {
        "source": "ecwid",
        "items": all_products,
        "total": total,
        "count": len(all_products),
    }


def fetch_catalog_products() -> Dict[str, Any]:
    now = time.time()
    config_fingerprint = get_external_cache_fingerprint()
    with catalog_products_cache_lock:
        cached_payload = catalog_products_cache.get("payload")
        cached_at = float(catalog_products_cache.get("cached_at") or 0.0)
        cached_fingerprint = catalog_products_cache.get("config_fingerprint")

    if cached_payload is not None and cached_fingerprint == config_fingerprint and now - cached_at < CACHE_TTL_SECONDS:
        payload = copy.deepcopy(cached_payload)
        payload["cachedAt"] = cached_at
        return payload

    payload = fetch_catalog_products_payload()
    payload["items"] = sorted(payload.get("items", []), key=lambda product: str(product.get("name", "")).lower())
    with catalog_products_cache_lock:
        catalog_products_cache["payload"] = copy.deepcopy(payload)
        catalog_products_cache["cached_at"] = now
        catalog_products_cache["config_fingerprint"] = config_fingerprint

    payload["cachedAt"] = now
    return payload


def build_order_item_product_key(item: Dict[str, Any]) -> str:
    product_id = item.get("productId")
    if product_id is not None and str(product_id).strip():
        return f"id:{str(product_id).strip()}"

    name = str(item.get("name", "") or "Naamloos product").strip().lower()
    sku = str(item.get("sku", "") or "").strip().lower()
    return f"fallback:{name}|{sku}"


def build_catalog_product_key(product: Dict[str, Any]) -> str:
    product_id = product.get("id")
    if product_id is not None and str(product_id).strip():
        return f"id:{str(product_id).strip()}"

    name = str(product.get("name", "") or "Naamloos product").strip().lower()
    sku = str(product.get("sku", "") or "").strip().lower()
    return f"fallback:{name}|{sku}"


def build_registrations_page_url() -> str:
    return url_for("registrations_page")


def build_leads_page_url() -> str:
    return url_for("leads_page")


def build_registration_detail_url(product_key: str) -> str:
    return url_for("registrations_detail_page", product_key=product_key.strip())


def build_registrations_overview_entries(products: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    completed_product_keys = load_completed_registration_event_keys()
    canceled_product_keys = load_canceled_registration_event_keys()
    entries = []

    for product in products:
        normalized_product = normalize_product(product)
        product_key = build_catalog_product_key(normalized_product)
        event_completed = product_key in completed_product_keys
        event_canceled = product_key in canceled_product_keys
        event_status_label = "Event geannuleerd" if event_canceled else ("Event afgerond" if event_completed else "Event open")
        event_status_sort = 2 if event_canceled else (1 if event_completed else 0)
        search_parts = [
            normalized_product["name"],
            normalized_product["sku"],
            normalized_product["id"],
        ]
        entries.append(
            {
                "productKey": product_key,
                "productId": normalized_product["id"],
                "name": normalized_product["name"],
                "sku": normalized_product["sku"],
                "enabled": normalized_product["enabled"],
                "searchText": " ".join(
                    str(part).strip().lower()
                    for part in search_parts
                    if str(part or "").strip()
                ),
                "detailUrl": build_registration_detail_url(product_key),
                "eventCompleted": event_completed,
                "eventCanceled": event_canceled,
                "eventStatusLabel": event_status_label,
                "eventStatusSort": event_status_sort,
            }
        )

    entries.sort(
        key=lambda item: (
            item["eventStatusSort"],
            item["name"].lower(),
            item["sku"].lower(),
            item["productKey"],
        )
    )
    return entries


def build_product_registration_summary(
    products: List[Dict[str, Any]],
    orders: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    entries_by_key: Dict[str, Dict[str, Any]] = {}

    for product in products:
        product_key = build_catalog_product_key(product)
        entries_by_key[product_key] = {
            "productKey": product_key,
            "productId": None if product.get("id") is None else str(product.get("id")),
            "name": str(product.get("name", "") or "Naamloos product"),
            "sku": str(product.get("sku", "") or ""),
            "price": float(product.get("price") or 0),
            "enabled": bool(product.get("enabled", True)),
            "orderCount": 0,
            "participantCount": 0,
            "emailCount": 0,
            "emails": [],
            "orders": [],
        }

    for order in sort_orders_desc(orders):
        created_at = parse_iso_datetime(order.get("createdAt", ""))
        display_date = created_at.strftime("%d-%m-%Y") if created_at else "-"
        display_time = created_at.strftime("%H:%M") if created_at else "-"
        for item in order.get("items", []):
            product_key = build_order_item_product_key(item)
            entry = entries_by_key.get(product_key)
            if entry is None:
                entry = {
                    "productKey": product_key,
                    "productId": None if item.get("productId") is None else str(item.get("productId")),
                    "name": str(item.get("name", "") or "Naamloos product"),
                    "sku": str(item.get("sku", "") or ""),
                    "price": float(item.get("price") or 0),
                    "enabled": True,
                    "orderCount": 0,
                    "participantCount": 0,
                    "emailCount": 0,
                    "emails": [],
                    "orders": [],
                }
                entries_by_key[product_key] = entry

            email = str(order.get("email", "") or "").strip()
            if email and email not in entry["emails"]:
                entry["emails"].append(email)

            entry["orderCount"] += 1
            entry["participantCount"] += max(int(item.get("quantity") or 0), 0)
            entry["orders"].append(
                {
                    "id": str(order.get("id", "") or ""),
                    "orderNumber": str(order.get("orderNumber", "") or order.get("id", "")),
                    "customerName": str(order.get("customerName", "") or "Onbekende klant"),
                    "email": email,
                    "status": str(order.get("status", "") or "-"),
                    "paymentStatus": str(order.get("paymentStatus", "") or "-"),
                    "displayDate": display_date,
                    "displayTime": display_time,
                    "quantity": max(int(item.get("quantity") or 0), 0),
                    "total": float(order.get("total") or 0),
                    "itemPrice": float(item.get("price") or 0),
                    "registrationDetails": extract_registration_details(order, str(order.get("customerName", "") or "")),
                }
            )

    entries = []
    for entry in entries_by_key.values():
        next_entry = dict(entry)
        next_entry["emailCount"] = len(next_entry["emails"])
        next_entry["emailList"] = ", ".join(next_entry["emails"])
        search_parts = [next_entry["name"], next_entry["sku"], next_entry["productId"]]
        for order in next_entry["orders"]:
            search_parts.extend(
                [
                    order.get("customerName", ""),
                    order.get("email", ""),
                    order.get("orderNumber", ""),
                ]
            )
        next_entry["searchText"] = " ".join(
            str(part).strip().lower()
            for part in search_parts
            if str(part or "").strip()
        )
        entries.append(next_entry)

    entries.sort(
        key=lambda item: (
            item["name"].lower(),
            item["sku"].lower(),
            item["productKey"],
        )
    )
    return entries


def build_registration_product_detail(
    products: List[Dict[str, Any]],
    orders: List[Dict[str, Any]],
    selected_product_key: str,
) -> Optional[Dict[str, Any]]:
    normalized_product_key = selected_product_key.strip()
    if not normalized_product_key:
        return None

    selected_product = next(
        (normalize_product(product) for product in products if build_catalog_product_key(product) == normalized_product_key),
        None,
    )

    detail_entry = None
    if selected_product is not None:
        detail_entry = {
            "productKey": normalized_product_key,
            "productId": selected_product["id"],
            "name": selected_product["name"],
            "sku": selected_product["sku"],
            "price": selected_product["price"],
            "enabled": selected_product["enabled"],
            "orderCount": 0,
            "participantCount": 0,
            "emailCount": 0,
            "emails": [],
            "orders": [],
        }

    known_order_ids: Set[str] = set()

    for order in sort_orders_desc(orders):
        created_at = parse_iso_datetime(order.get("createdAt", ""))
        display_date = created_at.strftime("%d-%m-%Y") if created_at else "-"
        display_time = created_at.strftime("%H:%M") if created_at else "-"
        for item in order.get("items", []):
            if build_order_item_product_key(item) != normalized_product_key:
                continue

            if detail_entry is None:
                detail_entry = {
                    "productKey": normalized_product_key,
                    "productId": None if item.get("productId") is None else str(item.get("productId")),
                    "name": str(item.get("name", "") or "Naamloos product"),
                    "sku": str(item.get("sku", "") or ""),
                    "price": float(item.get("price") or 0),
                    "enabled": True,
                    "orderCount": 0,
                    "participantCount": 0,
                    "emailCount": 0,
                    "emails": [],
                    "orders": [],
                }

            email = str(order.get("email", "") or "").strip()
            if email and email not in detail_entry["emails"]:
                detail_entry["emails"].append(email)

            order_id = str(order.get("id", "") or "")
            known_order_ids.add(order_id)
            item_quantity = max(int(item.get("quantity") or 0), 0)
            detail_entry["orderCount"] += 1
            detail_entry["participantCount"] += item_quantity
            registration_details = extract_registration_details(order, str(order.get("customerName", "") or ""))
            participant_rows = max(item_quantity, 1)
            for participant_index in range(participant_rows):
                detail_entry["orders"].append(
                    {
                        "id": order_id,
                        "rowId": f"{order_id}:{participant_index + 1}",
                        "orderNumber": str(order.get("orderNumber", "") or order.get("id", "")),
                        "customerName": str(order.get("customerName", "") or "Onbekende klant"),
                        "email": email,
                        "status": str(order.get("status", "") or "-"),
                        "paymentStatus": str(order.get("paymentStatus", "") or "-"),
                        "displayDate": display_date,
                        "displayTime": display_time,
                        "quantity": 1,
                        "originalQuantity": item_quantity,
                        "participantIndex": participant_index + 1,
                        "participantLabel": (
                            f"Aanmelding {participant_index + 1} van {item_quantity}"
                            if item_quantity > 1
                            else ""
                        ),
                        "total": float(order.get("total") or 0),
                        "itemPrice": float(item.get("price") or 0),
                        "registrationDetails": registration_details,
                    }
                )

    if detail_entry is None:
        return None

    email_settings = load_registration_event_email_settings(normalized_product_key)
    detail_entry["eventDate"] = email_settings.get("eventDate", "")
    detail_entry["eventDate2"] = email_settings.get("eventDate2", "")
    detail_entry["emailSubject"] = email_settings.get("emailSubject", "")
    detail_entry["emailBody"] = email_settings.get("emailBody", "")
    detail_entry["emailSettingsUpdatedAt"] = email_settings.get("updatedAt", "")
    detail_entry["eventCompleted"] = is_registration_event_completed(normalized_product_key)
    detail_entry["eventCanceled"] = is_registration_event_canceled(normalized_product_key)
    detail_entry["eventStatusLabel"] = (
        "Event geannuleerd"
        if detail_entry["eventCanceled"]
        else ("Event afgerond" if detail_entry["eventCompleted"] else "Event open")
    )
    detail_entry["eventCompletedLabel"] = detail_entry["eventStatusLabel"]
    emailed_order_ids = load_registration_emailed_order_ids(normalized_product_key, known_order_ids)
    pending_emails: List[str] = []
    pending_email_keys: Set[str] = set()
    emailed_order_count = len(known_order_ids.intersection(emailed_order_ids))

    for order in detail_entry["orders"]:
        order["emailed"] = order["id"] in emailed_order_ids
        if order["emailed"]:
            continue

        email = str(order.get("email", "") or "").strip()
        normalized_email = email.lower()
        if email and normalized_email not in pending_email_keys:
            pending_email_keys.add(normalized_email)
            pending_emails.append(email)

    detail_entry["emailCount"] = len(detail_entry["emails"])
    detail_entry["emailList"] = ", ".join(detail_entry["emails"])
    detail_entry["pendingEmailCount"] = len(pending_emails)
    detail_entry["pendingEmailList"] = ", ".join(pending_emails)
    detail_entry["emailedOrderCount"] = emailed_order_count
    detail_entry["pendingOrderCount"] = len(detail_entry["orders"]) - emailed_order_count
    return detail_entry


def normalize_registration_email_status_order_ids(order_ids: Any) -> List[str]:
    if not isinstance(order_ids, list):
        return []

    normalized_ids: List[str] = []
    seen_ids: Set[str] = set()
    for raw_order_id in order_ids:
        order_id = str(raw_order_id or "").strip()
        if not order_id or order_id in seen_ids:
            continue
        seen_ids.add(order_id)
        normalized_ids.append(order_id)
    return normalized_ids


def load_registration_emailed_order_ids(
    product_key: str,
    order_ids: Optional[Set[str]] = None,
) -> Set[str]:
    normalized_product_key = str(product_key or "").strip()
    if not normalized_product_key:
        return set()

    with get_db_connection() as connection:
        if order_ids:
            placeholders = ", ".join("?" for _ in order_ids)
            rows = connection.execute(
                f"""
                SELECT order_id
                FROM registration_email_statuses
                WHERE product_key = ?
                  AND order_id IN ({placeholders})
                """,
                (normalized_product_key, *sorted(order_ids)),
            ).fetchall()
        else:
            rows = connection.execute(
                """
                SELECT order_id
                FROM registration_email_statuses
                WHERE product_key = ?
                """,
                (normalized_product_key,),
            ).fetchall()

    return {str(row["order_id"] or "").strip() for row in rows if str(row["order_id"] or "").strip()}


def load_all_registration_emailed_order_ids() -> List[str]:
    with get_db_connection() as connection:
        rows = connection.execute(
            """
            SELECT DISTINCT order_id
            FROM registration_email_statuses
            WHERE trim(order_id) != ''
            ORDER BY order_id
            """
        ).fetchall()

    return [str(row["order_id"] or "").strip() for row in rows if str(row["order_id"] or "").strip()]


def load_registration_reminder_sent_order_ids(
    product_key: str,
    order_ids: Optional[Set[str]] = None,
) -> Set[str]:
    normalized_product_key = str(product_key or "").strip()
    if not normalized_product_key:
        return set()

    with get_db_connection() as connection:
        if order_ids:
            placeholders = ",".join("?" for _ in order_ids)
            rows = connection.execute(
                f"""
                SELECT order_id
                FROM registration_email_reminder_statuses
                WHERE product_key = ? AND order_id IN ({placeholders})
                """,
                (normalized_product_key, *sorted(order_ids)),
            ).fetchall()
        else:
            rows = connection.execute(
                """
                SELECT order_id
                FROM registration_email_reminder_statuses
                WHERE product_key = ?
                """,
                (normalized_product_key,),
            ).fetchall()

    return {str(row["order_id"] or "").strip() for row in rows if str(row["order_id"] or "").strip()}


def set_registration_reminders_sent(product_key: str, order_ids: List[str], sent: bool) -> List[str]:
    normalized_product_key = str(product_key or "").strip()
    normalized_order_ids = normalize_registration_email_status_order_ids(order_ids)
    if not normalized_product_key or not normalized_order_ids:
        return []

    with get_db_connection() as connection:
        if sent:
            timestamp = utcnow_iso()
            connection.executemany(
                """
                INSERT INTO registration_email_reminder_statuses (product_key, order_id, reminded_at)
                VALUES (?, ?, ?)
                ON CONFLICT(product_key, order_id) DO UPDATE SET reminded_at = excluded.reminded_at
                """,
                [(normalized_product_key, order_id, timestamp) for order_id in normalized_order_ids],
            )
        else:
            connection.executemany(
                """
                DELETE FROM registration_email_reminder_statuses
                WHERE product_key = ? AND order_id = ?
                """,
                [(normalized_product_key, order_id) for order_id in normalized_order_ids],
            )

    return normalized_order_ids


def set_registration_orders_emailed(product_key: str, order_ids: List[str], emailed: bool) -> List[str]:
    normalized_product_key = str(product_key or "").strip()
    normalized_order_ids = normalize_registration_email_status_order_ids(order_ids)
    if not normalized_product_key or not normalized_order_ids:
        return []

    with get_db_connection() as connection:
        if emailed:
            timestamp = utcnow_iso()
            connection.executemany(
                """
                INSERT INTO registration_email_statuses (product_key, order_id, emailed_at)
                VALUES (?, ?, ?)
                ON CONFLICT(product_key, order_id) DO UPDATE SET emailed_at = excluded.emailed_at
                """,
                [(normalized_product_key, order_id, timestamp) for order_id in normalized_order_ids],
            )
        else:
            connection.executemany(
                """
                DELETE FROM registration_email_statuses
                WHERE product_key = ?
                  AND order_id = ?
                """,
                [(normalized_product_key, order_id) for order_id in normalized_order_ids],
            )

    return normalized_order_ids


def normalize_registration_event_date(value: Any) -> str:
    normalized_value = str(value or "").strip()[:10]
    if not normalized_value:
        return ""
    parsed_date = parse_iso_date(normalized_value)
    return parsed_date.isoformat() if parsed_date is not None else ""


def load_registration_event_email_settings(product_key: str) -> Dict[str, str]:
    normalized_product_key = str(product_key or "").strip()
    if not normalized_product_key:
        return {
            "productKey": "",
            "productName": "",
            "eventDate": "",
            "eventDate2": "",
            "emailSubject": "",
            "emailBody": "",
            "updatedAt": "",
        }

    with get_db_connection() as connection:
        row = connection.execute(
            """
            SELECT product_key, product_name, event_date, event_date_2, email_subject, email_body, updated_at
            FROM registration_event_email_settings
            WHERE product_key = ?
            """,
            (normalized_product_key,),
        ).fetchone()

    if row is None:
        return {
            "productKey": normalized_product_key,
            "productName": "",
            "eventDate": "",
            "eventDate2": "",
            "emailSubject": "",
            "emailBody": "",
            "updatedAt": "",
        }

    return {
        "productKey": str(row["product_key"] or "").strip(),
        "productName": str(row["product_name"] or "").strip(),
        "eventDate": str(row["event_date"] or "").strip(),
        "eventDate2": str(row["event_date_2"] or "").strip(),
        "emailSubject": str(row["email_subject"] or "").strip(),
        "emailBody": str(row["email_body"] or "").strip(),
        "updatedAt": str(row["updated_at"] or "").strip(),
    }


def load_registration_event_email_templates(exclude_product_key: str = "") -> List[Dict[str, str]]:
    normalized_excluded_key = str(exclude_product_key or "").strip()
    with get_db_connection() as connection:
        rows = connection.execute(
            """
            SELECT product_key, product_name, event_date, event_date_2, email_subject, email_body, updated_at
            FROM registration_event_email_settings
            WHERE trim(email_body) != ''
            ORDER BY updated_at DESC, product_name COLLATE NOCASE
            """
        ).fetchall()

    templates: List[Dict[str, str]] = []
    for row in rows:
        product_key = str(row["product_key"] or "").strip()
        if product_key and product_key == normalized_excluded_key:
            continue
        product_name = str(row["product_name"] or "").strip() or product_key
        event_date = str(row["event_date"] or "").strip()
        event_date_2 = str(row["event_date_2"] or "").strip()
        date_label = format_registration_event_dates_label(event_date, event_date_2)
        templates.append(
            {
                "productKey": product_key,
                "productName": product_name,
                "eventDate": event_date,
                "eventDate2": event_date_2,
                "emailSubject": str(row["email_subject"] or "").strip(),
                "emailBody": str(row["email_body"] or "").strip(),
                "updatedAt": str(row["updated_at"] or "").strip(),
                "label": f"{product_name} ({date_label})" if date_label else product_name,
            }
        )
    return templates


def save_registration_event_email_settings(
    product_key: str,
    product_name: str,
    event_date: Any,
    event_date_2: Any,
    email_subject: Any,
    email_body: Any,
) -> Dict[str, str]:
    normalized_product_key = str(product_key or "").strip()
    if not normalized_product_key:
        raise ValueError("Product ontbreekt.")

    normalized_product_name = str(product_name or "").strip()[:300]
    normalized_event_date = normalize_registration_event_date(event_date)
    raw_event_date = str(event_date or "").strip()
    if raw_event_date and not normalized_event_date:
        raise ValueError("Vul een geldige eventdatum in.")
    normalized_event_date_2 = normalize_registration_event_date(event_date_2)
    raw_event_date_2 = str(event_date_2 or "").strip()
    if raw_event_date_2 and not normalized_event_date_2:
        raise ValueError("Vul een geldige tweede eventdatum in.")

    normalized_subject = str(email_subject or "").strip()[:300]
    normalized_body = str(email_body or "").strip()
    if len(normalized_body) > 20000:
        raise ValueError("De mailtekst is te lang.")

    updated_at = utcnow_iso()
    with get_db_connection() as connection:
        connection.execute(
            """
            INSERT INTO registration_event_email_settings
                (product_key, product_name, event_date, event_date_2, email_subject, email_body, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(product_key) DO UPDATE SET
                product_name = excluded.product_name,
                event_date = excluded.event_date,
                event_date_2 = excluded.event_date_2,
                email_subject = excluded.email_subject,
                email_body = excluded.email_body,
                updated_at = excluded.updated_at
            """,
            (
                normalized_product_key,
                normalized_product_name,
                normalized_event_date,
                normalized_event_date_2,
                normalized_subject,
                normalized_body,
                updated_at,
            ),
        )

    return load_registration_event_email_settings(normalized_product_key)


def registration_auto_email_is_configured() -> bool:
    return bool(
        get_env_bool("REGISTRATION_AUTO_EMAILS_ENABLED", False)
        and settings.EMAIL_HOST
        and settings.EMAIL_HOST_USER
        and settings.EMAIL_HOST_PASSWORD
        and settings.DEFAULT_FROM_EMAIL
    )


def registration_order_is_paid(order: Dict[str, Any]) -> bool:
    payment_status = str(order.get("paymentStatus", "") or "").strip().upper()
    order_status = str(order.get("status", "") or "").strip().upper()
    paid_statuses = {"PAID", "ACCEPTED", "COMPLETE", "COMPLETED"}
    return payment_status in paid_statuses or order_status in paid_statuses


def registration_order_is_after_auto_email_start(order: Dict[str, Any]) -> bool:
    start_date_value = get_env("REGISTRATION_AUTO_EMAILS_START_DATE")
    if not start_date_value:
        start_date = date.today()
    else:
        start_date = parse_iso_date(start_date_value[:10])
        if start_date is None:
            return False

    created_at = parse_iso_datetime(str(order.get("createdAt", "") or ""))
    if created_at is None:
        return False
    return created_at.date() >= start_date


def get_registration_confirmation_subject(product_name: str) -> str:
    configured_subject = get_env("REGISTRATION_EMAIL_SUBJECT")
    if configured_subject:
        return configured_subject
    clean_product_name = str(product_name or "").strip()
    if clean_product_name:
        return f"Bevestiging inschrijving {clean_product_name}"
    return "Bevestiging inschrijving HWS Voetbalschool"


def format_display_date(value: Any) -> str:
    parsed_date = parse_iso_date(str(value or "").strip()[:10])
    if parsed_date is None:
        return ""
    month_name = DUTCH_MONTH_NAMES[parsed_date.month - 1].lower()
    weekday_name = DUTCH_WEEKDAY_NAMES[parsed_date.weekday()].lower()
    return f"{weekday_name} {parsed_date.day} {month_name} {parsed_date.year}"


def format_registration_event_dates_label(event_date: Any, event_date_2: Any = "") -> str:
    first_label = format_display_date(event_date) or str(event_date or "").strip()
    second_label = format_display_date(event_date_2) or str(event_date_2 or "").strip()
    if first_label and second_label:
        return f"{first_label} en {second_label}"
    return first_label or second_label


def render_registration_email_template(template: str, order: Dict[str, Any], item: Dict[str, Any], settings_row: Dict[str, str]) -> str:
    customer_name = str(order.get("customerName", "") or "").strip() or "ouder/verzorger"
    product_name = str(item.get("name", "") or "").strip() or "de activiteit"
    order_number = str(order.get("orderNumber", "") or order.get("id", "") or "").strip()
    details = extract_registration_details(order, customer_name)
    participant_name = normalize_registration_person_name(
        str(details.get("firstName", "") or ""),
        str(details.get("lastName", "") or ""),
    )
    event_date = str(settings_row.get("eventDate", "") or "").strip()
    event_date_2 = str(settings_row.get("eventDate2", "") or "").strip()
    replacements = {
        "klant_naam": customer_name,
        "deelnemer_naam": participant_name or customer_name,
        "voornaam": str(details.get("firstName", "") or "").strip() or customer_name,
        "achternaam": str(details.get("lastName", "") or "").strip(),
        "product_naam": product_name,
        "event_datum": format_display_date(event_date) or event_date,
        "eventdatum": format_display_date(event_date) or event_date,
        "event_datum_2": format_display_date(event_date_2) or event_date_2,
        "eventdatum_2": format_display_date(event_date_2) or event_date_2,
        "event_datums": format_registration_event_dates_label(event_date, event_date_2),
        "ordernummer": order_number,
    }
    rendered = str(template or "")
    for key, value in replacements.items():
        rendered = rendered.replace("{" + key + "}", value)
    return rendered


def build_registration_confirmation_body(order: Dict[str, Any], item: Dict[str, Any]) -> str:
    customer_name = str(order.get("customerName", "") or "").strip() or "ouder/verzorger"
    product_name = str(item.get("name", "") or "").strip() or "de activiteit"
    order_number = str(order.get("orderNumber", "") or order.get("id", "") or "").strip()
    details = extract_registration_details(order, customer_name)
    participant_name = normalize_registration_person_name(
        str(details.get("firstName", "") or ""),
        str(details.get("lastName", "") or ""),
    )

    intro_name = participant_name or customer_name
    lines = [
        f"Beste {customer_name},",
        "",
        f"Bedankt voor je inschrijving voor {product_name}.",
    ]
    if participant_name:
        lines.append(f"We hebben de aanmelding van {intro_name} goed ontvangen.")
    else:
        lines.append("We hebben je aanmelding goed ontvangen.")
    if order_number:
        lines.append(f"Ordernummer: {order_number}.")

    extra_text = get_env("REGISTRATION_EMAIL_EXTRA_TEXT")
    if extra_text:
        lines.extend(["", extra_text])

    lines.extend(
        [
            "",
            "Je ontvangt later eventuele praktische informatie over de activiteit.",
        ]
    )
    return "\n".join(lines)


def build_registration_event_email_body(
    order: Dict[str, Any],
    item: Dict[str, Any],
    settings_row: Dict[str, str],
) -> str:
    configured_body = str(settings_row.get("emailBody", "") or "").strip()
    if configured_body:
        return render_registration_email_template(configured_body, order, item, settings_row)
    return build_registration_confirmation_body(order, item)


def render_registration_email_inline_html(text: str) -> str:
    escaped_text = html.escape(str(text or ""))
    escaped_text = re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", escaped_text)
    escaped_text = re.sub(r"(?<!\*)\*(?!\*)(.+?)(?<!\*)\*(?!\*)", r"<em>\1</em>", escaped_text)
    return escaped_text


def render_registration_email_body_html(body: str) -> str:
    lines = str(body or "").replace("\r\n", "\n").replace("\r", "\n").split("\n")
    html_parts: List[str] = []
    paragraph_lines: List[str] = []
    list_items: List[str] = []

    def flush_paragraph() -> None:
        if paragraph_lines:
            html_parts.append(f"<p>{'<br>'.join(paragraph_lines)}</p>")
            paragraph_lines.clear()

    def flush_list() -> None:
        if list_items:
            html_parts.append(f"<ul>{''.join(list_items)}</ul>")
            list_items.clear()

    for raw_line in lines:
        stripped_line = raw_line.strip()
        bullet_match = re.match(r"^(?:[-*]|•|&bull;)\s+(.+)$", stripped_line)
        if bullet_match:
            flush_paragraph()
            list_items.append(f"<li>{render_registration_email_inline_html(bullet_match.group(1))}</li>")
            continue
        if not stripped_line:
            flush_paragraph()
            flush_list()
            continue
        flush_list()
        paragraph_lines.append(render_registration_email_inline_html(raw_line))

    flush_paragraph()
    flush_list()
    return "\n".join(html_parts) or "<p></p>"


def render_registration_email_signature_html() -> str:
    logo_url = "https://www.workspace.hwsvoetbalschool.nl/static/assets/hws-logo.png"
    return f"""
<table role="presentation" cellpadding="0" cellspacing="0" style="margin-top:22px;border-collapse:collapse;font-family:Arial,Helvetica,sans-serif;color:#1f2933;">
  <tr>
    <td colspan="2" style="padding:0 0 14px 0;font-size:14px;line-height:20px;color:#1f2933;">Met vriendelijke groet,</td>
  </tr>
  <tr>
    <td style="padding:0 16px 0 0;vertical-align:top;">
      <img src="{logo_url}" alt="HWS Voetbalschool" width="92" style="display:block;width:92px;max-width:92px;height:auto;border:0;">
    </td>
    <td style="padding:0 0 0 16px;border-left:2px solid #f2c230;vertical-align:top;">
      <div style="font-size:16px;line-height:20px;font-weight:700;color:#101820;">David van Walstijn</div>
      <div style="font-size:13px;line-height:18px;font-weight:700;color:#f2c230;text-transform:uppercase;letter-spacing:0.4px;">HWS Voetbalschool</div>
      <div style="padding-top:10px;font-size:13px;line-height:20px;color:#344054;">
        <a href="tel:+31624845896" style="color:#344054;text-decoration:none;">06-24845896</a><br>
        <a href="mailto:info@hwsvoetbalschool.nl" style="color:#344054;text-decoration:none;">info@hwsvoetbalschool.nl</a>
      </div>
    </td>
  </tr>
</table>""".strip()


def render_registration_email_html(body: str) -> str:
    return f"{render_registration_email_body_html(body)}\n{render_registration_email_signature_html()}"


def parse_email_list(value: str) -> List[str]:
    emails: List[str] = []
    seen: Set[str] = set()
    for email in re.split(r"[\s,;]+", str(value or "")):
        normalized_email = email.strip()
        email_key = normalized_email.lower()
        if not normalized_email or email_key in seen:
            continue
        seen.add(email_key)
        emails.append(normalized_email)
    return emails


def append_unique_email(emails: List[str], email: str, excluded: Optional[Set[str]] = None) -> None:
    normalized_email = str(email or "").strip()
    email_key = normalized_email.lower()
    if not normalized_email or (excluded and email_key in excluded):
        return
    if email_key not in {existing.lower() for existing in emails}:
        emails.append(normalized_email)


def send_registration_confirmation_email(
    order: Dict[str, Any],
    item: Dict[str, Any],
    deliver_recipient_as_bcc: bool = False,
    subject_prefix: str = "",
) -> bool:
    recipient_email = str(order.get("email", "") or "").strip()
    if not recipient_email:
        return False

    product_name = str(item.get("name", "") or "").strip()
    product_key = build_order_item_product_key(item)
    email_settings = load_registration_event_email_settings(product_key)
    configured_subject = str(email_settings.get("emailSubject", "") or "").strip()
    from_name = get_env("REGISTRATION_EMAIL_FROM_NAME") or "HWS Voetbalschool"
    from_email = settings.DEFAULT_FROM_EMAIL
    sender = f"{from_name} <{from_email}>" if from_name else from_email
    bcc = parse_email_list(get_env("REGISTRATION_EMAIL_BCC"))
    reply_to = parse_email_list(get_env("REGISTRATION_EMAIL_REPLY_TO"))
    to = [recipient_email]

    if deliver_recipient_as_bcc:
        primary_recipient = (settings.DEFAULT_FROM_EMAIL or settings.EMAIL_HOST_USER or "info@hwsvoetbalschool.nl").strip()
        to = [primary_recipient]
        excluded_bcc = {email.lower() for email in to}
        manual_bcc: List[str] = []
        append_unique_email(manual_bcc, recipient_email, excluded_bcc)
        for email in bcc:
            append_unique_email(manual_bcc, email, excluded_bcc)
        append_unique_email(manual_bcc, "david.van.walstijn@gmail.com", excluded_bcc)
        bcc = manual_bcc

    body = build_registration_event_email_body(order, item, email_settings)
    subject = (
        render_registration_email_template(configured_subject, order, item, email_settings)
        if configured_subject
        else get_registration_confirmation_subject(product_name)
    )
    normalized_subject_prefix = str(subject_prefix or "").strip()
    if normalized_subject_prefix and not subject.lower().startswith(normalized_subject_prefix.lower()):
        subject = f"{normalized_subject_prefix} {subject}"

    email_message = EmailMessage(
        subject=subject,
        body=render_registration_email_html(body),
        from_email=sender,
        to=to,
        bcc=bcc,
        reply_to=reply_to or None,
    )
    email_message.content_subtype = "html"
    email_message.send(fail_silently=False)
    return True


def load_registration_events_due_for_reminder(reminder_date: date) -> Dict[str, Dict[str, str]]:
    due_event_date = (reminder_date + timedelta(days=7)).isoformat()
    with get_db_connection() as connection:
        rows = connection.execute(
            """
            SELECT product_key, product_name, event_date, event_date_2, email_subject, email_body, updated_at
            FROM registration_event_email_settings
            WHERE event_date = ?
            """,
            (due_event_date,),
        ).fetchall()

    due_events: Dict[str, Dict[str, str]] = {}
    for row in rows:
        product_key = str(row["product_key"] or "").strip()
        if not product_key:
            continue
        due_events[product_key] = {
            "productKey": product_key,
            "productName": str(row["product_name"] or "").strip(),
            "eventDate": str(row["event_date"] or "").strip(),
            "eventDate2": str(row["event_date_2"] or "").strip(),
            "emailSubject": str(row["email_subject"] or "").strip(),
            "emailBody": str(row["email_body"] or "").strip(),
            "updatedAt": str(row["updated_at"] or "").strip(),
        }
    return due_events


def send_registration_reminder_emails(
    orders: List[Dict[str, Any]],
    reminder_date: Optional[date] = None,
) -> Dict[str, Any]:
    result = {
        "reminderDate": (reminder_date or date.today()).isoformat(),
        "dueProductKeys": [],
        "sentOrderIds": [],
        "skippedOrderIds": [],
        "failedOrderIds": [],
    }
    if not registration_auto_email_is_configured():
        return result

    current_reminder_date = reminder_date or date.today()
    due_events = load_registration_events_due_for_reminder(current_reminder_date)
    result["dueProductKeys"] = sorted(due_events.keys())
    if not due_events:
        return result

    only_paid = get_env_bool("REGISTRATION_EMAIL_ONLY_PAID", True)
    sent_order_keys: Set[str] = set()
    skipped_order_keys: Set[str] = set()
    failed_order_keys: Set[str] = set()

    for order in orders:
        order_id = str(order.get("id", "") or "").strip()
        recipient_email = str(order.get("email", "") or "").strip()
        if not order_id or not recipient_email:
            continue
        if only_paid and not registration_order_is_paid(order):
            continue

        for item in order.get("items", []):
            product_key = build_order_item_product_key(item)
            if product_key not in due_events:
                continue

            order_key = f"{product_key}:{order_id}"
            if order_id not in load_registration_emailed_order_ids(product_key, {order_id}):
                skipped_order_keys.add(order_key)
                continue
            if order_id in load_registration_reminder_sent_order_ids(product_key, {order_id}):
                skipped_order_keys.add(order_key)
                continue

            try:
                email_sent = send_registration_confirmation_email(order, item, subject_prefix="Reminder: ")
            except Exception as exc:
                failed_order_keys.add(order_key)
                app.logger.warning("Automatische inschrijvingsreminder mislukt voor order %s: %s", order_id, exc)
                continue

            if not email_sent:
                skipped_order_keys.add(order_key)
                continue

            set_registration_reminders_sent(product_key, [order_id], True)
            sent_order_keys.add(order_key)

    result["sentOrderIds"] = sorted(sent_order_keys)
    result["skippedOrderIds"] = sorted(skipped_order_keys)
    result["failedOrderIds"] = sorted(failed_order_keys)
    return result


def auto_email_new_registration_orders(orders: List[Dict[str, Any]]) -> Dict[str, Any]:
    result = {"sentOrderIds": [], "failedOrderIds": []}
    if not registration_auto_email_is_configured():
        return result

    only_paid = get_env_bool("REGISTRATION_EMAIL_ONLY_PAID", True)
    sync_ecwid = get_env_bool("REGISTRATION_EMAIL_SYNC_ECWID_PROCESSING", True)
    sent_order_keys: Set[str] = set()
    failed_order_ids: Set[str] = set()

    for order in orders:
        order_id = str(order.get("id", "") or "").strip()
        if not order_id or not str(order.get("email", "") or "").strip():
            continue
        if not registration_order_is_after_auto_email_start(order):
            continue
        if only_paid and not registration_order_is_paid(order):
            continue

        for item in order.get("items", []):
            product_key = build_order_item_product_key(item)
            if not product_key:
                continue
            if order_id in load_registration_emailed_order_ids(product_key, {order_id}):
                continue

            try:
                email_sent = send_registration_confirmation_email(order, item)
            except Exception as exc:
                failed_order_ids.add(order_id)
                app.logger.warning("Automatische inschrijvingsmail mislukt voor order %s: %s", order_id, exc)
                continue

            if not email_sent:
                continue

            set_registration_orders_emailed(product_key, [order_id], True)
            sent_order_keys.add(f"{product_key}:{order_id}")
            if sync_ecwid:
                try:
                    update_ecwid_order_to_processing(order_id)
                except RuntimeError as exc:
                    app.logger.warning("Ecwid-status na automatische mail niet bijgewerkt voor order %s: %s", order_id, exc)

    result["sentOrderIds"] = sorted(sent_order_keys)
    result["failedOrderIds"] = sorted(failed_order_ids)
    return result


def send_registration_product_emails(product_key: str, orders: List[Dict[str, Any]]) -> Dict[str, Any]:
    normalized_product_key = str(product_key or "").strip()
    result = {
        "productKey": normalized_product_key,
        "sentOrderIds": [],
        "skippedOrderIds": [],
        "failedOrderIds": [],
        "ecwidUpdatedOrderIds": [],
    }
    if not normalized_product_key:
        return result

    sync_ecwid = get_env_bool("REGISTRATION_EMAIL_SYNC_ECWID_PROCESSING", True)
    seen_order_ids: Set[str] = set()
    sent_order_ids: Set[str] = set()
    skipped_order_ids: Set[str] = set()
    failed_order_ids: Set[str] = set()
    ecwid_updated_order_ids: Set[str] = set()

    for order in orders:
        order_id = str(order.get("id", "") or "").strip()
        recipient_email = str(order.get("email", "") or "").strip()
        if not order_id or order_id in seen_order_ids:
            continue
        if not recipient_email:
            continue

        matching_item = next(
            (item for item in order.get("items", []) if build_order_item_product_key(item) == normalized_product_key),
            None,
        )
        if matching_item is None:
            continue

        seen_order_ids.add(order_id)
        if order_id in load_registration_emailed_order_ids(normalized_product_key, {order_id}):
            skipped_order_ids.add(order_id)
            continue
        if get_env_bool("REGISTRATION_EMAIL_ONLY_PAID", True) and not registration_order_is_paid(order):
            skipped_order_ids.add(order_id)
            continue

        try:
            email_sent = send_registration_confirmation_email(order, matching_item, deliver_recipient_as_bcc=True)
        except Exception as exc:
            failed_order_ids.add(order_id)
            app.logger.warning("Handmatige inschrijvingsmail mislukt voor order %s: %s", order_id, exc)
            continue

        if not email_sent:
            skipped_order_ids.add(order_id)
            continue

        set_registration_orders_emailed(normalized_product_key, [order_id], True)
        sent_order_ids.add(order_id)
        if sync_ecwid:
            try:
                if update_ecwid_order_to_processing(order_id):
                    ecwid_updated_order_ids.add(order_id)
            except RuntimeError as exc:
                app.logger.warning("Ecwid-status na handmatige mail niet bijgewerkt voor order %s: %s", order_id, exc)

    result["sentOrderIds"] = sorted(sent_order_ids)
    result["skippedOrderIds"] = sorted(skipped_order_ids)
    result["failedOrderIds"] = sorted(failed_order_ids)
    result["ecwidUpdatedOrderIds"] = sorted(ecwid_updated_order_ids)
    return result


def load_completed_registration_event_keys() -> Set[str]:
    with get_db_connection() as connection:
        rows = connection.execute(
            """
            SELECT product_key
            FROM registration_event_statuses
            WHERE completed_at IS NOT NULL
              AND canceled_at IS NULL
              AND trim(product_key) != ''
            """
        ).fetchall()

    return {str(row["product_key"] or "").strip() for row in rows if str(row["product_key"] or "").strip()}


def load_canceled_registration_event_keys() -> Set[str]:
    with get_db_connection() as connection:
        rows = connection.execute(
            """
            SELECT product_key
            FROM registration_event_statuses
            WHERE canceled_at IS NOT NULL
              AND trim(product_key) != ''
            """
        ).fetchall()

    return {str(row["product_key"] or "").strip() for row in rows if str(row["product_key"] or "").strip()}


def is_registration_event_completed(product_key: str) -> bool:
    normalized_product_key = str(product_key or "").strip()
    if not normalized_product_key:
        return False
    return normalized_product_key in load_completed_registration_event_keys()


def is_registration_event_canceled(product_key: str) -> bool:
    normalized_product_key = str(product_key or "").strip()
    if not normalized_product_key:
        return False
    return normalized_product_key in load_canceled_registration_event_keys()


def set_registration_event_completed(product_key: str) -> bool:
    normalized_product_key = str(product_key or "").strip()
    if not normalized_product_key:
        return False

    with get_db_connection() as connection:
        connection.execute(
            """
            INSERT INTO registration_event_statuses (product_key, completed_at, canceled_at)
            VALUES (?, ?, NULL)
            ON CONFLICT(product_key) DO UPDATE SET
                completed_at = excluded.completed_at,
                canceled_at = NULL
            """,
            (normalized_product_key, utcnow_iso()),
        )

    return True


def set_registration_event_canceled(product_key: str) -> bool:
    normalized_product_key = str(product_key or "").strip()
    if not normalized_product_key:
        return False

    with get_db_connection() as connection:
        connection.execute(
            """
            INSERT INTO registration_event_statuses (product_key, completed_at, canceled_at)
            VALUES (?, ?, ?)
            ON CONFLICT(product_key) DO UPDATE SET
                completed_at = NULL,
                canceled_at = excluded.canceled_at
            """,
            (normalized_product_key, utcnow_iso(), utcnow_iso()),
        )

    return True


def build_orders_filter_options(orders: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, str]]]:
    statuses = sorted(
        {str(order.get("status", "")).strip() for order in orders if str(order.get("status", "")).strip()}
    )
    payment_statuses = sorted(
        {
            str(order.get("paymentStatus", "")).strip()
            for order in orders
            if str(order.get("paymentStatus", "")).strip()
        }
    )
    month_map: Dict[str, str] = {}
    for order in orders:
        created_at = parse_iso_datetime(order.get("createdAt", ""))
        if created_at is None:
            continue
        month_key = created_at.strftime("%Y-%m")
        month_map[month_key] = get_month_label(month_key).capitalize()

    return {
        "statuses": [{"value": value, "label": value.replace("_", " ").title()} for value in statuses],
        "payment_statuses": [
            {"value": value, "label": value.replace("_", " ").title()} for value in payment_statuses
        ],
        "months": [
            {"value": value, "label": month_map[value]}
            for value in sorted(month_map.keys(), reverse=True)
        ],
    }


def filter_orders(
    orders: List[Dict[str, Any]],
    search_query: str = "",
    status: str = "",
    payment_status: str = "",
    month: str = "",
) -> List[Dict[str, Any]]:
    normalized_query = search_query.strip().lower()
    normalized_status = status.strip()
    normalized_payment_status = payment_status.strip()
    normalized_month = month.strip()

    filtered_orders: List[Dict[str, Any]] = []
    for order in orders:
        created_at = parse_iso_datetime(order.get("createdAt", ""))
        searchable_parts = [
            str(order.get("orderNumber", "") or ""),
            str(order.get("customerName", "") or ""),
            str(order.get("email", "") or ""),
            str(order.get("paymentMethod", "") or ""),
            str(order.get("shippingMethod", "") or ""),
        ]
        searchable_parts.extend(str(item.get("name", "") or "") for item in order.get("items", []))
        searchable_text = " ".join(searchable_parts).lower()

        if normalized_query and normalized_query not in searchable_text:
            continue
        if normalized_status and str(order.get("status", "")).strip() != normalized_status:
            continue
        if normalized_payment_status and str(order.get("paymentStatus", "")).strip() != normalized_payment_status:
            continue
        if normalized_month:
            if created_at is None or created_at.strftime("%Y-%m") != normalized_month:
                continue

        filtered_orders.append(order)

    return filtered_orders


def build_orders_page_url(page: int = 1, search_query: str = "", status: str = "", payment_status: str = "", month: str = "") -> str:
    params: Dict[str, Any] = {"page": page}
    if search_query.strip():
        params["q"] = search_query.strip()
    if status.strip():
        params["status"] = status.strip()
    if payment_status.strip():
        params["payment_status"] = payment_status.strip()
    if month.strip():
        params["month"] = month.strip()
    return url_for("orders_page", **params)


def parse_selected_order_ids(raw_ids: List[str]) -> List[str]:
    selected_ids: List[str] = []
    for raw_id in raw_ids:
        for candidate in str(raw_id or "").split(","):
            cleaned = candidate.strip()
            if cleaned and cleaned not in selected_ids:
                selected_ids.append(cleaned)
    return selected_ids


def build_team_assignment_rows(orders: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for order in sort_orders_desc(orders):
        created_at = parse_iso_datetime(order.get("createdAt", ""))
        display_date = created_at.strftime("%d-%m-%Y") if created_at else "-"

        for item in order.get("items", []):
            quantity = max(int(item.get("quantity") or 0), 1)
            for participant_index in range(quantity):
                rows.append(
                    {
                        "date": display_date,
                        "orderNumber": str(order.get("orderNumber", "") or order.get("id", "")),
                        "customerName": str(order.get("customerName", "") or "-"),
                        "email": str(order.get("email", "") or "-"),
                        "product": str(item.get("name", "") or "Naamloos product"),
                        "sku": str(item.get("sku", "") or "-"),
                        "participantLabel": participant_index + 1 if quantity > 1 else "",
                        "team": "",
                    }
                )

    rows.sort(key=lambda row: (row["product"].lower(), row["date"], row["customerName"].lower(), row["orderNumber"]))
    return rows


def build_team_assignment_workbook(orders: List[Dict[str, Any]]) -> BytesIO:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Teamindeling"

    worksheet["A1"] = "Teamindeling geselecteerde bestellingen"
    worksheet["A1"].font = Font(size=14, bold=True)
    worksheet.merge_cells("A1:H1")
    worksheet["A2"] = f"Gegenereerd op {datetime.now().strftime('%d-%m-%Y %H:%M')}"
    worksheet.merge_cells("A2:H2")

    headers = ["Datum", "Ordernummer", "Naam", "E-mail", "Product", "SKU", "Deelnemer", "Team"]
    header_fill = PatternFill(fill_type="solid", fgColor="111111")
    header_font = Font(color="FFFFFF", bold=True)

    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=4, column=column_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, item in enumerate(build_team_assignment_rows(orders), start=5):
        worksheet.cell(row=row_index, column=1, value=item["date"])
        worksheet.cell(row=row_index, column=2, value=item["orderNumber"])
        worksheet.cell(row=row_index, column=3, value=item["customerName"])
        worksheet.cell(row=row_index, column=4, value=item["email"])
        worksheet.cell(row=row_index, column=5, value=item["product"])
        worksheet.cell(row=row_index, column=6, value=item["sku"])
        worksheet.cell(row=row_index, column=7, value=item["participantLabel"])
        worksheet.cell(row=row_index, column=8, value=item["team"])

    worksheet.freeze_panes = "A5"
    worksheet.auto_filter.ref = f"A4:H{max(5, worksheet.max_row)}"

    column_widths = {
        "A": 14,
        "B": 16,
        "C": 24,
        "D": 30,
        "E": 36,
        "F": 18,
        "G": 12,
        "H": 18,
    }
    for column, width in column_widths.items():
        worksheet.column_dimensions[column].width = width

    for row in worksheet.iter_rows(min_row=5, max_row=worksheet.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def parse_registration_birth_date(value: str) -> Optional[date]:
    raw_value = str(value or "").strip()
    if not raw_value:
        return None

    normalized_value = raw_value.replace("/", "-").replace(".", "-")
    for pattern in ("%d-%m-%Y", "%Y-%m-%d", "%d-%m-%y"):
        try:
            parsed = datetime.strptime(normalized_value, pattern).date()
        except ValueError:
            continue
        if parsed.year > datetime.now().year:
            parsed = parsed.replace(year=parsed.year - 100)
        return parsed

    return None


def normalize_registration_person_name(first_name: str, last_name: str) -> str:
    return normalize_match_text(f"{first_name} {last_name}")


def build_registration_participant_rows(selected_product: Dict[str, Any]) -> List[Dict[str, Any]]:
    participants: List[Dict[str, Any]] = []

    for order in selected_product.get("orders", []):
        details = order.get("registrationDetails", {}) if isinstance(order.get("registrationDetails"), dict) else {}
        quantity = max(int(order.get("quantity") or 0), 1)
        first_name = str(details.get("firstName", "") or "").strip()
        last_name = str(details.get("lastName", "") or "").strip()
        fallback_first_name, fallback_last_name = split_full_name(str(order.get("customerName", "") or ""))
        first_name = first_name or fallback_first_name
        last_name = last_name or fallback_last_name

        for participant_index in range(quantity):
            participant = {
                "id": f"{order.get('id', '')}:{participant_index}",
                "group": "",
                "firstName": first_name,
                "lastName": last_name,
                "gender": str(details.get("gender", "") or "").strip(),
                "birthDate": str(details.get("birthDate", "") or "").strip(),
                "birthDateParsed": parse_registration_birth_date(str(details.get("birthDate", "") or "")),
                "address": str(details.get("address", "") or "").strip(),
                "postalCode": str(details.get("postalCode", "") or "").strip(),
                "city": str(details.get("city", "") or "").strip(),
                "clubTeam": str(details.get("clubTeam", "") or "").strip(),
                "phone": str(details.get("phone", "") or "").strip(),
                "email": str(order.get("email", "") or "").strip(),
                "dietaryWishes": str(details.get("dietaryWishes", "") or "").strip(),
                "comments": str(details.get("comments", "") or "").strip(),
                "orderNumber": str(order.get("orderNumber", "") or order.get("id", "")),
            }
            participants.append(participant)

    return participants


def build_registration_team_clusters(participants: List[Dict[str, Any]]) -> List[List[int]]:
    parent = list(range(len(participants)))

    def find(index: int) -> int:
        while parent[index] != index:
            parent[index] = parent[parent[index]]
            index = parent[index]
        return index

    def union(left: int, right: int) -> None:
        left_root = find(left)
        right_root = find(right)
        if left_root != right_root:
            parent[right_root] = left_root

    club_team_indexes: Dict[str, List[int]] = {}
    name_candidates: List[Tuple[int, str, str, str]] = []

    for index, participant in enumerate(participants):
        club_team_key = normalize_match_text(str(participant.get("clubTeam", "") or ""))
        if club_team_key:
            club_team_indexes.setdefault(club_team_key, []).append(index)

        first_name = normalize_match_text(str(participant.get("firstName", "") or ""))
        last_name = normalize_match_text(str(participant.get("lastName", "") or ""))
        full_name = normalize_registration_person_name(
            str(participant.get("firstName", "") or ""),
            str(participant.get("lastName", "") or ""),
        )
        if first_name or last_name or full_name:
            name_candidates.append((index, first_name, last_name, full_name))

    for indexes in club_team_indexes.values():
        first_index = indexes[0]
        for index in indexes[1:]:
            union(first_index, index)

    for index, participant in enumerate(participants):
        comments = normalize_match_text(str(participant.get("comments", "") or ""))
        if not comments:
            continue

        comment_tokens = set(comments.split())
        for candidate_index, first_name, last_name, full_name in name_candidates:
            if candidate_index == index:
                continue
            if full_name and full_name in comments:
                union(index, candidate_index)
                continue
            if first_name and last_name and first_name in comment_tokens and last_name in comment_tokens:
                union(index, candidate_index)
                continue
            if first_name and len(first_name) >= 4 and first_name in comment_tokens:
                union(index, candidate_index)

    clusters_by_root: Dict[int, List[int]] = {}
    for index in range(len(participants)):
        clusters_by_root.setdefault(find(index), []).append(index)

    def cluster_sort_key(cluster: List[int]) -> Tuple[date, str]:
        parsed_dates = [
            participant["birthDateParsed"]
            for participant in (participants[index] for index in cluster)
            if participant.get("birthDateParsed") is not None
        ]
        youngest_date = max(parsed_dates) if parsed_dates else date.min
        names = " ".join(
            normalize_registration_person_name(
                str(participants[index].get("firstName", "") or ""),
                str(participants[index].get("lastName", "") or ""),
            )
            for index in cluster
        )
        return youngest_date, names

    return sorted(clusters_by_root.values(), key=cluster_sort_key, reverse=True)


def assign_registration_team_groups(participants: List[Dict[str, Any]], group_count: int) -> List[Dict[str, Any]]:
    if not participants:
        return []

    normalized_group_count = max(1, min(int(group_count or 1), len(participants)))
    clusters = build_registration_team_clusters(participants)
    target_sizes = [
        len(participants) // normalized_group_count + (1 if group_index < len(participants) % normalized_group_count else 0)
        for group_index in range(normalized_group_count)
    ]
    group_index = 0
    current_group_size = 0

    for cluster in clusters:
        target_size = target_sizes[group_index] if group_index < len(target_sizes) else len(participants)
        if (
            group_index < normalized_group_count - 1
            and current_group_size > 0
            and current_group_size + len(cluster) > target_size
        ):
            group_index += 1
            current_group_size = 0

        group_label = f"Groep {group_index + 1}"
        for participant_index in cluster:
            participants[participant_index]["group"] = group_label
        current_group_size += len(cluster)

    return sorted(
        participants,
        key=lambda participant: (
            int(str(participant.get("group", "Groep 999")).replace("Groep", "").strip() or 999),
            -(participant.get("birthDateParsed") or date.min).toordinal(),
            normalize_registration_person_name(
                str(participant.get("firstName", "") or ""),
                str(participant.get("lastName", "") or ""),
            ),
        ),
    )


def build_registration_team_assignment_workbook(
    selected_product: Dict[str, Any],
    group_count: int,
) -> BytesIO:
    participants = assign_registration_team_groups(
        build_registration_participant_rows(selected_product),
        group_count,
    )
    workbook = Workbook()
    worksheet = workbook.active
    sheet_title = re.sub(r"[\[\]\*:/\\?]", "", str(selected_product.get("name", "") or "Teamindeling")).strip()
    worksheet.title = (sheet_title or "Teamindeling")[:31]

    headers = [
        "Groep:",
        "Voornaam:",
        "Achternaam:",
        "Geslacht:",
        "Geboortedatum:",
        "Adres:",
        "Postcode:",
        "Plaats:",
        "Club/Team:",
        "06-nummer:",
        "E-mailadres:",
        "Dieetwensen:",
        "Opmerkingen:",
    ]
    header_fill = PatternFill(fill_type="solid", fgColor="FF595959")
    header_font = Font(name="Calibri", size=11, color="FFFFFFFF", bold=True)
    thin_gray_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = thin_gray_border

    for row_index, participant in enumerate(participants, start=2):
        values = [
            participant["group"],
            participant["firstName"],
            participant["lastName"],
            participant["gender"],
            participant["birthDate"],
            participant["address"],
            participant["postalCode"],
            participant["city"],
            participant["clubTeam"],
            participant["phone"],
            participant["email"],
            participant["dietaryWishes"],
            participant["comments"],
        ]
        for column_index, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_gray_border

    worksheet.freeze_panes = "A2"
    worksheet.row_dimensions[1].height = 24.75
    for row_index in range(2, max(2, worksheet.max_row) + 1):
        worksheet.row_dimensions[row_index].height = 24.75

    column_widths = {
        "A": 15.5,
        "B": 10.7,
        "C": 12.7,
        "D": 15.7,
        "E": 15.5,
        "F": 22,
        "G": 12,
        "H": 16,
        "I": 22,
        "J": 16,
        "K": 28,
        "L": 30.5,
        "M": 36,
    }
    for column, width in column_widths.items():
        worksheet.column_dimensions[column].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def build_summary(orders: List[Dict[str, Any]]) -> Dict[str, Any]:
    revenue = sum(
        float(order.get("total") or 0)
        for order in orders
        if order.get("paymentStatus") != "REFUNDED"
    )
    return {
        "orderCount": len(orders),
        "revenue": round(revenue, 2),
        "paidCount": sum(1 for order in orders if order.get("paymentStatus") == "PAID"),
        "openCount": sum(1 for order in orders if order.get("paymentStatus") != "PAID"),
        "refundedCount": sum(1 for order in orders if order.get("paymentStatus") == "REFUNDED"),
    }


def decimal_from_value(value: Any) -> Decimal:
    try:
        return Decimal(str(value or "0"))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def calculate_margin_percentage(revenue: Decimal, profit: Decimal) -> float:
    if revenue == 0:
        return 0.0

    return round(float((profit / revenue) * Decimal("100")), 1)


def get_mutation_search_text(mutation: Dict[str, Any]) -> str:
    sepa_fields = mutation.get("sepa_fields") or {}
    parts = [
        str(mutation.get("contra_account_name", "")).strip(),
        str(mutation.get("message", "")).strip(),
        str(mutation.get("description", "")).strip(),
        str(sepa_fields.get("remi", "")).strip(),
        str(sepa_fields.get("eref", "")).strip(),
        str(sepa_fields.get("sref", "")).strip(),
    ]
    return normalize_match_text(" ".join(part for part in parts if part))


def is_excluded_equity_mutation(mutation: Dict[str, Any]) -> bool:
    mutation_text = get_mutation_search_text(mutation)
    if not mutation_text:
        return False

    excluded_names = [
        "nick horst",
        "horst nick",
        "david van walstijn",
        "walstijn david",
    ]
    return any(normalize_match_text(name) in mutation_text for name in excluded_names)


def is_expense_payment(payment: Dict[str, Any]) -> bool:
    invoice_type = str(payment.get("invoice_type", "")).strip()
    return invoice_type in {"Document", "PurchaseInvoice", "Receipt", "PurchaseTransaction"}


def has_expense_booking(mutation: Dict[str, Any], ledger_account_types: Dict[str, str]) -> bool:
    for booking in mutation.get("ledger_account_bookings") or []:
        ledger_account_id = str(booking.get("ledger_account_id", "")).strip()
        if ledger_account_types.get(ledger_account_id) in {"expenses", "direct_costs"}:
            return True
    return False


def is_cost_mutation(mutation: Dict[str, Any], ledger_account_types: Dict[str, str]) -> bool:
    mutation_amount = decimal_from_value(mutation.get("amount"))
    if mutation_amount >= 0 or is_excluded_equity_mutation(mutation):
        return False

    if any(is_expense_payment(payment) for payment in (mutation.get("payments") or [])):
        return True

    return has_expense_booking(mutation, ledger_account_types)


def is_sales_invoice_payment(payment: Dict[str, Any]) -> bool:
    invoice_type = str(payment.get("invoice_type", "")).strip()
    return invoice_type in {"SalesInvoice", "Invoice"}


def is_external_sales_invoice_payment(payment: Dict[str, Any]) -> bool:
    invoice_type = str(payment.get("invoice_type", "")).strip()
    return invoice_type == "ExternalSalesInvoice"


def is_spaarpot_stripe_income_mutation(mutation: Dict[str, Any]) -> bool:
    amount = decimal_from_value(mutation.get("amount"))
    if amount == 0 or is_excluded_equity_mutation(mutation):
        return False

    mutation_text = get_mutation_search_text(mutation)
    if "stripe" not in mutation_text:
        return False

    payments = mutation.get("payments") or []
    if any(is_sales_invoice_payment(payment) for payment in payments):
        return False

    return amount > 0 or any(is_external_sales_invoice_payment(payment) for payment in payments)


def get_spaarpot_stripe_mutation_amount(mutation: Dict[str, Any]) -> Decimal:
    return abs(decimal_from_value(mutation.get("amount")))


def build_report_summary(ecwid_summary: Dict[str, Any], moneybird_summary: Dict[str, Any]) -> Dict[str, Any]:
    ecwid_revenue = decimal_from_value(ecwid_summary.get("revenue"))
    moneybird_revenue = decimal_from_value(moneybird_summary.get("revenue_received"))
    expenses_total = decimal_from_value(moneybird_summary.get("expenses_total"))
    total_revenue = ecwid_revenue + moneybird_revenue
    total_profit = total_revenue - expenses_total

    return {
        "ecwidRevenue": round(float(ecwid_revenue), 2),
        "moneybirdRevenue": round(float(moneybird_revenue), 2),
        "combinedRevenue": round(float(total_revenue), 2),
        "expensesTotal": round(float(expenses_total), 2),
        "profitTotal": round(float(total_profit), 2),
        "profitMarginPercentage": calculate_margin_percentage(total_revenue, total_profit),
        "moneybirdInvoiceCount": moneybird_summary.get("invoiceCount", 0),
        "moneybirdAdministrationName": moneybird_summary.get("administrationName", ""),
        "moneybirdLastSyncedAt": moneybird_summary.get("lastSyncedAt", ""),
    }


def get_default_dashboard_events() -> List[Dict[str, Any]]:
    return [
        {"label": "Voetbaldag VV Voorst", "matchTerms": ["vv voorst voetbaldag"]},
        {"label": "Voetbaldag SV Harfsen", "matchTerms": ["sv harfsen voetbaldag"]},
        {"label": "Voetbaldag WWNA", "matchTerms": ["wwna voetbaldag"]},
        {"label": "SummerCamp ABS", "matchTerms": ["apeldoornse boys"]},
        {
            "label": "SummerCamp SC Terschelling",
            "matchTerms": ["sc terschelling summercamp", "summercamp sc terschelling"],
        },
    ]


def get_db_connection() -> sqlite3.Connection:
    os.makedirs(DATA_DIR, exist_ok=True)
    connection = sqlite3.connect(DATABASE_PATH, timeout=30)
    connection.row_factory = sqlite3.Row
    return connection


def bootstrap_seed_data_files() -> None:
    os.makedirs(DATA_DIR, exist_ok=True)
    if os.path.abspath(DATA_DIR) == os.path.abspath(BUNDLED_DATA_DIR):
        return

    for filename in ("app.db", "dashboard_events.json", "agenda_trainings.json"):
        source_path = os.path.join(BUNDLED_DATA_DIR, filename)
        target_path = os.path.join(DATA_DIR, filename)
        if os.path.exists(target_path) or not os.path.exists(source_path):
            continue
        shutil.copy2(source_path, target_path)


def sync_seed_workspace_data() -> None:
    if os.path.abspath(DATA_DIR) == os.path.abspath(BUNDLED_DATA_DIR):
        return

    source_db_path = os.path.join(BUNDLED_DATA_DIR, "app.db")
    if not os.path.exists(source_db_path) or not os.path.exists(DATABASE_PATH):
        return

    with sqlite3.connect(DATABASE_PATH, timeout=30) as connection:
        connection.execute("ATTACH DATABASE ? AS seed", (source_db_path,))
        target_tables = {
            str(row[0])
            for row in connection.execute(
                "SELECT name FROM sqlite_master WHERE type = 'table'"
            ).fetchall()
        }
        seed_tables = {
            str(row[0])
            for row in connection.execute(
                "SELECT name FROM seed.sqlite_master WHERE type = 'table'"
            ).fetchall()
        }

        if {"trainer_profiles"} <= target_tables and {"trainer_profiles"} <= seed_tables:
            connection.execute(
                """
                INSERT INTO trainer_profiles (
                    id,
                    full_name,
                    email,
                    username,
                    role,
                    phone,
                    notes,
                    status,
                    created_at,
                    password_hash,
                    is_admin,
                    member_type,
                    system_role,
                    knvb_license,
                    education,
                    availability_days,
                    invite_token,
                    invite_expires_at,
                    invite_accepted_at
                )
                SELECT
                    seed_profile.id,
                    seed_profile.full_name,
                    seed_profile.email,
                    seed_profile.username,
                    seed_profile.role,
                    seed_profile.phone,
                    seed_profile.notes,
                    seed_profile.status,
                    seed_profile.created_at,
                    seed_profile.password_hash,
                    seed_profile.is_admin,
                    seed_profile.member_type,
                    seed_profile.system_role,
                    seed_profile.knvb_license,
                    seed_profile.education,
                    seed_profile.availability_days,
                    seed_profile.invite_token,
                    seed_profile.invite_expires_at,
                    seed_profile.invite_accepted_at
                FROM seed.trainer_profiles AS seed_profile
                WHERE NOT EXISTS (
                    SELECT 1
                    FROM trainer_profiles AS live_profile
                    WHERE live_profile.id = seed_profile.id
                       OR lower(live_profile.email) = lower(seed_profile.email)
                )
                """
            )

        if {"content_albums", "content_photos"} <= target_tables and {"content_albums", "content_photos"} <= seed_tables:
            connection.execute(
                """
                INSERT INTO content_albums (title, slug, created_at)
                SELECT seed_album.title, seed_album.slug, seed_album.created_at
                FROM seed.content_albums AS seed_album
                WHERE NOT EXISTS (
                    SELECT 1
                    FROM content_albums AS live_album
                    WHERE live_album.slug = seed_album.slug
                )
                """
            )
            connection.execute(
                """
                INSERT INTO content_photos (
                    album_id,
                    image_url,
                    remote_path,
                    file_name,
                    original_name,
                    content_type,
                    file_size,
                    storage_backend,
                    uploaded_at
                )
                SELECT
                    live_album.id,
                    seed_photo.image_url,
                    seed_photo.remote_path,
                    seed_photo.file_name,
                    seed_photo.original_name,
                    seed_photo.content_type,
                    seed_photo.file_size,
                    seed_photo.storage_backend,
                    seed_photo.uploaded_at
                FROM seed.content_photos AS seed_photo
                JOIN seed.content_albums AS seed_album
                    ON seed_album.id = seed_photo.album_id
                JOIN content_albums AS live_album
                    ON live_album.slug = seed_album.slug
                WHERE NOT EXISTS (
                    SELECT 1
                    FROM content_photos AS live_photo
                    WHERE live_photo.remote_path = seed_photo.remote_path
                )
                """
            )

        connection.commit()
        connection.execute("DETACH DATABASE seed")


def init_db() -> None:
    with get_db_connection() as connection:
        connection.executescript(
            """
            CREATE TABLE IF NOT EXISTS dashboard_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_id TEXT,
                label TEXT NOT NULL,
                match_terms TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS agenda_trainings (
                id TEXT PRIMARY KEY,
                title TEXT NOT NULL,
                date TEXT NOT NULL,
                time TEXT NOT NULL,
                end_time TEXT,
                location TEXT,
                training_type TEXT NOT NULL DEFAULT 'samenwerkende_amateurclub',
                status TEXT NOT NULL DEFAULT 'gepland',
                trainers_json TEXT NOT NULL DEFAULT '[]',
                notes TEXT
            );

            CREATE TABLE IF NOT EXISTS agenda_day_plans (
                date TEXT PRIMARY KEY,
                plan_type TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS agenda_api_credentials (
                id INTEGER PRIMARY KEY CHECK (id = 1),
                token_salt TEXT NOT NULL,
                created_at TEXT NOT NULL,
                last_used_at TEXT
            );

            CREATE TABLE IF NOT EXISTS exercises (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                category TEXT,
                duration TEXT,
                training_exercise TEXT,
                description TEXT,
                coaching TEXT,
                variation_easier TEXT,
                variation_harder TEXT,
                dimensions TEXT,
                materials TEXT,
                age_groups_json TEXT NOT NULL DEFAULT '[]',
                field_json TEXT NOT NULL DEFAULT '{}',
                source_slide INTEGER,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS training_sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                training_date TEXT,
                objective TEXT,
                notes TEXT,
                exercises_json TEXT NOT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS trainer_profiles (
                id TEXT PRIMARY KEY,
                full_name TEXT NOT NULL,
                email TEXT NOT NULL,
                username TEXT NOT NULL,
                password_hash TEXT,
                invite_token TEXT,
                invite_expires_at TEXT,
                invite_accepted_at TEXT,
                role TEXT NOT NULL,
                member_type TEXT,
                system_role TEXT,
                knvb_license TEXT,
                education TEXT,
                availability_days TEXT,
                phone TEXT,
                address TEXT,
                city TEXT,
                postal_code TEXT,
                bank_account_number TEXT,
                bank_account_name TEXT,
                notes TEXT,
                trainer_fees_json TEXT NOT NULL DEFAULT '[]',
                is_admin INTEGER NOT NULL DEFAULT 0,
                status TEXT NOT NULL,
                created_at TEXT NOT NULL
            );

            CREATE UNIQUE INDEX IF NOT EXISTS idx_trainer_profiles_username
            ON trainer_profiles (username COLLATE NOCASE);

            CREATE TABLE IF NOT EXISTS tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                due_date TEXT NOT NULL,
                is_done INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS football_days_playbook (
                id INTEGER PRIMARY KEY CHECK (id = 1),
                title TEXT NOT NULL,
                event_date TEXT,
                cycle_number TEXT,
                cycle_start_date TEXT,
                cycle_end_date TEXT,
                location TEXT,
                staff_json TEXT NOT NULL DEFAULT '[]',
                program_json TEXT NOT NULL DEFAULT '[]',
                contingencies TEXT,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS football_days_playbooks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                playbook_type TEXT NOT NULL DEFAULT 'voetbaldagen',
                title TEXT NOT NULL,
                event_date TEXT,
                cycle_number TEXT,
                location TEXT,
                ecwid_product_id TEXT,
                ecwid_product_name TEXT,
                ecwid_product_sku TEXT,
                staff_json TEXT NOT NULL DEFAULT '[]',
                program_json TEXT NOT NULL DEFAULT '[]',
                field_layout_json TEXT NOT NULL DEFAULT '[]',
                field_trainings_json TEXT NOT NULL DEFAULT '[]',
                cycle_no_training_dates_json TEXT NOT NULL DEFAULT '[]',
                contingencies TEXT,
                include_staff INTEGER NOT NULL DEFAULT 1,
                include_staff_setup_tasks INTEGER NOT NULL DEFAULT 1,
                include_program INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS planning_documents (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                planning_date TEXT,
                location TEXT,
                include_icons INTEGER NOT NULL DEFAULT 1,
                program_json TEXT NOT NULL DEFAULT '[]',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS contracts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                club_name TEXT NOT NULL,
                club_address TEXT,
                season TEXT,
                start_date TEXT,
                end_date TEXT,
                notice_period TEXT,
                training_lines_json TEXT NOT NULL DEFAULT '[]',
                training_execution_summary TEXT,
                training_execution_details TEXT,
                agenda_attachment_title TEXT,
                agenda_attachment_items_json TEXT NOT NULL DEFAULT '[]',
                hws_materials TEXT,
                club_materials TEXT,
                extra_activities TEXT,
                cost_lines_json TEXT NOT NULL DEFAULT '[]',
                price_per_training TEXT,
                training_count INTEGER NOT NULL DEFAULT 0,
                total_amount TEXT,
                min_players TEXT,
                weather_cancellation TEXT,
                hws_absence TEXT,
                liability TEXT,
                participation_risk TEXT,
                evaluation_moments TEXT,
                hws_signatory TEXT,
                club_signatory TEXT,
                signing_date TEXT,
                notes TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS proposals (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                club_name TEXT NOT NULL,
                proposal_type TEXT NOT NULL,
                season_start_year INTEGER NOT NULL,
                price_per_training TEXT NOT NULL,
                total_trainings INTEGER NOT NULL DEFAULT 0,
                total_amount REAL NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS automatic_invoice_settings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                club_name TEXT NOT NULL,
                standard_amount TEXT NOT NULL,
                training_amount TEXT NOT NULL,
                invoice_day INTEGER NOT NULL,
                repeat_enabled INTEGER NOT NULL DEFAULT 0,
                period_start TEXT,
                period_end TEXT,
                active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS automatic_invoice_runs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                setting_id INTEGER NOT NULL,
                invoice_month TEXT NOT NULL,
                moneybird_invoice_id TEXT,
                moneybird_draft_id TEXT,
                status TEXT NOT NULL,
                error_message TEXT,
                created_at TEXT NOT NULL,
                UNIQUE(setting_id, invoice_month),
                FOREIGN KEY (setting_id) REFERENCES automatic_invoice_settings(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS proposal_lines (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                proposal_id INTEGER NOT NULL,
                weekday_key TEXT NOT NULL,
                activity_description TEXT NOT NULL,
                training_count INTEGER NOT NULL DEFAULT 0,
                sort_order INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                FOREIGN KEY (proposal_id) REFERENCES proposals(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS dashboard_preferences (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS material_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                total_count INTEGER NOT NULL DEFAULT 0,
                sort_order INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS material_clubs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                sort_order INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS material_club_inventory (
                club_id INTEGER NOT NULL,
                material_id INTEGER NOT NULL,
                quantity INTEGER NOT NULL DEFAULT 0,
                PRIMARY KEY (club_id, material_id),
                FOREIGN KEY (club_id) REFERENCES material_clubs(id) ON DELETE CASCADE,
                FOREIGN KEY (material_id) REFERENCES material_items(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS social_media_ideas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                platform TEXT NOT NULL,
                content_type TEXT NOT NULL,
                priority TEXT NOT NULL DEFAULT 'Midden',
                is_scheduled INTEGER NOT NULL DEFAULT 0,
                notes TEXT,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS social_media_schedule (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                platform TEXT NOT NULL,
                publish_date TEXT NOT NULL,
                publish_time TEXT NOT NULL,
                status TEXT NOT NULL,
                notes TEXT,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS content_albums (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                slug TEXT NOT NULL,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS content_photos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                album_id INTEGER NOT NULL,
                image_url TEXT NOT NULL,
                remote_path TEXT NOT NULL,
                file_name TEXT NOT NULL,
                original_name TEXT,
                content_type TEXT,
                file_size INTEGER NOT NULL DEFAULT 0,
                storage_backend TEXT NOT NULL DEFAULT 'local',
                uploaded_at TEXT NOT NULL,
                FOREIGN KEY (album_id) REFERENCES content_albums(id)
            );

            CREATE TABLE IF NOT EXISTS rate_limit_attempts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                request_key TEXT NOT NULL,
                created_at REAL NOT NULL
            );

            CREATE TABLE IF NOT EXISTS registration_email_statuses (
                product_key TEXT NOT NULL,
                order_id TEXT NOT NULL,
                emailed_at TEXT NOT NULL,
                PRIMARY KEY (product_key, order_id)
            );

            CREATE INDEX IF NOT EXISTS idx_registration_email_statuses_product_key
            ON registration_email_statuses (product_key);

            CREATE TABLE IF NOT EXISTS registration_email_reminder_statuses (
                product_key TEXT NOT NULL,
                order_id TEXT NOT NULL,
                reminded_at TEXT NOT NULL,
                PRIMARY KEY (product_key, order_id)
            );

            CREATE INDEX IF NOT EXISTS idx_registration_email_reminder_statuses_product_key
            ON registration_email_reminder_statuses (product_key);

            CREATE TABLE IF NOT EXISTS registration_event_statuses (
                product_key TEXT PRIMARY KEY,
                completed_at TEXT,
                canceled_at TEXT
            );

            CREATE TABLE IF NOT EXISTS registration_event_email_settings (
                product_key TEXT PRIMARY KEY,
                product_name TEXT NOT NULL DEFAULT '',
                event_date TEXT,
                event_date_2 TEXT,
                email_subject TEXT NOT NULL DEFAULT '',
                email_body TEXT NOT NULL DEFAULT '',
                updated_at TEXT NOT NULL
            );

            CREATE INDEX IF NOT EXISTS idx_registration_event_email_settings_updated_at
            ON registration_event_email_settings (updated_at);

            CREATE TABLE IF NOT EXISTS spaarpot_manual_entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                year INTEGER NOT NULL,
                quarter INTEGER NOT NULL,
                description TEXT NOT NULL,
                amount REAL NOT NULL,
                created_at TEXT NOT NULL
            );

            CREATE INDEX IF NOT EXISTS idx_spaarpot_manual_entries_period
            ON spaarpot_manual_entries (year, quarter);

            CREATE TABLE IF NOT EXISTS external_api_cache (
                cache_key TEXT PRIMARY KEY,
                config_fingerprint TEXT NOT NULL,
                payload_json TEXT NOT NULL,
                cached_at REAL NOT NULL
            );

            CREATE TABLE IF NOT EXISTS trainer_fee_payment_statuses (
                trainer_id TEXT NOT NULL,
                season_start_year INTEGER NOT NULL,
                month INTEGER NOT NULL,
                paid INTEGER NOT NULL DEFAULT 0,
                updated_at TEXT NOT NULL,
                PRIMARY KEY (trainer_id, season_start_year, month)
            );

            CREATE TABLE IF NOT EXISTS budget_lines (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                season_start_year INTEGER NOT NULL,
                training_type TEXT NOT NULL,
                club TEXT NOT NULL,
                activity_title TEXT NOT NULL,
                income_amount TEXT NOT NULL DEFAULT '',
                trainer_amount TEXT NOT NULL DEFAULT '',
                trainer_payment_mode TEXT NOT NULL DEFAULT 'per_activity',
                trainer_bundle_count INTEGER NOT NULL DEFAULT 1,
                trainer_group TEXT NOT NULL DEFAULT '',
                trainer_id TEXT,
                sort_order INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE INDEX IF NOT EXISTS idx_budget_lines_season
            ON budget_lines (season_start_year, sort_order);

            CREATE TABLE IF NOT EXISTS web_push_subscriptions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id TEXT NOT NULL,
                endpoint TEXT NOT NULL UNIQUE,
                subscription_json TEXT NOT NULL,
                user_agent TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                last_sent_at TEXT,
                last_error TEXT
            );

            CREATE INDEX IF NOT EXISTS idx_web_push_subscriptions_user_id
            ON web_push_subscriptions (user_id);
            """
        )

        existing_columns = {
            row["name"]
            for row in connection.execute("PRAGMA table_info(trainer_profiles)").fetchall()
        }
        if "password_hash" not in existing_columns:
            connection.execute("ALTER TABLE trainer_profiles ADD COLUMN password_hash TEXT")
        if "invite_token" not in existing_columns:
            connection.execute("ALTER TABLE trainer_profiles ADD COLUMN invite_token TEXT")
        if "invite_expires_at" not in existing_columns:
            connection.execute("ALTER TABLE trainer_profiles ADD COLUMN invite_expires_at TEXT")
        if "invite_accepted_at" not in existing_columns:
            connection.execute("ALTER TABLE trainer_profiles ADD COLUMN invite_accepted_at TEXT")
        if "is_admin" not in existing_columns:
            connection.execute("ALTER TABLE trainer_profiles ADD COLUMN is_admin INTEGER NOT NULL DEFAULT 0")
        if "member_type" not in existing_columns:
            connection.execute("ALTER TABLE trainer_profiles ADD COLUMN member_type TEXT")
        if "system_role" not in existing_columns:
            connection.execute("ALTER TABLE trainer_profiles ADD COLUMN system_role TEXT")
        if "knvb_license" not in existing_columns:
            connection.execute("ALTER TABLE trainer_profiles ADD COLUMN knvb_license TEXT")
        if "education" not in existing_columns:
            connection.execute("ALTER TABLE trainer_profiles ADD COLUMN education TEXT")
        if "availability_days" not in existing_columns:
            connection.execute("ALTER TABLE trainer_profiles ADD COLUMN availability_days TEXT")
        if "address" not in existing_columns:
            connection.execute("ALTER TABLE trainer_profiles ADD COLUMN address TEXT")
        if "city" not in existing_columns:
            connection.execute("ALTER TABLE trainer_profiles ADD COLUMN city TEXT")
        if "postal_code" not in existing_columns:
            connection.execute("ALTER TABLE trainer_profiles ADD COLUMN postal_code TEXT")
        if "bank_account_number" not in existing_columns:
            connection.execute("ALTER TABLE trainer_profiles ADD COLUMN bank_account_number TEXT")
        if "bank_account_name" not in existing_columns:
            connection.execute("ALTER TABLE trainer_profiles ADD COLUMN bank_account_name TEXT")
        if "trainer_fees_json" not in existing_columns:
            connection.execute("ALTER TABLE trainer_profiles ADD COLUMN trainer_fees_json TEXT NOT NULL DEFAULT '[]'")

        registration_event_columns = {
            row["name"]
            for row in connection.execute("PRAGMA table_info(registration_event_statuses)").fetchall()
        }
        if "canceled_at" not in registration_event_columns:
            connection.execute("ALTER TABLE registration_event_statuses ADD COLUMN canceled_at TEXT")

        registration_email_settings_columns = {
            row["name"]
            for row in connection.execute("PRAGMA table_info(registration_event_email_settings)").fetchall()
        }
        if "event_date_2" not in registration_email_settings_columns:
            connection.execute("ALTER TABLE registration_event_email_settings ADD COLUMN event_date_2 TEXT")

        agenda_training_columns = {
            row["name"]
            for row in connection.execute("PRAGMA table_info(agenda_trainings)").fetchall()
        }
        if "trainers_json" not in agenda_training_columns:
            connection.execute("ALTER TABLE agenda_trainings ADD COLUMN trainers_json TEXT NOT NULL DEFAULT '[]'")
        if "training_type" not in agenda_training_columns:
            connection.execute("ALTER TABLE agenda_trainings ADD COLUMN training_type TEXT NOT NULL DEFAULT 'samenwerkende_amateurclub'")
        if "status" not in agenda_training_columns:
            connection.execute("ALTER TABLE agenda_trainings ADD COLUMN status TEXT NOT NULL DEFAULT 'gepland'")

        budget_line_columns = {
            row["name"]
            for row in connection.execute("PRAGMA table_info(budget_lines)").fetchall()
        }
        if "trainer_group" not in budget_line_columns:
            connection.execute("ALTER TABLE budget_lines ADD COLUMN trainer_group TEXT NOT NULL DEFAULT ''")

        football_playbook_columns = {
            row["name"]
            for row in connection.execute("PRAGMA table_info(football_days_playbooks)").fetchall()
        }
        for column_name, column_definition in (
            ("playbook_type", "playbook_type TEXT NOT NULL DEFAULT 'voetbaldagen'"),
            ("ecwid_product_id", "ecwid_product_id TEXT"),
            ("ecwid_product_name", "ecwid_product_name TEXT"),
            ("ecwid_product_sku", "ecwid_product_sku TEXT"),
            ("cycle_number", "cycle_number TEXT"),
            ("cycle_start_date", "cycle_start_date TEXT"),
            ("cycle_end_date", "cycle_end_date TEXT"),
            ("field_layout_json", "field_layout_json TEXT NOT NULL DEFAULT '[]'"),
            ("field_trainings_json", "field_trainings_json TEXT NOT NULL DEFAULT '[]'"),
            ("cycle_no_training_dates_json", "cycle_no_training_dates_json TEXT NOT NULL DEFAULT '[]'"),
            ("include_staff", "include_staff INTEGER NOT NULL DEFAULT 1"),
            ("include_staff_setup_tasks", "include_staff_setup_tasks INTEGER NOT NULL DEFAULT 1"),
            ("include_program", "include_program INTEGER NOT NULL DEFAULT 1"),
        ):
            if column_name in football_playbook_columns:
                continue
            try:
                connection.execute(f"ALTER TABLE football_days_playbooks ADD COLUMN {column_definition}")
            except sqlite3.OperationalError as exc:
                if "duplicate column name" not in str(exc).lower():
                    raise

        contract_columns = {
            row["name"]
            for row in connection.execute("PRAGMA table_info(contracts)").fetchall()
        }
        if "cost_lines_json" not in contract_columns:
            connection.execute("ALTER TABLE contracts ADD COLUMN cost_lines_json TEXT NOT NULL DEFAULT '[]'")
        if "club_address" not in contract_columns:
            connection.execute("ALTER TABLE contracts ADD COLUMN club_address TEXT")
        if "training_execution_summary" not in contract_columns:
            connection.execute("ALTER TABLE contracts ADD COLUMN training_execution_summary TEXT")
        if "training_execution_details" not in contract_columns:
            connection.execute("ALTER TABLE contracts ADD COLUMN training_execution_details TEXT")
        if "agenda_attachment_title" not in contract_columns:
            connection.execute("ALTER TABLE contracts ADD COLUMN agenda_attachment_title TEXT")
        if "agenda_attachment_items_json" not in contract_columns:
            connection.execute("ALTER TABLE contracts ADD COLUMN agenda_attachment_items_json TEXT NOT NULL DEFAULT '[]'")

        social_ideas_columns = {
            row["name"]
            for row in connection.execute("PRAGMA table_info(social_media_ideas)").fetchall()
        }
        if "priority" not in social_ideas_columns:
            connection.execute("ALTER TABLE social_media_ideas ADD COLUMN priority TEXT NOT NULL DEFAULT 'Midden'")
        if "is_scheduled" not in social_ideas_columns:
            connection.execute("ALTER TABLE social_media_ideas ADD COLUMN is_scheduled INTEGER NOT NULL DEFAULT 0")

        content_photo_columns = {
            row["name"]
            for row in connection.execute("PRAGMA table_info(content_photos)").fetchall()
        }
        if "original_name" not in content_photo_columns:
            connection.execute("ALTER TABLE content_photos ADD COLUMN original_name TEXT")
        if "content_type" not in content_photo_columns:
            connection.execute("ALTER TABLE content_photos ADD COLUMN content_type TEXT")
        if "file_size" not in content_photo_columns:
            connection.execute("ALTER TABLE content_photos ADD COLUMN file_size INTEGER NOT NULL DEFAULT 0")
        if "storage_backend" not in content_photo_columns:
            connection.execute("ALTER TABLE content_photos ADD COLUMN storage_backend TEXT NOT NULL DEFAULT 'local'")

        exercise_columns = {
            row["name"]
            for row in connection.execute("PRAGMA table_info(exercises)").fetchall()
        }
        if "title" not in exercise_columns:
            connection.execute("ALTER TABLE exercises ADD COLUMN title TEXT NOT NULL DEFAULT ''")
            if "name" in exercise_columns:
                connection.execute("UPDATE exercises SET title = name WHERE title = ''")
        if "duration" not in exercise_columns:
            connection.execute("ALTER TABLE exercises ADD COLUMN duration TEXT")
        if "training_exercise" not in exercise_columns:
            connection.execute("ALTER TABLE exercises ADD COLUMN training_exercise TEXT")
        if "variation_easier" not in exercise_columns:
            connection.execute("ALTER TABLE exercises ADD COLUMN variation_easier TEXT")
        if "variation_harder" not in exercise_columns:
            connection.execute("ALTER TABLE exercises ADD COLUMN variation_harder TEXT")
        if "dimensions" not in exercise_columns:
            connection.execute("ALTER TABLE exercises ADD COLUMN dimensions TEXT")
        if "materials" not in exercise_columns:
            connection.execute("ALTER TABLE exercises ADD COLUMN materials TEXT")
        if "age_groups_json" not in exercise_columns:
            connection.execute("ALTER TABLE exercises ADD COLUMN age_groups_json TEXT NOT NULL DEFAULT '[]'")
        if "field_json" not in exercise_columns:
            connection.execute("ALTER TABLE exercises ADD COLUMN field_json TEXT NOT NULL DEFAULT '{}'")
        if "source_slide" not in exercise_columns:
            connection.execute("ALTER TABLE exercises ADD COLUMN source_slide INTEGER")
        if "updated_at" not in exercise_columns:
            connection.execute("ALTER TABLE exercises ADD COLUMN updated_at TEXT NOT NULL DEFAULT ''")
            connection.execute("UPDATE exercises SET updated_at = created_at WHERE updated_at = '' AND created_at IS NOT NULL")
        for column_name, column_definition in (
            ("video_url", "video_url TEXT"),
            ("video_remote_path", "video_remote_path TEXT"),
            ("video_file_name", "video_file_name TEXT"),
            ("video_original_name", "video_original_name TEXT"),
            ("video_content_type", "video_content_type TEXT"),
            ("video_file_size", "video_file_size INTEGER NOT NULL DEFAULT 0"),
            ("video_storage_backend", "video_storage_backend TEXT"),
            ("video_uploaded_at", "video_uploaded_at TEXT"),
        ):
            if column_name in exercise_columns:
                continue
            try:
                connection.execute(f"ALTER TABLE exercises ADD COLUMN {column_definition}")
            except sqlite3.OperationalError as exc:
                if "duplicate column name" not in str(exc).lower():
                    raise

        training_session_columns = {
            row["name"]
            for row in connection.execute("PRAGMA table_info(training_sessions)").fetchall()
        }
        if "training_date" not in training_session_columns:
            try:
                connection.execute("ALTER TABLE training_sessions ADD COLUMN training_date TEXT")
            except sqlite3.OperationalError as exc:
                if "duplicate column name" not in str(exc).lower():
                    raise
        if "objective" not in training_session_columns:
            try:
                connection.execute("ALTER TABLE training_sessions ADD COLUMN objective TEXT")
            except sqlite3.OperationalError as exc:
                if "duplicate column name" not in str(exc).lower():
                    raise

        proposal_columns = {
            row["name"]
            for row in connection.execute("PRAGMA table_info(proposals)").fetchall()
        }
        if "total_trainings" not in proposal_columns:
            connection.execute("ALTER TABLE proposals ADD COLUMN total_trainings INTEGER NOT NULL DEFAULT 0")
        if "total_amount" not in proposal_columns:
            connection.execute("ALTER TABLE proposals ADD COLUMN total_amount REAL NOT NULL DEFAULT 0")
        if "updated_at" not in proposal_columns:
            connection.execute("ALTER TABLE proposals ADD COLUMN updated_at TEXT NOT NULL DEFAULT ''")
            connection.execute("UPDATE proposals SET updated_at = created_at WHERE updated_at = ''")

        proposal_line_columns = {
            row["name"]
            for row in connection.execute("PRAGMA table_info(proposal_lines)").fetchall()
        }
        if "line_time" not in proposal_line_columns:
            connection.execute("ALTER TABLE proposal_lines ADD COLUMN line_time TEXT NOT NULL DEFAULT ''")
        if "training_kind" not in proposal_line_columns:
            connection.execute("ALTER TABLE proposal_lines ADD COLUMN training_kind TEXT NOT NULL DEFAULT ''")
        if "training_count" not in proposal_line_columns:
            connection.execute("ALTER TABLE proposal_lines ADD COLUMN training_count INTEGER NOT NULL DEFAULT 0")
        if "sort_order" not in proposal_line_columns:
            connection.execute("ALTER TABLE proposal_lines ADD COLUMN sort_order INTEGER NOT NULL DEFAULT 0")
            connection.execute(
                """
                UPDATE proposal_lines
                SET sort_order = id
                WHERE sort_order = 0
                """
            )

        connection.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_content_photos_album_uploaded
            ON content_photos (album_id, uploaded_at ASC, id ASC)
            """
        )

        duplicate_email = connection.execute(
            """
            SELECT lower(email) AS email_key, COUNT(*) AS total
            FROM trainer_profiles
            GROUP BY lower(email)
            HAVING COUNT(*) > 1
            LIMIT 1
            """
        ).fetchone()
        if duplicate_email is None:
            connection.execute(
                """
                CREATE UNIQUE INDEX IF NOT EXISTS idx_trainer_profiles_email
                ON trainer_profiles (email COLLATE NOCASE)
                """
            )
        connection.execute(
            """
            CREATE UNIQUE INDEX IF NOT EXISTS idx_trainer_profiles_invite_token
            ON trainer_profiles (invite_token)
            """
        )
        connection.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_rate_limit_attempts_lookup
            ON rate_limit_attempts (request_key, created_at)
            """
        )
        connection.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_agenda_day_plans_date
            ON agenda_day_plans (date)
            """
        )
        connection.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_exercises_category_title
            ON exercises (category, title)
            """
        )
        connection.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_proposals_created
            ON proposals (created_at DESC, id DESC)
            """
        )
        connection.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_football_days_playbooks_type_date
            ON football_days_playbooks (playbook_type, event_date, updated_at, id)
            """
        )
        connection.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_planning_documents_date
            ON planning_documents (planning_date DESC, updated_at DESC, id DESC)
            """
        )
        connection.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_proposal_lines_proposal_sort
            ON proposal_lines (proposal_id, sort_order ASC, id ASC)
            """
        )
        connection.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_material_items_sort
            ON material_items (sort_order ASC, id ASC)
            """
        )
        connection.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_material_clubs_sort
            ON material_clubs (sort_order ASC, id ASC)
            """
        )


def table_has_rows(table_name: str) -> bool:
    with get_db_connection() as connection:
        row = connection.execute(f"SELECT 1 FROM {table_name} LIMIT 1").fetchone()
    return row is not None


def migrate_football_days_playbook_to_playbooks() -> None:
    if table_has_rows("football_days_playbooks"):
        return

    with get_db_connection() as connection:
        row = connection.execute(
            """
            SELECT title, event_date, location, staff_json, program_json, contingencies, updated_at
            FROM football_days_playbook
            WHERE id = 1
            """
        ).fetchone()
        if row is None:
            return

        updated_at = str(row["updated_at"] or utcnow_iso()).strip()
        connection.execute(
            """
            INSERT INTO football_days_playbooks (
                title, event_date, location, staff_json, program_json, contingencies, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                str(row["title"] or "Draaiboek Voetbaldagen").strip(),
                str(row["event_date"] or "").strip(),
                str(row["location"] or "").strip(),
                str(row["staff_json"] or "[]"),
                str(row["program_json"] or "[]"),
                str(row["contingencies"] or "").strip(),
                updated_at,
                updated_at,
            ),
        )


def migrate_dashboard_events_json_to_db() -> None:
    if table_has_rows("dashboard_events"):
        return

    events = get_default_dashboard_events()
    if os.path.exists(DASHBOARD_EVENTS_PATH):
        try:
            with open(DASHBOARD_EVENTS_PATH, "r", encoding="utf-8") as config_file:
                data = json.load(config_file)
            if isinstance(data, list) and data:
                events = data
        except (OSError, json.JSONDecodeError):
            pass

    save_dashboard_events_config(events)


def migrate_agenda_trainings_json_to_db() -> None:
    if table_has_rows("agenda_trainings") or not os.path.exists(AGENDA_TRAININGS_PATH):
        return

    try:
        with open(AGENDA_TRAININGS_PATH, "r", encoding="utf-8") as trainings_file:
            data = json.load(trainings_file)
    except (OSError, json.JSONDecodeError):
        return

    if not isinstance(data, list):
        return

    trainings = []
    for item in data:
        if not isinstance(item, dict):
            continue
        trainings.append(
            {
                "id": str(item.get("id", "")).strip() or str(int(time.time() * 1000)),
                "title": str(item.get("title", "")).strip(),
                "date": str(item.get("date", "")).strip(),
                "time": str(item.get("time", "")).strip(),
                "endTime": str(item.get("endTime", "")).strip(),
                "location": str(item.get("location", "")).strip(),
                "notes": str(item.get("notes", "")).strip(),
            }
        )

    if trainings:
        save_agenda_trainings(trainings)


def run_storage_migrations() -> None:
    bootstrap_seed_data_files()
    init_db()
    migrate_football_days_playbook_to_playbooks()
    migrate_dashboard_events_json_to_db()
    migrate_agenda_trainings_json_to_db()
    sync_seed_workspace_data()
    seed_workspace_tables()
    ensure_admin_account()


def load_dashboard_events_config() -> List[Dict[str, Any]]:
    def loader() -> List[Dict[str, Any]]:
        with get_db_connection() as connection:
            rows = connection.execute(
                """
                SELECT product_id, label, match_terms
                FROM dashboard_events
                ORDER BY id ASC
                """
            ).fetchall()

        if not rows:
            return []

        cleaned = []
        for row in rows:
            try:
                match_terms = json.loads(row["match_terms"] or "[]")
            except json.JSONDecodeError:
                match_terms = []
            cleaned.append(
                {
                    "productId": row["product_id"],
                    "label": row["label"] or "Onbekend event",
                    "matchTerms": match_terms if isinstance(match_terms, list) else [],
                }
            )

        return cleaned

    return get_cached_local_data("dashboard_events_config", (), loader)


def normalize_match_text(value: str) -> str:
    return "".join(char.lower() if char.isalnum() else " " for char in value).strip()


def matches_configured_event(item_name: str, configured_event: Dict[str, Any], product_id: Any) -> bool:
    configured_product_id = configured_event.get("productId")
    if configured_product_id is not None:
        if product_id is None:
            return False
        return str(configured_product_id) == str(product_id)

    normalized_item_name = normalize_match_text(item_name)
    item_tokens = {token for token in normalized_item_name.split() if token}

    for raw_term in configured_event.get("matchTerms", []):
        normalized_term = normalize_match_text(str(raw_term))
        term_tokens = {token for token in normalized_term.split() if token}
        if term_tokens and term_tokens.issubset(item_tokens):
            return True

    return False


def save_dashboard_events_config(events: List[Dict[str, Any]]) -> None:
    with get_db_connection() as connection:
        connection.execute("DELETE FROM dashboard_events")
        connection.executemany(
            """
            INSERT INTO dashboard_events (product_id, label, match_terms)
            VALUES (?, ?, ?)
            """,
            [
                (
                    None if item.get("productId") is None else str(item.get("productId")),
                    str(item.get("label", "")).strip() or "Onbekend event",
                    json.dumps(item.get("matchTerms", []), ensure_ascii=True),
                )
                for item in events
            ],
        )
    clear_local_data_cache()


def normalize_exercise_text(value: Any) -> str:
    normalized = str(value or "").replace("\r", "\n")
    normalized = re.sub(r"[ \t]+", " ", normalized)
    normalized = re.sub(r"\n{3,}", "\n\n", normalized)
    return normalized.strip(" \n\t")


def normalize_exercise_category(value: Any) -> str:
    normalized = normalize_exercise_text(value)
    if not normalized:
        return ""
    compact = re.sub(r"\s+", " ", normalized).strip().upper()
    compact_key = re.sub(r"[^A-Z0-9]+", "", compact)
    if compact in EXERCISE_CATEGORY_ALIASES:
        return EXERCISE_CATEGORY_ALIASES[compact]
    for alias, category in EXERCISE_CATEGORY_ALIASES.items():
        if re.sub(r"[^A-Z0-9]+", "", alias) == compact_key:
            return category
    for category in EXERCISE_CATEGORY_OPTIONS:
        if normalized.lower() == category.lower():
            return category
    return ""


def normalize_exercise_age_groups(value: Any) -> List[str]:
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
            value = parsed if isinstance(parsed, list) else value
        except json.JSONDecodeError:
            value = re.split(r"[,;\s]+", value)
    if not isinstance(value, (list, tuple, set)):
        return []

    allowed_lookup = {age.upper(): age for age in EXERCISE_AGE_GROUP_OPTIONS}
    age_groups: List[str] = []
    for item in value:
        normalized = normalize_exercise_text(item).upper().replace("JO", "O")
        normalized = re.sub(r"[^A-Z0-9]", "", normalized)
        if normalized in allowed_lookup and allowed_lookup[normalized] not in age_groups:
            age_groups.append(allowed_lookup[normalized])
    return age_groups


def normalize_exercise_title_key(value: Any) -> str:
    return re.sub(r"\s+", " ", normalize_exercise_text(value)).casefold()


def is_allowed_exercise_category(value: Any) -> bool:
    return normalize_exercise_category(value) in EXERCISE_CATEGORY_OPTIONS


def extract_pptx_slide_text(slide_root: XmlElementTree.Element) -> List[str]:
    lines: List[str] = []
    for text_node in slide_root.findall(".//a:t", PPTX_XML_NAMESPACES):
        value = normalize_exercise_text(text_node.text or "")
        if value:
            lines.append(value)
    return lines


def parse_exercise_text(lines: List[str]) -> Dict[str, str]:
    label_map = {
        "OEFENING:": "title",
        "TRAININGSOEFENING:": "trainingExercise",
        "ORGANISATIE:": "trainingExercise",
        "DUUR:": "duration",
        "OMSCHRIJVING OEFENING:": "description",
        "MATERIALEN:": "materials",
        "AFMETINGEN:": "dimensions",
        "COACHING:": "coaching",
        "BIJZONDERE SPELREGELS:": "specialRules",
        "VARIATIE MAKKELIJKER MAKEN:": "variationEasier",
        "VARIATIE MOEILIJKER MAKEN:": "variationHarder",
    }
    sections: Dict[str, List[str]] = {value: [] for value in label_map.values()}
    current_key = ""
    labels_by_length = sorted(label_map.items(), key=lambda item: len(item[0]), reverse=True)

    for raw_line in lines:
        line = normalize_exercise_text(raw_line)
        if not line:
            continue
        upper_line = line.upper()
        matched_label = ""
        for label, key in labels_by_length:
            if upper_line == label or upper_line.startswith(label):
                matched_label = label
                current_key = key
                inline_value = normalize_exercise_text(line[len(label):])
                if inline_value:
                    sections[current_key].append(inline_value)
                break
        if matched_label:
            continue
        if current_key:
            sections[current_key].append(line)

    parsed = {key: normalize_exercise_text("\n".join(value)) for key, value in sections.items()}
    duration = parsed.get("duration", "")
    if duration:
        parsed["duration"] = normalize_exercise_text(duration.replace("\n", " "))
    return parsed


def extract_docx_paragraph_text(paragraph: XmlElementTree.Element) -> str:
    return normalize_exercise_text(
        "".join(text_node.text or "" for text_node in paragraph.findall(".//w:t", DOCX_XML_NAMESPACES))
    )


def extract_docx_table_lines(table: XmlElementTree.Element) -> List[str]:
    lines: List[str] = []
    for row in table.findall("w:tr", DOCX_XML_NAMESPACES):
        for cell in row.findall("w:tc", DOCX_XML_NAMESPACES):
            for paragraph in cell.findall("w:p", DOCX_XML_NAMESPACES):
                text = extract_docx_paragraph_text(paragraph)
                if text:
                    lines.append(text)
    return lines


def extract_docx_relationship_targets(archive: zipfile.ZipFile) -> Dict[str, str]:
    try:
        relationships_root = XmlElementTree.fromstring(archive.read("word/_rels/document.xml.rels"))
    except KeyError:
        return {}

    targets: Dict[str, str] = {}
    for relationship in relationships_root.findall("rel:Relationship", DOCX_XML_NAMESPACES):
        relationship_id = str(relationship.get("Id") or "").strip()
        relationship_type = str(relationship.get("Type") or "")
        target = str(relationship.get("Target") or "").strip()
        if not relationship_id or "/image" not in relationship_type or not target:
            continue
        targets[relationship_id] = target if target.startswith("word/") else f"word/{target}"
    return targets


def extract_docx_table_image_data_url(
    archive: zipfile.ZipFile,
    table: XmlElementTree.Element,
    relationship_targets: Dict[str, str],
) -> str:
    image_names: List[str] = []
    for blip in table.findall(".//a:blip", DOCX_XML_NAMESPACES):
        relationship_id = str(blip.get(f"{{{DOCX_XML_NAMESPACES['r']}}}embed") or "").strip()
        target = relationship_targets.get(relationship_id, "")
        if target:
            image_names.append(target)
    if not image_names:
        return ""

    best_name = ""
    best_size = -1
    for image_name in image_names:
        try:
            size = archive.getinfo(image_name).file_size
        except KeyError:
            continue
        if size > best_size:
            best_name = image_name
            best_size = size
    if not best_name:
        return ""

    image_bytes = archive.read(best_name)
    content_type = mimetypes.guess_type(best_name)[0] or "image/png"
    encoded = base64.b64encode(image_bytes).decode("ascii")
    return f"data:{content_type};base64,{encoded}"


def parse_exercises_from_docx(file_bytes: bytes) -> List[Dict[str, Any]]:
    exercises: List[Dict[str, Any]] = []
    current_category = ""
    pending_title = ""

    with zipfile.ZipFile(BytesIO(file_bytes)) as archive:
        relationship_targets = extract_docx_relationship_targets(archive)
        document_root = XmlElementTree.fromstring(archive.read("word/document.xml"))
        body = document_root.find("w:body", DOCX_XML_NAMESPACES)
        if body is None:
            return exercises

        for child in body:
            tag_name = child.tag.rsplit("}", 1)[-1]
            if tag_name == "p":
                text = extract_docx_paragraph_text(child)
                if not text:
                    continue
                category = normalize_exercise_category(text)
                if category:
                    current_category = category
                    pending_title = ""
                else:
                    pending_title = text
                continue

            if tag_name != "tbl":
                continue

            parsed = parse_exercise_text(extract_docx_table_lines(child))
            title = normalize_exercise_text(parsed.get("title") or pending_title)
            if not title:
                continue

            image_data_url = extract_docx_table_image_data_url(archive, child, relationship_targets)
            description = parsed.get("description", "")
            special_rules = parsed.get("specialRules", "")
            if special_rules:
                description = normalize_exercise_text(
                    f"{description}\n\nBijzondere spelregels:\n{special_rules}"
                    if description
                    else f"Bijzondere spelregels:\n{special_rules}"
                )

            exercises.append(
                {
                    "title": title,
                    "category": normalize_exercise_category(current_category),
                    "trainingExercise": parsed.get("trainingExercise", ""),
                    "description": description,
                    "coaching": parsed.get("coaching", ""),
                    "variationEasier": parsed.get("variationEasier", ""),
                    "variationHarder": parsed.get("variationHarder", ""),
                    "dimensions": parsed.get("dimensions", ""),
                    "materials": parsed.get("materials", ""),
                    "field": (
                        {"imageDataUrl": image_data_url}
                        if image_data_url
                        else {"viewBox": [0, 0, PPTX_SLIDE_WIDTH, PPTX_SLIDE_HEIGHT], "elements": []}
                    ),
                    "sourceSlide": None,
                    "sourceLabel": f"Word #{len(exercises) + 1}",
                }
            )
            pending_title = ""

    return exercises


def pptx_shape_bounds(shape: XmlElementTree.Element) -> Optional[Dict[str, float]]:
    xfrm = shape.find(".//a:xfrm", PPTX_XML_NAMESPACES)
    if xfrm is None:
        return None
    offset = xfrm.find("a:off", PPTX_XML_NAMESPACES)
    extent = xfrm.find("a:ext", PPTX_XML_NAMESPACES)
    if offset is None or extent is None:
        return None
    try:
        x = float(offset.get("x") or 0)
        y = float(offset.get("y") or 0)
        width = float(extent.get("cx") or 0)
        height = float(extent.get("cy") or 0)
    except ValueError:
        return None
    if width <= 0 or height <= 0:
        return None
    return {"x": x, "y": y, "width": width, "height": height}


def pptx_shape_fill(shape: XmlElementTree.Element) -> str:
    color = shape.find(".//a:solidFill/a:srgbClr", PPTX_XML_NAMESPACES)
    if color is not None:
        value = str(color.get("val") or "").strip()
        if re.fullmatch(r"[0-9A-Fa-f]{6}", value):
            return f"#{value.upper()}"
    return "#111111"


def pptx_shape_text(shape: XmlElementTree.Element) -> str:
    return normalize_exercise_text(" ".join(text.text or "" for text in shape.findall(".//a:t", PPTX_XML_NAMESPACES)))


def is_exercise_field_shape(bounds: Dict[str, float]) -> bool:
    center_x = bounds["x"] + bounds["width"] / 2
    center_y = bounds["y"] + bounds["height"] / 2
    return (
        EXERCISE_FIELD_MIN_X <= center_x <= EXERCISE_FIELD_MAX_X
        and EXERCISE_FIELD_MIN_Y <= center_y <= EXERCISE_FIELD_MAX_Y
    )


def bounds_center(bounds: Dict[str, float]) -> Tuple[float, float]:
    return bounds["x"] + bounds["width"] / 2, bounds["y"] + bounds["height"] / 2


def bounds_contains(bounds: Dict[str, float], container: Dict[str, float], padding: float = 260000) -> bool:
    center_x, center_y = bounds_center(bounds)
    return (
        container["x"] - padding <= center_x <= container["x"] + container["width"] + padding
        and container["y"] - padding <= center_y <= container["y"] + container["height"] + padding
    )


def find_pptx_field_bounds(slide_root: XmlElementTree.Element) -> List[Dict[str, float]]:
    field_bounds: List[Dict[str, float]] = []
    for shape in slide_root.findall(".//p:sp", PPTX_XML_NAMESPACES):
        bounds = pptx_shape_bounds(shape)
        if bounds is None:
            continue
        if pptx_shape_fill(shape) != "#00B050":
            continue
        if bounds["width"] < 1800000 or bounds["height"] < 1200000:
            continue
        if bounds["width"] > 5000000 or bounds["height"] > 3600000:
            continue
        if pptx_shape_text(shape):
            continue
        field_bounds.append(bounds)

    if field_bounds:
        return field_bounds

    return [
        {
            "x": EXERCISE_FIELD_MIN_X,
            "y": EXERCISE_FIELD_MIN_Y,
            "width": EXERCISE_FIELD_MAX_X - EXERCISE_FIELD_MIN_X,
            "height": EXERCISE_FIELD_MAX_Y - EXERCISE_FIELD_MIN_Y,
        }
    ]


def is_shape_in_exercise_fields(bounds: Dict[str, float], field_bounds: List[Dict[str, float]]) -> bool:
    return any(bounds_contains(bounds, field_bound) for field_bound in field_bounds)


def extract_pptx_field_json(slide_root: XmlElementTree.Element) -> Dict[str, Any]:
    elements: List[Dict[str, Any]] = []
    field_bounds = find_pptx_field_bounds(slide_root)

    for shape in slide_root.findall(".//p:sp", PPTX_XML_NAMESPACES):
        bounds = pptx_shape_bounds(shape)
        if bounds is None or not is_shape_in_exercise_fields(bounds, field_bounds) or pptx_shape_text(shape):
            continue
        if bounds["width"] > 5000000 or bounds["height"] > 3900000:
            continue
        preset = ""
        geometry = shape.find(".//a:prstGeom", PPTX_XML_NAMESPACES)
        if geometry is not None:
            preset = str(geometry.get("prst") or "").strip()
        name_node = shape.find(".//p:cNvPr", PPTX_XML_NAMESPACES)
        name = str(name_node.get("name") or "") if name_node is not None else ""
        kind = "ellipse" if preset == "ellipse" or "Voetbal" in name or "DvW" in name else "rect"
        if "Trapezium" in name or preset in {"trapezoid", "parallelogram"}:
            kind = "cone"
        elements.append(
            {
                "type": kind,
                "x": bounds["x"],
                "y": bounds["y"],
                "width": bounds["width"],
                "height": bounds["height"],
                "fill": pptx_shape_fill(shape),
            }
        )

    for connector in slide_root.findall(".//p:cxnSp", PPTX_XML_NAMESPACES):
        bounds = pptx_shape_bounds(connector)
        if bounds is None or not is_shape_in_exercise_fields(bounds, field_bounds):
            continue
        elements.append(
            {
                "type": "line",
                "x": bounds["x"],
                "y": bounds["y"],
                "width": bounds["width"],
                "height": bounds["height"],
                "fill": pptx_shape_fill(connector),
            }
        )

    if not elements:
        return {"viewBox": [0, 0, PPTX_SLIDE_WIDTH, PPTX_SLIDE_HEIGHT], "elements": []}

    min_x = max(0, min(float(item["x"]) for item in elements) - 220000)
    min_y = max(0, min(float(item["y"]) for item in elements) - 220000)
    max_x = min(PPTX_SLIDE_WIDTH, max(float(item["x"]) + float(item["width"]) for item in elements) + 220000)
    max_y = min(PPTX_SLIDE_HEIGHT, max(float(item["y"]) + float(item["height"]) for item in elements) + 220000)
    return {"viewBox": [min_x, min_y, max_x - min_x, max_y - min_y], "elements": elements[:120]}


def parse_exercises_from_pptx(file_bytes: bytes) -> List[Dict[str, Any]]:
    exercises: List[Dict[str, Any]] = []
    current_category = ""

    with zipfile.ZipFile(BytesIO(file_bytes)) as archive:
        slide_names = sorted(
            [
                name
                for name in archive.namelist()
                if name.startswith("ppt/slides/slide") and name.endswith(".xml")
            ],
            key=lambda name: int(re.search(r"slide(\d+)\.xml$", name).group(1)),
        )

        for slide_name in slide_names:
            slide_number = int(re.search(r"slide(\d+)\.xml$", slide_name).group(1))
            slide_root = XmlElementTree.fromstring(archive.read(slide_name))
            lines = extract_pptx_slide_text(slide_root)
            has_exercise = any(line.upper() == "OEFENING:" for line in lines)
            if not has_exercise:
                category_candidates = [
                    line
                    for line in lines
                    if line and line.upper() not in EXERCISE_TEXT_LABELS and line != "-"
                ]
                if category_candidates:
                    current_category = category_candidates[0]
                continue

            parsed = parse_exercise_text(lines)
            title = normalize_exercise_text(parsed.get("title", ""))
            if not title:
                continue
            exercises.append(
                {
                    "title": title,
                    "category": normalize_exercise_category(current_category),
                    "trainingExercise": parsed.get("trainingExercise", ""),
                    "description": parsed.get("description", ""),
                    "coaching": parsed.get("coaching", ""),
                    "variationEasier": parsed.get("variationEasier", ""),
                    "variationHarder": parsed.get("variationHarder", ""),
                    "dimensions": parsed.get("dimensions", ""),
                    "materials": parsed.get("materials", ""),
                    "field": extract_pptx_field_json(slide_root),
                    "sourceSlide": slide_number,
                }
            )

    return exercises


def load_exercises() -> List[Dict[str, Any]]:
    def loader() -> List[Dict[str, Any]]:
        with get_db_connection() as connection:
            rows = connection.execute(
                """
                SELECT id, title, category, training_exercise, description, coaching,
                       variation_easier, variation_harder, dimensions, materials, age_groups_json, duration, field_json,
                       source_slide, updated_at, video_url, video_remote_path, video_file_name,
                       video_original_name, video_content_type, video_file_size, video_storage_backend,
                       video_uploaded_at
                FROM exercises
                ORDER BY
                    CASE
                        WHEN title GLOB '[A-Za-z]*' THEN 0
                        ELSE 1
                    END,
                    title COLLATE NOCASE,
                    id
                """
            ).fetchall()

        exercises = []
        for row in rows:
            try:
                field = json.loads(str(row["field_json"] or "{}"))
            except json.JSONDecodeError:
                field = {"viewBox": [0, 0, PPTX_SLIDE_WIDTH, PPTX_SLIDE_HEIGHT], "elements": []}
            exercises.append(
                {
                    "id": int(row["id"]),
                    "title": str(row["title"] or "").strip(),
                    "category": normalize_exercise_category(row["category"]),
                    "trainingExercise": str(row["training_exercise"] or "").strip(),
                    "description": str(row["description"] or "").strip(),
                    "coaching": str(row["coaching"] or "").strip(),
                    "variationEasier": str(row["variation_easier"] or "").strip(),
                    "variationHarder": str(row["variation_harder"] or "").strip(),
                    "dimensions": str(row["dimensions"] or "").strip(),
                    "materials": str(row["materials"] or "").strip(),
                    "ageGroups": normalize_exercise_age_groups(row["age_groups_json"]),
                    "duration": str(row["duration"] or "").strip(),
                    "field": field,
                    "sourceSlide": row["source_slide"],
                    "updatedAt": str(row["updated_at"] or "").strip(),
                    "videoUrl": str(row["video_url"] or "").strip(),
                    "videoRemotePath": str(row["video_remote_path"] or "").strip(),
                    "videoFileName": str(row["video_file_name"] or "").strip(),
                    "videoOriginalName": str(row["video_original_name"] or "").strip(),
                    "videoContentType": str(row["video_content_type"] or "").strip(),
                    "videoFileSize": int(row["video_file_size"] or 0),
                    "videoStorageBackend": str(row["video_storage_backend"] or "").strip(),
                    "videoUploadedAt": str(row["video_uploaded_at"] or "").strip(),
                }
            )
        return exercises

    return get_cached_local_data("exercises", (), loader)


def row_to_exercise(row: sqlite3.Row) -> Dict[str, Any]:
    try:
        field = json.loads(str(row["field_json"] or "{}"))
    except json.JSONDecodeError:
        field = {"viewBox": [0, 0, PPTX_SLIDE_WIDTH, PPTX_SLIDE_HEIGHT], "elements": []}
    return {
        "id": int(row["id"]),
        "title": str(row["title"] or "").strip(),
        "category": normalize_exercise_category(row["category"]),
        "trainingExercise": str(row["training_exercise"] or "").strip(),
        "description": str(row["description"] or "").strip(),
        "coaching": str(row["coaching"] or "").strip(),
        "variationEasier": str(row["variation_easier"] or "").strip(),
        "variationHarder": str(row["variation_harder"] or "").strip(),
        "dimensions": str(row["dimensions"] or "").strip(),
        "materials": str(row["materials"] or "").strip(),
        "ageGroups": normalize_exercise_age_groups(row["age_groups_json"]),
        "duration": str(row["duration"] or "").strip(),
        "field": field,
        "sourceSlide": row["source_slide"],
        "updatedAt": str(row["updated_at"] or "").strip(),
        "videoUrl": str(row["video_url"] or "").strip(),
        "videoRemotePath": str(row["video_remote_path"] or "").strip(),
        "videoFileName": str(row["video_file_name"] or "").strip(),
        "videoOriginalName": str(row["video_original_name"] or "").strip(),
        "videoContentType": str(row["video_content_type"] or "").strip(),
        "videoFileSize": int(row["video_file_size"] or 0),
        "videoStorageBackend": str(row["video_storage_backend"] or "").strip(),
        "videoUploadedAt": str(row["video_uploaded_at"] or "").strip(),
    }


def load_exercise_by_id(exercise_id: int) -> Optional[Dict[str, Any]]:
    with get_db_connection() as connection:
        row = connection.execute(
            """
        SELECT id, title, category, training_exercise, description, coaching,
               variation_easier, variation_harder, dimensions, materials, age_groups_json, duration, field_json,
               source_slide, updated_at, video_url, video_remote_path, video_file_name,
               video_original_name, video_content_type, video_file_size, video_storage_backend,
               video_uploaded_at
            FROM exercises
            WHERE id = ?
            """,
            (exercise_id,),
        ).fetchone()
    return row_to_exercise(row) if row is not None else None


def safe_svg_number(value: Any, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def safe_svg_color(value: Any, default: str = "#111111") -> str:
    normalized = str(value or "").strip()
    if re.fullmatch(r"#[0-9A-Fa-f]{6}", normalized):
        return normalized.upper()
    return default


def render_exercise_field_svg(field: Any, label: str = "Veldtekening") -> str:
    default_viewbox = [0.0, 0.0, 100.0, 70.0]
    if not isinstance(field, dict):
        field = {}
    image_data_url = str(field.get("imageDataUrl") or "").strip()
    if image_data_url.startswith("data:image/"):
        label_text = html.escape(str(label or "Veldtekening"), quote=True)
        image_src = html.escape(image_data_url, quote=True)
        return f'<img src="{image_src}" alt="{label_text}" loading="lazy">'

    raw_viewbox = field.get("viewBox")
    raw_elements = field.get("elements")
    if not isinstance(raw_viewbox, list) or len(raw_viewbox) != 4:
        raw_viewbox = default_viewbox
    if not isinstance(raw_elements, list):
        raw_elements = []

    viewbox = [safe_svg_number(value) for value in raw_viewbox]
    if viewbox[2] <= 0 or viewbox[3] <= 0:
        viewbox = default_viewbox

    label_text = html.escape(str(label or "Veldtekening"), quote=True)
    parts = [
        (
            f'<svg viewBox="{viewbox[0]} {viewbox[1]} {viewbox[2]} {viewbox[3]}" '
            f'role="img" aria-label="{label_text}" preserveAspectRatio="xMidYMid meet">'
        ),
        f'<rect x="{viewbox[0]}" y="{viewbox[1]}" width="{viewbox[2]}" height="{viewbox[3]}" fill="#159447"></rect>',
    ]

    for element in raw_elements[:140]:
        if not isinstance(element, dict):
            continue
        x = safe_svg_number(element.get("x"))
        y = safe_svg_number(element.get("y"))
        width = max(1.0, safe_svg_number(element.get("width"), 1.0))
        height = max(1.0, safe_svg_number(element.get("height"), 1.0))
        fill = safe_svg_color(element.get("fill"))
        element_type = str(element.get("type") or "").strip()

        if element_type == "ellipse":
            stroke = "#ffffff" if fill == "#000000" else "#111111"
            parts.append(
                f'<ellipse cx="{x + width / 2}" cy="{y + height / 2}" rx="{width / 2}" ry="{height / 2}" '
                f'fill="{fill}" stroke="{stroke}" stroke-width="9000"></ellipse>'
            )
        elif element_type == "cone":
            points = (
                f"{x + width * 0.18},{y + height} "
                f"{x + width * 0.82},{y + height} "
                f"{x + width * 0.62},{y} "
                f"{x + width * 0.38},{y}"
            )
            parts.append(f'<polygon points="{points}" fill="{fill}" stroke="#111111" stroke-width="9000"></polygon>')
        elif element_type == "line":
            parts.append(
                f'<line x1="{x}" y1="{y}" x2="{x + width}" y2="{y + height}" '
                f'stroke="{fill}" stroke-width="22000" stroke-linecap="round"></line>'
            )
        else:
            stroke = "#ffffff" if fill == "#00B050" else "#111111"
            parts.append(
                f'<rect x="{x}" y="{y}" width="{width}" height="{height}" '
                f'fill="{fill}" stroke="{stroke}" stroke-width="9000"></rect>'
            )

    parts.append("</svg>")
    return "".join(parts)


def normalize_exercise_field_image_upload(upload: Any) -> Tuple[Optional[Dict[str, str]], str]:
    if upload is None or not getattr(upload, "filename", ""):
        return None, "Kies eerst een afbeelding."

    file_bytes = upload.read()
    max_bytes = EXERCISE_FIELD_IMAGE_MAX_UPLOAD_MB * 1024 * 1024
    if not file_bytes:
        return None, "De afbeelding is leeg."
    if len(file_bytes) > max_bytes:
        return None, f"De afbeelding mag maximaal {EXERCISE_FIELD_IMAGE_MAX_UPLOAD_MB} MB zijn."

    filename = str(upload.filename or "").strip().lower()
    content_type = str(getattr(upload, "mimetype", "") or mimetypes.guess_type(filename)[0] or "").strip().lower()
    if content_type not in ALLOWED_IMAGE_EXTENSIONS:
        return None, "Upload een JPG-, PNG-, WebP- of AVIF-afbeelding."
    if not any(filename.endswith(extension) for extension in ALLOWED_IMAGE_EXTENSIONS[content_type]):
        return None, "De bestandsnaam past niet bij dit afbeeldingstype."
    if not validate_image_signature(content_type, file_bytes):
        return None, "Deze afbeelding kon niet worden gevalideerd."

    encoded = base64.b64encode(file_bytes).decode("ascii")
    return {"imageDataUrl": f"data:{content_type};base64,{encoded}"}, ""


def add_exercise_field_svgs(exercises: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    enriched = []
    for exercise in exercises:
        item = dict(exercise)
        item["fieldSvg"] = render_exercise_field_svg(item.get("field"), f"Veldtekening {item.get('title') or ''}".strip())
        enriched.append(item)
    return enriched


def replace_exercises(exercises: List[Dict[str, Any]]) -> None:
    now = utcnow_iso()
    with get_db_connection() as connection:
        exercise_columns = {
            row["name"]
            for row in connection.execute("PRAGMA table_info(exercises)").fetchall()
        }
        has_legacy_name = "name" in exercise_columns
        insert_columns = [
            "title",
            "category",
            "duration",
            "training_exercise",
            "description",
            "coaching",
            "variation_easier",
            "variation_harder",
            "dimensions",
            "materials",
            "age_groups_json",
            "field_json",
            "source_slide",
            "updated_at",
        ]
        if has_legacy_name:
            insert_columns.insert(0, "name")
        has_legacy_created_at = "created_at" in exercise_columns
        if has_legacy_created_at:
            insert_columns.append("created_at")

        placeholders = ", ".join("?" for _ in insert_columns)
        column_sql = ", ".join(insert_columns)
        connection.execute("DELETE FROM exercises")
        connection.executemany(
            f"INSERT INTO exercises ({column_sql}) VALUES ({placeholders})",
            [
                tuple(
                    [normalize_exercise_text(item.get("title"))] if has_legacy_name else []
                )
                + (
                    normalize_exercise_text(item.get("title")),
                    normalize_exercise_category(item.get("category")),
                    "",
                    normalize_exercise_text(item.get("trainingExercise")),
                    normalize_exercise_text(item.get("description")),
                    normalize_exercise_text(item.get("coaching")),
                    normalize_exercise_text(item.get("variationEasier")),
                    normalize_exercise_text(item.get("variationHarder")),
                    normalize_exercise_text(item.get("dimensions")),
                    normalize_exercise_text(item.get("materials")),
                    json.dumps(normalize_exercise_age_groups(item.get("ageGroups")), ensure_ascii=True),
                    json.dumps(item.get("field") or {}, ensure_ascii=True),
                    item.get("sourceSlide"),
                    now,
                )
                + ((now,) if has_legacy_created_at else ())
                for item in exercises
                if normalize_exercise_text(item.get("title"))
            ],
        )
    clear_local_data_cache()


def insert_exercises(exercises: List[Dict[str, Any]]) -> int:
    now = utcnow_iso()
    with get_db_connection() as connection:
        exercise_columns = {
            row["name"]
            for row in connection.execute("PRAGMA table_info(exercises)").fetchall()
        }
        has_legacy_name = "name" in exercise_columns
        has_legacy_created_at = "created_at" in exercise_columns
        insert_columns = [
            "title",
            "category",
            "duration",
            "training_exercise",
            "description",
            "coaching",
            "variation_easier",
            "variation_harder",
            "dimensions",
            "materials",
            "age_groups_json",
            "field_json",
            "source_slide",
            "updated_at",
        ]
        if has_legacy_name:
            insert_columns.insert(0, "name")
        if has_legacy_created_at:
            insert_columns.append("created_at")

        placeholders = ", ".join("?" for _ in insert_columns)
        column_sql = ", ".join(insert_columns)
        existing_title_keys = {
            normalize_exercise_title_key(row["title"])
            for row in connection.execute("SELECT title FROM exercises").fetchall()
            if normalize_exercise_title_key(row["title"])
        }
        seen_title_keys: Set[str] = set()
        rows = []
        for item in exercises:
            title = normalize_exercise_text(item.get("title"))
            title_key = normalize_exercise_title_key(title)
            if not title or title_key in existing_title_keys or title_key in seen_title_keys:
                continue
            seen_title_keys.add(title_key)
            rows.append(
                tuple([title] if has_legacy_name else [])
                + (
                    title,
                    normalize_exercise_category(item.get("category")),
                    "",
                    normalize_exercise_text(item.get("trainingExercise")),
                    normalize_exercise_text(item.get("description")),
                    normalize_exercise_text(item.get("coaching")),
                    normalize_exercise_text(item.get("variationEasier")),
                    normalize_exercise_text(item.get("variationHarder")),
                    normalize_exercise_text(item.get("dimensions")),
                    normalize_exercise_text(item.get("materials")),
                    json.dumps(normalize_exercise_age_groups(item.get("ageGroups")), ensure_ascii=True),
                    json.dumps(item.get("field") or {}, ensure_ascii=True),
                    item.get("sourceSlide"),
                    now,
                )
                + ((now,) if has_legacy_created_at else ())
            )
        if not rows:
            return 0
        connection.executemany(f"INSERT INTO exercises ({column_sql}) VALUES ({placeholders})", rows)
    clear_local_data_cache()
    return len(rows)


def filter_importable_exercises(exercises: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], int]:
    with get_db_connection() as connection:
        existing_title_keys = {
            normalize_exercise_title_key(row["title"])
            for row in connection.execute("SELECT title FROM exercises").fetchall()
            if normalize_exercise_title_key(row["title"])
        }

    importable: List[Dict[str, Any]] = []
    seen_title_keys: Set[str] = set()
    skipped_count = 0
    for exercise in exercises:
        title_key = normalize_exercise_title_key(exercise.get("title"))
        if not title_key or title_key in existing_title_keys or title_key in seen_title_keys:
            skipped_count += 1
            continue
        seen_title_keys.add(title_key)
        importable.append(exercise)
    return importable, skipped_count


def get_exercise_import_preview_path(preview_id: str) -> str:
    safe_id = re.sub(r"[^a-zA-Z0-9_-]", "", str(preview_id or ""))[:80]
    if not safe_id:
        raise ValueError("Ongeldige preview.")
    return os.path.join(EXERCISE_IMPORT_PREVIEW_DIR, f"{safe_id}.json")


def save_exercise_import_preview(exercises: List[Dict[str, Any]]) -> str:
    os.makedirs(EXERCISE_IMPORT_PREVIEW_DIR, exist_ok=True)
    preview_id = secrets.token_urlsafe(18)
    with open(get_exercise_import_preview_path(preview_id), "w", encoding="utf-8") as preview_file:
        json.dump(exercises, preview_file, ensure_ascii=True)
    return preview_id


def load_exercise_import_preview(preview_id: str) -> List[Dict[str, Any]]:
    path = get_exercise_import_preview_path(preview_id)
    with open(path, "r", encoding="utf-8") as preview_file:
        data = json.load(preview_file)
    if not isinstance(data, list):
        raise ValueError("Ongeldige preview.")
    return [item for item in data if isinstance(item, dict)]


def save_existing_exercise_import_preview(preview_id: str, exercises: List[Dict[str, Any]]) -> None:
    with open(get_exercise_import_preview_path(preview_id), "w", encoding="utf-8") as preview_file:
        json.dump(exercises, preview_file, ensure_ascii=True)


def clear_exercise_import_preview(preview_id: str) -> None:
    try:
        os.remove(get_exercise_import_preview_path(preview_id))
    except OSError:
        pass


def apply_submitted_exercise_import_edits(preview_exercises: List[Dict[str, Any]]) -> None:
    text_fields = {
        "title": "title",
        "training_exercise": "trainingExercise",
        "description": "description",
        "coaching": "coaching",
        "variation_easier": "variationEasier",
        "variation_harder": "variationHarder",
        "dimensions": "dimensions",
        "materials": "materials",
    }
    for index, preview_exercise in enumerate(preview_exercises):
        submitted_category = request.form.get(f"category_{index}")
        if submitted_category is not None:
            preview_exercise["category"] = normalize_exercise_category(submitted_category)
        for form_key, exercise_key in text_fields.items():
            submitted_value = request.form.get(f"{form_key}_{index}")
            if submitted_value is None:
                continue
            preview_exercise[exercise_key] = normalize_exercise_text(submitted_value)


def update_exercise_category(exercise_id: Any, category: Any) -> bool:
    try:
        normalized_id = int(exercise_id)
    except (TypeError, ValueError):
        return False
    normalized_category = normalize_exercise_category(category)
    if normalized_category not in EXERCISE_CATEGORY_OPTIONS:
        return False

    with get_db_connection() as connection:
        cursor = connection.execute(
            """
            UPDATE exercises
            SET category = ?, updated_at = ?
            WHERE id = ?
            """,
            (normalized_category, utcnow_iso(), normalized_id),
        )
    if cursor.rowcount > 0:
        clear_local_data_cache()
        return True
    return False


def update_exercise(exercise_id: Any, payload: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    try:
        normalized_id = int(exercise_id)
    except (TypeError, ValueError):
        return None

    existing_exercise = load_exercise_by_id(normalized_id)
    if existing_exercise is None:
        return None

    title = normalize_exercise_text(payload.get("title"))[:180] or str(existing_exercise.get("title") or "").strip()
    category = normalize_exercise_category(payload.get("category"))
    if category not in EXERCISE_CATEGORY_OPTIONS:
        category = normalize_exercise_category(existing_exercise.get("category"))
    if category not in EXERCISE_CATEGORY_OPTIONS:
        category = EXERCISE_CATEGORY_OPTIONS[0]
    if not title or category not in EXERCISE_CATEGORY_OPTIONS:
        return None

    cleaned = {
        "title": title,
        "category": category,
        "trainingExercise": normalize_exercise_text(payload.get("trainingExercise", existing_exercise.get("trainingExercise"))),
        "description": normalize_exercise_text(payload.get("description")),
        "coaching": normalize_exercise_text(payload.get("coaching")),
        "variationEasier": normalize_exercise_text(payload.get("variationEasier")),
        "variationHarder": normalize_exercise_text(payload.get("variationHarder")),
        "dimensions": normalize_exercise_text(payload.get("dimensions")),
        "materials": normalize_exercise_text(payload.get("materials")),
        "ageGroups": normalize_exercise_age_groups(payload.get("ageGroups")),
    }

    with get_db_connection() as connection:
        cursor = connection.execute(
            """
            UPDATE exercises
            SET title = ?,
                category = ?,
                training_exercise = ?,
                description = ?,
                coaching = ?,
                variation_easier = ?,
                variation_harder = ?,
                dimensions = ?,
                materials = ?,
                age_groups_json = ?,
                updated_at = ?
            WHERE id = ?
            """,
            (
                cleaned["title"],
                cleaned["category"],
                cleaned["trainingExercise"],
                cleaned["description"],
                cleaned["coaching"],
                cleaned["variationEasier"],
                cleaned["variationHarder"],
                cleaned["dimensions"],
                cleaned["materials"],
                json.dumps(cleaned["ageGroups"], ensure_ascii=True),
                utcnow_iso(),
                normalized_id,
            ),
        )
    if cursor.rowcount <= 0:
        return None
    clear_local_data_cache()
    return load_exercise_by_id(normalized_id)


def update_exercise_field_image(exercise_id: Any, field: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    try:
        normalized_id = int(exercise_id)
    except (TypeError, ValueError):
        return None
    if not isinstance(field, dict) or not str(field.get("imageDataUrl") or "").startswith("data:image/"):
        return None

    exercise = load_exercise_by_id(normalized_id)
    if exercise is None:
        return None
    current_field = dict(exercise.get("field") or {})
    current_field["imageDataUrl"] = str(field.get("imageDataUrl") or "")
    current_field["imageLayer"] = {"x": 50.0, "y": 50.0, "size": 100.0}
    if not isinstance(current_field.get("viewBox"), list) or len(current_field.get("viewBox") or []) != 4:
        current_field["viewBox"] = [0, 0, 100, 70]
    if not isinstance(current_field.get("elements"), list):
        current_field["elements"] = []
    if not isinstance(current_field.get("overlayItems"), list):
        current_field["overlayItems"] = []

    now = utcnow_iso()
    with get_db_connection() as connection:
        cursor = connection.execute(
            """
            UPDATE exercises
            SET field_json = ?, updated_at = ?
            WHERE id = ?
            """,
            (json.dumps(current_field, ensure_ascii=True), now, normalized_id),
        )
    if cursor.rowcount <= 0:
        return None
    clear_local_data_cache()
    return load_exercise_by_id(normalized_id)


def sanitize_exercise_field_overlay_item(item: Any) -> Optional[Dict[str, Any]]:
    if not isinstance(item, dict):
        return None
    item_type = str(item.get("type") or "").strip()
    if item_type == "cone":
        item_type = "small-cone"
    if item_type not in {"player", "big-cone", "small-cone", "goal", "ball", "line", "arrow", "text"}:
        return None

    def pct(value: Any, default: float = 50.0) -> float:
        try:
            numeric = float(value)
        except (TypeError, ValueError):
            numeric = default
        return min(100.0, max(0.0, numeric))

    cleaned = {
        "id": str(item.get("id") or secrets.token_urlsafe(8))[:40],
        "type": item_type,
        "x": pct(item.get("x")),
        "y": pct(item.get("y")),
        "color": safe_svg_color(item.get("color"), "#111111"),
    }
    try:
        size = float(item.get("size", 100))
    except (TypeError, ValueError):
        size = 100.0
    cleaned["size"] = min(220.0, max(45.0, size))
    if item_type in {"line", "arrow"}:
        cleaned["x2"] = pct(item.get("x2"), cleaned["x"] + 12)
        cleaned["y2"] = pct(item.get("y2"), cleaned["y"])
    if item_type == "text":
        text_value = normalize_exercise_text(item.get("text"))[:80]
        cleaned["text"] = text_value or "Tekst"
    return cleaned


def sanitize_exercise_field_image_layer(layer: Any) -> Optional[Dict[str, float]]:
    if not isinstance(layer, dict):
        return None

    def pct(value: Any, default: float = 50.0) -> float:
        try:
            numeric = float(value)
        except (TypeError, ValueError):
            numeric = default
        return min(100.0, max(0.0, numeric))

    try:
        size = float(layer.get("size", 100))
    except (TypeError, ValueError):
        size = 100.0
    return {
        "x": pct(layer.get("x")),
        "y": pct(layer.get("y")),
        "size": min(180.0, max(45.0, size)),
    }


def update_exercise_field_overlay(exercise_id: Any, payload: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    try:
        normalized_id = int(exercise_id)
    except (TypeError, ValueError):
        return None
    exercise = load_exercise_by_id(normalized_id)
    if exercise is None:
        return None

    field = dict(exercise.get("field") or {})
    if not isinstance(field.get("viewBox"), list) or len(field.get("viewBox") or []) != 4:
        field["viewBox"] = [0, 0, 100, 70]
    if not isinstance(field.get("elements"), list):
        field["elements"] = []
    raw_items = payload.get("overlayItems", [])
    if not isinstance(raw_items, list):
        raw_items = []
    field["overlayItems"] = [
        item
        for item in (sanitize_exercise_field_overlay_item(raw_item) for raw_item in raw_items[:160])
        if item is not None
    ]
    image_data_url = str(payload.get("imageDataUrl", field.get("imageDataUrl") or "") or "").strip()
    if image_data_url.startswith("data:image/"):
        field["imageDataUrl"] = image_data_url
        field["imageLayer"] = sanitize_exercise_field_image_layer(payload.get("imageLayer")) or sanitize_exercise_field_image_layer(field.get("imageLayer")) or {"x": 50.0, "y": 50.0, "size": 100.0}
    else:
        field.pop("imageDataUrl", None)
        field.pop("imageLayer", None)

    try:
        background_opacity = float(payload.get("backgroundOpacity", field.get("backgroundOpacity", 1)))
    except (TypeError, ValueError):
        background_opacity = 1.0
    field["backgroundOpacity"] = min(1.0, max(0.15, background_opacity))

    now = utcnow_iso()
    with get_db_connection() as connection:
        cursor = connection.execute(
            """
            UPDATE exercises
            SET field_json = ?, updated_at = ?
            WHERE id = ?
            """,
            (json.dumps(field, ensure_ascii=True), now, normalized_id),
        )
    if cursor.rowcount <= 0:
        return None
    clear_local_data_cache()
    return load_exercise_by_id(normalized_id)


def upload_exercise_video(exercise_id: Any, upload: Any) -> Tuple[Optional[Dict[str, Any]], str]:
    try:
        normalized_id = int(exercise_id)
    except (TypeError, ValueError):
        return None, "Kies eerst een geldige oefening."

    exercise = load_exercise_by_id(normalized_id)
    if exercise is None:
        return None, "Oefening niet gevonden."
    if upload is None or not getattr(upload, "filename", ""):
        return None, "Kies eerst een video."

    original_name = str(upload.filename or "").strip()
    upload_stream = upload.stream
    try:
        upload_stream.seek(0, os.SEEK_END)
        file_size = upload_stream.tell()
        upload_stream.seek(0)
    except (AttributeError, OSError):
        return None, "Deze video kon niet worden gelezen."
    config = get_exercise_video_storage_config()
    max_bytes = config["max_upload_mb"] * 1024 * 1024
    if file_size <= 0:
        return None, "De video is leeg."
    if file_size > max_bytes:
        return None, f"De video mag maximaal {config['max_upload_mb']} MB zijn."

    content_type = str(getattr(upload, "mimetype", "") or mimetypes.guess_type(original_name)[0] or "").strip().lower()
    if content_type not in config["allowed_types"] or content_type not in ALLOWED_VIDEO_EXTENSIONS:
        return None, "Upload een MP4-, WebM- of MOV-video."

    safe_name = sanitize_upload_filename(original_name)
    extension = os.path.splitext(safe_name)[1].lower()
    if not extension:
        extension = (mimetypes.guess_extension(content_type) or ".mp4").lower()
    if extension not in ALLOWED_VIDEO_EXTENSIONS[content_type]:
        return None, "De bestandsnaam past niet bij dit videotype."
    signature_bytes = upload_stream.read(32)
    upload_stream.seek(0)
    if not validate_video_signature(content_type, signature_bytes):
        return None, "Deze video kon niet worden gevalideerd."

    unique_name = f"{int(time.time() * 1000)}-{secrets.token_hex(4)}{extension}"
    remote_path = "/".join(
        [
            config["base_path"],
            f"{exercise['id']}-{slugify_value(exercise['title'])}",
            unique_name,
        ]
    )
    upload_result = upload_content_file(remote_path, upload_stream, file_size, content_type, config=config)
    now = utcnow_iso()
    old_remote_path = str(exercise.get("videoRemotePath") or "").strip()
    old_storage_backend = str(exercise.get("videoStorageBackend") or "").strip()

    with get_db_connection() as connection:
        cursor = connection.execute(
            """
            UPDATE exercises
            SET video_url = ?,
                video_remote_path = ?,
                video_file_name = ?,
                video_original_name = ?,
                video_content_type = ?,
                video_file_size = ?,
                video_storage_backend = ?,
                video_uploaded_at = ?,
                updated_at = ?
            WHERE id = ?
            """,
            (
                upload_result["url"],
                remote_path,
                unique_name,
                original_name,
                content_type,
                file_size,
                upload_result["storage_backend"],
                now,
                now,
                normalized_id,
            ),
        )
    if cursor.rowcount <= 0:
        delete_content_file(remote_path, upload_result["storage_backend"])
        return None, "Oefening niet gevonden."

    if old_remote_path:
        try:
            delete_content_file(old_remote_path, old_storage_backend or "local")
        except requests.RequestException:
            pass
    clear_local_data_cache()
    return load_exercise_by_id(normalized_id), ""


def delete_exercise_video(exercise_id: Any) -> bool:
    try:
        normalized_id = int(exercise_id)
    except (TypeError, ValueError):
        return False
    exercise = load_exercise_by_id(normalized_id)
    if exercise is None:
        return False
    remote_path = str(exercise.get("videoRemotePath") or "").strip()
    storage_backend = str(exercise.get("videoStorageBackend") or "local").strip()
    with get_db_connection() as connection:
        cursor = connection.execute(
            """
            UPDATE exercises
            SET video_url = NULL,
                video_remote_path = NULL,
                video_file_name = NULL,
                video_original_name = NULL,
                video_content_type = NULL,
                video_file_size = 0,
                video_storage_backend = NULL,
                video_uploaded_at = NULL,
                updated_at = ?
            WHERE id = ?
            """,
            (utcnow_iso(), normalized_id),
        )
    if cursor.rowcount <= 0:
        return False
    if remote_path:
        delete_content_file(remote_path, storage_backend or "local")
    clear_local_data_cache()
    return True


def delete_exercise(exercise_id: Any) -> bool:
    try:
        normalized_id = int(exercise_id)
    except (TypeError, ValueError):
        return False
    exercise = load_exercise_by_id(normalized_id)
    if exercise is None:
        return False
    with get_db_connection() as connection:
        cursor = connection.execute("DELETE FROM exercises WHERE id = ?", (normalized_id,))
    if cursor.rowcount > 0:
        remote_path = str(exercise.get("videoRemotePath") or "").strip()
        if remote_path:
            try:
                delete_content_file(remote_path, str(exercise.get("videoStorageBackend") or "local").strip())
            except requests.RequestException:
                pass
        clear_local_data_cache()
    return cursor.rowcount > 0


def normalize_training_session_exercises(items: Any) -> List[Dict[str, Any]]:
    if not isinstance(items, list):
        return []

    exercise_lookup = {int(exercise["id"]): exercise for exercise in load_exercises()}
    cleaned_items: List[Dict[str, Any]] = []
    for position, item in enumerate(items[:80], start=1):
        if not isinstance(item, dict):
            continue
        try:
            exercise_id = int(item.get("exerciseId") or item.get("id") or 0)
        except (TypeError, ValueError):
            continue
        exercise = exercise_lookup.get(exercise_id)
        if exercise is None:
            continue
        cleaned_items.append(
            {
                "exerciseId": exercise_id,
                "title": exercise.get("title", "Oefening"),
                "category": exercise.get("category", ""),
                "trainingExercise": exercise.get("trainingExercise", ""),
                "duration": normalize_exercise_text(item.get("duration") or exercise.get("duration")),
                "notes": normalize_exercise_text(item.get("notes"))[:800],
                "position": position,
            }
        )

    return cleaned_items


def normalize_training_session(row: sqlite3.Row) -> Dict[str, Any]:
    try:
        exercises = json.loads(str(row["exercises_json"] or "[]"))
    except json.JSONDecodeError:
        exercises = []
    if not isinstance(exercises, list):
        exercises = []

    return {
        "id": int(row["id"]),
        "title": str(row["title"] or "").strip(),
        "trainingDate": str(row["training_date"] or "").strip(),
        "objective": str(row["objective"] or "").strip(),
        "notes": str(row["notes"] or "").strip(),
        "exercises": exercises,
        "exerciseCount": len(exercises),
        "createdAt": str(row["created_at"] or "").strip(),
        "updatedAt": str(row["updated_at"] or "").strip(),
    }


def load_training_sessions() -> List[Dict[str, Any]]:
    def loader() -> List[Dict[str, Any]]:
        with get_db_connection() as connection:
            rows = connection.execute(
                """
                SELECT id, title, training_date, objective, notes, exercises_json, created_at, updated_at
                FROM training_sessions
                ORDER BY updated_at DESC, id DESC
                """
            ).fetchall()
        return [normalize_training_session(row) for row in rows]

    return get_cached_local_data("training_sessions", (), loader)


def save_training_session(payload: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    title = normalize_exercise_text(payload.get("title"))[:160]
    exercises = normalize_training_session_exercises(payload.get("exercises"))
    if not title or not exercises:
        return None

    notes = normalize_exercise_text(payload.get("notes"))[:1200]
    objective = normalize_exercise_text(payload.get("objective"))[:1200]
    training_date = str(payload.get("trainingDate") or "").strip()[:10]
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", training_date):
        training_date = ""
    now = utcnow_iso()
    with get_db_connection() as connection:
        cursor = connection.execute(
            """
            INSERT INTO training_sessions (title, training_date, objective, notes, exercises_json, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (title, training_date, objective, notes, json.dumps(exercises, ensure_ascii=True), now, now),
        )
        training_id = int(cursor.lastrowid)
    clear_local_data_cache()

    for training in load_training_sessions():
        if int(training["id"]) == training_id:
            return training
    return None


def load_agenda_trainings(start_date: Optional[str] = None, end_date: Optional[str] = None) -> List[Dict[str, Any]]:
    normalized_start_date = str(start_date or "").strip()
    normalized_end_date = str(end_date or "").strip()

    def loader() -> List[Dict[str, Any]]:
        query = """
            SELECT id, title, date, time, end_time, location, training_type, status, trainers_json, notes
            FROM agenda_trainings
        """
        params: List[str] = []
        conditions: List[str] = []

        if normalized_start_date:
            conditions.append("date >= ?")
            params.append(normalized_start_date)
        if normalized_end_date:
            conditions.append("date <= ?")
            params.append(normalized_end_date)
        if conditions:
            query += "\n        WHERE " + " AND ".join(conditions)
        query += "\n        ORDER BY date ASC, time ASC"

        with get_db_connection() as connection:
            rows = connection.execute(query, params).fetchall()

        trainings = []
        for row in rows:
            try:
                trainers_payload = json.loads(str(row["trainers_json"] or "[]"))
            except (json.JSONDecodeError, KeyError):
                trainers_payload = []
            trainers = [
                {
                    "id": str(item.get("id") or "").strip(),
                    "name": str(item.get("name") or "").strip(),
                }
                for item in trainers_payload
                if isinstance(item, dict) and str(item.get("id") or "").strip() and str(item.get("name") or "").strip()
            ] if isinstance(trainers_payload, list) else []
            type_option = get_agenda_training_type_option(row["training_type"])
            status_option = resolve_agenda_training_status(
                row["status"],
                row["date"],
                row["time"],
                row["end_time"],
            )
            trainings.append(
                {
                    "id": str(row["id"]),
                    "title": str(row["title"] or "").strip(),
                    "date": str(row["date"] or "").strip(),
                    "time": str(row["time"] or "").strip(),
                    "endTime": str(row["end_time"] or "").strip(),
                    "location": str(row["location"] or "").strip(),
                    "trainingType": type_option["value"],
                    "trainingTypeLabel": type_option["label"],
                    "trainingTypeClass": type_option["className"],
                    "status": status_option["value"],
                    "statusLabel": status_option["label"],
                    "statusClass": status_option["className"],
                    "trainers": trainers,
                    "trainerNames": ", ".join(item["name"] for item in trainers),
                    "notes": str(row["notes"] or "").strip(),
                }
            )

        return trainings

    return get_cached_local_data("agenda_trainings", (normalized_start_date, normalized_end_date), loader)


def save_agenda_trainings(trainings: List[Dict[str, Any]]) -> None:
    with get_db_connection() as connection:
        connection.execute("DELETE FROM agenda_trainings")
        connection.executemany(
            """
            INSERT INTO agenda_trainings (id, title, date, time, end_time, location, training_type, status, trainers_json, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    str(item.get("id", "")).strip(),
                    str(item.get("title", "")).strip(),
                    str(item.get("date", "")).strip(),
                    str(item.get("time", "")).strip(),
                    str(item.get("endTime", "")).strip(),
                    str(item.get("location", "")).strip(),
                    get_agenda_training_type_option(item.get("trainingType"))["value"],
                    normalize_agenda_training_status(item.get("status")),
                    json.dumps(item.get("trainers") if isinstance(item.get("trainers"), list) else [], ensure_ascii=False),
                    str(item.get("notes", "")).strip(),
                )
                for item in trainings
                if str(item.get("id", "")).strip()
            ],
        )
    clear_local_data_cache()


def build_agenda_training_signature(training: Dict[str, Any]) -> str:
    trainers = training.get("trainers") if isinstance(training.get("trainers"), list) else []
    normalized_trainers = [
        {
            "id": str(item.get("id") or "").strip(),
            "name": str(item.get("name") or "").strip(),
        }
        for item in trainers
        if isinstance(item, dict) and str(item.get("id") or "").strip()
    ]
    return json.dumps(
        {
            "title": str(training.get("title") or "").strip(),
            "time": str(training.get("time") or "").strip(),
            "endTime": str(training.get("endTime") or "").strip(),
            "location": str(training.get("location") or "").strip(),
            "trainingType": get_agenda_training_type_option(training.get("trainingType"))["value"],
            "status": normalize_agenda_training_status(training.get("status")),
            "trainers": normalized_trainers,
            "notes": str(training.get("notes") or "").strip(),
        },
        ensure_ascii=False,
        sort_keys=True,
    )


def update_agenda_training(
    training_id: str,
    update_scope: str,
    original_signature: str,
    title: str,
    date_value: str,
    time_value: str,
    end_time_value: str,
    location: str,
    training_type: str,
    status: str,
    trainers: List[Dict[str, str]],
    notes: str,
) -> int:
    normalized_training_id = str(training_id or "").strip()
    normalized_scope = str(update_scope or "").strip()
    normalized_title = title.strip()
    normalized_date = str(date_value or "").strip()
    normalized_time = time_value.strip()
    normalized_end_time = end_time_value.strip()
    normalized_location = location.strip()
    normalized_training_type = training_type.strip()
    normalized_status = normalize_agenda_training_status(status)

    if not normalized_training_id or not normalized_title or not normalized_time or not normalized_location or not normalized_training_type:
        return 0
    if normalized_scope != "matching" and parse_iso_date(normalized_date) is None:
        return 0

    trainings = load_agenda_trainings()
    updated_count = 0

    for training in trainings:
        is_target = str(training.get("id") or "").strip() == normalized_training_id
        if normalized_scope == "matching":
            is_target = build_agenda_training_signature(training) == original_signature
        if not is_target:
            continue

        training["title"] = normalized_title
        if normalized_scope != "matching":
            training["date"] = normalized_date
        training["time"] = normalized_time
        training["endTime"] = normalized_end_time
        training["location"] = normalized_location
        training["trainingType"] = normalized_training_type
        training["status"] = normalized_status
        training["trainers"] = trainers
        training["notes"] = notes.strip()
        updated_count += 1

    if updated_count:
        save_agenda_trainings(trainings)
    return updated_count


def delete_agenda_training(
    training_id: str,
    delete_scope: str,
    original_signature: str,
) -> int:
    normalized_training_id = str(training_id or "").strip()
    normalized_scope = str(delete_scope or "").strip()
    normalized_signature = str(original_signature or "").strip()

    if not normalized_training_id:
        return 0
    if normalized_scope == "matching" and not normalized_signature:
        return 0

    trainings = load_agenda_trainings()
    remaining_trainings: List[Dict[str, Any]] = []
    deleted_count = 0

    for training in trainings:
        is_target = str(training.get("id") or "").strip() == normalized_training_id
        if normalized_scope == "matching":
            is_target = build_agenda_training_signature(training) == normalized_signature

        if is_target:
            deleted_count += 1
            continue

        remaining_trainings.append(training)

    if deleted_count:
        save_agenda_trainings(remaining_trainings)
    return deleted_count


def add_agenda_training(
    title: str,
    date_value: str,
    time_value: str,
    end_time_value: str,
    location: str,
    training_type: str,
    trainers: List[Dict[str, str]],
    notes: str,
) -> None:
    trainings = load_agenda_trainings()
    trainings.append(
        {
            "id": str(int(time.time() * 1000)),
            "title": title.strip(),
            "date": date_value.strip(),
            "time": time_value.strip(),
            "endTime": end_time_value.strip(),
            "location": location.strip(),
            "trainingType": training_type.strip(),
            "status": "gepland",
            "trainers": trainers,
            "notes": notes.strip(),
        }
    )
    save_agenda_trainings(trainings)


def add_agenda_trainings_bulk(
    title: str,
    date_values: List[str],
    time_value: str,
    end_time_value: str,
    location: str,
    training_type: str,
    trainers: List[Dict[str, str]],
    notes: str,
) -> int:
    normalized_title = title.strip()
    normalized_time = time_value.strip()
    normalized_end_time = end_time_value.strip()
    normalized_location = location.strip()
    normalized_training_type = training_type.strip()
    valid_dates = []
    seen_dates = set()

    for date_value in date_values:
        normalized_date = str(date_value or "").strip()
        if not normalized_date or normalized_date in seen_dates:
            continue
        if parse_iso_date(normalized_date) is None:
            continue
        seen_dates.add(normalized_date)
        valid_dates.append(normalized_date)

    if not normalized_title or not normalized_time or not normalized_location or not normalized_training_type or not valid_dates:
        return 0

    trainings = load_agenda_trainings()
    base_id = int(time.time() * 1000)
    for index, date_value in enumerate(valid_dates):
        trainings.append(
            {
                "id": str(base_id + index),
                "title": normalized_title,
                "date": date_value,
                "time": normalized_time,
                "endTime": normalized_end_time,
                "location": normalized_location,
                "trainingType": normalized_training_type,
                "status": "gepland",
                "trainers": trainers,
                "notes": notes.strip(),
            }
        )
    save_agenda_trainings(trainings)
    return len(valid_dates)


def is_allowed_agenda_day_plan(plan_type: str) -> bool:
    return str(plan_type or "").strip() in AGENDA_DAY_PLAN_OPTIONS


def normalize_agenda_club(value: Any) -> str:
    normalized_value = str(value or "").strip()
    return normalized_value if normalized_value in AGENDA_CLUB_OPTIONS else ""


def get_agenda_club_class(value: Any) -> str:
    return AGENDA_CLUB_CLASS_NAMES.get(str(value or "").strip(), "")


def get_agenda_training_type_option(value: Any) -> Dict[str, str]:
    normalized_value = str(value or "").strip()
    for option in AGENDA_TRAINING_TYPE_OPTIONS:
        if option["value"] == normalized_value:
            return option
    return AGENDA_TRAINING_TYPE_OPTIONS[2]


def normalize_agenda_training_type(value: Any) -> str:
    normalized_value = str(value or "").strip()
    valid_values = {option["value"] for option in AGENDA_TRAINING_TYPE_OPTIONS}
    return normalized_value if normalized_value in valid_values else ""


def get_agenda_training_status_option(value: Any) -> Dict[str, str]:
    normalized_value = str(value or "").strip().lower()
    for option in AGENDA_TRAINING_STATUS_OPTIONS:
        if option["value"] == normalized_value:
            return option
    return AGENDA_TRAINING_STATUS_OPTIONS[0]


def normalize_agenda_training_status(value: Any) -> str:
    return get_agenda_training_status_option(value)["value"]


def resolve_agenda_training_status(value: Any, date_value: Any, time_value: Any, end_time_value: Any) -> Dict[str, str]:
    explicit_status = get_agenda_training_status_option(value)
    if explicit_status["value"] == "geannuleerd":
        return explicit_status

    normalized_date = str(date_value or "").strip()
    normalized_time = str(time_value or "").strip()
    normalized_end_time = str(end_time_value or "").strip() or compute_default_end_time(normalized_time)
    end_dt = combine_date_and_time(normalized_date, normalized_end_time)
    if end_dt and end_dt <= datetime.now():
        return get_agenda_training_status_option("gegeven")
    return explicit_status


def auto_mark_completed_agenda_trainings() -> int:
    with get_db_connection() as connection:
        rows = connection.execute(
            """
            SELECT id, date, time, end_time, status
            FROM agenda_trainings
            WHERE status = ?
            """,
            ("gepland",),
        ).fetchall()
        completed_ids = []
        for row in rows:
            normalized_time = str(row["time"] or "").strip()
            normalized_end_time = str(row["end_time"] or "").strip() or compute_default_end_time(normalized_time)
            end_dt = combine_date_and_time(str(row["date"] or "").strip(), normalized_end_time)
            if end_dt is not None and end_dt <= datetime.now():
                completed_ids.append(str(row["id"]))

        if not completed_ids:
            return 0

        connection.executemany(
            """
            UPDATE agenda_trainings
            SET status = ?
            WHERE id = ?
            """,
            [("gegeven", training_id) for training_id in completed_ids],
        )

    clear_local_data_cache()
    return len(completed_ids)


def build_agenda_trainer_options() -> List[Dict[str, str]]:
    options = []
    for profile in load_trainer_profiles():
        profile_id = str(profile.get("id") or "").strip()
        full_name = str(profile.get("fullName") or "").strip()
        if profile_id and full_name:
            options.append(
                {
                    "id": profile_id,
                    "name": full_name,
                }
            )
    return options


def normalize_agenda_trainers(trainer_ids: List[str]) -> List[Dict[str, str]]:
    selected_ids = []
    seen_ids = set()
    for trainer_id in trainer_ids:
        normalized_id = str(trainer_id or "").strip()
        if not normalized_id or normalized_id in seen_ids:
            continue
        seen_ids.add(normalized_id)
        selected_ids.append(normalized_id)

    if not selected_ids:
        return []

    trainer_options_by_id = {
        option["id"]: option
        for option in build_agenda_trainer_options()
    }
    return [
        {
            "id": trainer_options_by_id[trainer_id]["id"],
            "name": trainer_options_by_id[trainer_id]["name"],
        }
        for trainer_id in selected_ids
        if trainer_id in trainer_options_by_id
    ]


def load_agenda_day_plans(date_values: List[str]) -> Dict[str, str]:
    normalized_dates = [str(value or "").strip() for value in date_values if str(value or "").strip()]
    if not normalized_dates:
        return {}
    normalized_dates = sorted(set(normalized_dates))

    def loader() -> Dict[str, str]:
        placeholders = ",".join("?" for _ in normalized_dates)
        with get_db_connection() as connection:
            rows = connection.execute(
                f"""
                SELECT date, plan_type
                FROM agenda_day_plans
                WHERE date IN ({placeholders})
                """,
                normalized_dates,
            ).fetchall()

        return {
            str(row["date"] or "").strip(): str(row["plan_type"] or "").strip()
            for row in rows
            if str(row["date"] or "").strip() and str(row["plan_type"] or "").strip()
        }

    return get_cached_local_data("agenda_day_plans", tuple(normalized_dates), loader)


def load_all_agenda_day_plans() -> List[Dict[str, str]]:
    def loader() -> List[Dict[str, str]]:
        with get_db_connection() as connection:
            rows = connection.execute(
                """
                SELECT date, plan_type
                FROM agenda_day_plans
                ORDER BY date
                """
            ).fetchall()

        return [
            {
                "date": str(row["date"] or "").strip(),
                "planType": str(row["plan_type"] or "").strip(),
            }
            for row in rows
            if str(row["date"] or "").strip() and str(row["plan_type"] or "").strip()
        ]

    return get_cached_local_data("all_agenda_day_plans", (), loader)


AGENDA_API_CREDENTIAL_ID = 1
AGENDA_API_TOKEN_PREFIX = "hws_agenda_"
AGENDA_API_UID_DOMAIN = "workspace.hwsvoetbalschool.nl"


def get_agenda_api_signing_secret() -> str:
    return get_secret_env("AGENDA_API_SECRET") or str(settings.SECRET_KEY)


def derive_agenda_api_token(token_salt: str) -> str:
    digest = hmac.new(
        get_agenda_api_signing_secret().encode("utf-8"),
        str(token_salt or "").encode("utf-8"),
        hashlib.sha256,
    ).digest()
    encoded_digest = base64.urlsafe_b64encode(digest).decode("ascii").rstrip("=")
    return f"{AGENDA_API_TOKEN_PREFIX}{encoded_digest}"


def get_agenda_api_credential(create_if_missing: bool = True) -> Optional[Dict[str, str]]:
    with get_db_connection() as connection:
        if create_if_missing:
            now = utcnow_iso()
            connection.execute(
                """
                INSERT OR IGNORE INTO agenda_api_credentials (id, token_salt, created_at, last_used_at)
                VALUES (?, ?, ?, NULL)
                """,
                (AGENDA_API_CREDENTIAL_ID, secrets.token_urlsafe(32), now),
            )
        row = connection.execute(
            """
            SELECT token_salt, created_at, last_used_at
            FROM agenda_api_credentials
            WHERE id = ?
            """,
            (AGENDA_API_CREDENTIAL_ID,),
        ).fetchone()

    if row is None:
        return None
    token_salt = str(row["token_salt"] or "").strip()
    if not token_salt:
        return None
    return {
        "token": derive_agenda_api_token(token_salt),
        "createdAt": str(row["created_at"] or "").strip(),
        "lastUsedAt": str(row["last_used_at"] or "").strip(),
    }


def rotate_agenda_api_credential() -> Dict[str, str]:
    now = utcnow_iso()
    token_salt = secrets.token_urlsafe(32)
    with get_db_connection() as connection:
        connection.execute(
            """
            INSERT INTO agenda_api_credentials (id, token_salt, created_at, last_used_at)
            VALUES (?, ?, ?, NULL)
            ON CONFLICT(id) DO UPDATE SET
                token_salt = excluded.token_salt,
                created_at = excluded.created_at,
                last_used_at = NULL
            """,
            (AGENDA_API_CREDENTIAL_ID, token_salt, now),
        )
    return {
        "token": derive_agenda_api_token(token_salt),
        "createdAt": now,
        "lastUsedAt": "",
    }


def mark_agenda_api_credential_used() -> None:
    with get_db_connection() as connection:
        connection.execute(
            """
            UPDATE agenda_api_credentials
            SET last_used_at = ?
            WHERE id = ?
            """,
            (utcnow_iso(), AGENDA_API_CREDENTIAL_ID),
        )


def get_agenda_api_allowed_origins() -> Set[str]:
    return {
        origin.strip().rstrip("/")
        for origin in get_env("AGENDA_API_ALLOWED_ORIGINS").split(",")
        if origin.strip()
    }


def apply_agenda_api_response_headers(response: Any) -> Any:
    response["Cache-Control"] = "no-store"
    response["X-Robots-Tag"] = "noindex, nofollow"
    origin = str(request.headers.get("Origin", "") or "").strip().rstrip("/")
    if origin and origin in get_agenda_api_allowed_origins():
        response["Access-Control-Allow-Origin"] = origin
        response["Access-Control-Allow-Methods"] = "GET, OPTIONS"
        response["Access-Control-Allow-Headers"] = "Authorization, Content-Type"
        response["Access-Control-Max-Age"] = "3600"
        current_vary = str(response.get("Vary", "") or "").strip()
        if "Origin" not in {item.strip() for item in current_vary.split(",") if item.strip()}:
            response["Vary"] = f"{current_vary}, Origin".strip(", ")
    return response


def agenda_api_json_response(payload: Dict[str, Any], status_code: int = 200) -> Any:
    response = jsonify(payload)
    response.status_code = status_code
    return apply_agenda_api_response_headers(response)


def agenda_api_preflight_response() -> Any:
    origin = str(request.headers.get("Origin", "") or "").strip().rstrip("/")
    if not origin or origin not in get_agenda_api_allowed_origins():
        return agenda_api_json_response(
            {"error": {"code": "origin_not_allowed", "message": "Deze origin is niet toegestaan."}},
            403,
        )
    response = jsonify({})
    response.status_code = 204
    return apply_agenda_api_response_headers(response)


def get_agenda_api_request_token(allow_query_parameter: bool = False) -> str:
    authorization = str(request.headers.get("Authorization", "") or "").strip()
    scheme, separator, value = authorization.partition(" ")
    if separator and scheme.casefold() == "bearer":
        return value.strip()
    if allow_query_parameter:
        return str(request.args.get("token", "") or "").strip()
    return ""


def validate_agenda_api_request(allow_query_parameter: bool = False) -> Optional[Any]:
    credential = get_agenda_api_credential()
    request_token = get_agenda_api_request_token(allow_query_parameter=allow_query_parameter)
    expected_token = str(credential.get("token") or "") if credential else ""
    if not request_token or not expected_token or not hmac.compare_digest(request_token, expected_token):
        response = agenda_api_json_response(
            {
                "error": {
                    "code": "unauthorized",
                    "message": "Een geldige HWS-agenda API-sleutel is vereist.",
                }
            },
            401,
        )
        response["WWW-Authenticate"] = 'Bearer realm="HWS Agenda API"'
        return response
    mark_agenda_api_credential_used()
    return None


def parse_agenda_api_bool(value: Any, default: bool = False) -> bool:
    normalized_value = str(value or "").strip().casefold()
    if not normalized_value:
        return default
    if normalized_value in {"1", "true", "yes", "ja", "on"}:
        return True
    if normalized_value in {"0", "false", "no", "nee", "off"}:
        return False
    return default


def get_agenda_api_timezone() -> ZoneInfo:
    try:
        return ZoneInfo(str(settings.TIME_ZONE or "Europe/Amsterdam"))
    except (KeyError, ValueError):
        return ZoneInfo("Europe/Amsterdam")


def serialize_agenda_api_training(training: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    event_date = str(training.get("date") or "").strip()
    start_time = str(training.get("time") or "").strip()
    if parse_iso_date(event_date) is None or not start_time:
        return None
    end_time = str(training.get("endTime") or "").strip() or compute_default_end_time(start_time)
    try:
        start_local = datetime.fromisoformat(f"{event_date}T{start_time}")
        end_local = datetime.fromisoformat(f"{event_date}T{end_time}")
    except ValueError:
        return None
    if end_local <= start_local:
        end_local = start_local + timedelta(minutes=90)

    agenda_timezone = get_agenda_api_timezone()
    start_local = start_local.replace(tzinfo=agenda_timezone)
    end_local = end_local.replace(tzinfo=agenda_timezone)
    event_id = str(training.get("id") or "").strip()
    if not event_id:
        return None
    status = str(training.get("status") or "gepland").strip()
    trainers = [
        {
            "id": str(trainer.get("id") or "").strip(),
            "name": str(trainer.get("name") or "").strip(),
        }
        for trainer in training.get("trainers") or []
        if isinstance(trainer, dict) and str(trainer.get("id") or "").strip()
    ]
    return {
        "id": event_id,
        "uid": f"agenda-{event_id}@{AGENDA_API_UID_DOMAIN}",
        "title": str(training.get("title") or "").strip(),
        "start": start_local.isoformat(timespec="minutes"),
        "end": end_local.isoformat(timespec="minutes"),
        "allDay": False,
        "timezone": str(settings.TIME_ZONE or "Europe/Amsterdam"),
        "location": str(training.get("location") or "").strip(),
        "type": str(training.get("trainingType") or "").strip(),
        "typeLabel": str(training.get("trainingTypeLabel") or "").strip(),
        "status": status,
        "statusLabel": str(training.get("statusLabel") or "").strip(),
        "cancelled": status == "geannuleerd",
        "trainers": trainers,
        "notes": str(training.get("notes") or "").strip(),
    }


def serialize_agenda_api_day_plan(day_plan: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    event_date = str(day_plan.get("date") or "").strip()
    plan_type = str(day_plan.get("planType") or "").strip()
    parsed_date = parse_iso_date(event_date)
    if parsed_date is None or not plan_type:
        return None
    return {
        "id": f"day-plan-{event_date}",
        "uid": f"agenda-day-plan-{event_date}@{AGENDA_API_UID_DOMAIN}",
        "title": plan_type,
        "start": event_date,
        "end": (parsed_date + timedelta(days=1)).isoformat(),
        "allDay": True,
        "timezone": str(settings.TIME_ZONE or "Europe/Amsterdam"),
        "location": "",
        "type": "dagplanning",
        "typeLabel": "Dagplanning",
        "status": "gepland",
        "statusLabel": "Gepland",
        "cancelled": False,
        "trainers": [],
        "notes": "",
    }


def build_agenda_api_events(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    include_cancelled: bool = True,
    include_day_plans: bool = False,
) -> List[Dict[str, Any]]:
    events = [
        event
        for event in (
            serialize_agenda_api_training(training)
            for training in load_agenda_trainings(start_date, end_date)
        )
        if event is not None and (include_cancelled or not event["cancelled"])
    ]
    if include_day_plans:
        for day_plan in load_all_agenda_day_plans():
            event_date = str(day_plan.get("date") or "").strip()
            if start_date and event_date < start_date:
                continue
            if end_date and event_date > end_date:
                continue
            event = serialize_agenda_api_day_plan(day_plan)
            if event is not None:
                events.append(event)
    return sorted(events, key=lambda event: (str(event["start"]), str(event["id"])))


def escape_ical_text(value: Any) -> str:
    return (
        str(value or "")
        .replace("\\", "\\\\")
        .replace("\r\n", "\n")
        .replace("\r", "\n")
        .replace("\n", "\\n")
        .replace(";", "\\;")
        .replace(",", "\\,")
    )


def fold_ical_line(line: str) -> List[str]:
    folded_lines: List[str] = []
    current = ""
    byte_limit = 75
    for character in line:
        candidate = f"{current}{character}"
        if current and len(candidate.encode("utf-8")) > byte_limit:
            folded_lines.append(current)
            current = f" {character}"
            byte_limit = 75
        else:
            current = candidate
    folded_lines.append(current)
    return folded_lines


def build_agenda_icalendar(events: List[Dict[str, Any]]) -> str:
    utc_timezone = ZoneInfo("UTC")
    generated_at = datetime.now(tz=utc_timezone).strftime("%Y%m%dT%H%M%SZ")
    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//HWS Voetbalschool//Agenda API//NL",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
        "X-WR-CALNAME:HWS Voetbalschool",
        f"X-WR-TIMEZONE:{escape_ical_text(settings.TIME_ZONE)}",
        "REFRESH-INTERVAL;VALUE=DURATION:PT1H",
        "X-PUBLISHED-TTL:PT1H",
    ]
    for event in events:
        lines.extend(["BEGIN:VEVENT", f"UID:{escape_ical_text(event['uid'])}", f"DTSTAMP:{generated_at}"])
        if event.get("allDay"):
            lines.append(f"DTSTART;VALUE=DATE:{str(event['start']).replace('-', '')}")
            lines.append(f"DTEND;VALUE=DATE:{str(event['end']).replace('-', '')}")
        else:
            start_utc = datetime.fromisoformat(str(event["start"])).astimezone(utc_timezone)
            end_utc = datetime.fromisoformat(str(event["end"])).astimezone(utc_timezone)
            lines.append(f"DTSTART:{start_utc.strftime('%Y%m%dT%H%M%SZ')}")
            lines.append(f"DTEND:{end_utc.strftime('%Y%m%dT%H%M%SZ')}")
        description_parts = [
            str(event.get("typeLabel") or "").strip(),
            str(event.get("statusLabel") or "").strip(),
        ]
        trainer_names = ", ".join(
            str(trainer.get("name") or "").strip()
            for trainer in event.get("trainers") or []
            if isinstance(trainer, dict) and str(trainer.get("name") or "").strip()
        )
        if trainer_names:
            description_parts.append(f"Trainers: {trainer_names}")
        if event.get("notes"):
            description_parts.append(str(event["notes"]))
        lines.extend(
            [
                f"SUMMARY:{escape_ical_text(event.get('title'))}",
                f"LOCATION:{escape_ical_text(event.get('location'))}",
                f"DESCRIPTION:{escape_ical_text(chr(10).join(part for part in description_parts if part))}",
                f"STATUS:{'CANCELLED' if event.get('cancelled') else 'CONFIRMED'}",
                "TRANSP:OPAQUE",
                "END:VEVENT",
            ]
        )
    lines.append("END:VCALENDAR")
    return "\r\n".join(
        folded_line
        for line in lines
        for folded_line in fold_ical_line(line)
    ) + "\r\n"


def normalize_agenda_summary_filter(value: Any) -> str:
    normalized_value = str(value or "").strip().lower()
    valid_keys = {str(option.get("key") or "").strip() for option in AGENDA_SUMMARY_FILTER_OPTIONS}
    return normalized_value if normalized_value in valid_keys else "total"


def get_agenda_summary_filter_option(filter_key: str) -> Dict[str, Any]:
    normalized_key = normalize_agenda_summary_filter(filter_key)
    for option in AGENDA_SUMMARY_FILTER_OPTIONS:
        if option.get("key") == normalized_key:
            return option
    return AGENDA_SUMMARY_FILTER_OPTIONS[0]


def filter_agenda_day_plans_for_summary(day_plans: List[Dict[str, Any]], filter_key: str) -> List[Dict[str, Any]]:
    selected_filter = get_agenda_summary_filter_option(filter_key)
    start_date = selected_filter.get("start")
    end_date = selected_filter.get("end")
    if not isinstance(start_date, date) or not isinstance(end_date, date):
        return list(day_plans)

    filtered_day_plans: List[Dict[str, Any]] = []
    for day_plan in day_plans:
        current_date = day_plan.get("date")
        if isinstance(current_date, str):
            current_date = parse_iso_date(current_date.strip())
        if not isinstance(current_date, date):
            continue
        if start_date <= current_date <= end_date:
            filtered_day_plans.append(day_plan)

    return filtered_day_plans


def save_agenda_day_plans(day_plans: Dict[str, str], replace_dates: Optional[List[str]] = None) -> None:
    cleaned_rows: List[Tuple[str, str, str]] = []
    for raw_date, raw_plan_type in day_plans.items():
        date_value = str(raw_date or "").strip()
        plan_type = str(raw_plan_type or "").strip()
        if not date_value:
            continue
        if parse_iso_date(date_value) is None:
            raise ValueError("Ongeldige datum voor dagplanning.")
        if not plan_type:
            continue
        if not is_allowed_agenda_day_plan(plan_type):
            raise ValueError("Ongeldig agendatype gekozen.")
        cleaned_rows.append((date_value, plan_type, utcnow_iso()))

    target_dates = [
        str(value or "").strip()
        for value in (replace_dates or [row[0] for row in cleaned_rows])
        if str(value or "").strip()
    ]
    for date_value in target_dates:
        if parse_iso_date(date_value) is None:
            raise ValueError("Ongeldige datum voor dagplanning.")

    with get_db_connection() as connection:
        if target_dates:
            placeholders = ",".join("?" for _ in target_dates)
            connection.execute(
                f"DELETE FROM agenda_day_plans WHERE date IN ({placeholders})",
                target_dates,
            )
        connection.executemany(
            """
            INSERT INTO agenda_day_plans (date, plan_type, updated_at)
            VALUES (?, ?, ?)
            """,
            cleaned_rows,
        )
    clear_local_data_cache()


def utcnow_iso() -> str:
    return datetime.utcnow().replace(microsecond=0).isoformat()


def save_dashboard_preference(key: str, value: str) -> None:
    with get_db_connection() as connection:
        connection.execute(
            """
            INSERT INTO dashboard_preferences (key, value)
            VALUES (?, ?)
            ON CONFLICT(key) DO UPDATE SET value = excluded.value
            """,
            (key, value),
        )
    clear_local_data_cache()


def load_dashboard_preference(key: str, default: str = "") -> str:
    normalized_key = str(key or "").strip()

    def loader() -> str:
        with get_db_connection() as connection:
            row = connection.execute(
                "SELECT value FROM dashboard_preferences WHERE key = ?",
                (normalized_key,),
            ).fetchone()

        if row is None:
            return default
        return str(row["value"] or default)

    return get_cached_local_data("dashboard_preference", (normalized_key, default), loader)


def normalize_blocked_lead_emails(raw_value: Any) -> str:
    seen_emails: Set[str] = set()
    normalized_emails: List[str] = []
    for email in re.split(r"[\s,;]+", str(raw_value or "")):
        normalized_email = str(email or "").strip().lower()
        if not normalized_email or normalized_email in seen_emails:
            continue
        seen_emails.add(normalized_email)
        normalized_emails.append(normalized_email)
    return "\n".join(normalized_emails)


def load_blocked_lead_emails() -> str:
    return load_dashboard_preference("leads_blocked_emails", "")


def save_blocked_lead_emails(raw_value: Any) -> str:
    normalized_value = normalize_blocked_lead_emails(raw_value)
    save_dashboard_preference("leads_blocked_emails", normalized_value)
    return normalized_value


def load_dashboard_weather_settings() -> Dict[str, str]:
    defaults = {
        "weather_name": "Deventer",
        "weather_lat": "52.25",
        "weather_lon": "6.16",
    }

    def loader() -> Dict[str, str]:
        with get_db_connection() as connection:
            rows = connection.execute(
                "SELECT key, value FROM dashboard_preferences WHERE key IN ('weather_name', 'weather_lat', 'weather_lon')"
            ).fetchall()

        settings = dict(defaults)
        for row in rows:
            settings[str(row["key"])] = str(row["value"])
        return settings

    return get_cached_local_data("dashboard_weather_settings", (), loader)


def load_tasks() -> List[Dict[str, Any]]:
    def loader() -> List[Dict[str, Any]]:
        with get_db_connection() as connection:
            rows = connection.execute(
                """
                SELECT id, title, due_date, is_done, created_at
                FROM tasks
                ORDER BY is_done ASC, due_date ASC, created_at DESC
                """
            ).fetchall()

        return [
            {
                "id": int(row["id"]),
                "title": str(row["title"] or "").strip(),
                "dueDate": str(row["due_date"] or "").strip(),
                "isDone": bool(row["is_done"]),
                "createdAt": str(row["created_at"] or "").strip(),
            }
            for row in rows
        ]

    return get_cached_local_data("tasks", (), loader)


def add_task(title: str, due_date: str) -> None:
    with get_db_connection() as connection:
        connection.execute(
            "INSERT INTO tasks (title, due_date, is_done, created_at) VALUES (?, ?, 0, ?)",
            (title.strip(), due_date.strip(), utcnow_iso()),
        )
    clear_local_data_cache()


def toggle_task(task_id: int) -> None:
    with get_db_connection() as connection:
        connection.execute(
            """
            UPDATE tasks
            SET is_done = CASE WHEN is_done = 1 THEN 0 ELSE 1 END
            WHERE id = ?
            """,
            (task_id,),
        )
    clear_local_data_cache()


def delete_task(task_id: int) -> None:
    with get_db_connection() as connection:
        connection.execute("DELETE FROM tasks WHERE id = ?", (task_id,))
    clear_local_data_cache()


FOOTBALL_ACTIVITY_ICON_RULES: Tuple[Tuple[str, str], ...] = (
    ("ontvangst|aanmelden|inloop|registratie", "clipboard"),
    ("opstarten|kleedkamer|omkleden|shirt", "clipboard"),
    ("warming|warm-up|activatie", "flame"),
    ("training|techniek|oefening|dribbel|passen|partij|wedstrijd|fungames", "football"),
    ("lunch|eten|pauze|drinken", "utensils"),
    ("toernooi|finale|prijs|ceremonie|afsluiting|penalty|bokaal", "trophy"),
    ("foto|media|content", "camera"),
    ("ehbo|blessure|zorg", "medical"),
    ("materiaal|opbouw|afbouw|veld", "cones"),
    ("quiz", "clipboard"),
)
FOOTBALL_ACTIVITY_ICON_KEYS = {icon_key for _, icon_key in FOOTBALL_ACTIVITY_ICON_RULES} | {"clock"}


def infer_football_activity_icon(activity_name: str) -> str:
    normalized_name = activity_name.strip().lower()
    for pattern, icon_key in FOOTBALL_ACTIVITY_ICON_RULES:
        if re.search(pattern, normalized_name):
            return icon_key
    return "clock"


def clamp_float(value: Any, minimum: float, maximum: float, fallback: float) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError):
        number = fallback
    return max(minimum, min(maximum, number))


def normalize_hex_color(value: Any, fallback: str = "#D5EFD3") -> str:
    normalized = str(value or "").strip()
    if re.fullmatch(r"#[0-9A-Fa-f]{6}", normalized):
        return normalized.upper()
    return fallback


def normalize_football_field_layout(layout: Any) -> List[Dict[str, Any]]:
    normalized_layout: List[Dict[str, Any]] = []
    items = layout if isinstance(layout, list) else []
    for index, item in enumerate(items[:80]):
        if not isinstance(item, dict):
            continue
        if str(item.get("type") or "").strip() == "arrow" or str(item.get("id") or "").startswith("field-arrow-"):
            normalized_layout.append(
                {
                    "type": "arrow",
                    "id": str(item.get("id") or f"arrow-{index + 1}").strip()[:80],
                    "x1": round(clamp_float(item.get("x1"), 0.0, 100.0, 22.0), 3),
                    "y1": round(clamp_float(item.get("y1"), 0.0, 100.0, 24.0), 3),
                    "x2": round(clamp_float(item.get("x2"), 0.0, 100.0, 58.0), 3),
                    "y2": round(clamp_float(item.get("y2"), 0.0, 100.0, 34.0), 3),
                    "color": normalize_hex_color(item.get("color"), "#FFFFFF"),
                    "strokeWidth": round(clamp_float(item.get("strokeWidth") or item.get("width"), 2.0, 12.0, 5.0), 3),
                }
            )
            continue
        width = clamp_float(item.get("width"), 8.0, 100.0, 20.0)
        height = clamp_float(item.get("height"), 6.0, 100.0, 14.0)
        x = clamp_float(item.get("x"), 0.0, 100.0 - width, 8.0)
        y = clamp_float(item.get("y"), 0.0, 100.0 - height, 8.0)
        try:
            exercise_id = int(item.get("exerciseId") or 0)
        except (TypeError, ValueError):
            exercise_id = 0
        normalized_layout.append(
            {
                "type": "block",
                "id": str(item.get("id") or f"block-{index + 1}").strip()[:80],
                "x": round(x, 3),
                "y": round(y, 3),
                "width": round(width, 3),
                "height": round(height, 3),
                "title": str(item.get("title") or "").strip()[:120],
                "exerciseId": exercise_id,
                "exerciseTitle": str(item.get("exerciseTitle") or "").strip()[:180],
                "exerciseKind": str(item.get("exerciseKind") or "").strip()[:180],
                "category": str(item.get("category") or "").strip()[:120],
                "exerciseAgeGroups": normalize_exercise_age_groups(item.get("exerciseAgeGroups")),
                "sameExerciseExport": bool(item.get("sameExerciseExport")),
                "sameExerciseKey": str(item.get("sameExerciseKey") or "").strip()[:220],
                "color": normalize_hex_color(item.get("color")),
            }
        )
    return normalized_layout


def normalize_football_field_periods(periods: Any, fallback_layout: Any = None) -> List[Dict[str, Any]]:
    normalized_periods: List[Dict[str, Any]] = []
    items = periods if isinstance(periods, list) else []
    for index, item in enumerate(items[:8]):
        if not isinstance(item, dict):
            continue
        field_layout = normalize_football_field_layout(item.get("fieldLayout"))
        label = str(item.get("label") or item.get("name") or "").strip()
        start_time = str(item.get("startTime") or item.get("start") or "").strip()
        end_time = str(item.get("endTime") or item.get("end") or "").strip()
        if not label and not start_time and not end_time and not field_layout:
            continue
        normalized_periods.append(
            {
                "id": str(item.get("id") or f"field-period-{index + 1}").strip()[:80],
                "label": label[:120] or f"Plattegrond {index + 1}",
                "startTime": start_time[:20],
                "endTime": end_time[:20],
                "fieldLayout": field_layout,
            }
        )
    if not normalized_periods and fallback_layout:
        normalized_periods.append(
            {
                "id": "field-period-1",
                "label": "Plattegrond 1",
                "startTime": "",
                "endTime": "",
                "fieldLayout": normalize_football_field_layout(fallback_layout),
            }
        )
    return normalized_periods


def normalize_football_field_trainings(trainings: Any, fallback_layout: Any = None) -> List[Dict[str, Any]]:
    normalized_trainings: List[Dict[str, Any]] = []
    items = trainings if isinstance(trainings, list) else []
    for index, item in enumerate(items[:24]):
        if not isinstance(item, dict):
            continue
        field_layout = normalize_football_field_layout(item.get("fieldLayout"))
        field_periods = normalize_football_field_periods(item.get("fieldPeriods"), field_layout)
        if field_periods and not field_layout:
            field_layout = field_periods[0]["fieldLayout"]
        name = str(item.get("name") or item.get("title") or "").strip()
        date = str(item.get("date") or "").strip()
        age_groups = normalize_exercise_age_groups(item.get("ageGroups"))
        if not name and not date and not field_layout and not field_periods:
            continue
        normalized_trainings.append(
            {
                "id": str(item.get("id") or f"training-{index + 1}").strip()[:80],
                "name": name[:120] or f"Training {index + 1}",
                "date": date[:40],
                "dateLabel": format_football_days_date(date),
                "ageGroups": age_groups,
                "fieldPeriods": field_periods,
                "fieldLayout": field_layout,
            }
        )

    if not normalized_trainings and fallback_layout:
        normalized_trainings.append(
            {
                "id": "training-1",
                "name": "Training 1",
                "date": "",
                "dateLabel": format_football_days_date(""),
                "ageGroups": [],
                "fieldPeriods": normalize_football_field_periods([], fallback_layout),
                "fieldLayout": normalize_football_field_layout(fallback_layout),
            }
        )
    return normalized_trainings


def normalize_football_no_training_dates(value: Any) -> List[Dict[str, str]]:
    if isinstance(value, str):
        try:
            parsed_value = json.loads(value)
        except json.JSONDecodeError:
            raw_items = []
            matches = list(re.finditer(r"\d{4}-\d{2}-\d{2}", value))
            for match_index, match in enumerate(matches):
                next_start = matches[match_index + 1].start() if match_index + 1 < len(matches) else len(value)
                description = value[match.end() : next_start].strip(" \t\r\n,;:-–—")
                raw_items.append({"date": match.group(0), "description": description})
        else:
            raw_items = parsed_value if isinstance(parsed_value, list) else []
    else:
        raw_items = value if isinstance(value, list) else []

    normalized_dates: List[Dict[str, str]] = []
    seen_dates: Set[str] = set()
    for item in raw_items[:48]:
        raw_date = str(item.get("date") if isinstance(item, dict) else item or "").strip()
        raw_description = str(item.get("description") if isinstance(item, dict) else "").strip()
        if not raw_date:
            continue
        try:
            date_value = datetime.strptime(raw_date[:10], "%Y-%m-%d").date().isoformat()
        except ValueError:
            continue
        if date_value in seen_dates:
            continue
        seen_dates.add(date_value)
        normalized_dates.append(
            {
                "date": date_value,
                "dateLabel": format_football_days_date(date_value),
                "description": raw_description[:160] or "Geen training",
            }
        )
    return sorted(normalized_dates, key=lambda row: row["date"])


def sorted_football_cycle_trainings(trainings: Any) -> List[Dict[str, Any]]:
    items = trainings if isinstance(trainings, list) else []

    def sort_key(item: Tuple[int, Any]) -> Tuple[int, str, int]:
        index, training = item
        training_date = str(training.get("date") or "").strip() if isinstance(training, dict) else ""
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", training_date):
            return (0, training_date, index)
        return (1, "", index)

    return [
        training
        for _index, training in sorted(
            [(index, training) for index, training in enumerate(items) if isinstance(training, dict)],
            key=sort_key,
        )
    ]


def get_football_field_block_sort_key(block: Dict[str, Any], fallback_index: int) -> Tuple[int, int, str, int]:
    title = str(block.get("title") or "").strip()
    normalized_title = normalize_match_text(title)
    order_match = re.match(r"^o\s*0*(\d+)(?:\b|$)", normalized_title)
    if order_match:
        return (0, int(order_match.group(1)), normalized_title, fallback_index)
    return (1, fallback_index, normalized_title, fallback_index)


def sort_football_field_blocks(blocks: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    indexed_blocks = [
        (index, block)
        for index, block in enumerate(blocks)
        if isinstance(block, dict) and block.get("type") != "arrow"
    ]
    return [
        block
        for index, block in sorted(
            indexed_blocks,
            key=lambda item: get_football_field_block_sort_key(item[1], item[0]),
        )
    ]


def normalize_football_days_playbook(row: Optional[sqlite3.Row]) -> Dict[str, Any]:
    if row is None:
        return {
            "id": None,
            "playbookType": "voetbaldagen",
            "title": "Draaiboek Voetbaldagen",
            "eventDate": "",
            "cycleNumber": "",
            "cycleStartDate": "",
            "cycleEndDate": "",
            "location": "",
            "ecwidProductId": "",
            "ecwidProductName": "",
            "ecwidProductSku": "",
            "registrationCount": 0,
            "includeStaff": True,
            "includeStaffSetupTasks": True,
            "includeProgram": True,
            "staff": [],
            "program": [],
            "fieldLayout": [],
            "fieldTrainings": [],
            "cycleNoTrainingDates": [],
            "cycleNoTrainingDatesText": "",
            "contingencies": "",
            "createdAt": "",
            "updatedAt": "",
        }

    try:
        staff = json.loads(str(row["staff_json"] or "[]"))
    except json.JSONDecodeError:
        staff = []
    try:
        program = json.loads(str(row["program_json"] or "[]"))
    except json.JSONDecodeError:
        program = []
    try:
        field_layout = json.loads(str(row["field_layout_json"] or "[]"))
    except (KeyError, IndexError, json.JSONDecodeError):
        field_layout = []
    try:
        field_trainings = json.loads(str(row["field_trainings_json"] or "[]"))
    except (KeyError, IndexError, json.JSONDecodeError):
        field_trainings = []
    try:
        cycle_no_training_dates = json.loads(str(row["cycle_no_training_dates_json"] or "[]"))
    except (KeyError, IndexError, json.JSONDecodeError):
        cycle_no_training_dates = []

    normalized_program = []
    for item in program if isinstance(program, list) else []:
        if not isinstance(item, dict):
            continue
        activity_name = str(item.get("activity") or "").strip()
        if not activity_name:
            continue
        normalized_program.append(
            {
                "startTime": str(item.get("startTime") or "").strip(),
                "endTime": str(item.get("endTime") or "").strip(),
                "activity": activity_name,
                "icon": infer_football_activity_icon(activity_name),
            }
        )

    normalized_staff = []
    for item in staff if isinstance(staff, list) else []:
        if not isinstance(item, dict):
            continue
        name = str(item.get("name") or "").strip()
        role = str(item.get("role") or "").strip()
        setup_task = str(item.get("setupTask") or "").strip()
        if name or role or setup_task:
            normalized_staff.append({"name": name, "role": role, "setupTask": setup_task})

    normalized_field_layout = normalize_football_field_layout(field_layout)
    normalized_field_trainings = normalize_football_field_trainings(field_trainings, normalized_field_layout)
    normalized_no_training_dates = normalize_football_no_training_dates(cycle_no_training_dates)

    return {
        "id": int(row["id"]),
        "playbookType": str(row["playbook_type"] or "voetbaldagen").strip() if "playbook_type" in row.keys() else "voetbaldagen",
        "title": str(row["title"] or "Draaiboek Voetbaldagen").strip(),
        "eventDate": str(row["event_date"] or "").strip(),
        "cycleNumber": str(row["cycle_number"] or "").strip() if "cycle_number" in row.keys() else "",
        "cycleStartDate": str(row["cycle_start_date"] or "").strip() if "cycle_start_date" in row.keys() else "",
        "cycleEndDate": str(row["cycle_end_date"] or "").strip() if "cycle_end_date" in row.keys() else "",
        "location": str(row["location"] or "").strip(),
        "ecwidProductId": str(row["ecwid_product_id"] or "").strip(),
        "ecwidProductName": str(row["ecwid_product_name"] or "").strip(),
        "ecwidProductSku": str(row["ecwid_product_sku"] or "").strip(),
        "registrationCount": 0,
        "includeStaff": bool(row["include_staff"]) if "include_staff" in row.keys() else True,
        "includeStaffSetupTasks": bool(row["include_staff_setup_tasks"]) if "include_staff_setup_tasks" in row.keys() else True,
        "includeProgram": bool(row["include_program"]) if "include_program" in row.keys() else True,
        "staff": normalized_staff,
        "program": normalized_program,
        "fieldLayout": normalized_field_layout,
        "fieldTrainings": normalized_field_trainings,
        "cycleNoTrainingDates": normalized_no_training_dates,
        "cycleNoTrainingDatesText": "\n".join(
            f"{row['date']} - {row['description']}" if row.get("description") and row["description"] != "Geen training" else row["date"]
            for row in normalized_no_training_dates
        ),
        "contingencies": str(row["contingencies"] or "").strip(),
        "createdAt": str(row["created_at"] or "").strip(),
        "updatedAt": str(row["updated_at"] or "").strip(),
    }


def normalize_football_playbook_type(playbook_type: Any) -> str:
    value = str(playbook_type or "voetbaldagen").strip()
    return value if value in FOOTBALL_PLAYBOOK_CONTEXTS else "voetbaldagen"


def get_football_playbook_context(playbook_type: Any) -> Dict[str, str]:
    return FOOTBALL_PLAYBOOK_CONTEXTS[normalize_football_playbook_type(playbook_type)]


def load_football_days_playbooks(playbook_type: str = "voetbaldagen") -> List[Dict[str, Any]]:
    normalized_type = normalize_football_playbook_type(playbook_type)
    with get_db_connection() as connection:
        rows = connection.execute(
            """
            SELECT id, playbook_type, title, event_date, cycle_number, cycle_start_date, cycle_end_date, location, ecwid_product_id, ecwid_product_name, ecwid_product_sku, staff_json, program_json, field_layout_json, field_trainings_json, cycle_no_training_dates_json, contingencies, include_staff, include_staff_setup_tasks, include_program, created_at, updated_at
            FROM football_days_playbooks
            WHERE playbook_type = ?
            ORDER BY COALESCE(NULLIF(event_date, ''), updated_at) DESC, id DESC
            """,
            (normalized_type,),
        ).fetchall()
    return [normalize_football_days_playbook(row) for row in rows]


def load_football_days_playbook(playbook_id: int, playbook_type: str = "voetbaldagen") -> Optional[Dict[str, Any]]:
    normalized_type = normalize_football_playbook_type(playbook_type)
    with get_db_connection() as connection:
        row = connection.execute(
            """
            SELECT id, playbook_type, title, event_date, cycle_number, cycle_start_date, cycle_end_date, location, ecwid_product_id, ecwid_product_name, ecwid_product_sku, staff_json, program_json, field_layout_json, field_trainings_json, cycle_no_training_dates_json, contingencies, include_staff, include_staff_setup_tasks, include_program, created_at, updated_at
            FROM football_days_playbooks
            WHERE id = ? AND playbook_type = ?
            """,
            (playbook_id, normalized_type),
        ).fetchone()
    if row is None:
        return None
    return normalize_football_days_playbook(row)


def duplicate_football_days_playbook(playbook_id: int, playbook_type: str = "voetbaldagen") -> Optional[int]:
    context = get_football_playbook_context(playbook_type)
    playbook = load_football_days_playbook(playbook_id, context["playbookType"])
    if playbook is None:
        return None

    duplicate = copy.deepcopy(playbook)
    duplicate.pop("id", None)
    duplicate["playbookType"] = context["playbookType"]
    duplicate["title"] = f"Kopie van {str(playbook.get('title') or context['defaultTitle']).strip()}"
    duplicate["createdAt"] = ""
    duplicate["updatedAt"] = ""
    return save_football_days_playbook(duplicate, playbook_type=context["playbookType"])


def create_empty_football_days_playbook(playbook_type: str = "voetbaldagen") -> Dict[str, Any]:
    context = get_football_playbook_context(playbook_type)
    playbook = normalize_football_days_playbook(None)
    playbook["playbookType"] = context["playbookType"]
    playbook["title"] = context["defaultTitle"]
    playbook["includeStaff"] = True
    playbook["includeStaffSetupTasks"] = True
    playbook["includeProgram"] = True
    playbook["staff"] = [{"name": "", "role": "", "setupTask": ""}]
    playbook["program"] = [{"startTime": "", "endTime": "", "activity": "", "icon": "clock"}]
    playbook["fieldTrainings"] = [
        {
            "id": "training-1",
            "name": "Training 1",
            "date": "",
            "dateLabel": format_football_days_date(""),
            "ageGroups": [],
            "fieldPeriods": [{"id": "field-period-1", "label": "Plattegrond 1", "startTime": "", "endTime": "", "fieldLayout": []}],
            "fieldLayout": [],
        }
    ]
    return playbook


def count_ecwid_product_registrations(
    orders: List[Dict[str, Any]],
    product_id: str,
    product_name: str = "",
    product_sku: str = "",
) -> int:
    normalized_product_id = str(product_id or "").strip()
    normalized_product_name = normalize_match_text(product_name)
    normalized_product_sku = normalize_match_text(product_sku)
    if not normalized_product_id and not normalized_product_name and not normalized_product_sku:
        return 0

    if normalized_product_id:
        id_count = 0
        for order in orders:
            for item in order.get("items", []):
                if str(item.get("productId") or "").strip() == normalized_product_id:
                    id_count += max(int(item.get("quantity") or 0), 0)
        if id_count:
            return id_count

    registration_count = 0
    for order in orders:
        for item in order.get("items", []):
            item_name = normalize_match_text(item.get("name", ""))
            item_sku = normalize_match_text(item.get("sku", ""))
            is_match = False
            if normalized_product_sku:
                is_match = item_sku == normalized_product_sku
            if not is_match and normalized_product_name:
                is_match = item_name == normalized_product_name
            if is_match:
                registration_count += max(int(item.get("quantity") or 0), 0)
    return registration_count


def get_cached_ecwid_orders_payload() -> Optional[Dict[str, Any]]:
    config_fingerprint = get_external_cache_fingerprint()
    with ecwid_orders_cache_lock:
        cached_payload = ecwid_orders_cache.get("payload")
        cached_at = float(ecwid_orders_cache.get("cached_at") or 0.0)
        cached_fingerprint = ecwid_orders_cache.get("config_fingerprint")

    if cached_payload is None or cached_fingerprint != config_fingerprint:
        return None

    payload = dict(cached_payload)
    payload["cachedAt"] = cached_at
    return payload


def build_football_days_registration_counts(
    playbooks: List[Dict[str, Any]],
    orders: List[Dict[str, Any]],
) -> Dict[int, int]:
    counts: Dict[int, int] = {}
    for playbook in playbooks:
        playbook_id = int(playbook.get("id") or 0)
        if not playbook_id:
            continue
        counts[playbook_id] = count_ecwid_product_registrations(
            orders,
            str(playbook.get("ecwidProductId") or "").strip(),
            str(playbook.get("ecwidProductName") or "").strip(),
            str(playbook.get("ecwidProductSku") or "").strip(),
        )
    return counts


def attach_football_days_registration_counts(
    playbooks: List[Dict[str, Any]],
    *,
    cached_only: bool = False,
) -> List[Dict[str, Any]]:
    product_ids = {
        str(playbook.get("ecwidProductId") or "").strip()
        for playbook in playbooks
        if str(playbook.get("ecwidProductId") or "").strip()
    }
    if not product_ids:
        return playbooks

    try:
        orders_payload = get_cached_ecwid_orders_payload() if cached_only else fetch_ecwid_orders()
    except requests.RequestException:
        return playbooks
    if orders_payload is None:
        return playbooks
    orders = orders_payload.get("items", [])
    counts = build_football_days_registration_counts(playbooks, orders)

    for playbook in playbooks:
        playbook["registrationCount"] = counts.get(int(playbook.get("id") or 0), 0)
    return playbooks


def save_football_days_playbook(
    playbook: Dict[str, Any],
    playbook_id: Optional[int] = None,
    playbook_type: str = "voetbaldagen",
) -> int:
    now = utcnow_iso()
    normalized_type = normalize_football_playbook_type(playbook.get("playbookType") or playbook_type)
    context = get_football_playbook_context(normalized_type)
    payload = (
        normalized_type,
        str(playbook.get("title") or context["defaultTitle"]).strip(),
        str(playbook.get("eventDate") or "").strip(),
        str(playbook.get("cycleNumber") or "").strip() if context["supportsCycleDates"] else "",
        str(playbook.get("cycleStartDate") or "").strip() if context["supportsCycleDates"] else "",
        str(playbook.get("cycleEndDate") or "").strip() if context["supportsCycleDates"] else "",
        str(playbook.get("location") or "").strip(),
        str(playbook.get("ecwidProductId") or "").strip(),
        str(playbook.get("ecwidProductName") or "").strip(),
        str(playbook.get("ecwidProductSku") or "").strip(),
        json.dumps(playbook.get("staff") or [], ensure_ascii=False),
        json.dumps(playbook.get("program") or [], ensure_ascii=False),
        json.dumps(normalize_football_field_layout(playbook.get("fieldLayout")), ensure_ascii=False),
        json.dumps(normalize_football_field_trainings(playbook.get("fieldTrainings")), ensure_ascii=False),
        json.dumps(normalize_football_no_training_dates(playbook.get("cycleNoTrainingDates")), ensure_ascii=False),
        str(playbook.get("contingencies") or "").strip(),
        1 if playbook.get("includeStaff", True) else 0,
        1 if playbook.get("includeStaffSetupTasks", True) or not context["supportsStaffSetupTasks"] else 0,
        1 if playbook.get("includeProgram", True) else 0,
    )

    with get_db_connection() as connection:
        if playbook_id:
            connection.execute(
                """
                UPDATE football_days_playbooks
                SET playbook_type = ?,
                    title = ?,
                    event_date = ?,
                    cycle_number = ?,
                    cycle_start_date = ?,
                    cycle_end_date = ?,
                    location = ?,
                    ecwid_product_id = ?,
                    ecwid_product_name = ?,
                    ecwid_product_sku = ?,
                    staff_json = ?,
                    program_json = ?,
                    field_layout_json = ?,
                    field_trainings_json = ?,
                    cycle_no_training_dates_json = ?,
                    contingencies = ?,
                    include_staff = ?,
                    include_staff_setup_tasks = ?,
                    include_program = ?,
                    updated_at = ?
                WHERE id = ?
                """,
                (*payload, now, playbook_id),
            )
            return playbook_id

        cursor = connection.execute(
            """
            INSERT INTO football_days_playbooks (
                playbook_type, title, event_date, cycle_number, cycle_start_date, cycle_end_date, location, ecwid_product_id, ecwid_product_name, ecwid_product_sku, staff_json, program_json, field_layout_json, field_trainings_json, cycle_no_training_dates_json, contingencies, include_staff, include_staff_setup_tasks, include_program, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (*payload, now, now),
        )
        return int(cursor.lastrowid)


def build_football_days_playbook_from_form(playbook_type: str = "voetbaldagen") -> Dict[str, Any]:
    context = get_football_playbook_context(playbook_type)
    staff_names = request.form.getlist("staff_name")
    staff_roles = request.form.getlist("staff_role")
    staff_setup_tasks = request.form.getlist("staff_setup_task")
    program_starts = request.form.getlist("program_start")
    program_ends = request.form.getlist("program_end")
    program_activities = request.form.getlist("program_activity")

    staff = []
    for index, name in enumerate(staff_names):
        role = staff_roles[index] if index < len(staff_roles) else ""
        setup_task = staff_setup_tasks[index] if index < len(staff_setup_tasks) else ""
        row = {
            "name": str(name or "").strip(),
            "role": str(role or "").strip(),
            "setupTask": str(setup_task or "").strip(),
        }
        if row["name"] or row["role"] or row["setupTask"]:
            staff.append(row)

    program = []
    for index, activity in enumerate(program_activities):
        activity_name = str(activity or "").strip()
        if not activity_name:
            continue
        start_time = program_starts[index] if index < len(program_starts) else ""
        end_time = program_ends[index] if index < len(program_ends) else ""
        program.append(
            {
                "startTime": str(start_time or "").strip(),
                "endTime": str(end_time or "").strip(),
                "activity": activity_name,
                "icon": infer_football_activity_icon(activity_name),
            }
        )

    try:
        field_layout_payload = json.loads(request.form.get("field_layout_json", "[]"))
    except json.JSONDecodeError:
        field_layout_payload = []
    try:
        field_trainings_payload = json.loads(request.form.get("field_trainings_json", "[]"))
    except json.JSONDecodeError:
        field_trainings_payload = []

    include_staff = "1" in request.form.getlist("include_staff")
    include_staff_setup_tasks = "1" in request.form.getlist("include_staff_setup_tasks")
    include_program = "1" in request.form.getlist("include_program")
    if not context["supportsSectionToggles"]:
        include_staff = True
        include_program = True
    if not context["supportsStaffSetupTasks"]:
        include_staff_setup_tasks = True

    return {
        "playbookType": context["playbookType"],
        "title": request.form.get("title", context["defaultTitle"]).strip() or context["defaultTitle"],
        "eventDate": request.form.get("event_date", "").strip(),
        "cycleNumber": request.form.get("cycle_number", "").strip() if context["supportsCycleDates"] else "",
        "cycleStartDate": request.form.get("cycle_start_date", "").strip() if context["supportsCycleDates"] else "",
        "cycleEndDate": request.form.get("cycle_end_date", "").strip() if context["supportsCycleDates"] else "",
        "cycleNoTrainingDates": (
            normalize_football_no_training_dates(request.form.get("cycle_no_training_dates", ""))
            if context["supportsCycleDates"]
            else []
        ),
        "location": request.form.get("location", "").strip(),
        "ecwidProductId": request.form.get("ecwid_product_id", "").strip() if context["showEcwidProduct"] else "",
        "ecwidProductName": request.form.get("ecwid_product_name", "").strip() if context["showEcwidProduct"] else "",
        "ecwidProductSku": request.form.get("ecwid_product_sku", "").strip() if context["showEcwidProduct"] else "",
        "includeStaff": include_staff,
        "includeStaffSetupTasks": include_staff_setup_tasks,
        "includeProgram": include_program,
        "staff": staff,
        "program": program,
        "fieldLayout": normalize_football_field_layout(field_layout_payload),
        "fieldTrainings": normalize_football_field_trainings(field_trainings_payload, field_layout_payload),
        "contingencies": request.form.get("contingencies", "").strip() if context["showContingencies"] else "",
    }


def clean_football_days_club_name(value: Any) -> str:
    cleaned = re.sub(r"\|.*", "", str(value or ""))
    cleaned = re.sub(r"\bdraaiboek\b", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\bvoetbaldag(?:en)?\b", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\bsamenwerkende\s+amateurclubs?\b", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\bhws\b", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\b\d{1,2}[-/]\d{1,2}[-/]\d{2,4}\b", "", cleaned)
    cleaned = re.sub(r"\s{2,}", " ", cleaned).strip()
    return cleaned or "HWS"


def format_football_days_date(value: Any) -> str:
    raw_value = str(value or "").strip()
    if not raw_value:
        return "datum nog in te vullen"
    try:
        parsed_date = datetime.strptime(raw_value, "%Y-%m-%d").date()
    except ValueError:
        return raw_value
    weekday_labels = ["maandag", "dinsdag", "woensdag", "donderdag", "vrijdag", "zaterdag", "zondag"]
    month_labels = [
        "januari",
        "februari",
        "maart",
        "april",
        "mei",
        "juni",
        "juli",
        "augustus",
        "september",
        "oktober",
        "november",
        "december",
    ]
    weekday_label = weekday_labels[parsed_date.weekday()].capitalize()
    return f"{weekday_label} {parsed_date.day} {month_labels[parsed_date.month - 1]} {parsed_date.year}"


def format_football_cycle_date_range(start_value: Any, end_value: Any) -> str:
    start_label = format_football_days_date(start_value) if str(start_value or "").strip() else ""
    end_label = format_football_days_date(end_value) if str(end_value or "").strip() else ""
    if start_label and end_label:
        return f"{start_label} t/m {end_label}"
    return start_label or end_label or "cyclus nog in te vullen"


def normalize_football_days_export_payload(payload: Dict[str, Any], playbook_type: str = "voetbaldagen") -> Dict[str, Any]:
    data = payload if isinstance(payload, dict) else {}
    context = get_football_playbook_context(data.get("playbookType") or playbook_type)
    title = str(data.get("title") or context["defaultTitle"]).strip() or context["defaultTitle"]
    product_id = str(data.get("productId") or data.get("ecwidProductId") or "").strip() if context["showEcwidProduct"] else ""
    product_name = str(data.get("productName") or data.get("ecwidProductName") or "").strip() if context["showEcwidProduct"] else ""
    product_sku = str(data.get("productSku") or data.get("ecwidProductSku") or "").strip() if context["showEcwidProduct"] else ""
    submitted_club_name = clean_football_days_club_name(data.get("clubName"))
    club_name = (
        clean_football_days_club_name(product_name or data.get("location") or title)
        if submitted_club_name == "HWS"
        else submitted_club_name
    )
    registration_count = str(data.get("registrationCount") or "0").strip() or "0"
    include_staff = bool(data.get("includeStaff", True))
    include_staff_setup_tasks = bool(data.get("includeStaffSetupTasks", True))
    include_program = bool(data.get("includeProgram", True))
    if not context["supportsSectionToggles"]:
        include_staff = True
        include_program = True
    if not context["supportsStaffSetupTasks"]:
        include_staff_setup_tasks = True
    if product_id or product_name or product_sku:
        try:
            orders_payload = fetch_ecwid_orders()
            registration_count = str(
                count_ecwid_product_registrations(
                    orders_payload.get("items", []),
                    product_id,
                    product_name,
                    product_sku,
                )
            )
        except requests.RequestException:
            pass

    staff = []
    for item in data.get("staff") if isinstance(data.get("staff"), list) else []:
        if not isinstance(item, dict):
            continue
        row = {
            "name": str(item.get("name") or "").strip(),
            "role": str(item.get("role") or "").strip(),
            "setupTask": str(item.get("setupTask") or "").strip(),
        }
        if row["name"] or row["role"] or row["setupTask"]:
            staff.append(row)

    program = []
    for item in data.get("program") if isinstance(data.get("program"), list) else []:
        if not isinstance(item, dict):
            continue
        activity = str(item.get("activity") or "").strip()
        if not activity:
            continue
        program.append(
            {
                "startTime": str(item.get("startTime") or "").strip(),
                "endTime": str(item.get("endTime") or "").strip(),
                "activity": activity,
                "icon": (
                    str(item.get("icon") or "").strip()
                    if str(item.get("icon") or "").strip() in FOOTBALL_ACTIVITY_ICON_KEYS
                    else infer_football_activity_icon(activity)
                ),
            }
        )

    field_layout = normalize_football_field_layout(data.get("fieldLayout"))
    field_trainings = normalize_football_field_trainings(data.get("fieldTrainings"), field_layout)
    cycle_no_training_dates = normalize_football_no_training_dates(data.get("cycleNoTrainingDates")) if context["supportsCycleDates"] else []
    cycle_no_training_dates_text = str(data.get("cycleNoTrainingDatesText") or "").strip()
    if context["supportsCycleDates"] and cycle_no_training_dates_text:
        text_rows_by_date = {
            row["date"]: row
            for row in normalize_football_no_training_dates(cycle_no_training_dates_text)
            if row.get("description") and row.get("description") != "Geen training"
        }
        for row in cycle_no_training_dates:
            raw_description = str(row.get("description") or "").strip()
            text_row = text_rows_by_date.get(row.get("date"))
            if text_row and raw_description in {"", "-", "Geen training"}:
                row["description"] = text_row["description"]
    exercise_lookup = {int(exercise["id"]): exercise for exercise in load_exercises()}

    def enrich_field_layout(layout: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        for block in layout:
            if block.get("type") == "arrow":
                continue
            exercise = exercise_lookup.get(int(block.get("exerciseId") or 0))
            if exercise is None:
                continue
            block["exerciseTitle"] = exercise.get("title", "")
            block["exerciseKind"] = exercise.get("trainingExercise", "")
            block["category"] = exercise.get("category", "")
            block["exerciseAgeGroups"] = exercise.get("ageGroups", [])
            block["exerciseDetails"] = {
                "description": exercise.get("description", ""),
                "coaching": exercise.get("coaching", ""),
                "variationEasier": exercise.get("variationEasier", ""),
                "variationHarder": exercise.get("variationHarder", ""),
                "dimensions": exercise.get("dimensions", ""),
                "materials": exercise.get("materials", ""),
            }
            block["exerciseField"] = exercise.get("field") if isinstance(exercise.get("field"), dict) else {}
        return layout

    field_layout = enrich_field_layout(field_layout)
    for training in field_trainings:
        field_periods = training.get("fieldPeriods") if isinstance(training.get("fieldPeriods"), list) else []
        enriched_periods = []
        for period in field_periods:
            if not isinstance(period, dict):
                continue
            enriched_period = dict(period)
            enriched_period["fieldLayout"] = enrich_field_layout(period.get("fieldLayout") if isinstance(period.get("fieldLayout"), list) else [])
            enriched_periods.append(enriched_period)
        if enriched_periods:
            training["fieldPeriods"] = enriched_periods
            training["fieldLayout"] = enriched_periods[0]["fieldLayout"]
        else:
            training["fieldLayout"] = enrich_field_layout(training.get("fieldLayout") if isinstance(training.get("fieldLayout"), list) else [])

    cycle_date_range_label = (
        format_football_cycle_date_range(data.get("cycleStartDate"), data.get("cycleEndDate"))
        if context["supportsCycleDates"]
        else ""
    )
    cycle_number = str(data.get("cycleNumber") or "").strip() if context["supportsCycleDates"] else ""
    cycle_cover_label = f"CYCLUS {cycle_number}".strip() if cycle_number else "CYCLUS"
    cover_meta = f"{club_name.upper()} | {registration_count} AANMELDINGEN" if context["showEcwidProduct"] else club_name.upper()
    if context["supportsCycleDates"]:
        cover_meta = f"{club_name.upper()} | {cycle_cover_label} | {cycle_date_range_label}"
    cover_title = title if context["playbookType"] == "samenwerkende-amateurclubs" else context["pdfCoverTitle"]

    return {
        "title": title,
        "playbookType": context["playbookType"],
        "eventDate": str(data.get("eventDate") or "").strip(),
        "eventDateLabel": format_football_days_date(data.get("eventDate")),
        "cycleNumber": cycle_number,
        "cycleCoverLabel": cycle_cover_label,
        "cycleStartDate": str(data.get("cycleStartDate") or "").strip() if context["supportsCycleDates"] else "",
        "cycleEndDate": str(data.get("cycleEndDate") or "").strip() if context["supportsCycleDates"] else "",
        "cycleStartDateLabel": format_football_days_date(data.get("cycleStartDate")) if context["supportsCycleDates"] else "",
        "cycleEndDateLabel": format_football_days_date(data.get("cycleEndDate")) if context["supportsCycleDates"] else "",
        "cycleDateRangeLabel": cycle_date_range_label,
        "cycleNoTrainingDates": cycle_no_training_dates,
        "location": str(data.get("location") or "").strip(),
        "clubName": club_name,
        "coverTitle": cover_title,
        "introSubject": context["introSubject"],
        "coverMeta": cover_meta,
        "includeStaff": include_staff,
        "includeStaffSetupTasks": include_staff_setup_tasks,
        "includeProgram": include_program,
        "staff": staff,
        "program": program,
        "fieldLayout": field_layout,
        "fieldTrainings": field_trainings,
        "contingencies": str(data.get("contingencies") or "").strip() if context["showContingencies"] else "",
        "registrationCount": registration_count,
    }


def football_days_pdf_filename(data: Dict[str, Any]) -> str:
    context = get_football_playbook_context(data.get("playbookType"))
    if context["playbookType"] == "samenwerkende-amateurclubs":
        title = str(data.get("title") or context["defaultTitle"]).strip()
        normalized_title = unicodedata.normalize("NFKD", title).encode("ascii", "ignore").decode("ascii")
        filename_title = re.sub(r"[/\\]+", "-", normalized_title)
        filename_title = re.sub(r'[<>:"|?*\x00-\x1f]+', "", filename_title)
        filename_title = re.sub(r"\s{2,}", " ", filename_title).strip(" .-_")
        return f"{filename_title or 'samenwerkende-amateurclubs-draaiboek'}.pdf"
    date_part = data.get("eventDate") or "draaiboek"
    if context["supportsCycleDates"]:
        date_part = data.get("cycleStartDate") or data.get("eventDate") or "cyclus"
    base = slugify_value(f"{data.get('clubName') or context['playbookType']}-{date_part}")
    return f"{base}.pdf"


def chunk_items(items: List[Any], size: int) -> List[List[Any]]:
    return [items[index : index + size] for index in range(0, len(items), size)] or [[]]


def football_days_background_paths() -> List[str]:
    root = os.path.join(os.path.dirname(__file__), "static", "assets", "football-days-pdf")
    return [os.path.join(root, f"background-{index:02d}.png") for index in range(1, 11)]


def create_football_days_pdf(data: Dict[str, Any]) -> bytes:
    try:
        from reportlab.lib import colors
        from reportlab.lib.utils import ImageReader
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.pdfmetrics import stringWidth
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.pdfgen import canvas
    except ImportError as exc:
        raise RuntimeError("De PDF-library ontbreekt. Installeer de packages uit requirements.txt.") from exc

    font_root = os.path.join(os.path.dirname(__file__), "static", "assets", "fonts")
    pdf_fonts = {
        "regular": "PoppinsPDF",
        "bold": "PoppinsPDF-Bold",
        "extra_bold": "PoppinsPDF-ExtraBold",
        "black": "PoppinsPDF-Black",
    }
    font_files = {
        "regular": "Poppins-Regular.ttf",
        "bold": "Poppins-Bold.ttf",
        "extra_bold": "Poppins-ExtraBold.ttf",
        "black": "Poppins-Black.ttf",
    }
    try:
        registered_fonts = set(pdfmetrics.getRegisteredFontNames())
        for key, font_name in pdf_fonts.items():
            if font_name not in registered_fonts:
                pdfmetrics.registerFont(TTFont(font_name, os.path.join(font_root, font_files[key])))
    except Exception as exc:
        raise RuntimeError("De Poppins-fontbestanden ontbreken of kunnen niet worden geladen.") from exc

    regular_font = pdf_fonts["regular"]
    bold_font = pdf_fonts["bold"]
    extra_bold_font = pdf_fonts["extra_bold"]
    black_font = pdf_fonts["black"]
    pdf_white = colors.white

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=(FOOTBALL_DAYS_PDF_WIDTH, FOOTBALL_DAYS_PDF_HEIGHT))
    backgrounds = football_days_background_paths()
    def draw_background(page_index: int, shade_alpha: float = 0.18) -> None:
        background = backgrounds[page_index % len(backgrounds)]
        if os.path.exists(background):
            pdf.drawImage(ImageReader(background), 0, 0, FOOTBALL_DAYS_PDF_WIDTH, FOOTBALL_DAYS_PDF_HEIGHT)
        else:
            pdf.setFillColor(colors.HexColor("#0d0d0d"))
            pdf.rect(0, 0, FOOTBALL_DAYS_PDF_WIDTH, FOOTBALL_DAYS_PDF_HEIGHT, fill=1, stroke=0)
        pdf.saveState()
        pdf.setFillColor(colors.Color(0, 0, 0, alpha=shade_alpha))
        pdf.rect(0, 0, FOOTBALL_DAYS_PDF_WIDTH, FOOTBALL_DAYS_PDF_HEIGHT, fill=1, stroke=0)
        pdf.restoreState()

    def draw_text(text: Any, x: float, y: float, size: float, color: Any = pdf_white, font: str = None) -> None:
        pdf.setFillColor(color)
        font = font or bold_font
        pdf.setFont(font, size)
        pdf.drawString(x, y, str(text or ""))

    def split_text(text: Any, max_width: float, size: float, font: str = None) -> List[str]:
        font = font or regular_font
        words = str(text or "").replace("\r", "").split()
        lines: List[str] = []
        current = ""
        for word in words:
            candidate = f"{current} {word}".strip()
            if stringWidth(candidate, font, size) <= max_width or not current:
                current = candidate
            else:
                lines.append(current)
                current = word
        if current:
            lines.append(current)
        return lines

    def trim_text_to_width(text: Any, max_width: float, size: float, font: str) -> str:
        value = str(text or "").strip()
        if stringWidth(value, font, size) <= max_width:
            return value
        ellipsis = "..."
        while value and stringWidth(f"{value}{ellipsis}", font, size) > max_width:
            value = value[:-1].rstrip()
        return f"{value}{ellipsis}" if value else ellipsis

    def fit_text_lines(text: Any, max_width: float, max_lines: int, max_size: float, min_size: float, font: str) -> Tuple[float, List[str]]:
        value = str(text or "").strip()
        size = max_size
        while size >= min_size:
            lines = split_text(value, max_width, size, font) or [value]
            if len(lines) <= max_lines and all(stringWidth(line, font, size) <= max_width for line in lines):
                return size, lines
            size -= 1
        lines = split_text(value, max_width, min_size, font) or [value]
        fitted_lines = lines[:max_lines]
        if len(lines) > max_lines and fitted_lines:
            fitted_lines[-1] = trim_text_to_width(fitted_lines[-1], max_width, min_size, font)
        return min_size, fitted_lines

    def draw_centered_fitted_title(
        title: Any,
        y_top: float = 452,
        y_bottom: float = 394,
        max_width: float = 850,
        max_size: float = 54,
        min_size: float = 24,
    ) -> None:
        size, lines = fit_text_lines(str(title or "").upper(), max_width, 2, max_size, min_size, black_font)
        leading = size * 0.98
        text_height = size + ((len(lines) - 1) * leading)
        first_baseline = y_bottom + ((y_top - y_bottom - text_height) / 2) + size * 0.76 + ((len(lines) - 1) * leading)
        pdf.setFillColor(pdf_white)
        pdf.setFont(black_font, size)
        for index, line in enumerate(lines):
            line_width = stringWidth(line, black_font, size)
            pdf.drawString((FOOTBALL_DAYS_PDF_WIDTH - line_width) / 2, first_baseline - (index * leading), line)

    def draw_wrapped(text: Any, x: float, y: float, max_width: float, size: float, leading: float, color: Any = pdf_white, font: str = None) -> float:
        font = font or regular_font
        pdf.setFillColor(color)
        pdf.setFont(font, size)
        current_y = y
        for paragraph in str(text or "").splitlines() or [""]:
            lines = split_text(paragraph, max_width, size, font) or [""]
            for line in lines:
                pdf.drawString(x, current_y, line)
                current_y -= leading
            current_y -= leading * 0.35
        return current_y

    def draw_header(title: str, page_index: int) -> None:
        draw_background(page_index)
        draw_centered_fitted_title(title)

    def draw_exercise_header(title: str, page_index: int) -> None:
        draw_background(page_index)
        draw_centered_fitted_title(title)

    def draw_training_cover_page(training: Dict[str, Any], page_index: int) -> None:
        draw_background(page_index, 0.24)
        training_name = str(training.get("name") or "Training").strip()
        training_date = str(training.get("dateLabel") or format_football_days_date(training.get("date"))).strip().upper()
        draw_centered_fitted_title(training_name, 360, 294, 820, 56, 24)
        pdf.setFillColor(pdf_white)
        pdf.setFont(bold_font, 22)
        pdf.drawCentredString(FOOTBALL_DAYS_PDF_WIDTH / 2, 250, training_date)

    def draw_training_page_footer(training: Optional[Dict[str, Any]]) -> None:
        if not isinstance(training, dict):
            return
        training_name = str(training.get("name") or "Training").strip()
        training_date = str(training.get("dateLabel") or format_football_days_date(training.get("date"))).strip()
        footer_text = " | ".join(part for part in (training_name, training_date) if part)
        if not footer_text:
            return
        pdf.saveState()
        pdf.setFillColor(pdf_white)
        pdf.setFont(regular_font, 8)
        pdf.drawCentredString(FOOTBALL_DAYS_PDF_WIDTH / 2, 22, footer_text)
        pdf.restoreState()

    def draw_panel(x: float, y: float, width: float, height: float, alpha: float = 0.78) -> None:
        pdf.saveState()
        pdf.setFillColor(colors.Color(0, 0, 0, alpha=min(0.82, max(0.24, alpha))))
        pdf.setStrokeColor(colors.Color(1, 1, 1, alpha=0.45))
        pdf.roundRect(x, y, width, height, 4, fill=1, stroke=1)
        pdf.restoreState()

    def draw_table(headers: List[str], rows: List[List[str]], widths: List[float], x: float, y: float, row_height: float, font_size: float = 12) -> None:
        table_width = sum(widths)
        header_height = 31
        pdf.saveState()
        pdf.setFillColor(colors.Color(0, 0, 0, alpha=0.74))
        pdf.rect(x, y - header_height, table_width, header_height, fill=1, stroke=0)
        pdf.setFillColor(pdf_white)
        pdf.setFont(bold_font, 10)
        cursor_x = x
        for index, header in enumerate(headers):
            pdf.drawString(cursor_x + 12, y - 20, header.upper())
            cursor_x += widths[index]
        pdf.restoreState()

        current_y = y - header_height
        for row_index, row in enumerate(rows):
            pdf.saveState()
            pdf.setFillColor(colors.Color(0, 0, 0, alpha=0.56 if row_index % 2 == 0 else 0.46))
            pdf.rect(x, current_y - row_height, table_width, row_height, fill=1, stroke=0)
            pdf.setStrokeColor(colors.Color(1, 1, 1, alpha=0.18))
            pdf.line(x, current_y - row_height, x + table_width, current_y - row_height)
            pdf.restoreState()
            cursor_x = x
            for cell_index, cell in enumerate(row):
                draw_wrapped(cell or "-", cursor_x + 12, current_y - 18, widths[cell_index] - 22, font_size, font_size + 3)
                cursor_x += widths[cell_index]
            current_y -= row_height

    def readable_text_color(hex_color: Any) -> Any:
        return pdf_white

    def draw_program_icon(icon_key: str, cx: float, cy: float, size: float = 21) -> None:
        key = icon_key if icon_key in {"clipboard", "flame", "football", "utensils", "trophy", "camera", "medical", "cones", "clock"} else "clock"
        scale = size / 21
        outer_radius = size * 0.72
        pdf.saveState()
        pdf.setStrokeColor(colors.Color(1, 1, 1, alpha=0.94))
        pdf.setFillColor(colors.Color(1, 1, 1, alpha=0.08))
        pdf.setLineWidth(max(1, 1.6 * scale))
        pdf.circle(cx, cy, outer_radius, stroke=1, fill=1)
        pdf.translate(cx, cy)
        pdf.scale(scale, scale)
        pdf.setLineWidth(1.9)
        if key == "football":
            pdf.circle(0, 0, 7.6, stroke=1, fill=0)
            pdf.circle(0, 0, 2.2, stroke=1, fill=0)
            for dx, dy in ((0, 7.3), (6.9, 2.3), (4.4, -6.1), (-4.4, -6.1), (-6.9, 2.3)):
                pdf.line(0, 0, dx, dy)
        elif key == "trophy":
            pdf.roundRect(-7, 1, 14, 10, 2, stroke=1, fill=0)
            pdf.arc(-16, 2, -4, 14, 260, 130)
            pdf.arc(4, 2, 16, 14, -30, 130)
            pdf.line(0, 1, 0, -8)
            pdf.line(-8, -9, 8, -9)
        elif key == "utensils":
            for offset in (-4, 0, 4):
                pdf.line(-7 + offset / 2, 10, -7 + offset / 2, -3)
            pdf.line(-10, -3, -5, -3)
            pdf.line(7, 10, 7, -10)
            pdf.arc(3, 1, 13, 13, 90, 180)
        elif key == "clipboard":
            pdf.roundRect(-8, -10, 16, 20, 2, stroke=1, fill=0)
            pdf.roundRect(-5, 7, 10, 5, 1.4, stroke=1, fill=0)
            pdf.line(-4, 2, 5, 2)
            pdf.line(-4, -3, 5, -3)
        elif key == "flame":
            path = pdf.beginPath()
            path.moveTo(0, -10)
            path.curveTo(-14, -4, -7, 7, -1, 12)
            path.curveTo(-1, 5, 6, 5, 3, 12)
            path.curveTo(12, 4, 13, -7, 0, -10)
            pdf.drawPath(path, stroke=1, fill=0)
        elif key == "camera":
            pdf.roundRect(-11, -7, 22, 16, 2, stroke=1, fill=0)
            pdf.line(-5, 9, 4, 9)
            pdf.circle(0, 1, 4.2, stroke=1, fill=0)
        elif key == "medical":
            pdf.rect(-3, -10, 6, 20, stroke=1, fill=0)
            pdf.rect(-10, -3, 20, 6, stroke=1, fill=0)
        elif key == "cones":
            pdf.line(-8, -10, -2, 10)
            pdf.line(8, -10, 2, 10)
            pdf.line(-11, -10, 11, -10)
            pdf.line(-5, 1, 5, 1)
            pdf.line(-3, 7, 3, 7)
        else:
            pdf.circle(0, 0, 7.6, stroke=1, fill=0)
            pdf.line(0, 0, 0, 6)
            pdf.line(0, 0, 6, -4)
        pdf.restoreState()

    def draw_program_page(rows: List[Dict[str, Any]], page_index: int, layout_row_count: Optional[int] = None) -> None:
        draw_header("Programma", page_index)
        row_count = max(1, int(layout_row_count or len(rows)))
        top_y = 385
        bottom_y = 42
        available_height = top_y - bottom_y
        row_height = min(37, available_height / row_count)
        row_box_height = max(18, row_height - 4)
        icon_size = 18 if row_height >= 32 else 14
        start_font_size = 14 if row_height >= 32 else 11.5
        end_font_size = 10 if row_height >= 32 else 8.5
        activity_font_size = 16.5 if row_height >= 32 else 12.4
        activity_leading = 17.5 if row_height >= 32 else 11.4
        max_activity_lines = 2 if row_box_height >= 22 else 1
        x = 78
        width = 804
        for row_index, item in enumerate(rows):
            row_y = top_y - (row_index * row_height)
            pdf.saveState()
            pdf.setFillColor(colors.Color(0, 0, 0, alpha=0.55 if row_index % 2 == 0 else 0.45))
            pdf.setStrokeColor(colors.Color(1, 1, 1, alpha=0.18))
            pdf.roundRect(x, row_y - row_box_height + 2, width, row_box_height, 5, fill=1, stroke=1)
            pdf.restoreState()

            icon_key = str(item.get("icon") or infer_football_activity_icon(str(item.get("activity") or "")))
            row_center_y = row_y - (row_box_height / 2) + 2
            draw_program_icon(icon_key, x + 28, row_center_y, icon_size)

            start = str(item.get("startTime") or "").strip() or "--:--"
            end = str(item.get("endTime") or "").strip() or "--:--"
            pdf.setFillColor(pdf_white)
            pdf.setFont(black_font, start_font_size)
            pdf.drawString(x + 62, row_center_y + (3 if row_height >= 32 else 2), start)
            pdf.setFont(regular_font, end_font_size)
            pdf.setFillColor(colors.Color(1, 1, 1, alpha=0.74))
            pdf.drawString(x + 62, row_center_y - (12 if row_height >= 32 else 9), end)

            pdf.setFillColor(pdf_white)
            activity = str(item.get("activity") or "Nog in te vullen").strip()
            activity_width = width - 205
            fitted_activity_font_size = activity_font_size
            fitted_activity_leading = activity_leading
            activity_lines = []
            for candidate_font_size in (activity_font_size, 11.4, 10.5, 9.8):
                candidate_leading = min(fitted_activity_leading, candidate_font_size + 1)
                candidate_lines = split_text(activity, activity_width, candidate_font_size, bold_font)
                if len(candidate_lines) <= max_activity_lines or candidate_font_size == 9.8:
                    fitted_activity_font_size = candidate_font_size
                    fitted_activity_leading = candidate_leading
                    activity_lines = candidate_lines[:max_activity_lines]
                    break
            pdf.setFont(bold_font, fitted_activity_font_size)
            ascent = pdfmetrics.getAscent(bold_font) / 1000 * fitted_activity_font_size
            descent = abs(pdfmetrics.getDescent(bold_font) / 1000 * fitted_activity_font_size)
            text_block_height = ascent + descent + (max(0, len(activity_lines) - 1) * fitted_activity_leading)
            text_y = row_center_y + (text_block_height / 2) - ascent
            for line in activity_lines:
                pdf.drawString(x + 168, text_y, line)
                text_y -= fitted_activity_leading

    def draw_overview_page(page_index: int) -> None:
        draw_header("Overzicht", page_index)
        intro_subject = str(data.get("introSubject") or "voetbaldag")
        include_staff = bool(data.get("includeStaff", True))
        include_program = bool(data.get("includeProgram", True))
        has_contingencies = bool(str(data.get("contingencies") or "").strip())
        is_amateur_club_playbook = str(data.get("playbookType")) == "samenwerkende-amateurclubs"
        date_sentence = (
            f"voor {str(data.get('cycleCoverLabel') or 'de cyclus').lower()}: {data['cycleDateRangeLabel']}"
            if is_amateur_club_playbook
            else f"op {data['eventDateLabel']}"
        )
        visible_parts = []
        if include_staff:
            visible_parts.append("de taakverdeling")
        if include_program:
            visible_parts.append("het programma")
        visible_parts.append("de veldplattegrond")
        if has_contingencies:
            visible_parts.append("de afspraken voor onvoorziene situaties")
        visible_parts_text = ", ".join(visible_parts[:-1]) + (" en " + visible_parts[-1] if len(visible_parts) > 1 else visible_parts[0])
        intro = (
            f"Dit draaiboek bundelt alle praktische informatie voor de {intro_subject} bij {data['clubName']} "
            f"{date_sentence}. Het document bevat {visible_parts_text}."
        )
        panel_width = 670
        padding_x = 34
        padding_top = 34
        padding_bottom = 32
        intro_font_size = 17
        intro_leading = 23
        intro_lines: List[str] = []
        for paragraph in intro.splitlines() or [""]:
            intro_lines.extend(split_text(paragraph, panel_width - (padding_x * 2), intro_font_size, regular_font) or [""])
        intro_height = max(1, len(intro_lines)) * intro_leading

        tile_width = 292
        tile_height = 48
        tile_gap_x = 30
        tile_gap_y = 18
        detail_rows = 2
        detail_height = (detail_rows * tile_height) + ((detail_rows - 1) * tile_gap_y)
        intro_to_tiles_gap = 28
        panel_height = padding_top + intro_height + intro_to_tiles_gap + detail_height + padding_bottom
        panel_x = (FOOTBALL_DAYS_PDF_WIDTH - panel_width) / 2
        panel_y = (FOOTBALL_DAYS_PDF_HEIGHT - panel_height) / 2 - 8

        pdf.saveState()
        pdf.setFillColor(colors.Color(1, 1, 1, alpha=0.78))
        pdf.setStrokeColor(colors.Color(1, 1, 1, alpha=0.55))
        pdf.roundRect(panel_x, panel_y, panel_width, panel_height, 5, fill=1, stroke=1)
        pdf.restoreState()

        text_x = panel_x + padding_x
        text_y = panel_y + panel_height - padding_top
        pdf.setFillColor(colors.HexColor("#171717"))
        pdf.setFont(regular_font, intro_font_size)
        for line in intro_lines:
            pdf.drawString(text_x, text_y, line)
            text_y -= intro_leading

        if is_amateur_club_playbook:
            details = [
                ("Club", data["clubName"]),
                ("Cyclus", data.get("cycleNumber") or "Nog in te vullen"),
                ("Start cyclus", data["cycleStartDateLabel"]),
                ("Einde cyclus", data["cycleEndDateLabel"]),
            ]
        else:
            details = [
                ("Club", data["clubName"]),
                ("Datum", data["eventDateLabel"]),
                ("Locatie", data["location"] or "Nog in te vullen"),
                ("Aanmeldingen", data["registrationCount"]),
            ]
        detail_start_y = text_y - intro_to_tiles_gap + 7
        for index, (label, value) in enumerate(details):
            column = index % 2
            row = index // 2
            box_x = text_x + column * (tile_width + tile_gap_x)
            box_y = detail_start_y - (row * (tile_height + tile_gap_y)) - tile_height
            pdf.saveState()
            pdf.setFillColor(colors.Color(1, 1, 1, alpha=0.58))
            pdf.setStrokeColor(colors.Color(1, 1, 1, alpha=0.52))
            pdf.roundRect(box_x, box_y, tile_width, tile_height, 4, fill=1, stroke=1)
            pdf.restoreState()
            pdf.setFillColor(colors.HexColor("#303030"))
            pdf.setFont(bold_font, 11.5)
            pdf.drawString(box_x + 16, box_y + tile_height - 18, label.upper())
            pdf.setFillColor(colors.HexColor("#5f5f5f"))
            pdf.setFont(regular_font, 15)
            value_lines = split_text(value, tile_width - 32, 15, regular_font)[:2]
            value_y = box_y + tile_height - 33
            for line in value_lines:
                pdf.drawString(box_x + 16, value_y, line)
                value_y -= 14

    def draw_training_dates_page(rows: List[Dict[str, Any]], no_training_rows: List[Dict[str, str]], page_index: int) -> None:
        draw_header("Trainingsdata", page_index)
        sections = [
            {
                "title": "Trainingen",
                "rows": [
                    {
                        "date": str(training.get("date") or "").strip(),
                        "dateLabel": str(training.get("dateLabel") or format_football_days_date(training.get("date"))).strip(),
                        "label": str(training.get("name") or f"Training {index + 1}").strip(),
                    }
                    for index, training in enumerate(sorted_football_cycle_trainings(rows))
                ],
            },
            {
                "title": "Geen training",
                "rows": [
                    {
                        "date": row["date"],
                        "dateLabel": row["dateLabel"],
                        "label": row.get("description") or "Geen training",
                    }
                    for row in normalize_football_no_training_dates(no_training_rows)
                ],
            },
        ]
        sections = [section for section in sections if section["rows"]]
        total_rows = sum(len(section["rows"]) for section in sections)
        panel_width = 700
        panel_x = (FOOTBALL_DAYS_PDF_WIDTH - panel_width) / 2
        panel_y = 62
        panel_height = 306
        padding_x = 34
        top_y = panel_y + panel_height - 32
        heading_height = 24
        section_gap = 18
        row_height = min(
            36,
            (panel_height - 50 - (len(sections) * heading_height) - (max(0, len(sections) - 1) * section_gap))
            / max(1, total_rows),
        )
        font_size = 12.2 if row_height >= 28 else max(7.2, row_height - 5)
        max_lines = 2 if row_height >= 28 else 1
        date_width = 248

        pdf.saveState()
        pdf.setFillColor(colors.Color(1, 1, 1, alpha=0.78))
        pdf.setStrokeColor(colors.Color(1, 1, 1, alpha=0.55))
        pdf.roundRect(panel_x, panel_y, panel_width, panel_height, 5, fill=1, stroke=1)
        pdf.restoreState()

        cursor_y = top_y
        for section_index, section in enumerate(sections):
            if section_index:
                cursor_y -= section_gap
            pdf.setFillColor(colors.HexColor("#303030"))
            pdf.setFont(black_font, 11)
            pdf.drawString(panel_x + padding_x, cursor_y, f"{section['title']} datum".upper())
            pdf.drawString(panel_x + padding_x + date_width, cursor_y, "OMSCHRIJVING")
            table_top = cursor_y - 18

            def draw_centered_row_text(lines: List[str], x: float, box_y: float, box_height: float, font: str, size: float, color: Any) -> None:
                visible_lines = lines or ["-"]
                line_gap = 1.2
                line_height = size + line_gap
                ascent = pdfmetrics.getAscent(font) / 1000 * size
                descent = abs(pdfmetrics.getDescent(font) / 1000 * size)
                text_height = ascent + descent + ((len(visible_lines) - 1) * line_height)
                baseline = box_y + (box_height - text_height) / 2 + text_height - ascent
                pdf.setFillColor(color)
                pdf.setFont(font, size)
                for line in visible_lines:
                    pdf.drawString(x, baseline, line)
                    baseline -= line_height

            for row_index, row in enumerate(section["rows"]):
                row_top = table_top - (row_index * row_height)
                row_y = row_top - row_height + 4
                box_height = row_height - 4
                pdf.saveState()
                pdf.setFillColor(colors.Color(1, 1, 1, alpha=0.58 if row_index % 2 == 0 else 0.46))
                pdf.setStrokeColor(colors.Color(1, 1, 1, alpha=0.42))
                pdf.roundRect(panel_x + padding_x - 4, row_y, panel_width - (padding_x * 2) + 8, box_height, 4, fill=1, stroke=1)
                pdf.restoreState()

                date_lines = split_text(row["dateLabel"], date_width - 22, font_size, bold_font)[:max_lines]
                name_lines = split_text(row["label"], panel_width - (padding_x * 2) - date_width - 24, font_size, regular_font)[:max_lines]
                draw_centered_row_text(date_lines, panel_x + padding_x + 8, row_y, box_height, bold_font, font_size, colors.HexColor("#171717"))
                draw_centered_row_text(name_lines, panel_x + padding_x + date_width, row_y, box_height, regular_font, font_size, colors.HexColor("#5f5f5f"))
            cursor_y = table_top - (len(section["rows"]) * row_height)

    def draw_staff_page(rows: List[Dict[str, Any]], page_index: int) -> None:
        draw_header("Taakverdeling", page_index)
        include_setup_tasks = bool(data.get("includeStaffSetupTasks", True))
        x = 78
        width = 804
        top_y = 382
        header_height = 31
        row_height = 37
        if include_setup_tasks:
            columns = [
                ("Naam", x + 18, 210, bold_font),
                ("Rol", x + 250, 190, regular_font),
                ("Taak bij uitzetten", x + 468, 315, regular_font),
            ]
        else:
            columns = [
                ("Naam", x + 18, 360, bold_font),
                ("Rol", x + 420, 390, regular_font),
            ]

        pdf.saveState()
        pdf.setFillColor(colors.Color(0, 0, 0, alpha=0.72))
        pdf.roundRect(x, top_y - header_height, width, header_height, 5, fill=1, stroke=0)
        pdf.restoreState()
        pdf.setFillColor(pdf_white)
        pdf.setFont(black_font, 10.5)
        for label, column_x, _column_width, _font in columns:
            pdf.drawString(column_x, top_y - 20, label.upper())

        for row_index, member in enumerate(rows):
            row_top = top_y - header_height - (row_index * row_height)
            row_y = row_top - row_height
            pdf.saveState()
            pdf.setFillColor(colors.Color(0, 0, 0, alpha=0.55 if row_index % 2 == 0 else 0.45))
            pdf.setStrokeColor(colors.Color(1, 1, 1, alpha=0.16))
            pdf.roundRect(x, row_y, width, row_height - 2, 4, fill=1, stroke=1)
            pdf.restoreState()

            values = [
                str(member.get("name") or "-").strip() or "-",
                str(member.get("role") or "-").strip() or "-",
            ]
            if include_setup_tasks:
                values.append(str(member.get("setupTask") or "-").strip() or "-")
            for value, (_label, column_x, column_width, font) in zip(values, columns):
                font_size = 13.5 if font == bold_font else 12.5
                line_height = 14
                lines = split_text(value, column_width, font_size, font)[:2]
                text_y = row_y + 21 if len(lines) == 1 else row_y + 26
                pdf.setFillColor(pdf_white if font == bold_font else colors.Color(1, 1, 1, alpha=0.82))
                pdf.setFont(font, font_size)
                for line in lines:
                    pdf.drawString(column_x, text_y, line)
                    text_y -= line_height

    def draw_contingencies_page(rows: List[List[str]], page_index: int) -> None:
        draw_exercise_header("Onvoorziene omstandigheden", page_index)
        panel_width = 720
        panel_x = (FOOTBALL_DAYS_PDF_WIDTH - panel_width) / 2
        panel_y = 92
        panel_height = 285
        padding_x = 34
        padding_top = 30
        padding_bottom = 28
        body_width = panel_width - (padding_x * 2)

        font_size = 16
        leading = 20
        label_size = 10.5
        label_gap = 13
        entry_gap = 16
        entries: List[Dict[str, Any]] = []
        while font_size >= 10:
            entries = []
            total_height = 0.0
            for scenario, solution in rows:
                scenario_label = str(scenario or "").strip()
                solution_text = str(solution or "").strip() or "Nog in te vullen"
                show_label = bool(scenario_label and scenario_label.lower() not in {"scenario", "algemeen"})
                lines: List[str] = []
                for paragraph in solution_text.splitlines() or [solution_text]:
                    lines.extend(split_text(paragraph, body_width, font_size, regular_font) or [""])
                entry_height = (label_gap if show_label else 0) + (max(1, len(lines)) * leading)
                entries.append({"label": scenario_label, "showLabel": show_label, "lines": lines, "height": entry_height})
                total_height += entry_height
            total_height += max(0, len(entries) - 1) * entry_gap
            if total_height <= panel_height - padding_top - padding_bottom:
                break
            font_size -= 1
            leading = max(13, font_size + 4)
            entry_gap = max(8, font_size)

        pdf.saveState()
        pdf.setFillColor(colors.Color(1, 1, 1, alpha=0.78))
        pdf.setStrokeColor(colors.Color(1, 1, 1, alpha=0.55))
        pdf.roundRect(panel_x, panel_y, panel_width, panel_height, 5, fill=1, stroke=1)
        pdf.restoreState()

        text_y = panel_y + panel_height - padding_top
        for entry in entries:
            if entry["showLabel"]:
                pdf.setFillColor(colors.HexColor("#303030"))
                pdf.setFont(black_font, label_size)
                pdf.drawString(panel_x + padding_x, text_y, str(entry["label"]).upper())
                text_y -= label_gap
            pdf.setFillColor(colors.HexColor("#171717"))
            pdf.setFont(regular_font, font_size)
            available_lines = max(1, int((text_y - panel_y - padding_bottom) // leading))
            lines = entry["lines"][:available_lines]
            if len(entry["lines"]) > available_lines and lines:
                lines[-1] = trim_text_to_width(lines[-1], body_width, font_size, regular_font)
            for line in lines:
                pdf.drawString(panel_x + padding_x, text_y, line)
                text_y -= leading
            text_y -= entry_gap

    def draw_field_exercise_table(rows: List[List[str]]) -> None:
        x = 360
        width = 535
        top_y = 386
        bottom_y = 58
        row_count = max(1, len(rows))
        available_height = max(80, top_y - bottom_y)
        header_height = 28 if row_count <= 10 else max(16, min(24, available_height * 0.08))
        row_height = min(34, (available_height - header_height) / row_count)
        header_font_size = max(5.4, min(9.2, header_height * 0.34))
        row_font_size = max(4.2, min(10.5, row_height * 0.36))
        block_font_size = max(row_font_size, min(10.5, row_font_size + 0.8))
        line_gap = max(4.2, row_font_size + 1.2)
        max_lines = 2 if row_height >= 27 else 1
        row_radius = 3 if row_height >= 18 else 1.5
        columns = [
            ("#", x + 14, 34, black_font, max(4.4, row_font_size)),
            ("Naam blok", x + 58, 140, bold_font, block_font_size),
            ("Oefening", x + 214, 300, regular_font, row_font_size),
        ]

        pdf.saveState()
        pdf.setFillColor(colors.Color(0, 0, 0, alpha=0.72))
        pdf.roundRect(x, top_y - header_height, width, header_height, 4, fill=1, stroke=0)
        pdf.restoreState()
        pdf.setFillColor(pdf_white)
        pdf.setFont(black_font, header_font_size)
        for label, column_x, _column_width, _font, _font_size in columns:
            pdf.drawString(column_x, top_y - max(9, header_height * 0.66), label.upper())

        for row_index, row in enumerate(rows):
            row_top = top_y - header_height - (row_index * row_height)
            row_y = row_top - row_height
            pdf.saveState()
            pdf.setFillColor(colors.Color(0, 0, 0, alpha=0.55 if row_index % 2 == 0 else 0.45))
            pdf.setStrokeColor(colors.Color(1, 1, 1, alpha=0.16))
            pdf.roundRect(x, row_y, width, max(2.5, row_height - 2), row_radius, fill=1, stroke=1)
            pdf.restoreState()
            for value, (_label, column_x, column_width, font, font_size) in zip(row, columns):
                lines = split_text(value or "-", column_width, font_size, font)[:max_lines]
                if len(lines) > 1:
                    text_y = row_y + (row_height / 2) + (line_gap / 2) - 1
                else:
                    text_y = row_y + max(2, (row_height - font_size) / 2)
                pdf.setFillColor(pdf_white if font == bold_font or font == black_font else colors.Color(1, 1, 1, alpha=0.82))
                pdf.setFont(font, font_size)
                for line in lines:
                    pdf.drawString(column_x, text_y, line)
                    text_y -= line_gap

    def draw_exercise_field_preview(field: Any, x: float, y: float, width: float, height: float, label: str) -> None:
        pdf.saveState()
        pdf.setFillColor(colors.HexColor("#159447"))
        pdf.setStrokeColor(colors.Color(1, 1, 1, alpha=0.32))
        pdf.roundRect(x, y, width, height, 5, fill=1, stroke=1)
        pdf.restoreState()
        if not isinstance(field, dict):
            field = {"viewBox": [0, 0, 100, 70], "elements": [], "overlayItems": []}
            return

        def draw_overlay_items(items: Any, stage_x: float, stage_y: float, stage_width: float, stage_height: float) -> None:
            overlay_items = items if isinstance(items, list) else []

            def overlay_x(value: Any, default: float = 50.0) -> float:
                return stage_x + stage_width * (clamp_float(value, 0.0, 100.0, default) / 100)

            def overlay_y(value: Any, default: float = 50.0) -> float:
                return stage_y + stage_height - stage_height * (clamp_float(value, 0.0, 100.0, default) / 100)

            for item in overlay_items[:160]:
                if not isinstance(item, dict):
                    continue
                item_type = str(item.get("type") or "").strip()
                item_color = safe_svg_color(item.get("color"), "#111111")
                red = int(item_color[1:3], 16) / 255
                green = int(item_color[3:5], 16) / 255
                blue = int(item_color[5:7], 16) / 255
                item_x = overlay_x(item.get("x"))
                item_y = overlay_y(item.get("y"))
                item_size = clamp_float(item.get("size"), 45.0, 220.0, 100.0) / 100.0
                pdf.saveState()
                if item_type == "player":
                    pdf.setFillColor(colors.Color(red, green, blue, alpha=0.96))
                    pdf.setStrokeColor(pdf_white)
                    pdf.setLineWidth(1.2)
                    pdf.circle(item_x, item_y, 5.8 * item_size, stroke=1, fill=1)
                    pdf.setFillColor(colors.Color(max(red * 0.82, 0), max(green * 0.82, 0), max(blue * 0.82, 0), alpha=0.96))
                    pdf.roundRect(item_x - 6.0 * item_size, item_y - 13.5 * item_size, 12.0 * item_size, 6.2 * item_size, 2.2 * item_size, fill=1, stroke=0)
                elif item_type == "ball":
                    pdf.setFillColor(colors.Color(red, green, blue, alpha=0.96))
                    pdf.setStrokeColor(colors.HexColor("#111111"))
                    pdf.setLineWidth(1.1)
                    pdf.circle(item_x, item_y, 4.1 * item_size, stroke=1, fill=1)
                    pdf.line(item_x - 3.0 * item_size, item_y, item_x + 3.0 * item_size, item_y)
                    pdf.line(item_x, item_y - 3.0 * item_size, item_x, item_y + 3.0 * item_size)
                elif item_type in {"cone", "small-cone", "big-cone"}:
                    cone_scale = item_size * (1.25 if item_type == "big-cone" else 1.0)
                    pdf.setFillColor(colors.Color(red, green, blue, alpha=0.96))
                    pdf.setStrokeColor(colors.HexColor("#111111"))
                    pdf.setLineWidth(0.9)
                    path = pdf.beginPath()
                    path.moveTo(item_x, item_y + 6.5 * cone_scale)
                    path.lineTo(item_x - 6.0 * cone_scale, item_y - 6.5 * cone_scale)
                    path.lineTo(item_x + 6.0 * cone_scale, item_y - 6.5 * cone_scale)
                    path.close()
                    pdf.drawPath(path, stroke=1, fill=1)
                elif item_type == "goal":
                    goal_width = 28.0 * item_size
                    goal_height = 17.0 * item_size
                    pdf.setStrokeColor(colors.Color(red, green, blue, alpha=0.96))
                    pdf.setLineWidth(max(1.2, 2.2 * item_size))
                    pdf.rect(item_x - goal_width / 2, item_y - goal_height / 2, goal_width, goal_height, fill=0, stroke=1)
                    pdf.line(item_x - goal_width / 2, item_y - goal_height / 2, item_x + goal_width / 2, item_y - goal_height / 2)
                elif item_type in {"line", "arrow"}:
                    end_x = overlay_x(item.get("x2"), clamp_float(item.get("x"), 0.0, 100.0, 50.0) + 12)
                    end_y = overlay_y(item.get("y2"), clamp_float(item.get("y"), 0.0, 100.0, 50.0))
                    length = ((end_x - item_x) ** 2 + (end_y - item_y) ** 2) ** 0.5
                    if length >= 0.5:
                        line_width = 2.4 * item_size
                        pdf.setStrokeColor(colors.Color(red, green, blue, alpha=0.96))
                        pdf.setLineWidth(line_width)
                        pdf.setLineCap(1)
                        pdf.line(item_x, item_y, end_x, end_y)
                        if item_type == "arrow":
                            angle = atan2(end_y - item_y, end_x - item_x)
                            head_length = min(10.0, max(6.0, length * 0.35)) * item_size
                            head_angle = 0.62
                            pdf.line(end_x, end_y, end_x - head_length * cos(angle - head_angle), end_y - head_length * sin(angle - head_angle))
                            pdf.line(end_x, end_y, end_x - head_length * cos(angle + head_angle), end_y - head_length * sin(angle + head_angle))
                elif item_type == "text":
                    text = normalize_exercise_text(item.get("text"))[:80] or "Tekst"
                    pdf.setFillColor(colors.Color(1, 1, 1, alpha=0.9))
                    text_size = 7.5 * item_size
                    text_width = min(110.0, max(28.0, stringWidth(text, bold_font, text_size) + 8))
                    pdf.roundRect(item_x - text_width / 2, item_y - 7 * item_size, text_width, 14 * item_size, 3, fill=1, stroke=0)
                    pdf.setFillColor(colors.Color(red, green, blue, alpha=1))
                    pdf.setFont(bold_font, text_size)
                    pdf.drawCentredString(item_x, item_y - 2.5 * item_size, text)
                pdf.restoreState()

        image_data_url = str(field.get("imageDataUrl") or "").strip()
        if image_data_url.startswith("data:image/") and "," in image_data_url:
            try:
                image_bytes = base64.b64decode(image_data_url.split(",", 1)[1])
                stage_x = x + 10
                stage_y = y + 10
                stage_width = width - 20
                stage_height = height - 20
                pdf.drawImage(ImageReader(BytesIO(image_bytes)), stage_x, stage_y, stage_width, stage_height, preserveAspectRatio=True, anchor="c", mask="auto")
                draw_overlay_items(field.get("overlayItems"), stage_x, stage_y, stage_width, stage_height)
                return
            except Exception:
                pass

        raw_viewbox = field.get("viewBox")
        raw_elements = field.get("elements")
        if not isinstance(raw_viewbox, list) or len(raw_viewbox) != 4 or not isinstance(raw_elements, list) or not raw_elements:
            draw_overlay_items(field.get("overlayItems"), x + 10, y + 10, width - 20, height - 20)
            return

        viewbox = [safe_svg_number(value) for value in raw_viewbox]
        if viewbox[2] <= 0 or viewbox[3] <= 0:
            draw_wrapped("Geen veldtekening beschikbaar", x + 26, y + height / 2, width - 52, 12, 15, pdf_white, bold_font)
            return

        scale = min((width - 20) / viewbox[2], (height - 20) / viewbox[3])
        draw_width = viewbox[2] * scale
        draw_height = viewbox[3] * scale
        offset_x = x + (width - draw_width) / 2 - (viewbox[0] * scale)
        offset_y = y + (height - draw_height) / 2 + draw_height + (viewbox[1] * scale)

        def map_x(value: float) -> float:
            return offset_x + value * scale

        def map_y(value: float) -> float:
            return offset_y - value * scale

        for element in raw_elements[:140]:
            if not isinstance(element, dict):
                continue
            element_x = safe_svg_number(element.get("x"))
            element_y = safe_svg_number(element.get("y"))
            element_width = max(1.0, safe_svg_number(element.get("width"), 1.0))
            element_height = max(1.0, safe_svg_number(element.get("height"), 1.0))
            fill = safe_svg_color(element.get("fill"))
            fill_color = colors.HexColor(fill)
            element_type = str(element.get("type") or "").strip()
            mapped_x = map_x(element_x)
            mapped_y = map_y(element_y + element_height)
            mapped_width = element_width * scale
            mapped_height = element_height * scale
            pdf.saveState()
            pdf.setFillColor(fill_color)
            pdf.setStrokeColor(colors.white if fill in {"#000000", "#00B050"} else colors.HexColor("#111111"))
            pdf.setLineWidth(max(0.45, 9000 * scale))
            if element_type == "ellipse":
                pdf.ellipse(mapped_x, mapped_y, mapped_x + mapped_width, mapped_y + mapped_height, fill=1, stroke=1)
            elif element_type == "cone":
                points = [
                    map_x(element_x + element_width * 0.18),
                    map_y(element_y + element_height),
                    map_x(element_x + element_width * 0.82),
                    map_y(element_y + element_height),
                    map_x(element_x + element_width * 0.62),
                    map_y(element_y),
                    map_x(element_x + element_width * 0.38),
                    map_y(element_y),
                ]
                path = pdf.beginPath()
                path.moveTo(points[0], points[1])
                path.lineTo(points[2], points[3])
                path.lineTo(points[4], points[5])
                path.lineTo(points[6], points[7])
                path.close()
                pdf.drawPath(path, stroke=1, fill=1)
            elif element_type == "line":
                pdf.setStrokeColor(fill_color)
                pdf.setLineWidth(max(1.2, 22000 * scale))
                pdf.line(map_x(element_x), map_y(element_y), map_x(element_x + element_width), map_y(element_y + element_height))
            else:
                pdf.rect(mapped_x, mapped_y, mapped_width, mapped_height, fill=1, stroke=1)
            pdf.restoreState()

        draw_overlay_items(field.get("overlayItems"), x + (width - draw_width) / 2, y + (height - draw_height) / 2, draw_width, draw_height)

    def draw_exercise_text_panel(label: str, value: Any, x: float, y: float, width: float, height: float, max_lines: int = 5) -> None:
        pdf.saveState()
        pdf.setFillColor(colors.Color(0, 0, 0, alpha=0.5))
        pdf.setStrokeColor(colors.Color(1, 1, 1, alpha=0.2))
        pdf.roundRect(x, y, width, height, 5, fill=1, stroke=1)
        pdf.restoreState()
        pdf.setFillColor(colors.Color(1, 1, 1, alpha=0.78))
        pdf.setFont(black_font, 9.5)
        pdf.drawString(x + 12, y + height - 17, label.upper())

        body_top = y + height - 34
        body_bottom = y + 10
        body_height = max(8.0, body_top - body_bottom)
        bullet_pattern = re.compile(r"^\s*(?:[•*-])\s+(.+)$")

        def build_panel_lines(font_size: float) -> List[Dict[str, Any]]:
            panel_lines: List[Dict[str, Any]] = []
            bullet_indent = max(6.0, font_size * 1.35)
            for paragraph in str(value or "Niet ingevuld").splitlines() or ["Niet ingevuld"]:
                raw_paragraph = str(paragraph or "").strip()
                if not raw_paragraph:
                    continue
                bullet_match = bullet_pattern.match(raw_paragraph)
                if bullet_match:
                    wrapped_lines = split_text(bullet_match.group(1), width - 24 - bullet_indent, font_size, regular_font) or [""]
                    panel_lines.extend(
                        {"text": line, "bullet": line_index == 0, "indent": bullet_indent}
                        for line_index, line in enumerate(wrapped_lines)
                    )
                    continue
                panel_lines.extend(
                    {"text": line, "bullet": False, "indent": 0}
                    for line in split_text(raw_paragraph, width - 24, font_size, regular_font) or [""]
                )
            if not panel_lines:
                panel_lines = [{"text": "Niet ingevuld", "bullet": False, "indent": 0}]
            return panel_lines

        font_size = 10.2
        lines = build_panel_lines(font_size)
        leading = font_size * 1.16
        while font_size > 2.4:
            lines = build_panel_lines(font_size)
            leading = max(font_size * 1.12, font_size + 0.35)
            if len(lines) * leading <= body_height:
                break
            font_size -= 0.2
        if len(lines) * leading > body_height:
            font_size = max(1.6, min(font_size, (body_height / max(1, len(lines))) * 0.9))
            for _fit_attempt in range(8):
                lines = build_panel_lines(font_size)
                leading = max(font_size * 1.08, font_size + 0.2)
                if len(lines) * leading <= body_height or font_size <= 0.8:
                    break
                font_size = max(0.8, font_size * 0.86)

        pdf.setFillColor(pdf_white)
        pdf.setFont(regular_font, font_size)
        text_y = body_top
        bullet_radius = max(0.7, font_size * 0.16)
        for line in lines:
            text_x = x + 12 + float(line["indent"])
            if line["bullet"]:
                pdf.saveState()
                pdf.setFillColor(pdf_white)
                pdf.circle(x + 15, text_y + font_size * 0.32, bullet_radius, stroke=0, fill=1)
                pdf.restoreState()
            pdf.drawString(text_x, text_y, str(line["text"]))
            text_y -= leading

    def draw_field_line(x1: float, y1: float, x2: float, y2: float, width: float = 2.2) -> None:
        pdf.saveState()
        pdf.setStrokeColor(colors.Color(1, 1, 1, alpha=0.92))
        pdf.setLineWidth(width)
        pdf.line(x1, y1, x2, y2)
        pdf.restoreState()

    def draw_football_field(x: float, y: float, width: float, height: float, blocks: List[Dict[str, Any]]) -> None:
        pdf.saveState()
        pdf.setFillColor(colors.HexColor("#168736"))
        pdf.roundRect(x - 10, y - 10, width + 20, height + 20, 4, fill=1, stroke=0)
        pdf.setFillColor(colors.HexColor("#168736"))
        pdf.rect(x, y, width, height, fill=1, stroke=0)
        pdf.setStrokeColor(colors.Color(1, 1, 1, alpha=0.92))
        pdf.setLineWidth(2.2)
        pdf.rect(x, y, width, height, fill=0, stroke=1)
        pdf.restoreState()

        draw_field_line(x, y + height / 2, x + width, y + height / 2)
        pdf.saveState()
        pdf.setStrokeColor(colors.Color(1, 1, 1, alpha=0.92))
        pdf.setLineWidth(2.2)
        center_radius = width * 0.18
        pdf.circle(x + width / 2, y + height / 2, center_radius, stroke=1, fill=0)
        pdf.setFillColor(colors.Color(1, 1, 1, alpha=0.94))
        pdf.circle(x + width / 2, y + height / 2, 2.4, stroke=0, fill=1)

        penalty_width = width * 0.516
        penalty_height = height * 0.134
        penalty_x = x + (width - penalty_width) / 2
        goal_box_width = width * 0.236
        goal_box_height = height * 0.052
        goal_box_x = x + (width - goal_box_width) / 2
        goal_width = width * 0.11
        goal_depth = height * 0.024
        goal_x = x + (width - goal_width) / 2

        pdf.rect(penalty_x, y + height - penalty_height, penalty_width, penalty_height, fill=0, stroke=1)
        pdf.rect(penalty_x, y, penalty_width, penalty_height, fill=0, stroke=1)
        pdf.rect(goal_box_x, y + height - goal_box_height, goal_box_width, goal_box_height, fill=0, stroke=1)
        pdf.rect(goal_box_x, y, goal_box_width, goal_box_height, fill=0, stroke=1)
        pdf.rect(goal_x, y + height, goal_width, goal_depth, fill=0, stroke=1)
        pdf.rect(goal_x, y - goal_depth, goal_width, goal_depth, fill=0, stroke=1)
        pdf.circle(x + width / 2, y + height * 0.895, 2.1, stroke=0, fill=1)
        pdf.circle(x + width / 2, y + height * 0.105, 2.1, stroke=0, fill=1)
        corner_radius = width * 0.032
        pdf.arc(x - corner_radius, y + height - corner_radius, x + corner_radius, y + height + corner_radius, 270, 90)
        pdf.arc(x + width - corner_radius, y + height - corner_radius, x + width + corner_radius, y + height + corner_radius, 180, 90)
        pdf.arc(x - corner_radius, y - corner_radius, x + corner_radius, y + corner_radius, 0, 90)
        pdf.arc(x + width - corner_radius, y - corner_radius, x + width + corner_radius, y + corner_radius, 90, 90)
        pdf.restoreState()

        for arrow in [item for item in blocks if item.get("type") == "arrow"]:
            start_x = x + width * (float(arrow.get("x1") or 0) / 100)
            start_y = y + height - height * (float(arrow.get("y1") or 0) / 100)
            end_x = x + width * (float(arrow.get("x2") or 0) / 100)
            end_y = y + height - height * (float(arrow.get("y2") or 0) / 100)
            arrow_length = ((end_x - start_x) ** 2 + (end_y - start_y) ** 2) ** 0.5
            if arrow_length < 0.5:
                continue
            arrow_color = normalize_hex_color(arrow.get("color"), "#FFFFFF")
            red = int(arrow_color[1:3], 16) / 255
            green = int(arrow_color[3:5], 16) / 255
            blue = int(arrow_color[5:7], 16) / 255
            line_width = max(1.2, min(6.0, float(arrow.get("strokeWidth") or 5) * 0.55))
            angle = atan2(end_y - start_y, end_x - start_x)
            head_length = min(13.0, max(7.0, 7.5 + line_width * 1.15), arrow_length * 0.45)
            head_angle = 0.62
            left_head_x = end_x - head_length * cos(angle - head_angle)
            left_head_y = end_y - head_length * sin(angle - head_angle)
            right_head_x = end_x - head_length * cos(angle + head_angle)
            right_head_y = end_y - head_length * sin(angle + head_angle)
            pdf.saveState()
            pdf.setStrokeColor(colors.Color(red, green, blue, alpha=0.95))
            pdf.setLineWidth(line_width)
            pdf.setLineCap(1)
            pdf.setLineJoin(1)
            pdf.line(start_x, start_y, end_x, end_y)
            pdf.line(end_x, end_y, left_head_x, left_head_y)
            pdf.line(end_x, end_y, right_head_x, right_head_y)
            pdf.restoreState()

        block_items = [item for item in blocks if item.get("type") != "arrow"]
        for index, block in enumerate(block_items):
            block_width = max(14.0, width * (float(block.get("width") or 0) / 100))
            block_height = max(10.0, height * (float(block.get("height") or 0) / 100))
            block_x = x + width * (float(block.get("x") or 0) / 100)
            block_y = y + height - height * ((float(block.get("y") or 0) + float(block.get("height") or 0)) / 100)
            block_color = normalize_hex_color(block.get("color"))
            text_color = readable_text_color(block_color)

            pdf.saveState()
            red = int(block_color[1:3], 16) / 255
            green = int(block_color[3:5], 16) / 255
            blue = int(block_color[5:7], 16) / 255
            pdf.setFillColor(colors.Color(red, green, blue, alpha=0.6))
            pdf.setStrokeColor(colors.Color(1, 1, 1, alpha=0.72))
            pdf.roundRect(block_x, block_y, block_width, block_height, 2, fill=1, stroke=1)
            pdf.setFillColor(pdf_white)
            title = str(block.get("title") or f"Blok {index + 1}").strip()
            exercise_title = str(block.get("exerciseTitle") or "Geen oefening").strip()
            font_size = max(5.5, min(9.5, block_height * 0.2, block_width * 0.08))
            pdf.setFont(bold_font, font_size)
            title_lines = split_text(title.upper(), max(8, block_width - 8), font_size, bold_font)[:1]
            exercise_lines = split_text(exercise_title, max(8, block_width - 8), max(5, font_size - 1.5), bold_font)[:1]
            text_y = block_y + block_height / 2 + (font_size * 0.45)
            for line in title_lines:
                text_width = stringWidth(line, bold_font, font_size)
                pdf.drawString(block_x + max(3, (block_width - text_width) / 2), text_y, line)
                text_y -= font_size + 2
            pdf.setFont(bold_font, max(5, font_size - 1.5))
            for line in exercise_lines:
                text_width = stringWidth(line, bold_font, max(5, font_size - 1.5))
                pdf.drawString(block_x + max(3, (block_width - text_width) / 2), text_y, line)
            pdf.restoreState()

    def get_field_period_label(period: Optional[Dict[str, Any]]) -> str:
        if not period:
            return ""
        label = str(period.get("label") or "").strip()
        start_time = str(period.get("startTime") or "").strip()
        end_time = str(period.get("endTime") or "").strip()
        time_label = f"{start_time or '--:--'}-{end_time or '--:--'}" if start_time or end_time else ""
        if label and time_label:
            return f"{label} | {time_label}"
        return label or time_label

    def draw_field_layout_page(
        blocks: List[Dict[str, Any]],
        page_index: int,
        training: Optional[Dict[str, Any]] = None,
        period: Optional[Dict[str, Any]] = None,
    ) -> None:
        draw_header("Veldplattegrond", page_index)
        period_label = get_field_period_label(period)
        if period_label:
            pdf.setFont(bold_font, 13)
            pdf.setFillColor(pdf_white)
            pdf.drawString(94, 424, period_label)
        draw_football_field(94, 58, 230, 355, blocks)

        selected_blocks = sort_football_field_blocks(blocks)
        rows = []
        listed_exercise_keys: Set[str] = set()
        for index, block in enumerate(selected_blocks, start=1):
            if block.get("sameExerciseExport") and (block.get("exerciseTitle") or block.get("exerciseId")):
                export_key = get_exercise_detail_export_key(block, index)
                if export_key in listed_exercise_keys:
                    continue
                listed_exercise_keys.add(export_key)
            block_title = str(block.get("title") or f"Blok {index}").strip()
            exercise_title = str(block.get("exerciseTitle") or "Geen oefening geselecteerd").strip()
            rows.append([str(index), block_title, exercise_title])
        if not rows:
            rows = [["-", "Nog geen blokken", "Nog geen oefening geselecteerd"]]

        draw_field_exercise_table(rows)
        draw_training_page_footer(training)

    def draw_exercise_detail_page(
        block: Dict[str, Any],
        page_index: int,
        fallback_index: int,
        training: Optional[Dict[str, Any]] = None,
    ) -> None:
        block_title = str(block.get("title") or f"Blok {fallback_index}").strip()
        exercise_title = str(block.get("exerciseTitle") or "Oefening").strip()
        exercise_kind = str(block.get("exerciseKind") or block.get("category") or "").strip()
        header_title = f"{block_title} - {exercise_title}"
        draw_exercise_header(header_title, page_index)

        details = block.get("exerciseDetails") if isinstance(block.get("exerciseDetails"), dict) else {}
        content_y = 58
        content_height = 328
        left_x = 50
        column_width = 268
        center_x = 340
        center_width = 280
        right_x = 642
        panel_gap = 12
        top_row_height = 228
        bottom_row_height = content_height - top_row_height - panel_gap
        top_row_y = content_y + bottom_row_height + panel_gap
        variation_height = (top_row_height - panel_gap) / 2

        draw_exercise_text_panel("Omschrijving oefening", details.get("description", ""), left_x, top_row_y, column_width, top_row_height, 7)
        draw_exercise_field_preview(block.get("exerciseField"), center_x, top_row_y, center_width, top_row_height, exercise_title)
        draw_exercise_text_panel("Variatie makkelijker maken", details.get("variationEasier", ""), right_x, top_row_y + variation_height + panel_gap, column_width, variation_height, 4)
        draw_exercise_text_panel("Variatie moeilijker maken", details.get("variationHarder", ""), right_x, top_row_y, column_width, variation_height, 4)

        coaching_label = f"Coaching - {exercise_kind}" if exercise_kind else "Coaching"
        draw_exercise_text_panel(coaching_label, details.get("coaching", ""), left_x, content_y, column_width, bottom_row_height, 4)
        draw_exercise_text_panel("Materialen", details.get("materials", ""), center_x, content_y, center_width, bottom_row_height, 4)
        draw_exercise_text_panel("Afmetingen", details.get("dimensions", ""), right_x, content_y, column_width, bottom_row_height, 4)
        draw_training_page_footer(training)

    page_index = 0
    draw_background(page_index, 0.24)
    logo_path = os.path.join(os.path.dirname(__file__), "static", "assets", "hws-logo.png")
    if os.path.exists(logo_path):
        cover_logo_size = 260
        pdf.drawImage(
            ImageReader(logo_path),
            (FOOTBALL_DAYS_PDF_WIDTH - cover_logo_size) / 2,
            242,
            cover_logo_size,
            cover_logo_size,
            mask="auto",
            preserveAspectRatio=True,
            anchor="c",
        )
    pdf.setFillColor(pdf_white)
    pdf.setFont(extra_bold_font, 42)
    cover_title = str(data.get("coverTitle") or "HWS VOETBALDAG").upper()
    title_width = stringWidth(cover_title, extra_bold_font, 42)
    pdf.drawString((FOOTBALL_DAYS_PDF_WIDTH - title_width) / 2, 156, cover_title)
    pdf.setFont(bold_font, 20)
    meta = str(data.get("coverMeta") or f"{data['clubName'].upper()} | {data['registrationCount']} AANMELDINGEN").upper()
    pdf.drawString((FOOTBALL_DAYS_PDF_WIDTH - stringWidth(meta, bold_font, 20)) / 2, 122, meta)
    pdf.showPage()

    page_index += 1
    draw_overview_page(page_index)
    pdf.showPage()

    field_training_rows = data.get("fieldTrainings") if isinstance(data.get("fieldTrainings"), list) else []
    no_training_rows = data.get("cycleNoTrainingDates") if isinstance(data.get("cycleNoTrainingDates"), list) else []
    if data.get("playbookType") == "samenwerkende-amateurclubs" and (len(field_training_rows) > 1 or no_training_rows):
        page_index += 1
        draw_training_dates_page(field_training_rows, no_training_rows, page_index)
        pdf.showPage()

    if data.get("includeStaff", True):
        staff_rows = data["staff"] or [{"name": "Nog in te vullen", "role": "", "setupTask": ""}]
        for chunk in chunk_items(staff_rows, 8):
            page_index += 1
            draw_staff_page(chunk, page_index)
            pdf.showPage()

    if data.get("includeProgram", True):
        program_rows = data["program"] or [{"startTime": "", "endTime": "", "activity": "Nog in te vullen"}]
        program_chunk_size = 14 if len(program_rows) <= 14 else 12
        program_layout_row_count = program_chunk_size if len(program_rows) > program_chunk_size else len(program_rows)
        for chunk in chunk_items(program_rows, program_chunk_size):
            page_index += 1
            draw_program_page(chunk, page_index, program_layout_row_count)
            pdf.showPage()

    if str(data.get("contingencies") or "").strip():
        contingency_lines = [
            line.strip()
            for line in str(data.get("contingencies") or "").splitlines()
            if line.strip()
        ]
        contingency_rows = []
        for line in contingency_lines:
            if ":" in line:
                scenario, solution = line.split(":", 1)
                contingency_rows.append([scenario.strip(), solution.strip()])
            else:
                contingency_rows.append(["Scenario", line])
        for chunk in chunk_items(contingency_rows, 7):
            page_index += 1
            draw_contingencies_page(chunk, page_index)
            pdf.showPage()

    def get_exercise_detail_export_key(block: Dict[str, Any], fallback_index: int) -> str:
        export_key = str(block.get("sameExerciseKey") or "").strip()
        if not export_key:
            exercise_id = int(block.get("exerciseId") or 0)
            exercise_title = normalize_match_text(block.get("exerciseTitle"))
            export_key = f"exercise:{exercise_id}" if exercise_id else f"title:{exercise_title}"
        if not export_key or export_key == "title:":
            export_key = f"block:{block.get('id') or fallback_index}"
        return export_key

    def select_exercise_detail_blocks(blocks: List[Dict[str, Any]], sort_blocks: bool = True) -> List[Dict[str, Any]]:
        exported_keys: Set[str] = set()
        selected_blocks: List[Dict[str, Any]] = []
        ordered_blocks = sort_football_field_blocks(blocks) if sort_blocks else blocks
        for exercise_index, block in enumerate(ordered_blocks, start=1):
            if not (block.get("exerciseTitle") or block.get("exerciseId")):
                continue
            if block.get("sameExerciseExport"):
                export_key = get_exercise_detail_export_key(block, exercise_index)
                if export_key in exported_keys:
                    continue
                exported_keys.add(export_key)
            selected_blocks.append(block)
        return selected_blocks

    if data.get("playbookType") == "samenwerkende-amateurclubs" and field_training_rows:
        for training_index, training in enumerate(field_training_rows, start=1):
            training_name = str(training.get("name") or f"Training {training_index}").strip()
            field_period_rows = training.get("fieldPeriods") if isinstance(training.get("fieldPeriods"), list) else []
            if not field_period_rows:
                field_period_rows = [{"label": "Plattegrond 1", "startTime": "", "endTime": "", "fieldLayout": training.get("fieldLayout") if isinstance(training.get("fieldLayout"), list) else []}]
            page_index += 1
            draw_training_cover_page(training, page_index)
            pdf.showPage()
            all_training_blocks: List[Dict[str, Any]] = []
            for period in field_period_rows:
                field_layout_rows = period.get("fieldLayout") if isinstance(period.get("fieldLayout"), list) else []
                page_index += 1
                draw_field_layout_page(field_layout_rows, page_index, training, period)
                pdf.showPage()
                all_training_blocks.extend(sort_football_field_blocks(field_layout_rows))
            selected_exercise_blocks = select_exercise_detail_blocks(all_training_blocks, sort_blocks=False)
            for exercise_index, block in enumerate(selected_exercise_blocks, start=1):
                page_index += 1
                draw_exercise_detail_page(block, page_index, exercise_index, training)
                pdf.showPage()
    else:
        field_layout_rows = data.get("fieldLayout") if isinstance(data.get("fieldLayout"), list) else []
        page_index += 1
        draw_field_layout_page(field_layout_rows, page_index)
        pdf.showPage()

        selected_exercise_blocks = select_exercise_detail_blocks(field_layout_rows)
        for exercise_index, block in enumerate(selected_exercise_blocks, start=1):
            page_index += 1
            draw_exercise_detail_page(block, page_index, exercise_index)
            pdf.showPage()

    pdf.save()
    buffer.seek(0)
    return buffer.read()


@app.post("/api/voetbaldagen/export-pdf")
def api_football_days_export_pdf():
    return export_football_playbook_pdf("voetbaldagen")


@app.post("/api/samenwerkende-amateurclubs/export-pdf")
def api_amateur_clubs_export_pdf():
    return export_football_playbook_pdf("samenwerkende-amateurclubs")


def export_football_playbook_pdf(playbook_type: str):
    context = get_football_playbook_context(playbook_type)
    access_redirect = require_page_access(context["pageKey"])
    if access_redirect is not None:
        return access_redirect

    payload = request.get_json(silent=True) or {}
    data = normalize_football_days_export_payload(payload, context["playbookType"])
    try:
        pdf_bytes = create_football_days_pdf(data)
    except RuntimeError as exc:
        return jsonify({"error": str(exc)}), 500

    filename = football_days_pdf_filename(data)
    return (
        pdf_bytes,
        200,
        {
            "Content-Type": "application/pdf",
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Cache-Control": "no-store",
        },
    )


CONTRACT_DEFAULTS = {
    "title": "Overeenkomst van opdracht",
    "clubName": "",
    "clubAddress": "",
    "season": "",
    "startDate": "",
    "endDate": "",
    "noticePeriod": "30 dagen (schriftelijk)",
    "trainingExecutionSummary": "Trainingen volgens schema",
    "trainingExecutionDetails": "De invulling wordt per club afgestemd op basis van het trainingsschema.",
    "agendaAttachmentTitle": "",
    "agendaAttachmentItems": [],
    "hwsMaterials": "Trainingsmaterialen",
    "clubMaterials": "Ballen, doelen, veld en opslag",
    "extraActivities": "Voetbaldagen en techniektrainingen mogelijk zonder veldhuur",
    "pricePerTraining": "",
    "trainingCount": 0,
    "totalAmount": "",
    "costLines": [{"description": "", "pricePerTraining": "", "trainingCount": 0, "totalAmount": ""}],
    "minPlayers": "Minimaal 4 spelers per training",
    "weatherCancellation": "Bij afgelasting door weer worden geen kosten gerekend",
    "hwsAbsence": "Bij afwezigheid vanuit HWS vindt verrekening plaats",
    "liability": "Alleen bij opzet of grove nalatigheid",
    "participationRisk": "Deelname op eigen risico",
    "evaluationMoments": "Herfstvakantie telefonisch\nApril/mei evaluatie",
    "hwsSignatory": "HWS Voetbalschool",
    "clubSignatory": "",
    "signingDate": "",
    "notes": "",
}


def normalize_contract_line(item: Any) -> Optional[Dict[str, str]]:
    if not isinstance(item, dict):
        return None
    line = {
        "day": str(item.get("day") or "").strip(),
        "time": str(item.get("time") or "").strip(),
        "team": str(item.get("team") or "").strip(),
        "trainingType": str(item.get("trainingType") or item.get("type") or "").strip(),
    }
    if not any(line.values()):
        return None
    return line


def normalize_contract_cost_line(item: Any) -> Optional[Dict[str, Any]]:
    if not isinstance(item, dict):
        return None
    description = str(item.get("description") or item.get("label") or "").strip()
    price_per_training = str(item.get("pricePerTraining") or item.get("price") or "").strip()
    total_amount = str(item.get("totalAmount") or item.get("total") or "").strip()
    try:
        training_count = max(0, int(str(item.get("trainingCount") or item.get("count") or "0").strip() or 0))
    except ValueError:
        training_count = 0
    if not total_amount and price_per_training and training_count:
        total_amount = format_contract_money(parse_decimal_amount(price_per_training) * Decimal(training_count))
    line = {
        "description": description,
        "pricePerTraining": price_per_training,
        "trainingCount": training_count,
        "totalAmount": total_amount,
    }
    if not (description or price_per_training or training_count or total_amount):
        return None
    return line


def normalize_contract_agenda_attachment_item(item: Any) -> Optional[Dict[str, Any]]:
    if not isinstance(item, dict):
        return None
    key = str(item.get("key") or "").strip()
    label = str(item.get("label") or "").strip()
    plan_type = str(item.get("planType") or item.get("plan_type") or "").strip()
    weekday = str(item.get("weekday") or "").strip()
    count = int(item.get("count") or 0)
    copy_text = str(item.get("copyText") or item.get("copy_text") or "").strip()
    days = item.get("days") if isinstance(item.get("days"), list) else []
    normalized_days = []
    for day in days:
        if not isinstance(day, dict):
            continue
        day_date = str(day.get("date") or "").strip()
        day_label = str(day.get("label") or "").strip()
        if day_date or day_label:
            normalized_days.append({"date": day_date, "label": day_label})
    if not key and plan_type and weekday:
        key = f"{plan_type}|{weekday}"
    if not label and plan_type:
        label = f"{plan_type} - {weekday}" if weekday else plan_type
    if not (key or label or copy_text or normalized_days):
        return None
    return {
        "key": key,
        "label": label,
        "planType": plan_type,
        "weekday": weekday,
        "count": count or len(normalized_days),
        "copyText": copy_text,
        "days": normalized_days,
    }


def normalize_contract(row: Optional[sqlite3.Row]) -> Dict[str, Any]:
    contract = dict(CONTRACT_DEFAULTS)
    contract["id"] = None
    contract["createdAt"] = ""
    contract["updatedAt"] = ""
    contract["trainingLines"] = [{"day": "", "time": "", "team": "", "trainingType": ""}]
    if row is None:
        return contract

    try:
        training_lines_payload = json.loads(str(row["training_lines_json"] or "[]"))
    except json.JSONDecodeError:
        training_lines_payload = []
    training_lines = [
        normalized_line
        for normalized_line in (normalize_contract_line(item) for item in training_lines_payload if isinstance(training_lines_payload, list))
        if normalized_line is not None
    ]
    if "cost_lines_json" in row.keys():
        try:
            cost_lines_payload = json.loads(str(row["cost_lines_json"] or "[]"))
        except json.JSONDecodeError:
            cost_lines_payload = []
    else:
        cost_lines_payload = []
    cost_lines = [
        normalized_line
        for normalized_line in (normalize_contract_cost_line(item) for item in cost_lines_payload if isinstance(cost_lines_payload, list))
        if normalized_line is not None
    ]
    if not cost_lines and (str(row["price_per_training"] or "").strip() or int(row["training_count"] or 0) or str(row["total_amount"] or "").strip()):
        cost_lines = [
            {
                "description": "",
                "pricePerTraining": str(row["price_per_training"] or "").strip(),
                "trainingCount": int(row["training_count"] or 0),
                "totalAmount": str(row["total_amount"] or "").strip(),
            }
        ]
    if "agenda_attachment_items_json" in row.keys():
        try:
            attachment_payload = json.loads(str(row["agenda_attachment_items_json"] or "[]"))
        except json.JSONDecodeError:
            attachment_payload = []
    else:
        attachment_payload = []
    agenda_attachment_items = [
        normalized_item
        for normalized_item in (normalize_contract_agenda_attachment_item(item) for item in attachment_payload if isinstance(attachment_payload, list))
        if normalized_item is not None
    ]

    contract.update(
        {
            "id": int(row["id"]),
            "title": str(row["title"] or CONTRACT_DEFAULTS["title"]).strip(),
            "clubName": str(row["club_name"] or "").strip(),
            "clubAddress": str(row["club_address"] or "").strip() if "club_address" in row.keys() else "",
            "season": str(row["season"] or "").strip(),
            "startDate": str(row["start_date"] or "").strip(),
            "endDate": str(row["end_date"] or "").strip(),
            "noticePeriod": str(row["notice_period"] or "").strip(),
            "trainingLines": training_lines or [{"day": "", "time": "", "team": "", "trainingType": ""}],
            "trainingExecutionSummary": str(row["training_execution_summary"] or CONTRACT_DEFAULTS["trainingExecutionSummary"]).strip() if "training_execution_summary" in row.keys() else CONTRACT_DEFAULTS["trainingExecutionSummary"],
            "trainingExecutionDetails": str(row["training_execution_details"] or CONTRACT_DEFAULTS["trainingExecutionDetails"]).strip() if "training_execution_details" in row.keys() else CONTRACT_DEFAULTS["trainingExecutionDetails"],
            "agendaAttachmentTitle": str(row["agenda_attachment_title"] or "").strip() if "agenda_attachment_title" in row.keys() else "",
            "agendaAttachmentItems": agenda_attachment_items,
            "hwsMaterials": str(row["hws_materials"] or "").strip(),
            "clubMaterials": str(row["club_materials"] or "").strip(),
            "extraActivities": str(row["extra_activities"] or "").strip(),
            "pricePerTraining": str(row["price_per_training"] or "").strip(),
            "trainingCount": int(row["training_count"] or 0),
            "totalAmount": str(row["total_amount"] or "").strip(),
            "costLines": cost_lines or [{"description": "", "pricePerTraining": "", "trainingCount": 0, "totalAmount": ""}],
            "minPlayers": str(row["min_players"] or "").strip(),
            "weatherCancellation": str(row["weather_cancellation"] or "").strip(),
            "hwsAbsence": str(row["hws_absence"] or "").strip(),
            "liability": str(row["liability"] or "").strip(),
            "participationRisk": str(row["participation_risk"] or "").strip(),
            "evaluationMoments": str(row["evaluation_moments"] or "").strip(),
            "hwsSignatory": str(row["hws_signatory"] or "").strip(),
            "clubSignatory": str(row["club_signatory"] or "").strip(),
            "signingDate": str(row["signing_date"] or "").strip(),
            "notes": str(row["notes"] or "").strip(),
            "createdAt": str(row["created_at"] or "").strip(),
            "updatedAt": str(row["updated_at"] or "").strip(),
        }
    )
    return contract


def load_contracts() -> List[Dict[str, Any]]:
    with get_db_connection() as connection:
        rows = connection.execute(
            """
            SELECT *
            FROM contracts
            ORDER BY COALESCE(NULLIF(start_date, ''), updated_at) DESC, id DESC
            """
        ).fetchall()
    return [normalize_contract(row) for row in rows]


def load_contract(contract_id: int) -> Optional[Dict[str, Any]]:
    with get_db_connection() as connection:
        row = connection.execute("SELECT * FROM contracts WHERE id = ?", (contract_id,)).fetchone()
    if row is None:
        return None
    return normalize_contract(row)


def parse_contract_season_start_year(value: Any) -> Optional[int]:
    match = re.search(r"(20\d{2})", str(value or ""))
    if not match:
        return None
    try:
        return int(match.group(1))
    except ValueError:
        return None


def get_contract_agenda_attachment_bounds(contract: Optional[Dict[str, Any]] = None) -> Tuple[date, Optional[date]]:
    contract = contract or {}
    season_start_year = parse_contract_season_start_year(contract.get("season"))
    start_date = parse_iso_date(str(contract.get("startDate") or "").strip())
    end_date = parse_iso_date(str(contract.get("endDate") or "").strip())

    if start_date is None:
        start_date = date(season_start_year or date.today().year, 8, 1)
    if end_date is None and season_start_year is not None:
        end_date = date(season_start_year + 1, 7, 31)
    return start_date, end_date


def filter_contract_agenda_day_plans_for_bounds(
    day_plans: List[Dict[str, Any]],
    contract: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, Any]]:
    start_date, end_date = get_contract_agenda_attachment_bounds(contract)
    filtered_day_plans: List[Dict[str, Any]] = []
    for day_plan in day_plans:
        current_date = day_plan.get("date")
        if isinstance(current_date, str):
            current_date = parse_iso_date(current_date.strip())
        if not isinstance(current_date, date):
            continue
        if current_date < start_date:
            continue
        if end_date is not None and current_date > end_date:
            continue
        filtered_day_plans.append(day_plan)
    return filtered_day_plans


def filter_contract_agenda_trainings_for_bounds(
    trainings: List[Dict[str, Any]],
    contract: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, Any]]:
    start_date, end_date = get_contract_agenda_attachment_bounds(contract)
    filtered_trainings: List[Dict[str, Any]] = []
    for training in trainings:
        current_date = training.get("date")
        if isinstance(current_date, str):
            current_date = parse_iso_date(current_date.strip())
        if not isinstance(current_date, date):
            continue
        if current_date < start_date:
            continue
        if end_date is not None and current_date > end_date:
            continue
        filtered_trainings.append(training)
    return filtered_trainings


def build_contract_agenda_attachment_options(contract: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
    day_plans = load_all_agenda_day_plans()
    trainings = load_agenda_trainings()
    filtered_day_plans = filter_contract_agenda_day_plans_for_bounds(day_plans, contract)
    filtered_trainings = filter_contract_agenda_trainings_for_bounds(trainings, contract)
    summary_items = build_agenda_day_plan_summary(add_football_day_only_no_activity_days(filtered_day_plans, filtered_trainings))
    options: List[Dict[str, Any]] = []
    for summary_item in summary_items:
        plan_type = str(summary_item.get("label") or "").strip()
        for detail in summary_item.get("details") or []:
            weekday = str(detail.get("label") or "").strip()
            key = f"{plan_type}|{weekday}"
            count = int(detail.get("count") or 0)
            days = detail.get("days") if isinstance(detail.get("days"), list) else []
            label = f"{plan_type} - {weekday} ({count} {'dag' if count == 1 else 'dagen'})"
            option = normalize_contract_agenda_attachment_item(
                {
                    "key": key,
                    "label": label,
                    "planType": plan_type,
                    "weekday": weekday,
                    "count": count,
                    "copyText": detail.get("copyText") or "",
                    "days": days,
                }
            )
            if option is not None:
                options.append(option)
    return options


def parse_contract_agenda_attachment_items_from_form(contract: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
    selected_keys = {str(key or "").strip() for key in request.form.getlist("agenda_attachment_keys") if str(key or "").strip()}
    if not selected_keys:
        return []
    option_map = {option["key"]: option for option in build_contract_agenda_attachment_options(contract)}
    selected_items = [
        normalize_contract_agenda_attachment_item(option_map[key])
        for key in sorted(selected_keys)
        if key in option_map
    ]
    return [item for item in selected_items if item is not None]


def parse_contract_lines_from_form() -> List[Dict[str, str]]:
    days = request.form.getlist("line_day")
    times = request.form.getlist("line_time")
    teams = request.form.getlist("line_team")
    training_types = request.form.getlist("line_training_type")
    lines: List[Dict[str, str]] = []
    for index, day in enumerate(days):
        line = normalize_contract_line(
            {
                "day": day,
                "time": times[index] if index < len(times) else "",
                "team": teams[index] if index < len(teams) else "",
                "trainingType": training_types[index] if index < len(training_types) else "",
            }
        )
        if line is not None:
            lines.append(line)
    return lines


def parse_contract_cost_lines_from_form() -> List[Dict[str, Any]]:
    descriptions = request.form.getlist("cost_description")
    prices = request.form.getlist("cost_price_per_training")
    counts = request.form.getlist("cost_training_count")
    totals = request.form.getlist("cost_total_amount")
    lines: List[Dict[str, Any]] = []
    max_length = max(len(descriptions), len(prices), len(counts), len(totals), 0)
    for index in range(max_length):
        line = normalize_contract_cost_line(
            {
                "description": descriptions[index] if index < len(descriptions) else "",
                "pricePerTraining": prices[index] if index < len(prices) else "",
                "trainingCount": counts[index] if index < len(counts) else "",
                "totalAmount": totals[index] if index < len(totals) else "",
            }
        )
        if line is not None:
            lines.append(line)
    return lines


def parse_decimal_amount(value: Any) -> Decimal:
    raw_value = str(value or "").strip()
    cleaned = re.sub(r"[^0-9,.-]", "", raw_value).replace(".", "").replace(",", ".")
    if not cleaned:
        return Decimal("0")
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return Decimal("0")


def format_contract_money(value: Decimal) -> str:
    return f"{value.quantize(Decimal('0.01'))}".replace(".", ",")


def build_contract_from_form() -> Dict[str, Any]:
    training_lines = parse_contract_lines_from_form()
    cost_lines = parse_contract_cost_lines_from_form()
    contract_context = {
        "season": request.form.get("season", "").strip(),
        "startDate": request.form.get("start_date", "").strip(),
        "endDate": request.form.get("end_date", "").strip(),
    }
    agenda_attachment_items = parse_contract_agenda_attachment_items_from_form(contract_context)
    first_cost_line = cost_lines[0] if cost_lines else {"pricePerTraining": "", "trainingCount": 0, "totalAmount": ""}
    training_count = sum(int(line.get("trainingCount") or 0) for line in cost_lines)
    total_decimal = sum(parse_decimal_amount(line.get("totalAmount")) for line in cost_lines)
    total_amount = format_contract_money(total_decimal) if total_decimal else str(first_cost_line.get("totalAmount") or "")

    return {
        "title": request.form.get("title", "").strip() or CONTRACT_DEFAULTS["title"],
        "clubName": request.form.get("club_name", "").strip(),
        "clubAddress": request.form.get("club_address", "").strip(),
        "season": contract_context["season"],
        "startDate": contract_context["startDate"],
        "endDate": contract_context["endDate"],
        "noticePeriod": request.form.get("notice_period", "").strip(),
        "trainingLines": training_lines,
        "trainingExecutionSummary": request.form.get("training_execution_summary", "").strip() or CONTRACT_DEFAULTS["trainingExecutionSummary"],
        "trainingExecutionDetails": request.form.get("training_execution_details", "").strip() or CONTRACT_DEFAULTS["trainingExecutionDetails"],
        "agendaAttachmentTitle": request.form.get("agenda_attachment_title", "").strip(),
        "agendaAttachmentItems": agenda_attachment_items,
        "costLines": cost_lines,
        "hwsMaterials": CONTRACT_DEFAULTS["hwsMaterials"],
        "clubMaterials": CONTRACT_DEFAULTS["clubMaterials"],
        "extraActivities": request.form.get("extra_activities", "").strip(),
        "pricePerTraining": str(first_cost_line.get("pricePerTraining") or ""),
        "trainingCount": training_count or int(first_cost_line.get("trainingCount") or 0),
        "totalAmount": total_amount,
        "minPlayers": CONTRACT_DEFAULTS["minPlayers"],
        "weatherCancellation": CONTRACT_DEFAULTS["weatherCancellation"],
        "hwsAbsence": CONTRACT_DEFAULTS["hwsAbsence"],
        "liability": CONTRACT_DEFAULTS["liability"],
        "participationRisk": CONTRACT_DEFAULTS["participationRisk"],
        "evaluationMoments": CONTRACT_DEFAULTS["evaluationMoments"],
        "hwsSignatory": CONTRACT_DEFAULTS["hwsSignatory"],
        "clubSignatory": "",
        "signingDate": "",
        "notes": "",
    }


def save_contract(contract: Dict[str, Any], contract_id: Optional[int] = None) -> int:
    now = utcnow_iso()
    cost_lines = get_contract_cost_lines(contract)
    first_cost_line = cost_lines[0] if cost_lines else {"pricePerTraining": "", "trainingCount": 0, "totalAmount": ""}
    training_count = sum(int(line.get("trainingCount") or 0) for line in cost_lines)
    total_decimal = sum(parse_decimal_amount(line.get("totalAmount")) for line in cost_lines)
    total_amount = format_contract_money(total_decimal) if total_decimal else str(contract.get("totalAmount") or first_cost_line.get("totalAmount") or "").strip()
    payload = (
        str(contract.get("title") or CONTRACT_DEFAULTS["title"]).strip(),
        str(contract.get("clubName") or "").strip(),
        str(contract.get("clubAddress") or "").strip(),
        str(contract.get("season") or "").strip(),
        str(contract.get("startDate") or "").strip(),
        str(contract.get("endDate") or "").strip(),
        str(contract.get("noticePeriod") or "").strip(),
        json.dumps(contract.get("trainingLines") or [], ensure_ascii=False),
        str(contract.get("trainingExecutionSummary") or CONTRACT_DEFAULTS["trainingExecutionSummary"]).strip(),
        str(contract.get("trainingExecutionDetails") or CONTRACT_DEFAULTS["trainingExecutionDetails"]).strip(),
        str(contract.get("agendaAttachmentTitle") or "").strip(),
        json.dumps(contract.get("agendaAttachmentItems") or [], ensure_ascii=False),
        str(contract.get("hwsMaterials") or "").strip(),
        str(contract.get("clubMaterials") or "").strip(),
        str(contract.get("extraActivities") or "").strip(),
        json.dumps(cost_lines, ensure_ascii=False),
        str(contract.get("pricePerTraining") or first_cost_line.get("pricePerTraining") or "").strip(),
        training_count or int(contract.get("trainingCount") or first_cost_line.get("trainingCount") or 0),
        total_amount,
        str(contract.get("minPlayers") or "").strip(),
        str(contract.get("weatherCancellation") or "").strip(),
        str(contract.get("hwsAbsence") or "").strip(),
        str(contract.get("liability") or "").strip(),
        str(contract.get("participationRisk") or "").strip(),
        str(contract.get("evaluationMoments") or "").strip(),
        str(contract.get("hwsSignatory") or "").strip(),
        str(contract.get("clubSignatory") or "").strip(),
        str(contract.get("signingDate") or "").strip(),
        str(contract.get("notes") or "").strip(),
    )
    with get_db_connection() as connection:
        if contract_id:
            connection.execute(
                """
                UPDATE contracts
                SET title = ?, club_name = ?, club_address = ?, season = ?, start_date = ?, end_date = ?,
                    notice_period = ?, training_lines_json = ?, training_execution_summary = ?,
                    training_execution_details = ?, agenda_attachment_title = ?,
                    agenda_attachment_items_json = ?, hws_materials = ?,
                    club_materials = ?, extra_activities = ?, cost_lines_json = ?,
                    price_per_training = ?, training_count = ?, total_amount = ?, min_players = ?,
                    weather_cancellation = ?, hws_absence = ?, liability = ?,
                    participation_risk = ?, evaluation_moments = ?, hws_signatory = ?,
                    club_signatory = ?, signing_date = ?, notes = ?, updated_at = ?
                WHERE id = ?
                """,
                (*payload, now, contract_id),
            )
            return contract_id

        cursor = connection.execute(
            """
            INSERT INTO contracts (
                title, club_name, club_address, season, start_date, end_date, notice_period,
                training_lines_json, training_execution_summary, training_execution_details,
                agenda_attachment_title, agenda_attachment_items_json,
                hws_materials, club_materials, extra_activities,
                cost_lines_json, price_per_training, training_count, total_amount, min_players,
                weather_cancellation, hws_absence, liability, participation_risk,
                evaluation_moments, hws_signatory, club_signatory, signing_date,
                notes, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (*payload, now, now),
        )
        return int(cursor.lastrowid)


def format_contract_date(value: Any) -> str:
    raw_value = str(value or "").strip()
    if not raw_value:
        return ""
    try:
        parsed = datetime.strptime(raw_value, "%Y-%m-%d").date()
    except ValueError:
        return raw_value
    return parsed.strftime("%d-%m-%Y")


def contract_training_schedule_text(contract: Dict[str, Any]) -> str:
    lines = []
    for item in contract.get("trainingLines") or []:
        if not isinstance(item, dict):
            continue
        values = [
            str(item.get("day") or "").strip() or "-",
            str(item.get("time") or "").strip() or "-",
            str(item.get("team") or "").strip() or "-",
            str(item.get("trainingType") or "").strip() or "-",
        ]
        lines.append(" | ".join(values))
    return "\n".join(lines) or "Dag | Tijd | Team | Type"


def contract_training_schedule_lines(contract: Dict[str, Any]) -> List[str]:
    lines = ["Dag | Tijd | Team | Type"]
    for item in contract.get("trainingLines") or []:
        if not isinstance(item, dict):
            continue
        values = [
            str(item.get("day") or "").strip() or "-",
            str(item.get("time") or "").strip() or "-",
            str(item.get("team") or "").strip() or "-",
            str(item.get("trainingType") or "").strip() or "-",
        ]
        lines.append(" | ".join(values))
    return lines


def get_contract_cost_lines(contract: Dict[str, Any]) -> List[Dict[str, Any]]:
    cost_lines = [
        normalized_line
        for normalized_line in (normalize_contract_cost_line(item) for item in contract.get("costLines") or [])
        if normalized_line is not None
    ]
    if cost_lines:
        return cost_lines
    fallback = normalize_contract_cost_line(
        {
            "pricePerTraining": contract.get("pricePerTraining"),
            "trainingCount": contract.get("trainingCount"),
            "totalAmount": contract.get("totalAmount"),
        }
    )
    return [fallback] if fallback is not None else []


def contract_cost_lines_for_docx(contract: Dict[str, Any]) -> List[str]:
    lines: List[str] = []
    cost_lines = get_contract_cost_lines(contract)
    if not cost_lines:
        return ["Tarief: €0,00", "Aantal trainingen: 0", "Totaal: €0,00"]
    for index, line in enumerate(cost_lines):
        description = str(line.get("description") or "").strip()
        suffix = f" ({description})" if description else ""
        if index > 0:
            lines.append("")
        lines.extend(
            [
                f"Tarief{suffix}: €{str(line.get('pricePerTraining') or '0,00').strip()}",
                f"Aantal trainingen{suffix}: {int(line.get('trainingCount') or 0)}",
                f"Totaal{suffix}: €{str(line.get('totalAmount') or '0,00').strip()}",
            ]
        )
    return lines


def build_contract_replacements(contract: Dict[str, Any]) -> Dict[str, str]:
    cost_lines = get_contract_cost_lines(contract)
    first_cost_line = cost_lines[0] if cost_lines else {}
    training_count = sum(int(line.get("trainingCount") or 0) for line in cost_lines)
    total_decimal = sum(parse_decimal_amount(line.get("totalAmount")) for line in cost_lines)
    total_amount = format_contract_money(total_decimal) if total_decimal else str(contract.get("totalAmount") or first_cost_line.get("totalAmount") or "").strip()
    training_types = sorted(
        {
            str(line.get("trainingType") or "").strip()
            for line in contract.get("trainingLines") or []
            if isinstance(line, dict) and str(line.get("trainingType") or "").strip()
        }
    )
    return {
        "[NAAM CLUB]": str(contract.get("clubName") or "").strip() or "Naam club",
        "[ADRES CLUB]": str(contract.get("clubAddress") or "").strip() or "Adres club",
        "[SEIZOEN]": str(contract.get("season") or "").strip() or "Seizoen",
        "[STARTDATUM]": format_contract_date(contract.get("startDate")) or "Startdatum",
        "[EINDDATUM]": format_contract_date(contract.get("endDate")) or "Einddatum",
        "[BEDRAG]": str(first_cost_line.get("pricePerTraining") or contract.get("pricePerTraining") or "").strip() or "0,00",
        "[AANTAL]": str(training_count or contract.get("trainingCount") or 0),
        "[TOTAAL]": total_amount or "0,00",
        "[TOTAALBEDRAG]": total_amount or "0,00",
        "[Type training]": str(contract.get("trainingExecutionSummary") or "").strip() or ", ".join(training_types) or CONTRACT_DEFAULTS["trainingExecutionSummary"],
        "[Variabel per club]": str(contract.get("trainingExecutionDetails") or "").strip() or CONTRACT_DEFAULTS["trainingExecutionDetails"],
        "[Club:]": f"{str(contract.get('clubName') or '').strip() or 'Club'}:",
    }


def build_docx_line_run(lines: List[Any]) -> str:
    normalized_lines = [str(line or "").strip() for line in lines]
    normalized_lines = [line for line in normalized_lines if line]
    if not normalized_lines:
        normalized_lines = [""]
    return "<w:br/>".join(f"<w:t>{html.escape(line)}</w:t>" for line in normalized_lines)


def set_docx_container_text(container: XmlElementTree.Element, value: Any) -> None:
    text_nodes = container.findall(".//w:t", DOCX_XML_NAMESPACES)
    if not text_nodes:
        return
    text_nodes[0].text = str(value or "")
    for node in text_nodes[1:]:
        node.text = ""


def ensure_docx_child(parent: XmlElementTree.Element, tag_name: str) -> XmlElementTree.Element:
    word_namespace = f"{{{DOCX_XML_NAMESPACES['w']}}}"
    child = parent.find(f"w:{tag_name}", DOCX_XML_NAMESPACES)
    if child is None:
        child = XmlElementTree.SubElement(parent, f"{word_namespace}{tag_name}")
    return child


def set_docx_training_cell_layout(cell: XmlElementTree.Element, width: int, no_wrap: bool = True) -> None:
    word_namespace = f"{{{DOCX_XML_NAMESPACES['w']}}}"
    cell_properties = ensure_docx_child(cell, "tcPr")
    cell_width = ensure_docx_child(cell_properties, "tcW")
    cell_width.set(f"{word_namespace}w", str(width))
    cell_width.set(f"{word_namespace}type", "dxa")
    cell_margin = ensure_docx_child(cell_properties, "tcMar")
    for side_name, side_width in (("top", "40"), ("left", "90"), ("bottom", "40"), ("right", "90")):
        side = ensure_docx_child(cell_margin, side_name)
        side.set(f"{word_namespace}w", side_width)
        side.set(f"{word_namespace}type", "dxa")
    if no_wrap and cell_properties.find("w:noWrap", DOCX_XML_NAMESPACES) is None:
        XmlElementTree.SubElement(cell_properties, f"{word_namespace}noWrap")
    for paragraph in cell.findall(".//w:p", DOCX_XML_NAMESPACES):
        paragraph_properties = ensure_docx_child(paragraph, "pPr")
        justification = ensure_docx_child(paragraph_properties, "jc")
        justification.set(f"{word_namespace}val", "left")
        indentation = ensure_docx_child(paragraph_properties, "ind")
        indentation.set(f"{word_namespace}left", "0")
        indentation.set(f"{word_namespace}firstLine", "0")
        indentation.set(f"{word_namespace}hanging", "0")


def set_docx_training_row_layout(row: XmlElementTree.Element, widths: List[int], no_wrap_columns: Set[int]) -> None:
    cells = row.findall("w:tc", DOCX_XML_NAMESPACES)
    for index, cell in enumerate(cells):
        width = widths[index] if index < len(widths) else widths[-1]
        set_docx_training_cell_layout(cell, width, index in no_wrap_columns)


def remove_docx_paragraphs_by_text(root: XmlElementTree.Element, texts_to_remove: Set[str]) -> None:
    for paragraph in list(root.findall(".//w:p", DOCX_XML_NAMESPACES)):
        paragraph_text = "".join(node.text or "" for node in paragraph.findall(".//w:t", DOCX_XML_NAMESPACES)).strip()
        if paragraph_text not in texts_to_remove:
            continue
        parent = None
        for candidate in root.iter():
            if paragraph in list(candidate):
                parent = candidate
                break
        if parent is not None:
            parent.remove(paragraph)


def fill_contract_training_table_xml(xml_bytes: bytes, contract: Dict[str, Any]) -> bytes:
    root = XmlElementTree.fromstring(xml_bytes)
    table = root.find(".//w:tbl", DOCX_XML_NAMESPACES)
    if table is None:
        return xml_bytes
    word_namespace = f"{{{DOCX_XML_NAMESPACES['w']}}}"
    column_widths = [1500, 2350, 1900, 3320]

    table_properties = ensure_docx_child(table, "tblPr")
    table_width = ensure_docx_child(table_properties, "tblW")
    table_width.set(f"{word_namespace}w", str(sum(column_widths)))
    table_width.set(f"{word_namespace}type", "dxa")
    table_layout = ensure_docx_child(table_properties, "tblLayout")
    table_layout.set(f"{word_namespace}type", "fixed")

    table_grid = table.find("w:tblGrid", DOCX_XML_NAMESPACES)
    if table_grid is None:
        table_grid = XmlElementTree.Element(f"{word_namespace}tblGrid")
        table.insert(1 if table.find("w:tblPr", DOCX_XML_NAMESPACES) is not None else 0, table_grid)
    for grid_column in list(table_grid):
        table_grid.remove(grid_column)
    for width in column_widths:
        grid_column = XmlElementTree.SubElement(table_grid, f"{word_namespace}gridCol")
        grid_column.set(f"{word_namespace}w", str(width))

    rows = table.findall("w:tr", DOCX_XML_NAMESPACES)
    if len(rows) < 2:
        return xml_bytes
    set_docx_training_row_layout(rows[0], column_widths, {0, 1, 2})

    training_lines = [
        line
        for line in (normalize_contract_line(item) for item in contract.get("trainingLines") or [])
        if line is not None
    ]
    if not training_lines:
        training_lines = [{"day": "-", "time": "-", "team": "-", "trainingType": "-"}]

    template_row = rows[1]
    data_rows = rows[1:]
    while len(data_rows) < len(training_lines):
        new_row = copy.deepcopy(template_row)
        table.append(new_row)
        data_rows.append(new_row)

    for row in data_rows[len(training_lines) :]:
        table.remove(row)

    for row, line in zip(data_rows, training_lines):
        cells = row.findall("w:tc", DOCX_XML_NAMESPACES)
        values = [
            line.get("day") or "-",
            line.get("time") or "-",
            line.get("team") or "-",
            line.get("trainingType") or "-",
        ]
        set_docx_training_row_layout(row, column_widths, {0, 1, 2})
        for cell, value in zip(cells, values):
            set_docx_container_text(cell, value)

    return XmlElementTree.tostring(root, encoding="utf-8", xml_declaration=True)


def apply_contract_word_page_breaks_xml(xml_bytes: bytes) -> bytes:
    root = XmlElementTree.fromstring(xml_bytes)
    remove_docx_paragraphs_by_text(root, {"(Aanpasbaar per club)"})
    word_namespace = f"{{{DOCX_XML_NAMESPACES['w']}}}"
    for paragraph in root.findall(".//w:p", DOCX_XML_NAMESPACES):
        if not paragraph.findall(".//w:lastRenderedPageBreak", DOCX_XML_NAMESPACES):
            continue
        paragraph_properties = paragraph.find("w:pPr", DOCX_XML_NAMESPACES)
        if paragraph_properties is None:
            paragraph_properties = XmlElementTree.Element(f"{word_namespace}pPr")
            paragraph.insert(0, paragraph_properties)
        if paragraph_properties.find("w:pageBreakBefore", DOCX_XML_NAMESPACES) is None:
            paragraph_properties.insert(0, XmlElementTree.Element(f"{word_namespace}pageBreakBefore"))
    return XmlElementTree.tostring(root, encoding="utf-8", xml_declaration=True)


def adjust_contract_header_logo_xml(xml_bytes: bytes) -> bytes:
    root = XmlElementTree.fromstring(xml_bytes)
    logo_width = "505000"
    logo_height = "633500"
    anchor = root.find(".//wp:anchor", DOCX_XML_NAMESPACES)
    if anchor is None:
        return xml_bytes

    vertical_offset = anchor.find("wp:positionV/wp:posOffset", DOCX_XML_NAMESPACES)
    if vertical_offset is not None:
        vertical_offset.text = "40000"

    extent = anchor.find("wp:extent", DOCX_XML_NAMESPACES)
    if extent is not None:
        extent.set("cx", logo_width)
        extent.set("cy", logo_height)

    graphic_extent = anchor.find(".//a:xfrm/a:ext", DOCX_XML_NAMESPACES)
    if graphic_extent is not None:
        graphic_extent.set("cx", logo_width)
        graphic_extent.set("cy", logo_height)

    word_namespace = f"{{{DOCX_XML_NAMESPACES['w']}}}"
    for paragraph in root.findall(".//w:p", DOCX_XML_NAMESPACES):
        text = "".join(node.text or "" for node in paragraph.findall(".//w:t", DOCX_XML_NAMESPACES))
        if text != "HWS VOETBALSCHOOL":
            continue
        for run_properties in paragraph.findall(".//w:rPr", DOCX_XML_NAMESPACES):
            position = run_properties.find("w:position", DOCX_XML_NAMESPACES)
            if position is None:
                position = XmlElementTree.Element(f"{word_namespace}position")
                run_properties.append(position)
            position.set(f"{word_namespace}val", "10")

    return XmlElementTree.tostring(root, encoding="utf-8", xml_declaration=True)


def add_contract_watermark_xml(xml_bytes: bytes) -> bytes:
    root = XmlElementTree.fromstring(xml_bytes)
    body = root.find("w:body", DOCX_XML_NAMESPACES)
    source_anchor = root.find(".//wp:anchor", DOCX_XML_NAMESPACES)
    if body is None or source_anchor is None:
        return xml_bytes

    word_namespace = f"{{{DOCX_XML_NAMESPACES['w']}}}"
    drawing_namespace = f"{{{DOCX_XML_NAMESPACES['wp']}}}"
    a_namespace = f"{{{DOCX_XML_NAMESPACES['a']}}}"
    watermark_anchor = copy.deepcopy(source_anchor)
    watermark_anchor.set("behindDoc", "1")
    watermark_anchor.set("relativeHeight", "1")
    watermark_anchor.set("distT", "0")
    watermark_anchor.set("distB", "0")
    watermark_anchor.set("distL", "0")
    watermark_anchor.set("distR", "0")
    for attr_name in list(watermark_anchor.attrib):
        if attr_name.endswith("anchorId"):
            watermark_anchor.set(attr_name, "0A0B0C0D")
        if attr_name.endswith("editId"):
            watermark_anchor.set(attr_name, "0E0F0102")

    position_h = watermark_anchor.find("wp:positionH", DOCX_XML_NAMESPACES)
    if position_h is not None:
        position_h.set("relativeFrom", "page")
        for child in list(position_h):
            position_h.remove(child)
        align = XmlElementTree.Element(f"{drawing_namespace}align")
        align.text = "center"
        position_h.append(align)

    position_v = watermark_anchor.find("wp:positionV", DOCX_XML_NAMESPACES)
    if position_v is not None:
        position_v.set("relativeFrom", "page")
        for child in list(position_v):
            position_v.remove(child)
        align = XmlElementTree.Element(f"{drawing_namespace}align")
        align.text = "center"
        position_v.append(align)

    for wrap_node in watermark_anchor.findall("wp:wrapTight", DOCX_XML_NAMESPACES):
        child_index = list(watermark_anchor).index(wrap_node)
        watermark_anchor.remove(wrap_node)
        watermark_anchor.insert(child_index, XmlElementTree.Element(f"{drawing_namespace}wrapNone"))

    watermark_width = "3500000"
    watermark_height = "4392857"
    extent = watermark_anchor.find("wp:extent", DOCX_XML_NAMESPACES)
    if extent is not None:
        extent.set("cx", watermark_width)
        extent.set("cy", watermark_height)
    graphic_extent = watermark_anchor.find(".//a:xfrm/a:ext", DOCX_XML_NAMESPACES)
    if graphic_extent is not None:
        graphic_extent.set("cx", watermark_width)
        graphic_extent.set("cy", watermark_height)

    doc_pr = watermark_anchor.find("wp:docPr", DOCX_XML_NAMESPACES)
    if doc_pr is not None:
        doc_pr.set("id", "1978967948")
        doc_pr.set("name", "HWS watermerk")
        doc_pr.set("descr", "HWS logo watermerk")

    for element in watermark_anchor.iter():
        if element.tag.endswith("cNvPr"):
            element.set("id", "1978967948")
            element.set("name", "HWS watermerk")
            element.set("descr", "HWS logo watermerk")

    blip = watermark_anchor.find(".//a:blip", DOCX_XML_NAMESPACES)
    if blip is not None:
        blip.set(f"{{{DOCX_XML_NAMESPACES['r']}}}embed", CONTRACT_WATERMARK_REL_ID)
        for alpha_node in blip.findall("a:alphaModFix", DOCX_XML_NAMESPACES):
            blip.remove(alpha_node)

    watermark_paragraph = XmlElementTree.Element(f"{word_namespace}p")
    watermark_run = XmlElementTree.SubElement(watermark_paragraph, f"{word_namespace}r")
    watermark_drawing = XmlElementTree.SubElement(watermark_run, f"{word_namespace}drawing")
    watermark_drawing.append(watermark_anchor)
    body.insert(0, watermark_paragraph)

    return XmlElementTree.tostring(root, encoding="utf-8", xml_declaration=True)


def append_contract_agenda_attachment_xml(xml_bytes: bytes, contract: Dict[str, Any]) -> bytes:
    attachment_items = [
        item
        for item in (normalize_contract_agenda_attachment_item(item) for item in contract.get("agendaAttachmentItems") or [])
        if item is not None
    ]
    if not attachment_items:
        return xml_bytes

    root = XmlElementTree.fromstring(xml_bytes)
    body = root.find("w:body", DOCX_XML_NAMESPACES)
    if body is None:
        return xml_bytes

    word_namespace = f"{{{DOCX_XML_NAMESPACES['w']}}}"
    sect_pr = body.find("w:sectPr", DOCX_XML_NAMESPACES)
    insert_index = list(body).index(sect_pr) if sect_pr is not None else len(list(body))

    def make_paragraph(text: str = "", *, bold: bool = False, size: Optional[str] = None, page_break: bool = False) -> XmlElementTree.Element:
        paragraph = XmlElementTree.Element(f"{word_namespace}p")
        run = XmlElementTree.SubElement(paragraph, f"{word_namespace}r")
        if page_break:
            br = XmlElementTree.SubElement(run, f"{word_namespace}br")
            br.set(f"{word_namespace}type", "page")
        if text:
            run_properties = XmlElementTree.SubElement(run, f"{word_namespace}rPr")
            fonts = XmlElementTree.SubElement(run_properties, f"{word_namespace}rFonts")
            fonts.set(f"{word_namespace}ascii", "Poppins Medium")
            fonts.set(f"{word_namespace}hAnsi", "Poppins Medium")
            if bold:
                XmlElementTree.SubElement(run_properties, f"{word_namespace}b")
                XmlElementTree.SubElement(run_properties, f"{word_namespace}bCs")
            if size:
                font_size = XmlElementTree.SubElement(run_properties, f"{word_namespace}sz")
                font_size.set(f"{word_namespace}val", size)
                font_size_cs = XmlElementTree.SubElement(run_properties, f"{word_namespace}szCs")
                font_size_cs.set(f"{word_namespace}val", size)
            text_node = XmlElementTree.SubElement(run, f"{word_namespace}t")
            text_node.text = text
        return paragraph

    paragraphs = [
        make_paragraph(page_break=True),
        make_paragraph(str(contract.get("agendaAttachmentTitle") or "").strip() or "Bijlage agenda", bold=True, size="32"),
        make_paragraph(""),
    ]
    for item in attachment_items:
        paragraphs.append(make_paragraph(str(item.get("label") or "").strip(), bold=True, size="24"))
        copy_text = str(item.get("copyText") or "").strip()
        if copy_text:
            for line in copy_text.splitlines():
                paragraphs.append(make_paragraph(line, size="20"))
        else:
            for index, day in enumerate(item.get("days") or [], start=1):
                day_label = str(day.get("label") or day.get("date") or "").strip()
                if day_label:
                    paragraphs.append(make_paragraph(f"{index}. {day_label}", size="20"))
        paragraphs.append(make_paragraph(""))

    for offset, paragraph in enumerate(paragraphs):
        body.insert(insert_index + offset, paragraph)
    return XmlElementTree.tostring(root, encoding="utf-8", xml_declaration=True)


def add_contract_watermark_relationship_xml(xml_bytes: bytes) -> bytes:
    root = XmlElementTree.fromstring(xml_bytes)
    relationship_namespace = DOCX_XML_NAMESPACES["rel"]
    relationship_tag = f"{{{relationship_namespace}}}Relationship"
    for relationship in root.findall(f".//{relationship_tag}"):
        if relationship.get("Id") == CONTRACT_WATERMARK_REL_ID:
            return xml_bytes
    XmlElementTree.SubElement(
        root,
        relationship_tag,
        {
            "Id": CONTRACT_WATERMARK_REL_ID,
            "Type": "http://schemas.openxmlformats.org/officeDocument/2006/relationships/image",
            "Target": "media/HWS_watermark.png",
        },
    )
    return XmlElementTree.tostring(root, encoding="utf-8", xml_declaration=True)


def fill_contract_document_xml(xml_bytes: bytes, contract: Dict[str, Any]) -> bytes:
    xml_text = xml_bytes.decode("utf-8")
    xml_text = xml_text.replace(
        "<w:t>Tarief: €[BEDRAG]</w:t><w:br/><w:t>Aantal trainingen: [AANTAL]</w:t><w:br/><w:t>Totaal: €[TOTAAL]</w:t>",
        build_docx_line_run(contract_cost_lines_for_docx(contract)),
    )
    replacements = build_contract_replacements(contract)
    for placeholder, value in replacements.items():
        xml_text = xml_text.replace(placeholder, html.escape(value))
    xml_text = xml_text.replace(
        "<w:t>Opzegtermijn: 30 dagen (schriftelijk)</w:t>",
        f"<w:t>Opzegtermijn: {html.escape(str(contract.get('noticePeriod') or ''))}</w:t>",
    )
    xml_text = xml_text.replace(
        "<w:t>Voeg hier eenvoudig je schema toe:</w:t><w:br/><w:t>Dag | Tijd | Team | Type</w:t>",
        build_docx_line_run(["Voeg hier eenvoudig je schema toe:", *contract_training_schedule_lines(contract)]),
    )
    execution_details = str(contract.get("trainingExecutionDetails") or CONTRACT_DEFAULTS["trainingExecutionDetails"]).strip()
    xml_text = xml_text.replace("<w:t>[Variabel per club]</w:t>", build_docx_line_run(execution_details.splitlines()))
    xml_text = xml_text.replace("Voetbaldagen en techniektrainingen mogelijk zonder veldhuur", html.escape(str(contract.get("extraActivities") or "")))
    xml_text = xml_text.replace(
        "Bij afwezigheid van een HWS-trainer wordt de training verrekend",
        "Bij afwezigheid van een HWS-trainer wordt de training verrekend op de eerstvolgende factuur",
    )
    xml_output = fill_contract_training_table_xml(xml_text.encode("utf-8"), contract)
    xml_output = apply_contract_word_page_breaks_xml(xml_output)
    xml_output = adjust_contract_header_logo_xml(xml_output)
    xml_output = append_contract_agenda_attachment_xml(xml_output, contract)
    # LibreOffice renders Word's floating logo behind the paragraph border, which makes
    # the title underline cut through the shield in exported PDFs. Word visually keeps
    # the logo on top, so force that stacking order only for the export document.
    return xml_output.replace(b'behindDoc="1"', b'behindDoc="0"')


def create_contract_docx(contract: Dict[str, Any]) -> bytes:
    if not os.path.exists(CONTRACT_TEMPLATE_PATH):
        raise RuntimeError("Het standaard Word-template ontbreekt op de server.")
    output = BytesIO()
    with zipfile.ZipFile(CONTRACT_TEMPLATE_PATH, "r") as source_archive:
        with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as target_archive:
            for item in source_archive.infolist():
                data = source_archive.read(item.filename)
                if item.filename == "word/document.xml":
                    data = fill_contract_document_xml(data, contract)
                target_archive.writestr(item, data)
    output.seek(0)
    return output.read()


def contract_filename(contract: Dict[str, Any], extension: str) -> str:
    club_name = re.sub(r"[^a-zA-Z0-9_-]+", "-", str(contract.get("clubName") or "overeenkomst")).strip("-")
    season = re.sub(r"[^a-zA-Z0-9_-]+", "-", str(contract.get("season") or "")).strip("-")
    base = "-".join(part for part in ["overeenkomst", club_name, season] if part)
    return f"{base.lower()}.{extension}"


def find_office_converter() -> str:
    converter = shutil.which("libreoffice") or shutil.which("soffice")
    if not converter:
        raise RuntimeError("PDF-export vereist LibreOffice/soffice op de server.")
    return converter


def convert_contract_docx_to_pdf(docx_bytes: bytes) -> bytes:
    converter = find_office_converter()
    with tempfile.TemporaryDirectory(prefix="contract-export-") as temp_dir:
        docx_path = os.path.join(temp_dir, "overeenkomst.docx")
        pdf_path = os.path.join(temp_dir, "overeenkomst.pdf")
        with open(docx_path, "wb") as docx_file:
            docx_file.write(docx_bytes)
        office_profile_dir = os.path.join(temp_dir, "office-profile")
        os.makedirs(office_profile_dir, exist_ok=True)
        office_env = os.environ.copy()
        office_env["HOME"] = temp_dir
        office_env["XDG_RUNTIME_DIR"] = temp_dir
        completed = subprocess.run(
            [
                converter,
                "--headless",
                f"-env:UserInstallation=file://{office_profile_dir}",
                "--convert-to",
                "pdf",
                "--outdir",
                temp_dir,
                docx_path,
            ],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            env=office_env,
            timeout=60,
            check=False,
        )
        if completed.returncode != 0 or not os.path.exists(pdf_path):
            error_output = (completed.stderr or completed.stdout or b"").decode("utf-8", errors="ignore").strip()
            raise RuntimeError(f"PDF-export via Word-template mislukt. {error_output}".strip())
        with open(pdf_path, "rb") as pdf_file:
            return pdf_file.read()


def create_contract_pdf(contract: Dict[str, Any]) -> bytes:
    return add_contract_pdf_watermark(convert_contract_docx_to_pdf(create_contract_docx(contract)))


def add_contract_pdf_watermark(pdf_bytes: bytes) -> bytes:
    if not os.path.exists(CONTRACT_WATERMARK_PATH):
        raise RuntimeError("Het HWS watermerk ontbreekt op de server.")
    try:
        from pypdf import PdfReader, PdfWriter
        from reportlab.lib.utils import ImageReader
        from reportlab.pdfgen import canvas
    except Exception as exc:  # pragma: no cover - depends on server packages
        raise RuntimeError("PDF-watermerk vereist pypdf en reportlab op de server.") from exc

    reader = PdfReader(BytesIO(pdf_bytes))
    writer = PdfWriter()
    watermark_image = ImageReader(CONTRACT_WATERMARK_PATH)

    for page in reader.pages:
        page_width = float(page.mediabox.width)
        page_height = float(page.mediabox.height)
        image_width, image_height = watermark_image.getSize()
        watermark_width = page_width * 0.42
        watermark_height = watermark_width * image_height / image_width
        x = (page_width - watermark_width) / 2
        y = (page_height - watermark_height) / 2

        overlay_buffer = BytesIO()
        overlay = canvas.Canvas(overlay_buffer, pagesize=(page_width, page_height))
        overlay.drawImage(
            watermark_image,
            x,
            y,
            width=watermark_width,
            height=watermark_height,
            preserveAspectRatio=True,
            mask="auto",
        )
        overlay.save()
        overlay_buffer.seek(0)
        watermark_page = PdfReader(overlay_buffer).pages[0]
        page.merge_page(watermark_page, over=False)
        writer.add_page(page)

    output = BytesIO()
    writer.write(output)
    return output.getvalue()


def create_contract_pdf_reportlab_fallback(contract: Dict[str, Any]) -> bytes:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError as exc:
        raise RuntimeError("PDF-export is niet beschikbaar omdat ReportLab ontbreekt.") from exc

    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=20 * mm,
        rightMargin=20 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        title=contract.get("title") or "Overeenkomst",
    )
    styles = {
        "title": ParagraphStyle("ContractTitle", fontName="Helvetica-Bold", fontSize=17, leading=21, spaceAfter=12),
        "heading": ParagraphStyle("ContractHeading", fontName="Helvetica-Bold", fontSize=11.5, leading=15, spaceBefore=10, spaceAfter=5),
        "body": ParagraphStyle("ContractBody", fontName="Helvetica", fontSize=9.5, leading=13, spaceAfter=4),
        "small": ParagraphStyle("ContractSmall", fontName="Helvetica", fontSize=8.8, leading=12),
    }

    def p(text: Any, style: str = "body") -> Paragraph:
        escaped = html.escape(str(text or "")).replace("\n", "<br/>")
        return Paragraph(escaped or "-", styles[style])

    story = [
        p(f"OVEREENKOMST VAN OPDRACHT HWS VOETBALSCHOOL - {contract.get('clubName') or 'Naam club'} - {contract.get('season') or 'Seizoen'}", "title"),
        p("1. Duur van de Overeenkomst", "heading"),
        p(f"Startdatum: {format_contract_date(contract.get('startDate')) or '-'}\nEinddatum: {format_contract_date(contract.get('endDate')) or '-'}\nOpzegtermijn: {contract.get('noticePeriod') or '-'}"),
        p("2. Trainingen", "heading"),
    ]

    schedule_rows = [["Dag", "Tijd", "Team", "Type"]]
    for line in contract.get("trainingLines") or []:
        schedule_rows.append([
            str(line.get("day") or "-"),
            str(line.get("time") or "-"),
            str(line.get("team") or "-"),
            str(line.get("trainingType") or "-"),
        ])
    schedule_table = Table(schedule_rows, colWidths=[34 * mm, 30 * mm, 55 * mm, 47 * mm], repeatRows=1)
    schedule_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111111")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#C8C8C8")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.extend(
        [
            schedule_table,
            Spacer(1, 5 * mm),
            p("3. Materialen en faciliteiten", "heading"),
            p(f"HWS: {contract.get('hwsMaterials') or '-'}\nClub: {contract.get('clubMaterials') or '-'}"),
            p("4. Extra activiteiten", "heading"),
            p(contract.get("extraActivities") or "-"),
            p("5. Kosten", "heading"),
            p(f"Tarief: EUR {contract.get('pricePerTraining') or '-'}\nAantal trainingen: {contract.get('trainingCount') or 0}\nTotaal: EUR {contract.get('totalAmount') or '-'}"),
            p("6. Afgelastingen", "heading"),
            p(f"{contract.get('minPlayers') or '-'}\nWeer: {contract.get('weatherCancellation') or '-'}\nHWS afwezig: {contract.get('hwsAbsence') or '-'}"),
            p("7. Aansprakelijkheid", "heading"),
            p(f"{contract.get('liability') or '-'}\n{contract.get('participationRisk') or '-'}"),
            p("8. Evaluatie", "heading"),
            p(CONTRACT_DEFAULTS["evaluationMoments"]),
            p("9. Ondertekening", "heading"),
            p("HWS en Club ondertekenen hieronder"),
        ]
    )

    document.build(story)
    buffer.seek(0)
    return buffer.read()


def normalize_proposal_type(value: Any) -> str:
    normalized_value = str(value or "").strip().lower()
    valid_values = {str(option["value"]) for option in PROPOSAL_TYPE_OPTIONS}
    return normalized_value if normalized_value in valid_values else ""


def get_proposal_type_option(value: Any) -> Optional[Dict[str, Any]]:
    normalized_value = normalize_proposal_type(value)
    for option in PROPOSAL_TYPE_OPTIONS:
        if option["value"] == normalized_value:
            return option
    return None


def normalize_proposal_weekday(value: Any) -> str:
    normalized_value = str(value or "").strip().lower()
    valid_values = {str(option["value"]) for option in PROPOSAL_WEEKDAY_OPTIONS}
    return normalized_value if normalized_value in valid_values else ""


def get_proposal_weekday_option(value: Any) -> Optional[Dict[str, Any]]:
    normalized_value = normalize_proposal_weekday(value)
    for option in PROPOSAL_WEEKDAY_OPTIONS:
        if option["value"] == normalized_value:
            return option
    return None


def normalize_proposal_training_kind(value: Any) -> str:
    normalized_value = str(value or "").strip().lower()
    valid_values = {str(option["value"]) for option in PROPOSAL_TRAINING_KIND_OPTIONS}
    return normalized_value if normalized_value in valid_values else ""


def get_proposal_training_kind_option(value: Any) -> Optional[Dict[str, Any]]:
    normalized_value = normalize_proposal_training_kind(value)
    for option in PROPOSAL_TRAINING_KIND_OPTIONS:
        if option["value"] == normalized_value:
            return option
    return None


def normalize_proposal_line_time(value: Any) -> str:
    normalized_value = str(value or "").strip()
    if not normalized_value:
        return ""
    try:
        return datetime.strptime(normalized_value, "%H:%M").strftime("%H:%M")
    except ValueError:
        return ""


def normalize_price_input(value: Any) -> str:
    normalized = str(value or "").strip().replace(" ", "")
    if not normalized:
        return ""
    if "," in normalized and "." in normalized:
        normalized = normalized.replace(".", "").replace(",", ".")
    else:
        normalized = normalized.replace(",", ".")
    return normalized


def format_decimal_price(value: Decimal) -> str:
    formatted = f"{value.quantize(Decimal('0.01')):.2f}"
    return formatted


def build_proposal_form_state(
    club_name: str = "",
    proposal_type: str = "",
    season_start_year: str = "",
    price_per_training: str = "",
    lines: Optional[List[Dict[str, str]]] = None,
) -> Dict[str, Any]:
    cleaned_lines = []
    for line in (lines or []):
        cleaned_lines.append(
            {
                "weekday": normalize_proposal_weekday(line.get("weekday", "")),
                "time": normalize_proposal_line_time(line.get("time", "")),
                "trainingKind": normalize_proposal_training_kind(line.get("trainingKind", "")),
                "team": str(line.get("team", "")).strip(),
            }
        )

    if not cleaned_lines:
        cleaned_lines = [{"weekday": "", "time": "", "trainingKind": "", "team": ""}]

    return {
        "clubName": str(club_name or "").strip(),
        "proposalType": normalize_proposal_type(proposal_type),
        "seasonStartYear": str(season_start_year or "").strip(),
        "pricePerTraining": str(price_per_training or "").strip(),
        "lines": cleaned_lines,
    }


def parse_proposal_lines_from_form(form: Any) -> List[Dict[str, str]]:
    weekdays = form.getlist("line_weekday")
    times = form.getlist("line_time")
    training_kinds = form.getlist("line_training_kind")
    teams = form.getlist("line_team")
    line_count = max(len(weekdays), len(times), len(training_kinds), len(teams))
    lines: List[Dict[str, str]] = []

    for index in range(line_count):
        weekday = weekdays[index] if index < len(weekdays) else ""
        time_value = times[index] if index < len(times) else ""
        training_kind = training_kinds[index] if index < len(training_kinds) else ""
        team = teams[index] if index < len(teams) else ""
        lines.append(
            {
                "weekday": normalize_proposal_weekday(weekday),
                "time": normalize_proposal_line_time(time_value),
                "trainingKind": normalize_proposal_training_kind(training_kind),
                "team": str(team or "").strip(),
            }
        )

    return lines


def validate_proposal_input(
    club_name: str,
    proposal_type: str,
    season_start_year: str,
    price_per_training: str,
    lines: List[Dict[str, str]],
) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
    normalized_club_name = str(club_name or "").strip()
    normalized_type = normalize_proposal_type(proposal_type)
    normalized_price = normalize_price_input(price_per_training)

    if not normalized_club_name:
        return None, "Vul een clubnaam in."
    if not normalized_type:
        return None, "Kies of het om een samenwerkende amateurclub of techniektrainingen gaat."

    try:
        parsed_season_start_year = int(str(season_start_year or "").strip())
    except ValueError:
        return None, "Kies een geldig seizoen."

    available_seasons = {
        int(option["value"])
        for option in build_football_season_options(start_year=PROPOSAL_MIN_SEASON_START_YEAR)
        if str(option.get("value", "")).isdigit()
    }
    if parsed_season_start_year not in available_seasons:
        return None, "Kies een seizoen uit de lijst."

    if not normalized_price:
        return None, "Vul een bedrag per training in."

    price_decimal = decimal_from_value(normalized_price)
    if price_decimal < 0:
        return None, "Het bedrag per training mag niet negatief zijn."

    cleaned_lines: List[Dict[str, str]] = []
    has_partial_line = False
    for line in lines:
        weekday = normalize_proposal_weekday(line.get("weekday", ""))
        time_value = normalize_proposal_line_time(line.get("time", ""))
        training_kind = normalize_proposal_training_kind(line.get("trainingKind", ""))
        team = str(line.get("team", "")).strip()
        if not weekday and not time_value and not training_kind and not team:
            continue
        if not weekday or not time_value or not training_kind or not team:
            has_partial_line = True
            continue
        cleaned_lines.append(
            {
                "weekday": weekday,
                "time": time_value,
                "trainingKind": training_kind,
                "team": team,
            }
        )

    if has_partial_line:
        return None, "Vul per regel dag, tijd, soort en team in."
    if not cleaned_lines:
        return None, "Voeg minimaal een regel toe met dag, tijd, soort en team."

    return {
        "clubName": normalized_club_name,
        "proposalType": normalized_type,
        "seasonStartYear": parsed_season_start_year,
        "pricePerTraining": format_decimal_price(price_decimal),
        "lines": cleaned_lines,
    }, None


def calculate_training_counts_for_proposal(
    season_start_year: int,
    proposal_type: str,
    lines: List[Dict[str, Any]],
) -> Dict[str, Any]:
    proposal_type_option = get_proposal_type_option(proposal_type)
    if proposal_type_option is None:
        return {
            "countsByWeekday": {},
            "lineCounts": {},
            "totalTrainings": 0,
        }

    counts_by_weekday = build_proposal_weekday_counts(
        season_start_year,
        proposal_type_option["agenda_plan_type"],
    )
    line_counts: Dict[int, int] = {}
    total_trainings = 0
    for line in lines:
        weekday_key = normalize_proposal_weekday(line.get("weekdayKey", line.get("weekday", "")))
        line_id = int(line.get("id") or 0)
        count = counts_by_weekday.get(weekday_key, 0)
        if line_id:
            line_counts[line_id] = count
        total_trainings += count

    return {
        "countsByWeekday": counts_by_weekday,
        "lineCounts": line_counts,
        "totalTrainings": total_trainings,
    }


def build_proposal_weekday_counts(
    season_start_year: int,
    agenda_plan_type: Optional[str] = None,
) -> Dict[str, int]:
    season_range = get_football_season_range(season_start_year)
    with get_db_connection() as connection:
        query = """
            SELECT date
            FROM agenda_day_plans
            WHERE date >= ? AND date <= ?
        """
        params: List[str] = [
            season_range["start"].isoformat(),
            season_range["end"].isoformat(),
        ]
        if agenda_plan_type:
            query += " AND plan_type = ?"
            params.append(str(agenda_plan_type))
        query += " ORDER BY date ASC"
        rows = connection.execute(query, params).fetchall()

    counts_by_weekday = {
        str(option["value"]): 0
        for option in PROPOSAL_WEEKDAY_OPTIONS
    }
    weekday_lookup = {
        int(option["python_weekday"]): str(option["value"])
        for option in PROPOSAL_WEEKDAY_OPTIONS
    }

    for row in rows:
        current_date = parse_iso_date(str(row["date"] or "").strip())
        if current_date is None:
            continue
        weekday_key = weekday_lookup.get(current_date.weekday())
        if weekday_key:
            counts_by_weekday[weekday_key] = counts_by_weekday.get(weekday_key, 0) + 1

    return counts_by_weekday


def create_proposal(
    club_name: str,
    proposal_type: str,
    season_start_year: int,
    price_per_training: str,
    lines: List[Dict[str, str]],
) -> int:
    timestamp = utcnow_iso()
    with get_db_connection() as connection:
        cursor = connection.execute(
            """
            INSERT INTO proposals (
                club_name,
                proposal_type,
                season_start_year,
                price_per_training,
                total_trainings,
                total_amount,
                created_at,
                updated_at
            )
            VALUES (?, ?, ?, ?, 0, 0, ?, ?)
            """,
            (
                club_name.strip(),
                normalize_proposal_type(proposal_type),
                int(season_start_year),
                normalize_price_input(price_per_training),
                timestamp,
                timestamp,
            ),
        )
        proposal_id = int(cursor.lastrowid)
        connection.executemany(
            """
            INSERT INTO proposal_lines (
                proposal_id,
                weekday_key,
                line_time,
                training_kind,
                activity_description,
                training_count,
                sort_order,
                created_at
            )
            VALUES (?, ?, ?, ?, ?, 0, ?, ?)
            """,
            [
                (
                    proposal_id,
                    normalize_proposal_weekday(line.get("weekday", "")),
                    normalize_proposal_line_time(line.get("time", "")),
                    normalize_proposal_training_kind(line.get("trainingKind", "")),
                    str(line.get("team", "")).strip(),
                    index,
                    timestamp,
                )
                for index, line in enumerate(lines)
            ],
        )

    refresh_proposal_metrics(proposal_id)
    return proposal_id


def refresh_proposal_metrics(proposal_id: int) -> None:
    with get_db_connection() as connection:
        proposal_row = connection.execute(
            """
            SELECT id, proposal_type, season_start_year, price_per_training, total_trainings, total_amount
            FROM proposals
            WHERE id = ?
            """,
            (proposal_id,),
        ).fetchone()
        if proposal_row is None:
            return

        line_rows = connection.execute(
            """
            SELECT id, weekday_key
            FROM proposal_lines
            WHERE proposal_id = ?
            ORDER BY sort_order ASC, id ASC
            """,
            (proposal_id,),
        ).fetchall()

        counts_payload = calculate_training_counts_for_proposal(
            int(proposal_row["season_start_year"]),
            str(proposal_row["proposal_type"] or "").strip(),
            [
                {
                    "id": int(row["id"]),
                    "weekdayKey": str(row["weekday_key"] or "").strip(),
                }
                for row in line_rows
            ],
        )
        line_counts = counts_payload["lineCounts"]
        total_trainings = int(counts_payload["totalTrainings"])
        price_decimal = decimal_from_value(proposal_row["price_per_training"])
        total_amount = round(float(price_decimal * Decimal(total_trainings)), 2)

        connection.executemany(
            """
            UPDATE proposal_lines
            SET training_count = ?
            WHERE id = ?
            """,
            [
                (int(line_counts.get(int(row["id"]), 0)), int(row["id"]))
                for row in line_rows
            ],
        )
        current_total_trainings = int(proposal_row["total_trainings"] or 0)
        current_total_amount = round(float(proposal_row["total_amount"] or 0), 2)
        if current_total_trainings != total_trainings or current_total_amount != total_amount:
            connection.execute(
                """
                UPDATE proposals
                SET total_trainings = ?, total_amount = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    total_trainings,
                    total_amount,
                    utcnow_iso(),
                    proposal_id,
                ),
            )


def build_proposal_payload(proposal_row: sqlite3.Row, line_rows: List[sqlite3.Row]) -> Dict[str, Any]:
    proposal_type_option = get_proposal_type_option(proposal_row["proposal_type"])
    season_start_year = int(proposal_row["season_start_year"])
    price_decimal = decimal_from_value(proposal_row["price_per_training"])

    lines = []
    for row in line_rows:
        weekday_option = get_proposal_weekday_option(row["weekday_key"])
        training_kind_option = get_proposal_training_kind_option(row["training_kind"])
        training_count = int(row["training_count"] or 0)
        line_total = round(float(price_decimal * Decimal(training_count)), 2)
        lines.append(
            {
                "id": int(row["id"]),
                "weekdayKey": str(row["weekday_key"] or "").strip(),
                "weekdayLabel": weekday_option["label"] if weekday_option else str(row["weekday_key"] or "").strip(),
                "time": str(row["line_time"] or "").strip(),
                "trainingKind": str(row["training_kind"] or "").strip(),
                "trainingKindLabel": training_kind_option["label"] if training_kind_option else str(row["training_kind"] or "").strip(),
                "teamName": str(row["activity_description"] or "").strip(),
                "trainingCount": training_count,
                "lineTotalAmount": line_total,
                "lineTotalAmountLabel": format_currency(line_total),
            }
        )

    total_amount = round(float(proposal_row["total_amount"] or 0), 2)
    return {
        "id": int(proposal_row["id"]),
        "clubName": str(proposal_row["club_name"] or "").strip(),
        "proposalType": str(proposal_row["proposal_type"] or "").strip(),
        "proposalTypeLabel": proposal_type_option["label"] if proposal_type_option else str(proposal_row["proposal_type"] or "").strip(),
        "agendaPlanType": proposal_type_option["agenda_plan_type"] if proposal_type_option else "",
        "seasonStartYear": season_start_year,
        "seasonLabel": get_football_season_label(season_start_year),
        "pricePerTraining": format_decimal_price(price_decimal),
        "pricePerTrainingLabel": format_currency(float(price_decimal)),
        "totalTrainings": int(proposal_row["total_trainings"] or 0),
        "totalAmount": total_amount,
        "totalAmountLabel": format_currency(total_amount),
        "createdAt": str(proposal_row["created_at"] or "").strip(),
        "updatedAt": str(proposal_row["updated_at"] or "").strip(),
        "createdAtLabel": format_datetime_display(str(proposal_row["created_at"] or "").strip()),
        "updatedAtLabel": format_datetime_display(str(proposal_row["updated_at"] or "").strip()),
        "lines": lines,
    }


def load_proposal_by_id(proposal_id: int, refresh_metrics: bool = True) -> Optional[Dict[str, Any]]:
    if refresh_metrics:
        refresh_proposal_metrics(proposal_id)

    with get_db_connection() as connection:
        proposal_row = connection.execute(
            """
            SELECT
                id,
                club_name,
                proposal_type,
                season_start_year,
                price_per_training,
                total_trainings,
                total_amount,
                created_at,
                updated_at
            FROM proposals
            WHERE id = ?
            """,
            (proposal_id,),
        ).fetchone()
        if proposal_row is None:
            return None

        line_rows = connection.execute(
            """
            SELECT id, weekday_key, line_time, training_kind, activity_description, training_count, sort_order, created_at
            FROM proposal_lines
            WHERE proposal_id = ?
            ORDER BY sort_order ASC, id ASC
            """,
            (proposal_id,),
        ).fetchall()

    return build_proposal_payload(proposal_row, list(line_rows))


def load_proposals(refresh_metrics: bool = True) -> List[Dict[str, Any]]:
    with get_db_connection() as connection:
        proposal_ids = [
            int(row["id"])
            for row in connection.execute(
                """
                SELECT id
                FROM proposals
                ORDER BY created_at DESC, id DESC
                """
            ).fetchall()
        ]

    proposals = []
    for proposal_id in proposal_ids:
        proposal = load_proposal_by_id(proposal_id, refresh_metrics=refresh_metrics)
        if proposal is not None:
            proposals.append(proposal)
    return proposals


def delete_proposal(proposal_id: int) -> None:
    with get_db_connection() as connection:
        connection.execute(
            """
            DELETE FROM proposal_lines
            WHERE proposal_id = ?
            """,
            (proposal_id,),
        )
        connection.execute(
            """
            DELETE FROM proposals
            WHERE id = ?
            """,
            (proposal_id,),
        )


def load_social_media_ideas() -> List[Dict[str, Any]]:
    with get_db_connection() as connection:
        rows = connection.execute(
            """
            SELECT id, title, platform, content_type, priority, is_scheduled, notes, created_at
            FROM social_media_ideas
            ORDER BY created_at DESC, id DESC
            """
        ).fetchall()

    return [
        {
            "platforms": parse_social_media_platforms(row["platform"]),
            "id": int(row["id"]),
            "title": str(row["title"] or "").strip(),
            "platform": format_social_media_platforms(parse_social_media_platforms(row["platform"])),
            "contentType": str(row["content_type"] or "").strip(),
            "priority": str(row["priority"] or "Midden").strip() or "Midden",
            "isScheduled": bool(row["is_scheduled"]),
            "notes": str(row["notes"] or "").strip(),
            "createdAt": str(row["created_at"] or "").strip(),
        }
        for row in rows
    ]


def parse_social_media_platforms(raw_value: Any) -> List[str]:
    values: List[str] = []
    seen: set[str] = set()

    for part in str(raw_value or "").split(","):
        platform = part.strip()
        if not platform or platform in seen:
            continue
        values.append(platform)
        seen.add(platform)

    return values


def format_social_media_platforms(platforms: List[str]) -> str:
    return ", ".join(platforms)


def add_social_media_idea(title: str, platforms: List[str], content_type: str, priority: str, notes: str) -> None:
    with get_db_connection() as connection:
        connection.execute(
            """
            INSERT INTO social_media_ideas (title, platform, content_type, priority, notes, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                title.strip(),
                format_social_media_platforms(platforms),
                content_type.strip(),
                priority.strip() or "Midden",
                notes.strip(),
                utcnow_iso(),
            ),
        )


def update_social_media_idea(idea_id: int, title: str, platforms: List[str], content_type: str, priority: str, notes: str) -> None:
    with get_db_connection() as connection:
        connection.execute(
            """
            UPDATE social_media_ideas
            SET
                title = ?,
                platform = ?,
                content_type = ?,
                priority = ?,
                notes = ?
            WHERE id = ?
            """,
            (
                title.strip(),
                format_social_media_platforms(platforms),
                content_type.strip(),
                priority.strip() or "Midden",
                notes.strip(),
                idea_id,
            ),
        )


def set_social_media_idea_scheduled(idea_id: int, is_scheduled: bool) -> None:
    with get_db_connection() as connection:
        connection.execute(
            "UPDATE social_media_ideas SET is_scheduled = ? WHERE id = ?",
            (1 if is_scheduled else 0, idea_id),
        )


def delete_social_media_idea(idea_id: int) -> None:
    with get_db_connection() as connection:
        connection.execute("DELETE FROM social_media_ideas WHERE id = ?", (idea_id,))


def load_social_media_schedule() -> List[Dict[str, Any]]:
    with get_db_connection() as connection:
        rows = connection.execute(
            """
            SELECT id, title, platform, publish_date, publish_time, status, notes, created_at
            FROM social_media_schedule
            ORDER BY publish_date ASC, publish_time ASC, id ASC
            """
        ).fetchall()

    return [
        {
            "id": int(row["id"]),
            "title": str(row["title"] or "").strip(),
            "platform": str(row["platform"] or "").strip(),
            "publishDate": str(row["publish_date"] or "").strip(),
            "publishTime": str(row["publish_time"] or "").strip(),
            "status": str(row["status"] or "").strip(),
            "notes": str(row["notes"] or "").strip(),
            "createdAt": str(row["created_at"] or "").strip(),
        }
        for row in rows
    ]


def build_social_media_week_events(schedule_items: List[Dict[str, Any]], week_start: date) -> List[Dict[str, Any]]:
    calendar_start_minutes = 0
    pixels_per_hour = 56
    week_end = week_start + timedelta(days=6)
    events = []

    for item in schedule_items:
        publish_date = str(item.get("publishDate", "")).strip()
        publish_time = str(item.get("publishTime", "")).strip()
        if not publish_date or not publish_time:
            continue

        item_date = date.fromisoformat(publish_date)
        if item_date < week_start or item_date > week_end:
            continue

        start_dt = combine_date_and_time(publish_date, publish_time)
        end_dt = start_dt + timedelta(minutes=60)
        start_minutes = start_dt.hour * 60 + start_dt.minute
        end_minutes = end_dt.hour * 60 + end_dt.minute
        top = max(((start_minutes - calendar_start_minutes) / 60) * pixels_per_hour, 0)
        height = max(((end_minutes - start_minutes) / 60) * pixels_per_hour, 48)

        events.append(
            {
                "id": item["id"],
                "title": item["title"],
                "date": publish_date,
                "time": publish_time,
                "endTime": end_dt.strftime("%H:%M"),
                "location": item.get("platform", ""),
                "notes": item.get("notes", ""),
                "status": item.get("status", ""),
                "dayIndex": (item_date - week_start).days,
                "top": round(top, 1),
                "height": round(height, 1),
            }
        )

    return events


def add_social_media_schedule_item(
    title: str,
    platform: str,
    publish_date: str,
    publish_time: str,
    status: str,
    notes: str,
) -> None:
    with get_db_connection() as connection:
        connection.execute(
            """
            INSERT INTO social_media_schedule (title, platform, publish_date, publish_time, status, notes, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                title.strip(),
                platform.strip(),
                publish_date.strip(),
                publish_time.strip(),
                status.strip(),
                notes.strip(),
                utcnow_iso(),
            ),
        )


def update_social_media_schedule_item(
    item_id: int,
    title: str,
    platform: str,
    publish_date: str,
    publish_time: str,
    status: str,
    notes: str,
) -> None:
    with get_db_connection() as connection:
        connection.execute(
            """
            UPDATE social_media_schedule
            SET
                title = ?,
                platform = ?,
                publish_date = ?,
                publish_time = ?,
                status = ?,
                notes = ?
            WHERE id = ?
            """,
            (
                title.strip(),
                platform.strip(),
                publish_date.strip(),
                publish_time.strip(),
                status.strip(),
                notes.strip(),
                item_id,
            ),
        )


def delete_social_media_schedule_item(item_id: int) -> None:
    with get_db_connection() as connection:
        connection.execute("DELETE FROM social_media_schedule WHERE id = ?", (item_id,))


def slugify_value(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value or "")
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", ascii_value).strip("-").lower()
    return slug or "album"


def sanitize_upload_filename(file_name: str) -> str:
    normalized = unicodedata.normalize("NFKD", file_name or "")
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
    sanitized = re.sub(r"[^a-zA-Z0-9._-]+", "-", ascii_value).strip(".-_")
    return sanitized or f"bestand-{int(time.time())}.bin"


def format_datetime_display(value: str, fallback: str = "-") -> str:
    parsed = parse_iso_datetime(value)
    if parsed is None:
        return fallback
    return parsed.strftime("%d-%m-%Y %H:%M")


def can_manage_content(user: Optional[Dict[str, Any]]) -> bool:
    if not user:
        return False
    return bool(user.get("isAdmin"))


def derive_recovered_album_title(album_id: int, remote_path: str) -> str:
    normalized_path = str(remote_path or "").strip().strip("/")
    path_parts = [part for part in normalized_path.split("/") if part]
    if len(path_parts) >= 3:
        candidate = path_parts[2]
        if "-" in candidate:
            candidate = candidate.split("-", 1)[1]
        candidate = candidate.replace("-", " ").strip()
        if candidate:
            return candidate.title()
    return f"Hersteld album {album_id}"


def ensure_content_album_records_exist() -> int:
    with get_db_connection() as connection:
        orphan_rows = connection.execute(
            """
            SELECT
                cp.album_id,
                MIN(cp.uploaded_at) AS first_uploaded_at,
                MIN(cp.remote_path) AS sample_remote_path
            FROM content_photos cp
            LEFT JOIN content_albums ca ON ca.id = cp.album_id
            WHERE ca.id IS NULL
            GROUP BY cp.album_id
            ORDER BY cp.album_id ASC
            """
        ).fetchall()

        if not orphan_rows:
            return 0

        repaired_total = 0
        for row in orphan_rows:
            album_id = int(row["album_id"])
            title = derive_recovered_album_title(album_id, str(row["sample_remote_path"] or "").strip())
            connection.execute(
                """
                INSERT INTO content_albums (id, title, slug, created_at)
                VALUES (?, ?, ?, ?)
                """,
                (
                    album_id,
                    title,
                    slugify_value(title),
                    str(row["first_uploaded_at"] or "").strip() or datetime.utcnow().isoformat(),
                ),
            )
            repaired_total += 1

    return repaired_total


def build_content_storage_status() -> Dict[str, Any]:
    config = get_content_storage_config()
    return {
        "mode_label": "Bunny.net" if config["bunny_enabled"] else "Lokale opslag",
        "is_bunny_enabled": config["bunny_enabled"],
        "missing_config": config["missing_config"],
        "base_path": config["base_path"],
        "max_upload_mb": config["max_upload_mb"],
        "max_request_mb": config["max_request_mb"],
        "max_upload_files": config["max_upload_files"],
        "allowed_types": config["allowed_types"],
    }


def request_prefers_json() -> bool:
    accept = str(request.headers.get("Accept", "") or "").lower()
    requested_with = str(request.headers.get("X-Requested-With", "") or "").lower()
    return "application/json" in accept or requested_with == "xmlhttprequest"


def get_bunny_storage_host(region: str) -> str:
    normalized_region = (region or "storage").strip().lower()
    if normalized_region == "storage":
        return "https://storage.bunnycdn.com"
    return f"https://{normalized_region}.storage.bunnycdn.com"


def upload_content_bytes(remote_path: str, content: bytes, content_type: str, config: Optional[Dict[str, Any]] = None) -> Dict[str, str]:
    if config is None:
        config = get_content_storage_config()
    normalized_remote_path = remote_path.strip().strip("/")
    if not normalized_remote_path:
        raise ValueError("Geen uploadpad opgegeven.")

    if config["bunny_enabled"]:
        upload_url = f"{get_bunny_storage_host(config['region'])}/{config['zone']}/{normalized_remote_path}"
        checksum = hashlib.sha256(content).hexdigest().upper()
        response = requests.put(
            upload_url,
            headers={
                "AccessKey": config["access_key"],
                "Content-Type": content_type,
                "Checksum": checksum,
                "Content-Length": str(len(content)),
            },
            data=content,
            timeout=60,
        )
        response.raise_for_status()
        return {
            "url": f"{config['public_base']}/{normalized_remote_path}",
            "storage_backend": "bunny",
        }

    local_root = config["local_upload_root"]
    local_path = os.path.join(local_root, *normalized_remote_path.split("/"))
    os.makedirs(os.path.dirname(local_path), exist_ok=True)
    with open(local_path, "wb") as local_file:
        local_file.write(content)
    return {
        "url": f"/static/uploads/{normalized_remote_path}",
        "storage_backend": "local",
    }


def upload_content_file(
    remote_path: str,
    file_obj: Any,
    file_size: int,
    content_type: str,
    config: Optional[Dict[str, Any]] = None,
) -> Dict[str, str]:
    if config is None:
        config = get_content_storage_config()
    normalized_remote_path = remote_path.strip().strip("/")
    if not normalized_remote_path:
        raise ValueError("Geen uploadpad opgegeven.")

    try:
        file_obj.seek(0)
    except (AttributeError, OSError):
        pass

    if config["bunny_enabled"]:
        upload_url = f"{get_bunny_storage_host(config['region'])}/{config['zone']}/{normalized_remote_path}"
        response = requests.put(
            upload_url,
            headers={
                "AccessKey": config["access_key"],
                "Content-Type": content_type,
                "Content-Length": str(file_size),
            },
            data=file_obj,
            timeout=(30, 7200),
        )
        response.raise_for_status()
        return {
            "url": f"{config['public_base']}/{normalized_remote_path}",
            "storage_backend": "bunny",
        }

    local_root = config["local_upload_root"]
    local_path = os.path.join(local_root, *normalized_remote_path.split("/"))
    os.makedirs(os.path.dirname(local_path), exist_ok=True)
    with open(local_path, "wb") as local_file:
        shutil.copyfileobj(file_obj, local_file, length=1024 * 1024)
    return {
        "url": f"/static/uploads/{normalized_remote_path}",
        "storage_backend": "local",
    }


def delete_content_file(remote_path: str, storage_backend: str) -> None:
    normalized_remote_path = str(remote_path or "").strip().strip("/")
    if not normalized_remote_path:
        return

    if storage_backend == "bunny":
        config = get_content_storage_config()
        if not config["bunny_enabled"]:
            return
        delete_url = f"{get_bunny_storage_host(config['region'])}/{config['zone']}/{normalized_remote_path}"
        response = requests.delete(
            delete_url,
            headers={"AccessKey": config["access_key"]},
            timeout=30,
        )
        if response.status_code not in {200, 201, 202, 204, 404}:
            response.raise_for_status()
        if response.status_code != 404:
            try:
                purge_content_url(f"{config['public_base']}/{normalized_remote_path}")
            except requests.RequestException:
                pass
        return

    local_root = get_content_storage_config()["local_upload_root"]
    local_path = os.path.join(local_root, *normalized_remote_path.split("/"))
    if os.path.exists(local_path):
        os.remove(local_path)


def purge_content_url(public_url: str) -> bool:
    config = get_content_storage_config()
    api_access_key = str(config.get("api_access_key") or "").strip()
    normalized_public_url = str(public_url or "").strip()
    if not api_access_key or not normalized_public_url:
        return False

    response = requests.post(
        "https://api.bunny.net/purge",
        headers={"AccessKey": api_access_key},
        params={
            "url": normalized_public_url,
            "async": "false",
        },
        timeout=30,
    )
    response.raise_for_status()
    return True


def create_content_album(title: str) -> int:
    slug = slugify_value(title)
    created_at = datetime.utcnow().isoformat()
    with get_db_connection() as connection:
        cursor = connection.execute(
            """
            INSERT INTO content_albums (title, slug, created_at)
            VALUES (?, ?, ?)
            """,
            (title.strip(), slug, created_at),
        )
        return int(cursor.lastrowid)


def find_content_album_by_title(title: str) -> Optional[Dict[str, Any]]:
    normalized_title = str(title or "").strip()
    if not normalized_title:
        return None

    with get_db_connection() as connection:
        row = connection.execute(
            """
            SELECT id
            FROM content_albums
            WHERE lower(trim(title)) = lower(trim(?))
            ORDER BY created_at ASC, id ASC
            LIMIT 1
            """,
            (normalized_title,),
        ).fetchone()

    if row is None:
        return None
    return load_content_album(int(row["id"]))


def load_content_album(album_id: int) -> Optional[Dict[str, Any]]:
    ensure_content_album_records_exist()
    with get_db_connection() as connection:
        album_row = connection.execute(
            """
            SELECT id, title, slug, created_at
            FROM content_albums
            WHERE id = ?
            """,
            (album_id,),
        ).fetchone()
        if album_row is None:
            return None

        stats_row = connection.execute(
            """
            SELECT
                COUNT(*) AS photo_count,
                MIN(uploaded_at) AS first_uploaded_at,
                MAX(uploaded_at) AS last_uploaded_at
            FROM content_photos
            WHERE album_id = ?
            """,
            (album_id,),
        ).fetchone()

        cover_row = connection.execute(
            """
            SELECT image_url
            FROM content_photos
            WHERE album_id = ?
            ORDER BY uploaded_at ASC, id ASC
            LIMIT 1
            """,
            (album_id,),
        ).fetchone()

    uploaded_at = (
        str(stats_row["first_uploaded_at"] or "").strip()
        if stats_row is not None
        else ""
    ) or str(album_row["created_at"] or "").strip()
    return {
        "id": int(album_row["id"]),
        "title": str(album_row["title"] or "").strip(),
        "slug": str(album_row["slug"] or "").strip(),
        "createdAt": str(album_row["created_at"] or "").strip(),
        "uploadedAt": uploaded_at,
        "uploadedAtDisplay": format_datetime_display(uploaded_at),
        "photoCount": int(stats_row["photo_count"] or 0) if stats_row is not None else 0,
        "coverUrl": str(cover_row["image_url"] or "").strip() if cover_row is not None else "",
    }


def load_content_album_summaries() -> List[Dict[str, Any]]:
    ensure_content_album_records_exist()
    with get_db_connection() as connection:
        rows = connection.execute(
            """
            SELECT
                a.id,
                a.title,
                a.slug,
                a.created_at,
                COUNT(p.id) AS photo_count,
                MIN(p.uploaded_at) AS first_uploaded_at,
                MAX(p.uploaded_at) AS last_uploaded_at
            FROM content_albums a
            LEFT JOIN content_photos p ON p.album_id = a.id
            GROUP BY a.id
            ORDER BY COALESCE(MAX(p.uploaded_at), a.created_at) DESC, a.id DESC
            """
        ).fetchall()

        cover_rows = connection.execute(
            """
            SELECT cp.album_id, cp.image_url
            FROM content_photos cp
            INNER JOIN (
                SELECT album_id, MIN(id) AS first_photo_id
                FROM content_photos
                GROUP BY album_id
            ) first_photos
                ON first_photos.album_id = cp.album_id
               AND first_photos.first_photo_id = cp.id
            """
        ).fetchall()

    cover_map = {int(row["album_id"]): str(row["image_url"] or "").strip() for row in cover_rows}
    albums = []
    for row in rows:
        uploaded_at = str(row["first_uploaded_at"] or "").strip() or str(row["created_at"] or "").strip()
        albums.append(
            {
                "id": int(row["id"]),
                "title": str(row["title"] or "").strip(),
                "slug": str(row["slug"] or "").strip(),
                "createdAt": str(row["created_at"] or "").strip(),
                "uploadedAt": uploaded_at,
                "uploadedAtDisplay": format_datetime_display(uploaded_at),
                "lastUploadedAt": str(row["last_uploaded_at"] or "").strip(),
                "photoCount": int(row["photo_count"] or 0),
                "coverUrl": cover_map.get(int(row["id"]), ""),
            }
        )
    return albums


def load_content_album_photos(album_id: int) -> List[Dict[str, Any]]:
    ensure_content_album_records_exist()
    with get_db_connection() as connection:
        rows = connection.execute(
            """
            SELECT
                id,
                album_id,
                image_url,
                remote_path,
                file_name,
                original_name,
                content_type,
                file_size,
                storage_backend,
                uploaded_at
            FROM content_photos
            WHERE album_id = ?
            ORDER BY uploaded_at ASC, id ASC
            """,
            (album_id,),
        ).fetchall()

    photos = []
    for row in rows:
        photos.append(
            {
                "id": int(row["id"]),
                "albumId": int(row["album_id"]),
                "imageUrl": str(row["image_url"] or "").strip(),
                "remotePath": str(row["remote_path"] or "").strip(),
                "fileName": str(row["file_name"] or "").strip(),
                "originalName": str(row["original_name"] or "").strip(),
                "contentType": str(row["content_type"] or "").strip(),
                "fileSize": int(row["file_size"] or 0),
                "storageBackend": str(row["storage_backend"] or "local").strip(),
                "uploadedAt": str(row["uploaded_at"] or "").strip(),
                "uploadedAtDisplay": format_datetime_display(str(row["uploaded_at"] or "").strip()),
            }
        )
    return photos


def store_content_photo(
    album_id: int,
    original_name: str,
    file_name: str,
    content_type: str,
    file_size: int,
    remote_path: str,
    image_url: str,
    storage_backend: str,
) -> None:
    uploaded_at = datetime.utcnow().isoformat()
    with get_db_connection() as connection:
        connection.execute(
            """
            INSERT INTO content_photos (
                album_id,
                image_url,
                remote_path,
                file_name,
                original_name,
                content_type,
                file_size,
                storage_backend,
                uploaded_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                album_id,
                image_url,
                remote_path,
                file_name,
                original_name,
                content_type,
                file_size,
                storage_backend,
                uploaded_at,
            ),
        )


def store_content_photos(album_id: int, uploaded_items: List[Dict[str, Any]]) -> None:
    if not uploaded_items:
        return

    uploaded_at = datetime.utcnow().isoformat()
    rows: List[Tuple[Any, ...]] = []
    for item in uploaded_items:
        rows.append(
            (
                album_id,
                item["image_url"],
                item["remote_path"],
                item["file_name"],
                item["original_name"],
                item["content_type"],
                item["file_size"],
                item["storage_backend"],
                uploaded_at,
            )
        )

    with get_db_connection() as connection:
        connection.executemany(
            """
            INSERT INTO content_photos (
                album_id,
                image_url,
                remote_path,
                file_name,
                original_name,
                content_type,
                file_size,
                storage_backend,
                uploaded_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            rows,
        )


def delete_content_photo(photo_id: int, album_id: int) -> bool:
    with get_db_connection() as connection:
        row = connection.execute(
            """
            SELECT id, remote_path, storage_backend
            FROM content_photos
            WHERE id = ? AND album_id = ?
            """,
            (photo_id, album_id),
        ).fetchone()
        if row is None:
            return False
        connection.execute("DELETE FROM content_photos WHERE id = ?", (photo_id,))

    delete_content_file(
        str(row["remote_path"] or "").strip(),
        str(row["storage_backend"] or "local").strip(),
    )
    return True


def delete_empty_content_album(album_id: int) -> None:
    with get_db_connection() as connection:
        row = connection.execute(
            "SELECT 1 FROM content_photos WHERE album_id = ? LIMIT 1",
            (album_id,),
        ).fetchone()
        if row is None:
            connection.execute("DELETE FROM content_albums WHERE id = ?", (album_id,))


def delete_content_album(album_id: int) -> bool:
    photos = load_content_album_photos(album_id)
    album = load_content_album(album_id)
    if album is None:
        return False

    for photo in photos:
        delete_content_file(photo["remotePath"], photo["storageBackend"])

    with get_db_connection() as connection:
        connection.execute("DELETE FROM content_photos WHERE album_id = ?", (album_id,))
        connection.execute("DELETE FROM content_albums WHERE id = ?", (album_id,))
    return True


def resolve_content_album_id(selected_album_id: int, new_album_title: str) -> Optional[int]:
    if selected_album_id:
        return selected_album_id
    normalized_title = new_album_title.strip()
    if normalized_title:
        with content_album_lock:
            existing_album = find_content_album_by_title(normalized_title)
            if existing_album is not None:
                return int(existing_album["id"])
            return create_content_album(normalized_title)
    return None


def collect_content_upload_files() -> List[Any]:
    uploaded_files: List[Any] = []
    for field_name in ("photos", "photo_folder"):
        uploaded_files.extend(request.files.getlist(field_name))
    return uploaded_files


def prepare_content_upload_entry(album: Dict[str, Any], uploaded_file: Any, config: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    original_name = str(getattr(uploaded_file, "filename", "") or "").strip()
    if not original_name:
        return None

    file_bytes = uploaded_file.read()
    if not file_bytes:
        return None

    content_type = str(getattr(uploaded_file, "mimetype", "") or "").strip().lower()
    if not content_type:
        guessed_type = mimetypes.guess_type(original_name)[0]
        content_type = str(guessed_type or "").strip().lower()

    if content_type not in config["allowed_types"]:
        raise ValueError(f"Bestandstype niet toegestaan: {original_name}")

    max_upload_bytes = config["max_upload_mb"] * 1024 * 1024
    if len(file_bytes) > max_upload_bytes:
        raise ValueError(
            f"Bestand is te groot: {original_name}. Maximaal {config['max_upload_mb']} MB toegestaan."
        )

    safe_name = sanitize_upload_filename(original_name)
    extension = os.path.splitext(safe_name)[1].lower()
    if not extension:
        guessed_extension = mimetypes.guess_extension(content_type) or ""
        extension = guessed_extension.lower()
    allowed_extensions = ALLOWED_IMAGE_EXTENSIONS.get(content_type, set())
    if allowed_extensions and extension not in allowed_extensions:
        raise ValueError(f"Bestandsextensie niet toegestaan: {original_name}")
    if not validate_image_signature(content_type, file_bytes):
        raise ValueError(f"Bestandsinhoud niet geldig voor type: {original_name}")

    unique_name = f"{int(time.time() * 1000)}-{secrets.token_hex(4)}{extension}"
    remote_path = "/".join(
        [
            config["base_path"],
            date.today().isoformat(),
            f"{album['id']}-{album['slug']}",
            unique_name,
        ]
    )
    return {
        "original_name": original_name,
        "file_name": unique_name,
        "content_type": content_type,
        "file_size": len(file_bytes),
        "remote_path": remote_path,
        "content": file_bytes,
    }


def upload_prepared_content_entry(entry: Dict[str, Any]) -> Dict[str, Any]:
    upload_result = upload_content_bytes(
        entry["remote_path"],
        entry["content"],
        entry["content_type"],
    )
    return {
        "original_name": entry["original_name"],
        "file_name": entry["file_name"],
        "content_type": entry["content_type"],
        "file_size": entry["file_size"],
        "remote_path": entry["remote_path"],
        "image_url": upload_result["url"],
        "storage_backend": upload_result["storage_backend"],
    }


def upload_files_to_content_album(album_id: int, uploaded_files: List[Any]) -> int:
    album = load_content_album(album_id)
    if album is None:
        raise ValueError("Het gekozen album bestaat niet meer.")

    config = get_content_storage_config()
    prepared_entries: List[Dict[str, Any]] = []
    for uploaded_file in uploaded_files:
        prepared_entry = prepare_content_upload_entry(album, uploaded_file, config)
        if prepared_entry is not None:
            prepared_entries.append(prepared_entry)

    if not prepared_entries:
        raise ValueError("Selecteer minimaal één foto om te uploaden.")

    max_workers = min(4, len(prepared_entries))
    uploaded_items: List[Dict[str, Any]] = []
    try:
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_map = {
                executor.submit(upload_prepared_content_entry, entry): entry
                for entry in prepared_entries
            }
            for future in as_completed(future_map):
                uploaded_items.append(future.result())
    except Exception:
        for uploaded_item in uploaded_items:
            try:
                delete_content_file(uploaded_item["remote_path"], uploaded_item["storage_backend"])
            except requests.RequestException:
                pass
        raise

    store_content_photos(album["id"], uploaded_items)
    return len(uploaded_items)


def derive_member_type_from_role(role: str) -> str:
    normalized_role = role.strip().lower()
    if normalized_role in {"vrijwilliger"}:
        return "Vrijwilliger"
    if normalized_role in {"stagiair"}:
        return "Stagiair"
    return "Medewerker"


def normalize_system_role(role: str) -> str:
    normalized_role = role.strip().lower()
    if normalized_role in {"admin", "beheerder"}:
        return "Admin"
    # Accounts with the retired role remain usable as regular trainer accounts.
    if normalized_role == "social media beheerder":
        return "Trainer"
    return role.strip()


def is_allowed_system_role(role: str) -> bool:
    return normalize_system_role(role) in {"Admin", "Trainer"}


def role_grants_admin_access(role: str) -> bool:
    return normalize_system_role(role) == "Admin"


def is_trainer_user(user: Optional[Dict[str, Any]]) -> bool:
    if not user:
        return False
    role = normalize_system_role(str(user.get("systemRole") or user.get("role") or ""))
    return role.casefold() == "trainer"


def filter_agenda_trainings_for_user(
    trainings: List[Dict[str, Any]],
    user: Optional[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    if not is_trainer_user(user):
        return trainings

    trainer_id = str(user.get("id") or "").strip() if user else ""
    if not trainer_id:
        return []

    return [
        training
        for training in trainings
        if trainer_id
        in {
            str(trainer.get("id") or "").strip()
            for trainer in training.get("trainers") or []
            if isinstance(trainer, dict)
        }
    ]


def get_visible_pages_for_user(user: Optional[Dict[str, Any]]) -> Set[str]:
    if not user:
        return set()
    if user.get("isAdmin"):
        return {
            "dashboard",
            "agenda",
            "draaiboeken",
            "voetbaldagen",
            "samenwerkende-amateurclubs",
            "management",
            "planning",
            "api",
            "materialen",
            "orders",
            "leads",
            "financien",
            "revenue",
            "spaarpot",
            "trainer-fees",
            "begroting",
            "voorstellen-maker",
            "overeenkomsten",
            "oefenstof",
            "oefeningen-bibliotheek",
            "trainingen",
            "exercise-videos",
            "marketing",
            "social-media",
            "content",
            "trainers",
            "profile",
        }
    visible_pages = {"materialen", "orders", "leads", "draaiboeken", "voetbaldagen", "samenwerkende-amateurclubs", "oefenstof", "oefeningen-bibliotheek", "trainingen", "profile"}
    if is_trainer_user(user):
        visible_pages.update({"dashboard", "agenda"})
    return visible_pages


def user_can_access_page(user: Optional[Dict[str, Any]], page_key: str) -> bool:
    return page_key in get_visible_pages_for_user(user)


def require_page_access(page_key: str) -> Optional[Any]:
    user = get_current_user()
    if user is None:
        return redirect(url_for("login_page", next=request.path))
    if user_can_access_page(user, page_key):
        return None
    fallback_page = "dashboard" if user_can_access_page(user, "dashboard") else "profile"
    if fallback_page == "dashboard":
        return redirect(url_for("index"))
    return redirect(url_for("personal_profile_page"))


def parse_non_negative_int(value: Any) -> int:
    try:
        number = int(str(value or "0").strip())
    except (TypeError, ValueError):
        return 0
    return max(0, number)


def load_materials_inventory() -> Dict[str, Any]:
    with get_db_connection() as connection:
        material_rows = connection.execute(
            """
            SELECT id, name, total_count
            FROM material_items
            ORDER BY sort_order ASC, id ASC
            """
        ).fetchall()
        club_rows = connection.execute(
            """
            SELECT id, name
            FROM material_clubs
            ORDER BY sort_order ASC, id ASC
            """
        ).fetchall()
        inventory_rows = connection.execute(
            """
            SELECT club_id, material_id, quantity
            FROM material_club_inventory
            """
        ).fetchall()

    quantities = {
        (int(row["club_id"]), int(row["material_id"])): parse_non_negative_int(row["quantity"])
        for row in inventory_rows
    }
    clubs = [
        {
            "id": int(row["id"]),
            "key": f"club-{int(row['id'])}",
            "name": str(row["name"] or "").strip(),
            "quantities": {},
        }
        for row in club_rows
    ]
    materials = []
    for row in material_rows:
        material_id = int(row["id"])
        total_count = parse_non_negative_int(row["total_count"])
        allocated_count = sum(quantities.get((club["id"], material_id), 0) for club in clubs)
        material = {
            "id": material_id,
            "key": f"material-{material_id}",
            "name": str(row["name"] or "").strip(),
            "totalCount": total_count,
            "allocatedCount": allocated_count,
            "availableCount": total_count - allocated_count,
        }
        materials.append(material)

    for club in clubs:
        club["totalCount"] = 0
        for material in materials:
            quantity = quantities.get((club["id"], material["id"]), 0)
            club["quantities"][material["key"]] = quantity
            club["totalCount"] += quantity

    return {
        "materials": materials,
        "clubs": clubs,
        "totalCount": sum(material["totalCount"] for material in materials),
        "allocatedCount": sum(material["allocatedCount"] for material in materials),
        "availableCount": sum(material["availableCount"] for material in materials),
    }


def _create_materials_club_pdf_portrait(club: Dict[str, Any], materials: List[Dict[str, Any]]) -> bytes:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.utils import ImageReader
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.pdfgen import canvas
    except ImportError as exc:
        raise RuntimeError("De PDF-library ontbreekt. Installeer de packages uit requirements.txt.") from exc

    font_root = os.path.join(os.path.dirname(__file__), "static", "assets", "fonts")
    font_names = {
        "regular": "MaterialsPoppins",
        "bold": "MaterialsPoppinsBold",
        "extra_bold": "MaterialsPoppinsExtraBold",
    }
    font_files = {
        "regular": "Poppins-Regular.ttf",
        "bold": "Poppins-Bold.ttf",
        "extra_bold": "Poppins-ExtraBold.ttf",
    }
    registered_fonts = set(pdfmetrics.getRegisteredFontNames())
    for key, font_name in font_names.items():
        if font_name not in registered_fonts:
            pdfmetrics.registerFont(TTFont(font_name, os.path.join(font_root, font_files[key])))

    rows = []
    quantities = club.get("quantities") if isinstance(club.get("quantities"), dict) else {}
    for material in materials:
        quantity = parse_non_negative_int(quantities.get(material.get("key"), 0))
        if quantity > 0:
            rows.append({"name": str(material.get("name") or "Materiaal").strip(), "quantity": quantity})

    buffer = BytesIO()
    page_width, page_height = A4
    pdf = canvas.Canvas(buffer, pagesize=A4)
    pdf.setTitle(f"Materialenkrat - {club.get('name') or 'Club'}")
    pdf.setAuthor("HWS Voetbalschool")
    black = colors.HexColor("#111111")
    charcoal = colors.HexColor("#252525")
    muted = colors.HexColor("#717171")
    line = colors.HexColor("#dddddd")
    soft = colors.HexColor("#f4f4f4")
    white = colors.white
    gold = colors.HexColor("#d6a34f")
    margin = 40

    pdf.setFillColor(black)
    pdf.rect(0, page_height - 177, page_width, 177, fill=1, stroke=0)
    pdf.setFillColor(gold)
    pdf.rect(0, page_height - 183, page_width, 6, fill=1, stroke=0)

    logo_path = os.path.join(os.path.dirname(__file__), "static", "assets", "hws-logo.png")
    if os.path.exists(logo_path):
        pdf.drawImage(ImageReader(logo_path), page_width - 143, page_height - 155, 105, 105, preserveAspectRatio=True, mask="auto", anchor="c")

    pdf.setFillColor(gold)
    pdf.setFont(font_names["bold"], 9)
    pdf.drawString(margin, page_height - 50, "HWS VOETBALSCHOOL")
    pdf.setFillColor(white)
    pdf.setFont(font_names["extra_bold"], 25)
    pdf.drawString(margin, page_height - 85, "MATERIALENKRAT")
    pdf.setFont(font_names["bold"], 18)
    club_name = str(club.get("name") or "Club").strip()
    pdf.drawString(margin, page_height - 116, club_name[:38])
    pdf.setFillColor(colors.HexColor("#cfcfcf"))
    pdf.setFont(font_names["regular"], 8.5)
    pdf.drawString(margin, page_height - 143, f"Vast overzicht  •  bijgewerkt {date.today().strftime('%d-%m-%Y')}")

    total_quantity = sum(row["quantity"] for row in rows)
    summary_y = page_height - 236
    card_gap = 12
    card_width = (page_width - (2 * margin) - card_gap) / 2
    for index, (label, value) in enumerate((("TOTAAL AANTAL", total_quantity), ("SOORTEN MATERIAAL", len(rows)))):
        x = margin + index * (card_width + card_gap)
        pdf.setFillColor(soft)
        pdf.roundRect(x, summary_y, card_width, 43, 6, fill=1, stroke=0)
        pdf.setFillColor(muted)
        pdf.setFont(font_names["bold"], 7.5)
        pdf.drawString(x + 13, summary_y + 25, label)
        pdf.setFillColor(black)
        pdf.setFont(font_names["extra_bold"], 15)
        pdf.drawRightString(x + card_width - 13, summary_y + 18, str(value))

    list_top = summary_y - 39
    pdf.setFillColor(black)
    pdf.setFont(font_names["extra_bold"], 12)
    pdf.drawString(margin, list_top, "INHOUD VAN DE KRAT")
    pdf.setFillColor(muted)
    pdf.setFont(font_names["regular"], 7.7)
    pdf.drawRightString(page_width - margin, list_top, "Controleer en vink af")

    list_bottom = 92
    list_height = list_top - 22 - list_bottom
    row_count = len(rows)
    column_count = 1 if row_count <= 13 else 2 if row_count <= 34 else 3
    rows_per_column = max(1, ceil(row_count / column_count))
    column_gap = 13
    column_width = (page_width - (2 * margin) - ((column_count - 1) * column_gap)) / column_count
    row_height = min(31, list_height / rows_per_column)
    name_size = 9 if column_count == 1 else 8 if column_count == 2 else 6.7

    if not rows:
        pdf.setFillColor(soft)
        pdf.roundRect(margin, list_top - 84, page_width - (2 * margin), 62, 7, fill=1, stroke=0)
        pdf.setFillColor(charcoal)
        pdf.setFont(font_names["bold"], 10)
        pdf.drawCentredString(page_width / 2, list_top - 52, "Voor deze club zijn nog geen materialen opgeslagen.")
    else:
        for index, row in enumerate(rows):
            column_index = index // rows_per_column
            row_index = index % rows_per_column
            x = margin + column_index * (column_width + column_gap)
            y_top = list_top - 22 - (row_index * row_height)
            pdf.setStrokeColor(line)
            pdf.setLineWidth(0.6)
            pdf.line(x, y_top - row_height, x + column_width, y_top - row_height)
            box_size = min(12, max(8, row_height - 12))
            box_y = y_top - ((row_height + box_size) / 2) + 1
            pdf.setStrokeColor(charcoal)
            pdf.setLineWidth(1)
            pdf.rect(x, box_y, box_size, box_size, fill=0, stroke=1)
            pdf.setFillColor(black)
            pdf.setFont(font_names["bold"], name_size)
            max_name_width = column_width - box_size - 48
            material_name = row["name"]
            while len(material_name) > 3 and pdfmetrics.stringWidth(material_name, font_names["bold"], name_size) > max_name_width:
                material_name = material_name[:-1]
            if material_name != row["name"]:
                material_name = f"{material_name.rstrip()}…"
            pdf.drawString(x + box_size + 9, y_top - (row_height / 2) - (name_size / 3), material_name)
            pdf.setFillColor(gold)
            pdf.setFont(font_names["extra_bold"], 10 if column_count < 3 else 8.5)
            pdf.drawRightString(x + column_width, y_top - (row_height / 2) - 3, f"{row['quantity']}×")

    pdf.setStrokeColor(line)
    pdf.line(margin, 68, page_width - margin, 68)
    pdf.setFillColor(muted)
    pdf.setFont(font_names["regular"], 7.2)
    pdf.drawString(margin, 49, "Controleer de krat na gebruik en meld ontbrekend of beschadigd materiaal bij HWS.")
    pdf.setFillColor(black)
    pdf.setFont(font_names["bold"], 7.2)
    pdf.drawRightString(page_width - margin, 49, "hwsvoetbalschool.nl")

    pdf.showPage()
    pdf.save()
    buffer.seek(0)
    return buffer.read()


def create_materials_club_pdf(
    club: Dict[str, Any],
    materials: List[Dict[str, Any]],
    *,
    _pdf: Optional[Any] = None,
) -> bytes:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.utils import ImageReader
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.pdfgen import canvas
    except ImportError as exc:
        raise RuntimeError("De PDF-library ontbreekt. Installeer de packages uit requirements.txt.") from exc

    font_root = os.path.join(os.path.dirname(__file__), "static", "assets", "fonts")
    font_names = {
        "regular": "MaterialsPoppins",
        "bold": "MaterialsPoppinsBold",
        "extra_bold": "MaterialsPoppinsExtraBold",
    }
    font_files = {
        "regular": "Poppins-Regular.ttf",
        "bold": "Poppins-Bold.ttf",
        "extra_bold": "Poppins-ExtraBold.ttf",
    }
    registered_fonts = set(pdfmetrics.getRegisteredFontNames())
    for key, font_name in font_names.items():
        if font_name not in registered_fonts:
            pdfmetrics.registerFont(TTFont(font_name, os.path.join(font_root, font_files[key])))

    rows = []
    quantities = club.get("quantities") if isinstance(club.get("quantities"), dict) else {}
    for material in materials:
        quantity = parse_non_negative_int(quantities.get(material.get("key"), 0))
        if quantity > 0:
            rows.append({"name": str(material.get("name") or "Materiaal").strip(), "quantity": quantity})

    buffer = BytesIO() if _pdf is None else None
    page_width, page_height = landscape(A4)
    pdf = _pdf or canvas.Canvas(buffer, pagesize=(page_width, page_height))
    pdf.setTitle(f"Materialenkrat - {club.get('name') or 'Club'}")
    pdf.setAuthor("HWS Voetbalschool")
    black = colors.HexColor("#111111")
    charcoal = colors.HexColor("#252525")
    muted = colors.HexColor("#6f6f6f")
    line = colors.HexColor("#dddddd")
    soft = colors.HexColor("#f4f4f4")
    white = colors.white
    gold = colors.HexColor("#d6a34f")
    margin = 36

    pdf.setFillColor(black)
    pdf.rect(0, page_height - 125, page_width, 125, fill=1, stroke=0)
    pdf.setFillColor(gold)
    pdf.rect(0, page_height - 130, page_width, 5, fill=1, stroke=0)

    logo_path = os.path.join(os.path.dirname(__file__), "static", "assets", "hws-logo.png")
    if os.path.exists(logo_path):
        pdf.drawImage(ImageReader(logo_path), page_width - 132, page_height - 119, 98, 98, preserveAspectRatio=True, mask="auto", anchor="c")

    pdf.setFillColor(gold)
    pdf.setFont(font_names["bold"], 8.5)
    pdf.drawString(margin, page_height - 34, "HWS VOETBALSCHOOL")
    pdf.setFillColor(white)
    pdf.setFont(font_names["extra_bold"], 23)
    title_y = page_height - 76
    club_name = str(club.get("name") or "Club").strip().upper()
    title_prefix = "MATERIALENKRAT "
    title_max_width = page_width - 150 - margin
    while len(club_name) > 3 and pdfmetrics.stringWidth(
        f"{title_prefix}{club_name}", font_names["extra_bold"], 23
    ) > title_max_width:
        club_name = club_name[:-1]
    pdf.drawString(margin, title_y, f"{title_prefix}{club_name}")

    column_gap = 24
    left_width = (page_width - (2 * margin) - column_gap) / 2
    right_x = margin + left_width + column_gap
    right_width = page_width - right_x - margin
    list_top = page_height - 160
    pdf.setFillColor(black)
    pdf.setFont(font_names["extra_bold"], 11)
    pdf.drawString(margin, list_top, "MATERIALEN IN DE KRAT")

    list_bottom = 176
    list_height = list_top - 17 - list_bottom
    row_count = len(rows)
    column_count = 1 if row_count <= 10 else 2 if row_count <= 22 else 3
    rows_per_column = max(1, ceil(row_count / column_count))
    material_column_gap = 14
    column_width = (left_width - ((column_count - 1) * material_column_gap)) / column_count
    row_height = min(25, list_height / rows_per_column)
    name_size = 8.5 if column_count == 1 else 7.5 if column_count == 2 else 6.5

    if not rows:
        pdf.setFillColor(soft)
        pdf.roundRect(margin, list_top - 70, left_width, 50, 7, fill=1, stroke=0)
        pdf.setFillColor(charcoal)
        pdf.setFont(font_names["bold"], 8.5)
        pdf.drawCentredString(margin + (left_width / 2), list_top - 48, "Voor deze club zijn nog geen materialen opgeslagen.")
    else:
        for index, row in enumerate(rows):
            column_index = index // rows_per_column
            row_index = index % rows_per_column
            x = margin + column_index * (column_width + material_column_gap)
            y_top = list_top - 17 - (row_index * row_height)
            pdf.setStrokeColor(line)
            pdf.setLineWidth(0.5)
            pdf.line(x, y_top - row_height, x + column_width, y_top - row_height)
            bullet_radius = 2.7
            bullet_y = y_top - (row_height / 2) - 1
            pdf.setFillColor(gold)
            pdf.circle(x + bullet_radius, bullet_y, bullet_radius, fill=1, stroke=0)
            pdf.setFillColor(black)
            pdf.setFont(font_names["bold"], name_size)
            max_name_width = column_width - 42
            material_name = row["name"]
            while len(material_name) > 3 and pdfmetrics.stringWidth(material_name, font_names["bold"], name_size) > max_name_width:
                material_name = material_name[:-1]
            if material_name != row["name"]:
                material_name = f"{material_name.rstrip()}…"
            pdf.drawString(x + 12, y_top - (row_height / 2) - (name_size / 3), material_name)
            pdf.setFillColor(gold)
            pdf.setFont(font_names["extra_bold"], 9 if column_count < 3 else 7.5)
            pdf.drawRightString(x + column_width, y_top - (row_height / 2) - 3, f"{row['quantity']}×")

    contact_y = 47
    contact_height = 103
    pdf.setFillColor(soft)
    pdf.roundRect(right_x, contact_y, right_width, contact_height, 8, fill=1, stroke=0)
    pdf.setFillColor(gold)
    pdf.roundRect(right_x, contact_y, 6, contact_height, 3, fill=1, stroke=0)
    pdf.setFillColor(black)
    pdf.setFont(font_names["extra_bold"], 9.5)
    pdf.drawString(right_x + 19, contact_y + contact_height - 24, "CONTACTPERSOON")
    pdf.setFont(font_names["bold"], 9)
    pdf.drawString(right_x + 19, contact_y + contact_height - 47, "David van Walstijn")
    pdf.setFillColor(muted)
    pdf.setFont(font_names["regular"], 8)
    pdf.drawString(right_x + 19, contact_y + 29, "info@hwsvoetbalschool.nl")
    pdf.setFillColor(black)
    pdf.setFont(font_names["bold"], 9)
    pdf.drawString(right_x + 19, contact_y + 13, "06-24845896")

    rules_y = contact_y + contact_height + 14
    rules_height = list_top - rules_y + 2
    pdf.setFillColor(soft)
    pdf.roundRect(right_x, rules_y, right_width, rules_height, 8, fill=1, stroke=0)
    pdf.setFillColor(black)
    pdf.setFont(font_names["extra_bold"], 11)
    pdf.drawString(right_x + 20, rules_y + rules_height - 31, "AFSPRAKEN RONDOM DE MATERIALEN")
    rule_lines = [
        ["Tel vóór en na iedere training de ballen en hoedjes."],
        ["Zet het trainingsveld vóór aanvang zorgvuldig uit.", "Ruim het veld na afloop volledig op."],
        ["Controleer na iedere training of al het materiaal", "compleet en onbeschadigd is."],
        ["Ontbreekt er materiaal of moeten de hesjes worden gewassen?", "Geef dit direct door aan de contactpersoon."],
    ]
    rule_y = rules_y + rules_height - 70
    for rule_index, lines in enumerate(rule_lines, start=1):
        pdf.setFillColor(gold)
        pdf.circle(right_x + 30, rule_y + 2, 10, fill=1, stroke=0)
        pdf.setFillColor(black)
        pdf.setFont(font_names["extra_bold"], 8)
        pdf.drawCentredString(right_x + 30, rule_y - 0.8, str(rule_index))
        for line_index, text in enumerate(lines):
            font_key = "bold" if line_index == 0 or rule_index == 3 else "regular"
            pdf.setFont(font_names[font_key], 8.3)
            pdf.drawString(right_x + 52, rule_y - (line_index * 13), text)
        rule_y -= 42 if len(lines) == 1 else 49

    pdf.showPage()
    if buffer is None:
        return b""
    pdf.save()
    buffer.seek(0)
    return buffer.read()


def create_materials_all_clubs_pdf(clubs: List[Dict[str, Any]], materials: List[Dict[str, Any]]) -> bytes:
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.pdfgen import canvas
    except ImportError as exc:
        raise RuntimeError("De PDF-library ontbreekt. Installeer de packages uit requirements.txt.") from exc

    buffer = BytesIO()
    page_width, page_height = landscape(A4)
    pdf = canvas.Canvas(buffer, pagesize=(page_width, page_height))
    for club in clubs:
        create_materials_club_pdf(club, materials, _pdf=pdf)
    pdf.setTitle("Materialenkratten - alle clubs")
    pdf.setAuthor("HWS Voetbalschool")
    pdf.save()
    buffer.seek(0)
    return buffer.read()


def build_materials_inventory_from_form() -> Dict[str, Any]:
    material_keys = request.form.getlist("material_key")
    material_names = request.form.getlist("material_name")
    material_totals = request.form.getlist("material_total")
    club_keys = request.form.getlist("club_key")
    club_names = request.form.getlist("club_name")

    materials = []
    seen_material_names = set()
    for index, material_key in enumerate(material_keys):
        name = str(material_names[index] if index < len(material_names) else "").strip()
        if not name:
            continue
        name_key = name.casefold()
        if name_key in seen_material_names:
            continue
        seen_material_names.add(name_key)
        materials.append(
            {
                "key": str(material_key or f"material-{index}").strip(),
                "name": name,
                "totalCount": parse_non_negative_int(material_totals[index] if index < len(material_totals) else 0),
            }
        )

    clubs = []
    seen_club_names = set()
    for index, club_key in enumerate(club_keys):
        name = str(club_names[index] if index < len(club_names) else "").strip()
        if not name:
            continue
        name_key = name.casefold()
        if name_key in seen_club_names:
            continue
        seen_club_names.add(name_key)
        clubs.append(
            {
                "key": str(club_key or f"club-{index}").strip(),
                "name": name,
            }
        )

    quantities = {}
    for club in clubs:
        for material in materials:
            field_name = f"quantity__{club['key']}__{material['key']}"
            quantities[(club["key"], material["key"])] = parse_non_negative_int(request.form.get(field_name, 0))

    return {"materials": materials, "clubs": clubs, "quantities": quantities}


def save_materials_inventory(inventory: Dict[str, Any]) -> None:
    now = utcnow_iso()
    with get_db_connection() as connection:
        connection.execute("DELETE FROM material_club_inventory")
        connection.execute("DELETE FROM material_items")
        connection.execute("DELETE FROM material_clubs")

        material_id_by_key = {}
        for index, material in enumerate(inventory["materials"]):
            cursor = connection.execute(
                """
                INSERT INTO material_items (name, total_count, sort_order, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?)
                """,
                (material["name"], material["totalCount"], index, now, now),
            )
            material_id_by_key[material["key"]] = int(cursor.lastrowid)

        club_id_by_key = {}
        for index, club in enumerate(inventory["clubs"]):
            cursor = connection.execute(
                """
                INSERT INTO material_clubs (name, sort_order, created_at, updated_at)
                VALUES (?, ?, ?, ?)
                """,
                (club["name"], index, now, now),
            )
            club_id_by_key[club["key"]] = int(cursor.lastrowid)

        for (club_key, material_key), quantity in inventory["quantities"].items():
            if quantity <= 0:
                continue
            club_id = club_id_by_key.get(club_key)
            material_id = material_id_by_key.get(material_key)
            if club_id is None or material_id is None:
                continue
            connection.execute(
                """
                INSERT INTO material_club_inventory (club_id, material_id, quantity)
                VALUES (?, ?, ?)
                """,
                (club_id, material_id, quantity),
            )
    clear_local_data_cache()


def normalize_trainer_fee_rows(raw_rows: Any) -> List[Dict[str, Any]]:
    rows = raw_rows if isinstance(raw_rows, list) else []
    normalized_rows = []
    valid_activities = {str(option["value"]): str(option["label"]) for option in TRAINER_FEE_ACTIVITY_OPTIONS}
    valid_activities[TRAINER_FEE_ALL_ACTIVITIES_VALUE] = TRAINER_FEE_ALL_ACTIVITIES_VALUE
    valid_types = {str(option["value"]): str(option["label"]) for option in TRAINER_FEE_TYPE_OPTIONS}
    valid_days = {
        "maandag": "Maandag",
        "dinsdag": "Dinsdag",
        "woensdag": "Woensdag",
        "donderdag": "Donderdag",
        "vrijdag": "Vrijdag",
        "zaterdag": "Zaterdag",
        "zondag": "Zondag",
    }
    for club_options in build_trainer_fee_agenda_activity_options().values():
        for option in club_options:
            valid_activities.setdefault(str(option["value"]), str(option["label"]))

    for item in rows:
        if not isinstance(item, dict):
            continue
        club = str(item.get("club") or "").strip()
        activity = str(item.get("activity") or item.get("activityType") or "").strip()
        amount = str(item.get("amount") or "").strip()
        day = str(item.get("day") or "").strip().lower()
        day = day if day in valid_days else ""
        time_value = str(item.get("time") or "").strip()
        time_value = time_value if re.fullmatch(r"(?:[01]\d|2[0-3]):[0-5]\d", time_value) else ""
        group = str(item.get("group") or item.get("trainerGroup") or "").strip()[:100]
        raw_types = item.get("types")
        if isinstance(raw_types, list):
            types = [str(value or "").strip() for value in raw_types]
        else:
            single_type = str(item.get("type") or "").strip()
            types = [single_type] if single_type else []
        types = [value for value in types if value in valid_types]
        fee_type = types[0] if types else ""
        if not club and not activity and not amount and not fee_type and not day and not time_value:
            continue
        valid_clubs = set(TRAINER_FEE_CLUB_OPTIONS_BY_TYPE.get(fee_type, AGENDA_CLUB_OPTIONS))
        if fee_type == "voetbaldag_summercamp":
            valid_clubs.add(TRAINER_FEE_ALL_CLUBS_VALUE)
        if club not in valid_clubs:
            club = ""
        if activity == TRAINER_FEE_ALL_ACTIVITIES_VALUE and fee_type != "voetbaldag_summercamp":
            activity = ""
        if activity not in valid_activities:
            activity = ""
        normalized_rows.append(
            {
                "day": day,
                "dayLabel": valid_days.get(day, ""),
                "time": time_value,
                "club": club,
                "type": fee_type,
                "typeLabel": valid_types.get(fee_type, fee_type),
                "types": [fee_type] if fee_type else [],
                "typeLabels": [valid_types[fee_type]] if fee_type else [],
                "activity": activity,
                "activityLabel": valid_activities.get(activity, activity),
                "amount": amount,
                "group": group,
            }
        )

    group_amounts: Dict[str, str] = {}
    for row in normalized_rows:
        group = str(row.get("group") or "").strip().casefold()
        if group and str(row.get("amount") or "").strip() and group not in group_amounts:
            group_amounts[group] = str(row["amount"])
    for row in normalized_rows:
        group = str(row.get("group") or "").strip().casefold()
        if group and group in group_amounts:
            row["amount"] = group_amounts[group]
    return normalized_rows


def parse_trainer_fee_rows_from_form(form_data: Any) -> List[Dict[str, Any]]:
    days = form_data.getlist("fee_day")
    times = form_data.getlist("fee_time")
    types = form_data.getlist("fee_type")
    clubs = form_data.getlist("fee_club")
    activities = form_data.getlist("fee_activity")
    amounts = form_data.getlist("fee_amount")
    groups = form_data.getlist("fee_group")
    max_length = max(len(days), len(times), len(types), len(clubs), len(activities), len(amounts), len(groups), 0)
    raw_rows = []
    for index in range(max_length):
        raw_rows.append(
            {
                "day": days[index] if index < len(days) else "",
                "time": times[index] if index < len(times) else "",
                "type": types[index] if index < len(types) else "",
                "club": clubs[index] if index < len(clubs) else "",
                "activity": activities[index] if index < len(activities) else "",
                "amount": amounts[index] if index < len(amounts) else "",
                "group": groups[index] if index < len(groups) else "",
            }
        )
    return normalize_trainer_fee_rows(raw_rows)


def trainer_fees_json_dumps(rows: List[Dict[str, Any]]) -> str:
    return json.dumps(normalize_trainer_fee_rows(rows), ensure_ascii=False)


def update_trainer_fee_rows(profile_id: str, trainer_fees: List[Dict[str, Any]]) -> None:
    with get_db_connection() as connection:
        connection.execute(
            "UPDATE trainer_profiles SET trainer_fees_json = ? WHERE id = ?",
            (trainer_fees_json_dumps(trainer_fees), profile_id.strip()),
        )
    clear_local_data_cache()


def build_trainer_fee_agenda_activity_options() -> Dict[str, List[Dict[str, str]]]:
    options_by_club: Dict[str, List[Dict[str, str]]] = {club: [] for club in AGENDA_CLUB_OPTIONS}
    seen_by_club: Dict[str, set[str]] = {club: set() for club in AGENDA_CLUB_OPTIONS}

    for training in load_agenda_trainings():
        club = normalize_agenda_club(training.get("location"))
        if not club:
            continue
        title = str(training.get("title") or "").strip()
        if not title:
            continue
        training_type_label = str(training.get("trainingTypeLabel") or "").strip()
        value = title
        label = f"{title} ({training_type_label})" if training_type_label else title
        if value in seen_by_club[club]:
            continue
        seen_by_club[club].add(value)
        options_by_club[club].append(
            {
                "value": value,
                "label": label,
            }
        )

    return options_by_club


def build_trainer_fee_club_options_by_type() -> Dict[str, List[str]]:
    return {
        key: list(value)
        for key, value in TRAINER_FEE_CLUB_OPTIONS_BY_TYPE.items()
    }


def build_agenda_club_options_by_training_type() -> Dict[str, List[str]]:
    return {
        key: list(value)
        for key, value in AGENDA_CLUB_OPTIONS_BY_TRAINING_TYPE.items()
    }


def build_trainer_fee_year_options(trainings: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    years = {
        parsed_date.year
        for training in trainings
        for parsed_date in [parse_iso_date(str(training.get("date") or ""))]
        if parsed_date is not None
    }
    years.add(date.today().year)
    return [{"value": str(year), "label": str(year)} for year in sorted(years, reverse=True)]


def parse_trainer_fee_amount(value: Any) -> Decimal:
    amount = parse_decimal_amount(str(value or ""))
    return amount if amount > 0 else Decimal("0")


def get_trainer_fee_type_for_training(training_type: Any) -> str:
    normalized_type = str(training_type or "").strip()
    if normalized_type in {"voetbaldag", "summercamp"}:
        return "voetbaldag_summercamp"
    if normalized_type in {"samenwerkende_amateurclub", "techniektraining"}:
        return normalized_type
    return ""


def trainer_fee_row_matches(fee_row: Dict[str, Any], fee_type: str, club: str, activity: str) -> bool:
    row_type = str(fee_row.get("type") or "").strip()
    row_club = str(fee_row.get("club") or "").strip()
    row_activity = str(fee_row.get("activity") or "").strip()

    if not row_club or not row_activity:
        return False
    if row_type and row_type != fee_type:
        return False
    if row_club and row_club != TRAINER_FEE_ALL_CLUBS_VALUE and row_club != club:
        return False
    if row_activity and row_activity != TRAINER_FEE_ALL_ACTIVITIES_VALUE and row_activity != activity:
        return False
    return True


def find_trainer_fee_amount(profile: Dict[str, Any], club: str, activity: str, fee_type: str = "") -> Decimal:
    for fee_row in profile.get("trainerFees") or []:
        if not trainer_fee_row_matches(fee_row, fee_type, club, activity):
            continue
        return parse_trainer_fee_amount(fee_row.get("amount"))
    return Decimal("0")


def find_trainer_fee_row(profile: Dict[str, Any], club: str, activity: str, fee_type: str = "") -> Optional[Dict[str, Any]]:
    for fee_row in profile.get("trainerFees") or []:
        if trainer_fee_row_matches(fee_row, fee_type, club, activity):
            return fee_row
    return None


def trainer_fee_training_counts_for_summary(training: Dict[str, Any]) -> bool:
    return str(training.get("status") or "").strip() == "gegeven"


def load_trainer_fee_payment_statuses(season_start_year: int) -> Dict[Tuple[str, int], bool]:
    with get_db_connection() as connection:
        rows = connection.execute(
            """
            SELECT trainer_id, month, paid
            FROM trainer_fee_payment_statuses
            WHERE season_start_year = ?
            """,
            (season_start_year,),
        ).fetchall()
    return {
        (str(row["trainer_id"] or "").strip(), int(row["month"])): bool(row["paid"])
        for row in rows
    }


def save_trainer_fee_payment_status(trainer_id: str, season_start_year: int, month: int, paid: bool) -> None:
    normalized_trainer_id = trainer_id.strip()
    if not normalized_trainer_id or month < 1 or month > 12:
        return
    updated_at = datetime.now().isoformat(timespec="seconds")
    with get_db_connection() as connection:
        connection.execute(
            """
            INSERT INTO trainer_fee_payment_statuses (trainer_id, season_start_year, month, paid, updated_at)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(trainer_id, season_start_year, month)
            DO UPDATE SET paid = excluded.paid, updated_at = excluded.updated_at
            """,
            (normalized_trainer_id, season_start_year, month, 1 if paid else 0, updated_at),
        )
    clear_local_data_cache()


def build_trainer_fee_season_options(trainings: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    season_years = {
        get_season_start_year_for_date(parsed_date)
        for training in trainings
        for parsed_date in [parse_iso_date(str(training.get("date") or ""))]
        if parsed_date is not None
    }
    current_season_start_year = get_season_start_year_for_date(date.today())
    season_years.add(current_season_start_year)
    season_years.add(current_season_start_year + 1)
    return [
        {"value": str(year), "label": get_football_season_label(year)}
        for year in sorted(season_years, reverse=True)
    ]


def build_trainer_fee_monthly_summary(season_start_year: int, selected_month: Optional[int] = None) -> Dict[str, Any]:
    season_range = get_football_season_range(season_start_year)
    season_months = get_season_months(season_start_year)
    profiles = [
        profile
        for profile in load_trainer_profiles()
        if str(profile.get("fullName") or "").strip().lower() != "david van walstijn"
    ]
    profiles_by_id = {str(profile.get("id") or "").strip(): profile for profile in profiles}
    month_totals: Dict[int, Decimal] = {month: Decimal("0") for month in range(1, 13)}
    trainer_totals: Dict[str, Dict[int, Decimal]] = {
        str(profile.get("id") or "").strip(): {month: Decimal("0") for month in range(1, 13)}
        for profile in profiles
    }
    trainer_counts: Dict[str, Dict[int, int]] = {
        str(profile.get("id") or "").strip(): {month: 0 for month in range(1, 13)}
        for profile in profiles
    }
    payment_statuses = load_trainer_fee_payment_statuses(season_start_year)

    trainings = load_agenda_trainings(season_range["start"].isoformat(), season_range["end"].isoformat())
    matched_training_count = 0
    missing_rate_count = 0
    paid_groups: set[Tuple[str, str, str]] = set()

    for training in trainings:
        if not trainer_fee_training_counts_for_summary(training):
            continue
        training_date = parse_iso_date(str(training.get("date") or ""))
        if training_date is None or training_date < season_range["start"] or training_date > season_range["end"]:
            continue
        club = normalize_agenda_club(training.get("location"))
        activity = str(training.get("title") or "").strip()
        fee_type = get_trainer_fee_type_for_training(training.get("trainingType"))
        if not club or not activity:
            continue
        for trainer in training.get("trainers") or []:
            trainer_id = str(trainer.get("id") or "").strip()
            profile = profiles_by_id.get(trainer_id)
            if profile is None:
                continue
            fee_row = find_trainer_fee_row(profile, club, activity, fee_type)
            amount = parse_trainer_fee_amount(fee_row.get("amount")) if fee_row else Decimal("0")
            if amount <= 0:
                missing_rate_count += 1
                continue
            fee_group = str(fee_row.get("group") or "").strip().casefold() if fee_row else ""
            payment_key = (trainer_id, training_date.isoformat(), fee_group)
            month = training_date.month
            trainer_counts[trainer_id][month] += 1
            matched_training_count += 1
            if fee_group and payment_key in paid_groups:
                continue
            if fee_group:
                paid_groups.add(payment_key)
            trainer_totals[trainer_id][month] += amount
            month_totals[month] += amount

    trainer_rows = []
    for profile in profiles:
        trainer_id = str(profile.get("id") or "").strip()
        monthly_amounts = trainer_totals.get(trainer_id, {})
        monthly_counts = trainer_counts.get(trainer_id, {})
        year_total = sum(monthly_amounts.values(), Decimal("0"))
        trainer_rows.append(
            {
                "id": trainer_id,
                "name": str(profile.get("fullName") or "").strip(),
                "role": str(profile.get("systemRole") or "").strip(),
                "months": [
                    {
                        "month": month_meta["month"],
                        "year": month_meta["year"],
                        "label": month_meta["label"],
                        "shortLabel": month_meta["shortLabel"],
                        "amount": round(float(monthly_amounts.get(month_meta["month"], Decimal("0"))), 2),
                        "amountLabel": format_currency(float(monthly_amounts.get(month_meta["month"], Decimal("0")))).replace("EUR", "€"),
                        "trainingCount": monthly_counts.get(month_meta["month"], 0),
                        "paid": payment_statuses.get((trainer_id, month_meta["month"]), False),
                    }
                    for month_meta in season_months
                ],
                "yearTotal": round(float(year_total), 2),
                "yearTotalLabel": format_currency(float(year_total)).replace("EUR", "€"),
            }
        )

    months = [
        {
            "month": month_meta["month"],
            "year": month_meta["year"],
            "label": month_meta["label"],
            "shortLabel": month_meta["shortLabel"],
            "total": round(float(month_totals[month_meta["month"]]), 2),
            "totalLabel": format_currency(float(month_totals[month_meta["month"]])).replace("EUR", "€"),
        }
        for month_meta in season_months
    ]
    visible_months = [month for month in months if month["month"] == selected_month] if selected_month else months
    selected_total = month_totals.get(selected_month, Decimal("0")) if selected_month else sum(month_totals.values(), Decimal("0"))

    return {
        "selectedSeason": str(season_start_year),
        "selectedSeasonLabel": get_football_season_label(season_start_year),
        "selectedMonth": selected_month,
        "selectedMonthLabel": DUTCH_FULL_MONTH_NAMES[selected_month - 1] if selected_month else "",
        "months": months,
        "visibleMonths": visible_months,
        "trainers": trainer_rows,
        "yearTotal": round(float(sum(month_totals.values(), Decimal("0"))), 2),
        "yearTotalLabel": format_currency(float(sum(month_totals.values(), Decimal("0")))).replace("EUR", "€"),
        "selectedTotal": round(float(selected_total), 2),
        "selectedTotalLabel": format_currency(float(selected_total)).replace("EUR", "€"),
        "matchedTrainingCount": matched_training_count,
        "missingRateCount": missing_rate_count,
    }


def build_budget_activity_key(training_type: Any, club: Any, activity_title: Any) -> str:
    return "||".join(
        [
            str(training_type or "").strip(),
            str(club or "").strip(),
            str(activity_title or "").strip(),
        ]
    )


def parse_budget_activity_key(value: Any) -> Dict[str, str]:
    parts = str(value or "").split("||", 2)
    if len(parts) != 3:
        return {"trainingType": "", "club": "", "activityTitle": ""}
    training_type, club, activity_title = (part.strip() for part in parts)
    return {
        "trainingType": training_type if get_agenda_training_type_option(training_type)["value"] == training_type else "",
        "club": normalize_agenda_club(club),
        "activityTitle": activity_title,
    }


def build_budget_activity_options(season_start_year: int) -> List[Dict[str, Any]]:
    season_range = get_football_season_range(season_start_year)
    grouped: Dict[str, Dict[str, Any]] = {}
    for training in load_agenda_trainings(season_range["start"].isoformat(), season_range["end"].isoformat()):
        if str(training.get("status") or "").strip() == "geannuleerd":
            continue
        training_type = str(training.get("trainingType") or "").strip()
        club = normalize_agenda_club(training.get("location"))
        activity_title = str(training.get("title") or "").strip()
        if not training_type or not club or not activity_title:
            continue
        key = build_budget_activity_key(training_type, club, activity_title)
        if key not in grouped:
            grouped[key] = {
                "key": key,
                "trainingType": training_type,
                "trainingTypeLabel": str(training.get("trainingTypeLabel") or "").strip(),
                "club": club,
                "activityTitle": activity_title,
                "label": f"{activity_title} - {club} - {training.get('trainingTypeLabel') or training_type}",
                "count": 0,
                "scheduleSlots": [],
            }
        grouped[key]["count"] += 1
        training_date = parse_iso_date(str(training.get("date") or "").strip())
        start_time = str(training.get("time") or "").strip()
        end_time = str(training.get("endTime") or "").strip()
        if training_date is not None and start_time:
            slot = {
                "weekday": DUTCH_WEEKDAY_NAMES[training_date.weekday()],
                "weekdayIndex": training_date.weekday(),
                "startTime": start_time,
                "endTime": end_time,
            }
            slot_key = f"{slot['weekdayIndex']}|{slot['startTime']}|{slot['endTime']}"
            existing_slot_keys = {
                f"{item['weekdayIndex']}|{item['startTime']}|{item['endTime']}"
                for item in grouped[key]["scheduleSlots"]
            }
            if slot_key not in existing_slot_keys:
                grouped[key]["scheduleSlots"].append(slot)
                grouped[key]["scheduleSlots"].sort(
                    key=lambda item: (int(item["weekdayIndex"]), str(item["startTime"]), str(item["endTime"]))
                )
    return sorted(grouped.values(), key=lambda item: (str(item["club"]).lower(), str(item["activityTitle"]).lower()))


def load_budget_lines(season_start_year: int) -> List[Dict[str, Any]]:
    with get_db_connection() as connection:
        rows = connection.execute(
            """
            SELECT id, season_start_year, training_type, club, activity_title, income_amount,
                   trainer_amount, trainer_payment_mode, trainer_bundle_count, trainer_group, trainer_id,
                   sort_order, created_at, updated_at
            FROM budget_lines
            WHERE season_start_year = ?
            ORDER BY sort_order ASC, id ASC
            """,
            (season_start_year,),
        ).fetchall()
    return [
        {
            "id": int(row["id"]),
            "seasonStartYear": int(row["season_start_year"]),
            "trainingType": str(row["training_type"] or "").strip(),
            "club": str(row["club"] or "").strip(),
            "activityTitle": str(row["activity_title"] or "").strip(),
            "activityKey": build_budget_activity_key(row["training_type"], row["club"], row["activity_title"]),
            "incomeAmount": str(row["income_amount"] or "").strip(),
            "trainerAmount": str(row["trainer_amount"] or "").strip(),
            "trainerPaymentMode": str(row["trainer_payment_mode"] or "per_activity").strip(),
            "trainerBundleCount": max(1, parse_non_negative_int(row["trainer_bundle_count"])),
            "trainerGroup": str(row["trainer_group"] or "").strip(),
            "trainerId": str(row["trainer_id"] or "").strip(),
        }
        for row in rows
    ]


def parse_budget_lines_from_form(form_data: Any, season_start_year: int) -> List[Dict[str, Any]]:
    activity_keys = form_data.getlist("activity_key")
    income_amounts = form_data.getlist("income_amount")
    trainer_amounts = form_data.getlist("trainer_amount")
    trainer_groups = form_data.getlist("trainer_group")
    trainer_ids = form_data.getlist("trainer_id")
    valid_trainer_ids = {option["id"] for option in build_agenda_trainer_options()}
    rows = []
    max_length = max(
        len(activity_keys),
        len(income_amounts),
        len(trainer_amounts),
        len(trainer_groups),
        len(trainer_ids),
        0,
    )
    for index in range(max_length):
        activity = parse_budget_activity_key(activity_keys[index] if index < len(activity_keys) else "")
        if not activity["trainingType"] or not activity["club"] or not activity["activityTitle"]:
            continue
        trainer_id = str(trainer_ids[index] if index < len(trainer_ids) else "").strip()
        trainer_group = str(trainer_groups[index] if index < len(trainer_groups) else "").strip()
        rows.append(
            {
                "seasonStartYear": season_start_year,
                "trainingType": activity["trainingType"],
                "club": activity["club"],
                "activityTitle": activity["activityTitle"],
                "incomeAmount": format_contract_money(parse_decimal_amount(income_amounts[index] if index < len(income_amounts) else "")),
                "trainerAmount": format_contract_money(parse_decimal_amount(trainer_amounts[index] if index < len(trainer_amounts) else "")),
                "trainerPaymentMode": "per_group" if trainer_group else "per_activity",
                "trainerBundleCount": 1,
                "trainerGroup": trainer_group,
                "trainerId": trainer_id if trainer_id in valid_trainer_ids else "",
            }
        )
    return rows


def save_budget_lines(season_start_year: int, rows: List[Dict[str, Any]]) -> None:
    now = utcnow_iso()
    with get_db_connection() as connection:
        connection.execute("DELETE FROM budget_lines WHERE season_start_year = ?", (season_start_year,))
        connection.executemany(
            """
            INSERT INTO budget_lines (
                season_start_year, training_type, club, activity_title, income_amount,
                trainer_amount, trainer_payment_mode, trainer_bundle_count, trainer_group, trainer_id,
                sort_order, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    season_start_year,
                    row["trainingType"],
                    row["club"],
                    row["activityTitle"],
                    row["incomeAmount"],
                    row["trainerAmount"],
                    row["trainerPaymentMode"],
                    row["trainerBundleCount"],
                    row["trainerGroup"],
                    row["trainerId"],
                    index,
                    now,
                    now,
                )
                for index, row in enumerate(rows)
            ],
        )
    clear_local_data_cache()


def forward_budget_rows_to_trainer_profiles(rows: List[Dict[str, Any]], selected_indexes: set[int], season_start_year: int) -> Tuple[int, int]:
    """Append selected budget information to trainer planning without replacing existing rows."""
    activity_options = {option["key"]: option for option in build_budget_activity_options(season_start_year)}
    profiles = {str(profile.get("id") or "").strip(): profile for profile in load_trainer_profiles()}
    pending_by_trainer: Dict[str, List[Dict[str, Any]]] = {}
    skipped = 0
    group_amounts: Dict[str, str] = {}
    group_trainer_ids: Dict[str, str] = {}
    for row in rows:
        group_key = str(row.get("trainerGroup") or "").strip().casefold()
        amount = str(row.get("trainerAmount") or "").strip()
        trainer_id = str(row.get("trainerId") or "").strip()
        if group_key and amount and group_key not in group_amounts:
            group_amounts[group_key] = amount
        if group_key and trainer_id and group_key not in group_trainer_ids:
            group_trainer_ids[group_key] = trainer_id

    selected_group_keys = {
        str(rows[index].get("trainerGroup") or "").strip().casefold()
        for index in selected_indexes
        if 0 <= index < len(rows) and str(rows[index].get("trainerGroup") or "").strip()
    }
    expanded_selected_indexes = {
        index
        for index, row in enumerate(rows)
        if index in selected_indexes
        or str(row.get("trainerGroup") or "").strip().casefold() in selected_group_keys
    }

    for index, row in enumerate(rows):
        if index not in expanded_selected_indexes:
            continue
        group = str(row.get("trainerGroup") or "").strip()
        group_key = group.casefold()
        trainer_id = group_trainer_ids.get(group_key, str(row.get("trainerId") or "").strip()) if group else str(row.get("trainerId") or "").strip()
        if not trainer_id or trainer_id not in profiles:
            skipped += 1
            continue
        option = activity_options.get(build_budget_activity_key(row["trainingType"], row["club"], row["activityTitle"]), {})
        slots = option.get("scheduleSlots") if isinstance(option.get("scheduleSlots"), list) else []
        if not slots:
            slots = [{}]
        fee_type = get_trainer_fee_type_for_training(row.get("trainingType"))
        trainer_amount = group_amounts.get(group_key, str(row.get("trainerAmount") or "")) if group else str(row.get("trainerAmount") or "")
        for slot in slots:
            weekday = str(slot.get("weekday") or "").strip().lower()
            day = weekday if weekday in {"maandag", "dinsdag", "woensdag", "donderdag", "vrijdag", "zaterdag", "zondag"} else ""
            pending_by_trainer.setdefault(trainer_id, []).append(
                {
                    "day": day,
                    "time": str(slot.get("startTime") or "").strip(),
                    "type": fee_type,
                    "club": row.get("club", ""),
                    "activity": row.get("activityTitle", ""),
                    "amount": trainer_amount,
                    "group": group,
                }
            )

    added = 0
    for trainer_id, new_rows in pending_by_trainer.items():
        existing = list(profiles[trainer_id].get("trainerFees") or [])
        existing_keys = {
            (item.get("day", ""), item.get("time", ""), item.get("type", ""), item.get("club", ""), item.get("activity", ""))
            for item in existing
        }
        for new_row in new_rows:
            key = (new_row["day"], new_row["time"], new_row["type"], new_row["club"], new_row["activity"])
            if key in existing_keys:
                continue
            existing.append(new_row)
            existing_keys.add(key)
            added += 1
        update_trainer_fee_rows(trainer_id, existing)
    return added, skipped


def build_budget_summary(season_start_year: int) -> Dict[str, Any]:
    activity_options = build_budget_activity_options(season_start_year)
    activity_counts = {option["key"]: int(option["count"]) for option in activity_options}
    activity_labels = {option["key"]: str(option["label"]) for option in activity_options}
    trainer_names = {option["id"]: option["name"] for option in build_agenda_trainer_options()}
    budget_lines = load_budget_lines(season_start_year)
    group_summaries: Dict[str, Dict[str, Any]] = {}
    for index, line in enumerate(budget_lines):
        trainer_group = str(line.get("trainerGroup") or "").strip()
        if not trainer_group:
            continue
        group_key = trainer_group.casefold()
        count = activity_counts.get(line["activityKey"], 0)
        income_total = parse_decimal_amount(line.get("incomeAmount")) * count
        if group_key not in group_summaries:
            group_summaries[group_key] = {
                "leaderIndex": index,
                "trainingCount": count,
                "incomeTotal": income_total,
                "trainerAmount": parse_decimal_amount(line.get("trainerAmount")),
            }
        else:
            group_summaries[group_key]["trainingCount"] = max(int(group_summaries[group_key]["trainingCount"]), count)
            group_summaries[group_key]["incomeTotal"] += income_total
    lines = []
    total_income = Decimal("0")
    total_trainer_costs = Decimal("0")

    for line_index, line in enumerate(budget_lines):
        key = line["activityKey"]
        count = activity_counts.get(key, 0)
        income_amount = parse_decimal_amount(line.get("incomeAmount"))
        trainer_amount = parse_decimal_amount(line.get("trainerAmount"))
        trainer_group = str(line.get("trainerGroup") or "").strip()
        income_total = income_amount * count
        total_income += income_total
        group_key = trainer_group.casefold() if trainer_group else ""
        group_summary = group_summaries.get(group_key) if group_key else None
        is_group_leader = bool(group_summary and int(group_summary["leaderIndex"]) == line_index)
        is_group_follower = bool(group_summary and int(group_summary["leaderIndex"]) != line_index)
        if group_summary:
            if is_group_leader:
                trainer_cost = group_summary["trainerAmount"] * int(group_summary["trainingCount"])
                result = group_summary["incomeTotal"] - trainer_cost
                display_income_total = group_summary["incomeTotal"]
                total_trainer_costs += trainer_cost
            else:
                trainer_cost = Decimal("0")
                result = Decimal("0")
                display_income_total = Decimal("0")
        else:
            trainer_cost = trainer_amount * count
            result = income_total - trainer_cost
            display_income_total = income_total
            total_trainer_costs += trainer_cost
        lines.append(
            {
                **line,
                "activityLabel": activity_labels.get(key) or f"{line['activityTitle']} - {line['club']}",
                "count": count,
                "incomeTotal": round(float(display_income_total), 2),
                "incomeTotalLabel": format_currency(float(display_income_total)).replace("EUR", "€"),
                "trainerCost": round(float(trainer_cost), 2),
                "trainerCostLabel": format_currency(float(trainer_cost)).replace("EUR", "€"),
                "result": round(float(result), 2),
                "resultLabel": format_currency(float(result)).replace("EUR", "€"),
                "trainerName": trainer_names.get(str(line.get("trainerId") or "").strip(), ""),
                "isGrouped": bool(group_summary),
                "isGroupLeader": is_group_leader,
                "isGroupFollower": is_group_follower,
            }
        )

    result_total = total_income - total_trainer_costs
    return {
        "selectedSeason": str(season_start_year),
        "selectedSeasonLabel": get_football_season_label(season_start_year),
        "activityOptions": activity_options,
        "lines": lines,
        "lineCount": len(lines),
        "totalIncome": round(float(total_income), 2),
        "totalIncomeLabel": format_currency(float(total_income)).replace("EUR", "€"),
        "totalTrainerCosts": round(float(total_trainer_costs), 2),
        "totalTrainerCostsLabel": format_currency(float(total_trainer_costs)).replace("EUR", "€"),
        "resultTotal": round(float(result_total), 2),
        "resultTotalLabel": format_currency(float(result_total)).replace("EUR", "€"),
    }


def build_user_payload(row: sqlite3.Row) -> Dict[str, Any]:
    system_role = normalize_system_role(str(row["system_role"] or row["role"] or ""))
    trainer_fees_json = str(row["trainer_fees_json"] or "[]") if "trainer_fees_json" in row.keys() else "[]"
    try:
        trainer_fees_payload = json.loads(trainer_fees_json)
    except (TypeError, json.JSONDecodeError):
        trainer_fees_payload = []
    return {
        "id": str(row["id"]),
        "fullName": str(row["full_name"] or "").strip(),
        "email": str(row["email"] or "").strip(),
        "username": str(row["username"] or "").strip(),
        "passwordHash": str(row["password_hash"] or "").strip(),
        "inviteToken": str(row["invite_token"] or "").strip(),
        "inviteExpiresAt": str(row["invite_expires_at"] or "").strip(),
        "inviteAcceptedAt": str(row["invite_accepted_at"] or "").strip(),
        "role": system_role,
        "memberType": str(row["member_type"] or "").strip(),
        "systemRole": system_role,
        "knvbLicense": str(row["knvb_license"] or "").strip(),
        "education": str(row["education"] or "").strip(),
        "availabilityDays": [day for day in str(row["availability_days"] or "").split(",") if day],
        "phone": str(row["phone"] or "").strip(),
        "address": str(row["address"] or "").strip() if "address" in row.keys() else "",
        "city": str(row["city"] or "").strip() if "city" in row.keys() else "",
        "postalCode": str(row["postal_code"] or "").strip() if "postal_code" in row.keys() else "",
        "bankAccountNumber": str(row["bank_account_number"] or "").strip() if "bank_account_number" in row.keys() else "",
        "bankAccountName": str(row["bank_account_name"] or "").strip() if "bank_account_name" in row.keys() else "",
        "notes": str(row["notes"] or "").strip(),
        "trainerFees": normalize_trainer_fee_rows(trainer_fees_payload),
        "isAdmin": bool(row["is_admin"]) or role_grants_admin_access(system_role),
        "status": str(row["status"] or "Actief").strip(),
        "createdAt": str(row["created_at"] or "").strip(),
    }


def normalize_username_seed(value: str) -> str:
    cleaned = []
    previous_separator = False
    for char in value.lower():
        if char.isalnum():
            cleaned.append(char)
            previous_separator = False
        elif not previous_separator:
            cleaned.append(".")
            previous_separator = True

    normalized = "".join(cleaned).strip(".")
    return normalized or "gebruiker"


def build_invite_expiry(days: int = 14) -> str:
    return (datetime.utcnow() + timedelta(days=days)).replace(microsecond=0).isoformat()


def create_invite_token() -> str:
    return secrets.token_urlsafe(32)


def update_trainer_profile(
    profile_id: str,
    full_name: str,
    email: str,
    username: str,
    member_type: str,
    system_role: str,
    knvb_license: str,
    education: str,
    phone: str,
    address: str,
    city: str,
    postal_code: str,
    bank_account_number: str,
    bank_account_name: str,
    notes: str,
    availability_days: List[str],
    is_admin: bool,
    trainer_fees: Optional[List[Dict[str, Any]]] = None,
) -> None:
    with get_db_connection() as connection:
        if trainer_fees is None:
            connection.execute(
                """
                UPDATE trainer_profiles
                SET
                    full_name = ?,
                    email = ?,
                    username = ?,
                    role = ?,
                    member_type = ?,
                    system_role = ?,
                    knvb_license = ?,
                    education = ?,
                    availability_days = ?,
                    phone = ?,
                    address = ?,
                    city = ?,
                    postal_code = ?,
                    bank_account_number = ?,
                    bank_account_name = ?,
                    notes = ?,
                    is_admin = ?
                WHERE id = ?
                """,
                (
                    full_name.strip(),
                    email.strip(),
                    username.strip(),
                    system_role.strip() or "Trainer",
                    member_type.strip(),
                    system_role.strip(),
                    knvb_license.strip(),
                    education.strip(),
                    ",".join(day.strip() for day in availability_days if day.strip()),
                    phone.strip(),
                    address.strip(),
                    city.strip(),
                    postal_code.strip(),
                    bank_account_number.strip(),
                    bank_account_name.strip(),
                    notes.strip(),
                    1 if is_admin else 0,
                    profile_id.strip(),
                ),
            )
        else:
            connection.execute(
                """
                UPDATE trainer_profiles
                SET
                    full_name = ?,
                    email = ?,
                    username = ?,
                    role = ?,
                    member_type = ?,
                    system_role = ?,
                    knvb_license = ?,
                    education = ?,
                    availability_days = ?,
                    phone = ?,
                    address = ?,
                    city = ?,
                    postal_code = ?,
                    bank_account_number = ?,
                    bank_account_name = ?,
                    notes = ?,
                    is_admin = ?,
                    trainer_fees_json = ?
                WHERE id = ?
                """,
                (
                    full_name.strip(),
                    email.strip(),
                    username.strip(),
                    system_role.strip() or "Trainer",
                    member_type.strip(),
                    system_role.strip(),
                    knvb_license.strip(),
                    education.strip(),
                    ",".join(day.strip() for day in availability_days if day.strip()),
                    phone.strip(),
                    address.strip(),
                    city.strip(),
                    postal_code.strip(),
                    bank_account_number.strip(),
                    bank_account_name.strip(),
                    notes.strip(),
                    1 if is_admin else 0,
                    trainer_fees_json_dumps(trainer_fees),
                    profile_id.strip(),
                ),
            )
    clear_local_data_cache()


def delete_trainer_profile(profile_id: str) -> None:
    with get_db_connection() as connection:
        connection.execute("DELETE FROM trainer_profiles WHERE id = ?", (profile_id.strip(),))
    clear_local_data_cache()


def seed_workspace_tables() -> None:
    settings = load_dashboard_weather_settings()
    for key, value in settings.items():
        save_dashboard_preference(key, value)

    if not table_has_rows("tasks"):
        add_task("Veldindeling controleren", "2026-04-01")
        add_task("Trainingsmateriaal klaarzetten", "2026-04-02")


def build_workspace_summary() -> Dict[str, int]:
    tasks = load_tasks()
    agenda_items = load_agenda_trainings()
    team_members = load_trainer_profiles()
    return {
        "openTasks": sum(1 for item in tasks if not item.get("isDone")),
        "doneTasks": sum(1 for item in tasks if item.get("isDone")),
        "agendaCount": len(agenda_items),
        "teamCount": len(team_members),
    }


def get_weather_description(code: int) -> Dict[str, str]:
    descriptions = {
        0: {"label": "Onbewolkt", "icon": "Sun"},
        1: {"label": "Licht bewolkt", "icon": "CloudSun"},
        2: {"label": "Half bewolkt", "icon": "Cloud"},
        3: {"label": "Bewolkt", "icon": "Cloud"},
        45: {"label": "Mist", "icon": "CloudFog"},
        48: {"label": "Rijp mist", "icon": "CloudFog"},
        51: {"label": "Lichte motregen", "icon": "CloudDrizzle"},
        53: {"label": "Motregen", "icon": "CloudDrizzle"},
        55: {"label": "Zware motregen", "icon": "CloudRain"},
        61: {"label": "Lichte regen", "icon": "CloudRain"},
        63: {"label": "Regen", "icon": "CloudRain"},
        65: {"label": "Zware regen", "icon": "CloudRain"},
        71: {"label": "Lichte sneeuw", "icon": "CloudSnow"},
        73: {"label": "Sneeuw", "icon": "CloudSnow"},
        75: {"label": "Zware sneeuw", "icon": "Snowflake"},
        80: {"label": "Lichte buien", "icon": "CloudRain"},
        81: {"label": "Buien", "icon": "CloudRain"},
        82: {"label": "Zware buien", "icon": "CloudLightning"},
        95: {"label": "Onweer", "icon": "CloudLightning"},
        96: {"label": "Onweer met hagel", "icon": "CloudLightning"},
        99: {"label": "Zwaar onweer", "icon": "CloudLightning"},
    }
    return descriptions.get(code, {"label": "Onbekend", "icon": "Cloud"})


def load_trainer_profiles() -> List[Dict[str, Any]]:
    def loader() -> List[Dict[str, Any]]:
        with get_db_connection() as connection:
            rows = connection.execute(
                """
                SELECT
                    id,
                    full_name,
                    email,
                    username,
                    password_hash,
                    invite_token,
                    invite_expires_at,
                    invite_accepted_at,
                    role,
                    member_type,
                    system_role,
                    knvb_license,
                    education,
                    availability_days,
                    phone,
                    address,
                    city,
                    postal_code,
                    bank_account_number,
                    bank_account_name,
                    notes,
                    trainer_fees_json,
                    is_admin,
                    status,
                    created_at
                FROM trainer_profiles
                ORDER BY full_name COLLATE NOCASE ASC, created_at DESC
                """
            ).fetchall()

        profiles = []
        for row in rows:
            profile = build_user_payload(row)
            profile["memberType"] = profile.get("memberType") or ("Medewerker" if profile.get("isAdmin") else "Vrijwilliger")
            profile["systemRole"] = profile.get("systemRole") or profile.get("role") or ""
            profiles.append(profile)

        return profiles

    return get_cached_local_data("trainer_profiles", (), loader)


def build_admin_account_debug_summary() -> Dict[str, Any]:
    profiles = load_trainer_profiles()
    return {
        "total": len(profiles),
        "admins": sum(1 for item in profiles if item.get("isAdmin")),
        "invited": sum(1 for item in profiles if item.get("status") == "Uitgenodigd"),
        "active": sum(1 for item in profiles if item.get("status") == "Actief"),
    }


def build_admin_content_debug_summary(repaired_albums: Optional[int] = None) -> Dict[str, Any]:
    if repaired_albums is None:
        repaired_albums = ensure_content_album_records_exist()
    with get_db_connection() as connection:
        counts_row = connection.execute(
            """
            SELECT
                (SELECT COUNT(*) FROM content_albums) AS album_count,
                (SELECT COUNT(*) FROM content_photos) AS photo_count,
                (
                    SELECT COUNT(*)
                    FROM content_photos cp
                    LEFT JOIN content_albums ca ON ca.id = cp.album_id
                    WHERE ca.id IS NULL
                ) AS orphan_photo_count,
                (SELECT COUNT(*) FROM exercises) AS exercise_count,
                (SELECT COUNT(*) FROM faq_items) AS faq_count,
                (SELECT COUNT(*) FROM training_plans) AS training_plan_count,
                (SELECT COUNT(*) FROM workflow_documents) AS workflow_document_count
            """
        ).fetchone()

    return {
        "albums": int(counts_row["album_count"] or 0) if counts_row is not None else 0,
        "photos": int(counts_row["photo_count"] or 0) if counts_row is not None else 0,
        "orphanPhotos": int(counts_row["orphan_photo_count"] or 0) if counts_row is not None else 0,
        "repairedAlbums": repaired_albums,
        "exercises": int(counts_row["exercise_count"] or 0) if counts_row is not None else 0,
        "faqItems": int(counts_row["faq_count"] or 0) if counts_row is not None else 0,
        "trainingPlans": int(counts_row["training_plan_count"] or 0) if counts_row is not None else 0,
        "workflowDocuments": int(counts_row["workflow_document_count"] or 0) if counts_row is not None else 0,
    }


def trainer_username_exists(username: str, exclude_profile_id: str = "") -> bool:
    with get_db_connection() as connection:
        if exclude_profile_id.strip():
            row = connection.execute(
                """
                SELECT 1
                FROM trainer_profiles
                WHERE lower(username) = lower(?) AND id != ?
                LIMIT 1
                """,
                (username.strip(), exclude_profile_id.strip()),
            ).fetchone()
        else:
            row = connection.execute(
                "SELECT 1 FROM trainer_profiles WHERE lower(username) = lower(?) LIMIT 1",
                (username.strip(),),
            ).fetchone()
    return row is not None


def trainer_email_exists(email: str, exclude_profile_id: str = "") -> bool:
    with get_db_connection() as connection:
        if exclude_profile_id.strip():
            row = connection.execute(
                """
                SELECT 1
                FROM trainer_profiles
                WHERE lower(email) = lower(?) AND id != ?
                LIMIT 1
                """,
                (email.strip(), exclude_profile_id.strip()),
            ).fetchone()
        else:
            row = connection.execute(
                "SELECT 1 FROM trainer_profiles WHERE lower(email) = lower(?) LIMIT 1",
                (email.strip(),),
            ).fetchone()
    return row is not None


def build_internal_username(full_name: str, email: str, exclude_profile_id: str = "") -> str:
    email_local_part = email.strip().split("@", 1)[0]
    base_value = normalize_username_seed(email_local_part or full_name or "gebruiker")
    candidate = base_value
    suffix = 2

    while trainer_username_exists(candidate, exclude_profile_id=exclude_profile_id):
        candidate = f"{base_value}.{suffix}"
        suffix += 1

    return candidate


def add_trainer_profile(
    full_name: str,
    email: str,
    password: str,
    role: str,
    member_type: str,
    system_role: str,
    knvb_license: str,
    education: str,
    availability_days: List[str],
    phone: str,
    address: str = "",
    city: str = "",
    postal_code: str = "",
    bank_account_number: str = "",
    bank_account_name: str = "",
    notes: str = "",
    is_admin: bool = False,
) -> None:
    created_at = datetime.now().isoformat(timespec="seconds")
    profile_id = f"trainer-{int(time.time() * 1000)}"
    username = build_internal_username(full_name, email)
    with get_db_connection() as connection:
        connection.execute(
            """
            INSERT INTO trainer_profiles (
                id, full_name, email, username, password_hash, invite_token, invite_expires_at, invite_accepted_at,
                role, member_type, system_role, knvb_license, education, availability_days, phone, address, city, postal_code,
                bank_account_number, bank_account_name, notes, is_admin, status, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                profile_id,
                full_name.strip(),
                email.strip(),
                username.strip(),
                hash_password(password),
                None,
                None,
                utcnow_iso(),
                role.strip() or system_role.strip() or "Trainer",
                member_type.strip() or "Vrijwilliger",
                system_role.strip() or role.strip() or "Trainer",
                knvb_license.strip(),
                education.strip(),
                ",".join(day.strip() for day in availability_days if day.strip()),
                phone.strip(),
                address.strip(),
                city.strip(),
                postal_code.strip(),
                bank_account_number.strip(),
                bank_account_name.strip(),
                notes.strip(),
                1 if is_admin else 0,
                "Actief",
                created_at,
            ),
        )
    clear_local_data_cache()


def create_trainer_invite_profile(
    full_name: str,
    email: str,
    role: str,
    member_type: str,
    system_role: str,
    knvb_license: str,
    education: str,
    availability_days: List[str],
    phone: str,
    address: str,
    city: str,
    postal_code: str,
    bank_account_number: str,
    bank_account_name: str,
    notes: str,
    is_admin: bool = False,
) -> Dict[str, str]:
    created_at = datetime.now().isoformat(timespec="seconds")
    profile_id = f"trainer-{int(time.time() * 1000)}"
    username = build_internal_username(full_name, email)
    invite_token = create_invite_token()
    invite_expires_at = build_invite_expiry()
    with get_db_connection() as connection:
        connection.execute(
            """
            INSERT INTO trainer_profiles (
                id, full_name, email, username, password_hash, invite_token, invite_expires_at, invite_accepted_at,
                role, member_type, system_role, knvb_license, education, availability_days, phone, address, city, postal_code,
                bank_account_number, bank_account_name, notes, is_admin, status, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                profile_id,
                full_name.strip(),
                email.strip(),
                username.strip(),
                None,
                invite_token,
                invite_expires_at,
                None,
                role.strip() or system_role.strip() or "Trainer",
                member_type.strip() or "Vrijwilliger",
                system_role.strip() or role.strip() or "Trainer",
                knvb_license.strip(),
                education.strip(),
                ",".join(day.strip() for day in availability_days if day.strip()),
                phone.strip(),
                address.strip(),
                city.strip(),
                postal_code.strip(),
                bank_account_number.strip(),
                bank_account_name.strip(),
                notes.strip(),
                1 if is_admin else 0,
                "Uitgenodigd",
                created_at,
            ),
        )
    clear_local_data_cache()

    return {
        "profileId": profile_id,
        "inviteToken": invite_token,
        "inviteExpiresAt": invite_expires_at,
    }


def get_user_by_id(user_id: str) -> Optional[Dict[str, Any]]:
    normalized_user_id = str(user_id or "").strip()

    def loader() -> Optional[Dict[str, Any]]:
        with get_db_connection() as connection:
            row = connection.execute(
                """
                SELECT
                    id, full_name, email, username, password_hash, invite_token, invite_expires_at, invite_accepted_at, role, member_type, system_role,
                    knvb_license, education, availability_days, phone, address, city, postal_code, bank_account_number,
                    bank_account_name, notes, trainer_fees_json, is_admin, status, created_at
                FROM trainer_profiles
                WHERE id = ?
                LIMIT 1
                """,
                (normalized_user_id,),
            ).fetchone()

        if row is None:
            return None

        return build_user_payload(row)

    return get_cached_local_data("user_by_id", (normalized_user_id,), loader)


def get_user_by_username(username: str) -> Optional[Dict[str, Any]]:
    normalized_username = str(username or "").strip()

    def loader() -> Optional[Dict[str, Any]]:
        with get_db_connection() as connection:
            row = connection.execute(
                """
                SELECT
                    id, full_name, email, username, password_hash, invite_token, invite_expires_at, invite_accepted_at, role, member_type, system_role,
                    knvb_license, education, availability_days, phone, address, city, postal_code, bank_account_number,
                    bank_account_name, notes, is_admin, status, created_at
                FROM trainer_profiles
                WHERE lower(username) = lower(?)
                LIMIT 1
                """,
                (normalized_username,),
            ).fetchone()

        if row is None:
            return None

        return build_user_payload(row)

    return get_cached_local_data("user_by_username", (normalized_username.lower(),), loader)


def get_user_by_login(login_value: str) -> Optional[Dict[str, Any]]:
    normalized_login = str(login_value or "").strip()

    def loader() -> Optional[Dict[str, Any]]:
        with get_db_connection() as connection:
            row = connection.execute(
                """
                SELECT
                    id, full_name, email, username, password_hash, invite_token, invite_expires_at, invite_accepted_at, role, member_type, system_role,
                    knvb_license, education, availability_days, phone, address, city, postal_code, bank_account_number,
                    bank_account_name, notes, is_admin, status, created_at
                FROM trainer_profiles
                WHERE lower(email) = lower(?) OR lower(username) = lower(?)
                ORDER BY is_admin DESC, created_at ASC
                LIMIT 1
                """,
                (normalized_login, normalized_login),
            ).fetchone()

        if row is None:
            return None

        return build_user_payload(row)

    return get_cached_local_data("user_by_login", (normalized_login.lower(),), loader)


def get_user_by_invite_token(invite_token: str) -> Optional[Dict[str, Any]]:
    normalized_invite_token = str(invite_token or "").strip()

    def loader() -> Optional[Dict[str, Any]]:
        with get_db_connection() as connection:
            row = connection.execute(
                """
                SELECT
                    id, full_name, email, username, password_hash, invite_token, invite_expires_at, invite_accepted_at, role, member_type, system_role,
                    knvb_license, education, availability_days, phone, address, city, postal_code, bank_account_number,
                    bank_account_name, notes, is_admin, status, created_at
                FROM trainer_profiles
                WHERE invite_token = ?
                LIMIT 1
                """,
                (normalized_invite_token,),
            ).fetchone()

        if row is None:
            return None

        return build_user_payload(row)

    return get_cached_local_data("user_by_invite_token", (normalized_invite_token,), loader)


def accept_trainer_invite(profile_id: str, password: str) -> None:
    with get_db_connection() as connection:
        connection.execute(
            """
            UPDATE trainer_profiles
            SET
                password_hash = ?,
                invite_token = NULL,
                invite_accepted_at = ?,
                status = 'Actief'
            WHERE id = ?
            """,
            (
                hash_password(password),
                utcnow_iso(),
                profile_id.strip(),
            ),
        )
    clear_local_data_cache()


def authenticate_user(login_value: str, password: str) -> Optional[Dict[str, Any]]:
    user = get_user_by_login(login_value)
    if user is None or not user.get("passwordHash"):
        return None
    if not check_password_hash(user["passwordHash"], password):
        return None
    if password_needs_rehash(user["passwordHash"]):
        update_user_password_hash(user["id"], password)
        user = get_user_by_id(user["id"]) or user
    return user


def ensure_admin_account() -> None:
    with get_db_connection() as connection:
        row = connection.execute(
            "SELECT 1 FROM trainer_profiles WHERE is_admin = 1 LIMIT 1"
        ).fetchone()
    if row is not None:
        return

    admin_password = get_env("ADMIN_PASSWORD")
    admin_email = get_env("ADMIN_EMAIL")
    if not admin_email or not admin_password or is_placeholder_value(admin_password):
        app.logger.warning("Geen automatisch admin-account aangemaakt: ADMIN_EMAIL/ADMIN_PASSWORD ontbreken.")
        return

    add_trainer_profile(
        full_name="Beheerder",
        email=admin_email,
        password=admin_password,
        role="Admin",
        member_type="Medewerker",
        system_role="Admin",
        knvb_license="",
        education="",
        availability_days=[],
        phone="",
        notes="Automatisch aangemaakt beheeraccount.",
        is_admin=True,
    )


def require_admin_user() -> Optional[Any]:
    user = get_current_user()
    if user is None or not user.get("isAdmin"):
        return require_page_access("dashboard")
    return None


def get_default_post_login_path(user: Dict[str, Any]) -> str:
    return "/"


def is_valid_email_address(value: str) -> bool:
    normalized = str(value or "").strip()
    return bool(re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", normalized))


def combine_date_and_time(date_value: str, time_value: str) -> datetime:
    return datetime.fromisoformat(f"{date_value}T{time_value}")


def compute_default_end_time(time_value: str) -> str:
    start = datetime.strptime(time_value, "%H:%M")
    end = start + timedelta(minutes=90)
    return end.strftime("%H:%M")


def get_week_days(week_start: date) -> List[Dict[str, Any]]:
    day_names = ["ma", "di", "wo", "do", "vr", "za", "zo"]
    days = []
    for index in range(7):
        current = week_start + timedelta(days=index)
        days.append(
            {
                "date": current,
                "key": current.isoformat(),
                "shortLabel": f"{day_names[index]} {current.day}-{current.month}",
                "isToday": current == date.today(),
            }
        )
    return days


def build_trainer_dashboard_week_schedule(
    user: Optional[Dict[str, Any]],
    reference_datetime: Optional[datetime] = None,
) -> List[Dict[str, Any]]:
    if not is_trainer_user(user):
        return []

    trainer_id = str(user.get("id") or "").strip()
    if not trainer_id:
        return []

    current_moment = reference_datetime or datetime.now(ZoneInfo(settings.TIME_ZONE)).replace(tzinfo=None)
    current_day = current_moment.date()
    week_end = current_day + timedelta(days=6 - current_day.weekday())
    schedule: List[Dict[str, Any]] = []

    for training in load_agenda_trainings(current_day.isoformat(), week_end.isoformat()):
        assigned_trainer_ids = {
            str(trainer.get("id") or "").strip()
            for trainer in training.get("trainers") or []
            if isinstance(trainer, dict)
        }
        if trainer_id not in assigned_trainer_ids:
            continue

        try:
            training_day = date.fromisoformat(str(training.get("date") or "").strip())
        except ValueError:
            continue
        if training_day < current_day or training_day > week_end:
            continue

        start_time = str(training.get("time") or "").strip()
        end_time = str(training.get("endTime") or "").strip()
        if training_day == current_day and (end_time or start_time):
            try:
                final_time = datetime.strptime(end_time or start_time, "%H:%M").time()
            except ValueError:
                final_time = None
            if final_time is not None and final_time <= current_moment.time():
                continue

        if training_day == current_day:
            date_label = "Vandaag"
        elif training_day == current_day + timedelta(days=1):
            date_label = "Morgen"
        else:
            date_label = (
                f"{DUTCH_WEEKDAY_NAMES[training_day.weekday()]} "
                f"{training_day.day} {DUTCH_FULL_MONTH_NAMES[training_day.month - 1]}"
            )

        time_label = start_time
        if start_time and end_time:
            time_label = f"{start_time} - {end_time}"

        schedule.append(
            {
                "id": str(training.get("id") or "").strip(),
                "date": training_day.isoformat(),
                "dateLabel": date_label,
                "time": start_time,
                "timeLabel": time_label,
                "title": str(training.get("title") or "Training").strip() or "Training",
                "location": str(training.get("location") or "").strip(),
                "clubClass": get_agenda_club_class(training.get("location", "")),
                "trainingTypeLabel": str(training.get("trainingTypeLabel") or "").strip(),
                "trainingTypeClass": str(training.get("trainingTypeClass") or "").strip(),
                "status": str(training.get("status") or "").strip(),
                "statusLabel": str(training.get("statusLabel") or "Gepland").strip() or "Gepland",
            }
        )

    schedule.sort(key=lambda item: (item["date"], item["time"], item["title"].casefold()))
    return schedule


def format_agenda_summary_day_label(day_value: date) -> str:
    return (
        f"{DUTCH_WEEKDAY_NAMES[day_value.weekday()]} "
        f"{day_value.day} {DUTCH_FULL_MONTH_NAMES[day_value.month - 1]} {day_value.year}"
    )


def format_agenda_day_copy_label(day_value: date, plan_type: str = "") -> str:
    label = format_agenda_summary_day_label(day_value)
    if plan_type == "Geen activiteit":
        reason = AGENDA_NO_ACTIVITY_COPY_REASONS.get(day_value.isoformat())
        if reason:
            label = f"{label} ({reason})"
    return label


def build_numbered_agenda_day_copy_text(days: List[date], plan_type: str = "") -> str:
    return "\n".join(
        f"{index}. {format_agenda_day_copy_label(day_value, plan_type)}"
        for index, day_value in enumerate(sorted(days), start=1)
    )


def is_football_day_agenda_training(training: Dict[str, Any]) -> bool:
    title = normalize_match_text(training.get("title", ""))
    return "voetbaldag" in title


def add_football_day_only_no_activity_days(
    day_plans: List[Dict[str, Any]],
    trainings: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    summary_day_plans = list(day_plans)
    plan_types_by_date: Dict[date, Set[str]] = {}
    no_activity_dates: Set[date] = set()
    trainings_by_date: Dict[date, List[Dict[str, Any]]] = {}

    for day_plan in day_plans:
        plan_type = str(day_plan.get("planType") or day_plan.get("plan_type") or "").strip()
        current_date = day_plan.get("date")
        if isinstance(current_date, str):
            current_date = parse_iso_date(current_date.strip())
        if isinstance(current_date, date):
            plan_types_by_date.setdefault(current_date, set()).add(plan_type)
            if plan_type == "Geen activiteit":
                no_activity_dates.add(current_date)

    for training in trainings:
        current_date = training.get("date")
        if isinstance(current_date, str):
            current_date = parse_iso_date(current_date.strip())
        if isinstance(current_date, date):
            trainings_by_date.setdefault(current_date, []).append(training)

    football_day_dates = {
        current_date
        for current_date, plan_types in plan_types_by_date.items()
        if plan_types == {"Voetbaldag"}
    }
    football_day_dates.update(
        current_date
        for current_date, day_trainings in trainings_by_date.items()
        if current_date not in plan_types_by_date
        and day_trainings
        and all(is_football_day_agenda_training(training) for training in day_trainings)
    )

    for current_date in sorted(football_day_dates):
        day_trainings = trainings_by_date.get(current_date, [])
        has_regular_training = any(
            not is_football_day_agenda_training(training)
            for training in day_trainings
        )
        if current_date not in no_activity_dates and not has_regular_training:
            summary_day_plans.append(
                {
                    "date": current_date,
                    "planType": "Geen activiteit",
                }
            )

    return summary_day_plans


def build_agenda_day_plan_summary(day_plans: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    summary: List[Dict[str, Any]] = []
    plan_counts = {option: 0 for option in AGENDA_DAY_PLAN_OPTIONS}
    weekday_counts = {
        option: {weekday: 0 for weekday in range(7)}
        for option in AGENDA_DAY_PLAN_OPTIONS
    }
    weekday_days = {
        option: {weekday: [] for weekday in range(7)}
        for option in AGENDA_DAY_PLAN_OPTIONS
    }

    for day_plan in day_plans:
        plan_type = str(day_plan.get("planType") or day_plan.get("plan_type") or "").strip()
        if plan_type not in plan_counts:
            continue

        current_date = day_plan.get("date")
        if isinstance(current_date, str):
            current_date = parse_iso_date(current_date.strip())
        if isinstance(current_date, date):
            plan_counts[plan_type] += 1
            weekday_counts[plan_type][current_date.weekday()] += 1
            weekday_days[plan_type][current_date.weekday()].append(current_date)

    for option in AGENDA_DAY_PLAN_OPTIONS:
        item = {
            "label": option,
            "count": plan_counts.get(option, 0),
            "details": [
                {
                    "label": DUTCH_WEEKDAY_NAMES[weekday],
                    "count": count,
                    "days": [
                        {
                            "date": day_value.isoformat(),
                            "label": format_agenda_summary_day_label(day_value),
                        }
                        for day_value in sorted(weekday_days[option][weekday])
                    ],
                    "copyText": build_numbered_agenda_day_copy_text(weekday_days[option][weekday], option),
                }
                for weekday, count in weekday_counts[option].items()
                if count > 0
            ],
        }
        summary.append(item)

    return summary


def build_week_label(week_start: date) -> str:
    week_end = week_start + timedelta(days=6)
    if week_start.month == week_end.month:
        return f"{week_start.day}-{week_end.day} {DUTCH_MONTH_NAMES[week_start.month - 1]} {week_start.year}"
    return (
        f"{week_start.day} {DUTCH_MONTH_NAMES[week_start.month - 1]} - "
        f"{week_end.day} {DUTCH_MONTH_NAMES[week_end.month - 1]} {week_start.year}"
    )


def add_months(base_date: date, month_offset: int) -> date:
    month_index = (base_date.month - 1) + month_offset
    year = base_date.year + (month_index // 12)
    month = (month_index % 12) + 1
    return date(year, month, 1)


def build_month_label(month_start: date) -> str:
    return f"{DUTCH_MONTH_NAMES[month_start.month - 1]} {month_start.year}"


def build_full_month_label(month_start: date) -> str:
    return f"{DUTCH_FULL_MONTH_NAMES[month_start.month - 1]} {month_start.year}"


def build_agenda_month_days(month_start: date) -> List[List[Dict[str, Any]]]:
    sunday_first_calendar = calendar.Calendar(firstweekday=6)
    month_weeks: List[List[Dict[str, Any]]] = []

    for week in sunday_first_calendar.monthdatescalendar(month_start.year, month_start.month):
        week_days: List[Dict[str, Any]] = []
        for current_date in week:
            week_days.append(
                {
                    "date": current_date,
                    "key": current_date.isoformat(),
                    "dayNumber": current_date.day,
                    "isCurrentMonth": current_date.month == month_start.month,
                    "isToday": current_date == date.today(),
                }
            )
        month_weeks.append(week_days)

    return month_weeks


def normalize_agenda_label(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalize_agenda_region(value: Any) -> str:
    return normalize_agenda_label(value).lower()


def expand_agenda_date_range(start_date: str, end_date: str) -> List[str]:
    if not start_date or not end_date:
        return []

    current_date = date.fromisoformat(start_date)
    final_date = date.fromisoformat(end_date)
    expanded_dates: List[str] = []

    while current_date <= final_date:
        expanded_dates.append(current_date.isoformat())
        current_date += timedelta(days=1)

    return expanded_dates


def fetch_school_holidays_for_schoolyear(school_year: str, region: str) -> Dict[str, Any]:
    normalized_school_year = normalize_agenda_label(school_year)
    normalized_region = normalize_agenda_region(region) or "all"
    cache_key = f"{normalized_school_year}:{normalized_region}"
    now = time.time()

    with agenda_school_holidays_cache_lock:
        cached_payload = agenda_school_holidays_cache.get(cache_key)
        if cached_payload and now - float(cached_payload.get("cached_at") or 0.0) < AGENDA_EXTERNAL_CACHE_TTL_SECONDS:
            return dict(cached_payload["payload"])

    response = requests.get(
        f"{RIJKSOVERHEID_SCHOOL_HOLIDAYS_API_BASE}/schoolyear/{normalized_school_year}",
        params={"output": "json"},
        timeout=12,
    )
    response.raise_for_status()
    payload = response.json()
    records = payload if isinstance(payload, list) else [payload]
    items: List[Dict[str, Any]] = []
    seen_items: Set[Tuple[str, str, str]] = set()

    for record in records:
        if not isinstance(record, dict):
            continue
        for content_item in record.get("content", []):
            if not isinstance(content_item, dict):
                continue
            parsed_school_year = normalize_agenda_label(content_item.get("schoolyear"))
            for vacation in content_item.get("vacations", []):
                if not isinstance(vacation, dict):
                    continue
                vacation_type = normalize_agenda_label(vacation.get("type"))
                for region_item in vacation.get("regions", []):
                    if not isinstance(region_item, dict):
                        continue
                    region_name = normalize_agenda_region(region_item.get("region"))
                    if normalized_region != "all" and region_name not in {normalized_region, "heel nederland"}:
                        continue
                    start_date = normalize_agenda_label(region_item.get("startdate"))[:10]
                    end_date = normalize_agenda_label(region_item.get("enddate"))[:10]
                    for date_key in expand_agenda_date_range(start_date, end_date):
                        dedupe_key = (date_key, vacation_type, region_name)
                        if not date_key or not vacation_type or dedupe_key in seen_items:
                            continue
                        seen_items.add(dedupe_key)
                        items.append(
                            {
                                "date": date_key,
                                "label": vacation_type,
                                "schoolyear": parsed_school_year,
                                "region": region_name,
                            }
                        )

    result = {
        "items": items,
        "schoolYear": normalized_school_year,
        "region": normalized_region,
        "cachedAt": now,
    }
    with agenda_school_holidays_cache_lock:
        agenda_school_holidays_cache[cache_key] = {
            "payload": result,
            "cached_at": now,
        }
    return dict(result)


def fetch_public_holidays_for_year(year: int) -> Dict[str, Any]:
    normalized_year = int(year)
    cache_key = str(normalized_year)
    now = time.time()

    with agenda_public_holidays_cache_lock:
        cached_payload = agenda_public_holidays_cache.get(cache_key)
        if cached_payload and now - float(cached_payload.get("cached_at") or 0.0) < AGENDA_EXTERNAL_CACHE_TTL_SECONDS:
            return dict(cached_payload["payload"])

    response = requests.get(
        f"{NAGER_PUBLIC_HOLIDAYS_API_BASE}/{normalized_year}/NL",
        timeout=12,
    )
    response.raise_for_status()
    payload = response.json()
    items: List[Dict[str, Any]] = []
    seen_items: Set[Tuple[str, str]] = set()

    for item in payload if isinstance(payload, list) else []:
        if not isinstance(item, dict):
            continue
        date_key = normalize_agenda_label(item.get("date"))
        label = normalize_agenda_label(item.get("localName")) or normalize_agenda_label(item.get("name"))
        dedupe_key = (date_key, label)
        if not date_key or not label or dedupe_key in seen_items:
            continue
        seen_items.add(dedupe_key)
        items.append(
            {
                "date": date_key,
                "label": label,
                "localName": normalize_agenda_label(item.get("localName")),
                "name": normalize_agenda_label(item.get("name")),
            }
        )

    result = {
        "items": items,
        "year": normalized_year,
        "cachedAt": now,
    }
    with agenda_public_holidays_cache_lock:
        agenda_public_holidays_cache[cache_key] = {
            "payload": result,
            "cached_at": now,
        }
    return dict(result)


def format_agenda_school_holiday_label(label: Any, region: Any) -> str:
    normalized_label = normalize_agenda_label(label)
    normalized_region = normalize_agenda_region(region)
    if not normalized_label:
        return ""
    if not normalized_region:
        return normalized_label
    if normalized_region == "heel nederland":
        return f"{normalized_label} (heel Nederland)"
    return f"{normalized_label} ({normalized_region})"


def get_agenda_school_holiday_region_order(region: Any) -> int:
    normalized_region = normalize_agenda_region(region)
    if normalized_region == "noord":
        return 1
    if normalized_region == "midden":
        return 2
    if normalized_region == "zuid":
        return 3
    return 99


def build_agenda_school_holiday_labels(items: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    grouped_items: Dict[Tuple[str, str], Dict[str, Any]] = {}

    for item in items:
        date_key = normalize_agenda_label(item.get("date"))[:10]
        base_label = normalize_agenda_label(item.get("label"))
        region_name = normalize_agenda_region(item.get("region"))
        if not date_key or not base_label or not region_name:
            continue

        group_key = (date_key, base_label)
        group = grouped_items.setdefault(
            group_key,
            {
                "date": date_key,
                "base_label": base_label,
                "regions": set(),
            },
        )
        group["regions"].add(region_name)

    labels: List[Dict[str, str]] = []
    for group in grouped_items.values():
        region_names = list(group["regions"])
        has_nationwide = "heel nederland" in region_names or all(
            region_name in group["regions"] for region_name in ("noord", "midden", "zuid")
        )
        if has_nationwide:
            formatted_label = format_agenda_school_holiday_label(group["base_label"], "heel nederland")
        else:
            sorted_regions = ", ".join(
                sorted(region_names, key=get_agenda_school_holiday_region_order)
            )
            formatted_label = f"{group['base_label']} ({sorted_regions})"

        labels.append(
            {
                "date": group["date"],
                "label": formatted_label,
            }
        )

    labels.sort(key=lambda item: (item["date"], item["label"]))
    return labels


def build_agenda_external_labels(day_keys: List[str], school_region: str = "all") -> Dict[str, List[str]]:
    labels_by_day: Dict[str, List[str]] = {day_key: [] for day_key in day_keys}
    seen_labels: Dict[str, Set[str]] = {day_key: set() for day_key in day_keys}
    valid_day_keys = set(day_keys)
    years = sorted(
        {
            int(day_key[:4])
            for day_key in day_keys
            if len(day_key) >= 4 and day_key[:4].isdigit()
        }
    )
    school_years = sorted(
        {
            f"{year - 1}-{year}"
            for year in years
        }
        | {
            f"{year}-{year + 1}"
            for year in years
        }
    )

    school_holiday_items: List[Dict[str, Any]] = []
    for school_year in school_years:
        payload = fetch_school_holidays_for_schoolyear(school_year, school_region)
        school_holiday_items.extend(payload.get("items", []))

    for holiday in build_agenda_school_holiday_labels(school_holiday_items):
        date_key = normalize_agenda_label(holiday.get("date"))[:10]
        label = normalize_agenda_label(holiday.get("label"))
        if date_key not in valid_day_keys or not label or label in seen_labels[date_key]:
            continue
        seen_labels[date_key].add(label)
        labels_by_day[date_key].append(label)

    for year in years:
        payload = fetch_public_holidays_for_year(year)
        for holiday in payload.get("items", []):
            date_key = normalize_agenda_label(holiday.get("date"))[:10]
            label = normalize_agenda_label(holiday.get("label"))
            if date_key not in valid_day_keys or not label or label in seen_labels[date_key]:
                continue
            seen_labels[date_key].add(label)
            labels_by_day[date_key].append(label)

    return labels_by_day


def build_agenda_week_events(trainings: List[Dict[str, Any]], week_start: date) -> List[Dict[str, Any]]:
    calendar_start_minutes = 0
    pixels_per_hour = 56
    week_end = week_start + timedelta(days=6)
    events = []

    for training in trainings:
        if not training.get("date") or not training.get("time"):
            continue

        training_date = date.fromisoformat(training["date"])
        if training_date < week_start or training_date > week_end:
            continue

        start_time = training["time"]
        end_time = training.get("endTime") or compute_default_end_time(start_time)
        start_dt = combine_date_and_time(training["date"], start_time)
        end_dt = combine_date_and_time(training["date"], end_time)
        if end_dt <= start_dt:
            end_dt = start_dt + timedelta(minutes=90)

        start_minutes = start_dt.hour * 60 + start_dt.minute
        end_minutes = end_dt.hour * 60 + end_dt.minute
        top = max(((start_minutes - calendar_start_minutes) / 60) * pixels_per_hour, 0)
        height = max(((end_minutes - start_minutes) / 60) * pixels_per_hour, 48)
        column = (training_date - week_start).days + 2

        events.append(
            {
                "id": training["id"],
                "title": training["title"],
                "date": training["date"],
                "time": start_time,
                "endTime": end_time,
                "location": training.get("location", ""),
                "clubClass": get_agenda_club_class(training.get("location", "")),
                "trainingType": training.get("trainingType", ""),
                "trainingTypeLabel": training.get("trainingTypeLabel", ""),
                "trainingTypeClass": training.get("trainingTypeClass", ""),
                "status": training.get("status", ""),
                "statusLabel": training.get("statusLabel", ""),
                "statusClass": training.get("statusClass", ""),
                "trainers": training.get("trainers", []),
                "trainerNames": training.get("trainerNames", ""),
                "notes": training.get("notes", ""),
                "signature": build_agenda_training_signature(training),
                "dayIndex": (training_date - week_start).days,
                "top": round(top, 1),
                "height": round(height, 1),
                "startMinutes": start_minutes,
                "endMinutes": end_minutes,
                "overlapIndex": 0,
                "overlapCount": 1,
                "overlapFontScale": 1,
                "textScale": calculate_agenda_week_event_text_scale(
                    training["title"],
                    training.get("trainingTypeLabel", ""),
                    training.get("location", ""),
                    training.get("trainerNames", ""),
                    height,
                ),
                "fontScale": 1,
                "isDenseText": False,
            }
        )

    assign_agenda_week_event_overlaps(events)
    return events


def assign_agenda_week_event_overlaps(events: List[Dict[str, Any]]) -> None:
    events_by_day: Dict[int, List[Dict[str, Any]]] = {}
    for event in events:
        events_by_day.setdefault(int(event.get("dayIndex", 0)), []).append(event)

    for day_events in events_by_day.values():
        sorted_events = sorted(
            day_events,
            key=lambda item: (
                int(item.get("startMinutes", 0)),
                int(item.get("endMinutes", 0)),
                normalize_agenda_label(item.get("title", "")),
            ),
        )
        component: List[Dict[str, Any]] = []
        component_end = -1

        for event in sorted_events:
            start_minutes = int(event.get("startMinutes", 0))
            end_minutes = int(event.get("endMinutes", start_minutes))
            if component and start_minutes >= component_end:
                assign_agenda_week_event_overlap_component(component)
                component = []
                component_end = -1

            component.append(event)
            component_end = max(component_end, end_minutes)

        if component:
            assign_agenda_week_event_overlap_component(component)


def assign_agenda_week_event_overlap_component(events: List[Dict[str, Any]]) -> None:
    active_columns: List[Tuple[int, int]] = []
    max_column = 0

    for event in events:
        start_minutes = int(event.get("startMinutes", 0))
        end_minutes = int(event.get("endMinutes", start_minutes))
        active_columns = [
            (active_end, active_column)
            for active_end, active_column in active_columns
            if active_end > start_minutes
        ]
        used_columns = {active_column for _, active_column in active_columns}
        column = 0
        while column in used_columns:
            column += 1

        event["overlapIndex"] = column
        active_columns.append((end_minutes, column))
        max_column = max(max_column, column + 1)

    overlap_count = max(max_column, 1)
    font_scale = max(0.58, round(1 - ((overlap_count - 1) * 0.14), 2))
    for event in events:
        text_scale = float(event.get("textScale", 1) or 1)
        combined_font_scale = max(0.48, round(font_scale * text_scale, 2))
        event["overlapCount"] = overlap_count
        event["overlapFontScale"] = font_scale
        event["fontScale"] = combined_font_scale
        event["isDenseText"] = combined_font_scale <= 0.72


def calculate_agenda_week_event_text_scale(
    title: Any,
    training_type_label: Any,
    location: Any,
    trainer_names: Any,
    height: float,
) -> float:
    title_length = len(normalize_agenda_label(title))

    if height <= 58 and title_length > 42:
        return 0.72
    if height <= 72 and title_length > 58:
        return 0.78
    if height <= 92 and title_length > 76:
        return 0.82
    if title_length > 96:
        return 0.86
    if title_length > 76:
        return 0.9
    return 1


def build_agenda_month_events(trainings: List[Dict[str, Any]], visible_day_keys: Set[str]) -> Dict[str, List[Dict[str, Any]]]:
    events_by_day: Dict[str, List[Dict[str, Any]]] = {day_key: [] for day_key in visible_day_keys}

    for training in trainings:
        training_date = normalize_agenda_label(training.get("date"))[:10]
        start_time = normalize_agenda_label(training.get("time"))
        if training_date not in visible_day_keys or not start_time:
            continue

        events_by_day.setdefault(training_date, []).append(
            {
                "id": training.get("id"),
                "title": normalize_agenda_label(training.get("title")),
                "date": training_date,
                "time": start_time,
                "endTime": normalize_agenda_label(training.get("endTime")) or compute_default_end_time(start_time),
                "location": normalize_agenda_label(training.get("location")),
                "clubClass": get_agenda_club_class(training.get("location", "")),
                "trainingType": normalize_agenda_label(training.get("trainingType")),
                "trainingTypeLabel": normalize_agenda_label(training.get("trainingTypeLabel")),
                "trainingTypeClass": normalize_agenda_label(training.get("trainingTypeClass")),
                "status": normalize_agenda_label(training.get("status")),
                "statusLabel": normalize_agenda_label(training.get("statusLabel")),
                "statusClass": normalize_agenda_label(training.get("statusClass")),
                "trainers": training.get("trainers", []),
                "trainerNames": normalize_agenda_label(training.get("trainerNames")),
                "notes": normalize_agenda_label(training.get("notes")),
                "signature": build_agenda_training_signature(training),
            }
        )

    for day_events in events_by_day.values():
        day_events.sort(key=lambda item: (item.get("time", ""), item.get("title", "")))

    return events_by_day


def build_product_summary(orders: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    configured_events = load_dashboard_events_config()
    summary = []

    for configured_event in configured_events:
        configured_product_id = configured_event.get("productId")
        sold_count = 0

        for order in orders:
            for item in order.get("items", []):
                if matches_configured_event(
                    item.get("name", ""),
                    configured_event,
                    item.get("productId"),
                ):
                    sold_count += int(item.get("quantity", 0) or 0)

        summary.append(
            {
                "productId": configured_product_id,
                "label": configured_event.get("label", "Onbekend event"),
                "soldCount": sold_count,
            }
        )

    return sorted(summary, key=lambda item: item["soldCount"], reverse=True)


def search_catalog_products(keyword: str) -> List[Dict[str, Any]]:
    if not keyword.strip():
        return []

    query = normalize_match_text(keyword)
    query_tokens = {token for token in query.split() if token}
    products = fetch_catalog_products().get("items", [])

    filtered_products = []
    for item in products:
        normalized_search_text = normalize_match_text(
            " ".join(
                str(value or "")
                for value in (item.get("name", ""), item.get("sku", ""), item.get("id", ""))
            )
        )
        item_tokens = {token for token in normalized_search_text.split() if token}
        if query_tokens and not all(
            any(item_token.startswith(query_token) or query_token in item_token for item_token in item_tokens)
            for query_token in query_tokens
        ):
            continue

        filtered_products.append(
            {
                "id": item.get("id"),
                "name": item.get("name", "Naamloos product"),
                "sku": item.get("sku", ""),
                "price": item.get("price", 0),
                "enabled": item.get("enabled", True),
            }
        )

    return filtered_products[:20]


def fetch_orders_from_ecwid(run_auto_email: bool = True) -> Dict[str, Any]:
    config = get_config()
    if not config["store_id"] or not config["secret_token"]:
        return {
            "source": "mock",
            "items": mock_orders(),
            "summary": build_summary(mock_orders()),
            "message": (
                "Live Ecwid-koppeling staat nog niet aan. "
                "Voeg ECWID_STORE_ID en ECWID_SECRET_TOKEN toe."
            ),
        }

    all_orders: List[Dict[str, Any]] = []
    offset = 0
    limit = 100
    total = 0

    try:
        while True:
            response = requests.get(
                f"{ECWID_API_BASE}/{config['store_id']}/orders",
                headers={"Authorization": f"Bearer {config['secret_token']}"},
                params={
                    "limit": limit,
                    "offset": offset,
                    "responseFields": ECWID_RESPONSE_FIELDS,
                },
                timeout=20,
            )
            response.raise_for_status()
            payload = response.json()

            batch = payload.get("items", [])
            total = payload.get("total", total)
            all_orders.extend(batch)

            if not batch or len(batch) < limit or len(all_orders) >= total:
                break

            offset += limit
    except requests.RequestException:
        return {
            "source": "mock",
            "items": mock_orders(),
            "summary": build_summary(mock_orders()),
            "message": (
                "Ecwid kon nu niet worden geladen. "
                "Controleer ECWID_STORE_ID en ECWID_SECRET_TOKEN; tijdelijke voorbeelddata wordt getoond."
            ),
        }

    automatic_return_sync = sync_refunded_orders_to_returned(all_orders)
    if automatic_return_sync["syncedOrderIds"]:
        app.logger.info(
            "%s terugbetaalde Ecwid-bestelling(en) automatisch op geretourneerd gezet.",
            len(automatic_return_sync["syncedOrderIds"]),
        )
    if automatic_return_sync["failedOrderIds"]:
        app.logger.warning(
            "%s terugbetaalde Ecwid-bestelling(en) konden niet automatisch op geretourneerd worden gezet.",
            len(automatic_return_sync["failedOrderIds"]),
        )

    normalized_orders = [normalize_order(order) for order in all_orders]
    if run_auto_email:
        auto_email_result = auto_email_new_registration_orders(normalized_orders)
        if auto_email_result["sentOrderIds"]:
            app.logger.info(
                "%s automatische inschrijvingsmail(s) verzonden.",
                len(auto_email_result["sentOrderIds"]),
            )

    return {
        "source": "ecwid",
        "items": normalized_orders,
        "summary": build_summary(normalized_orders),
        "total": total,
        "count": len(all_orders),
        "automaticReturnSync": automatic_return_sync,
        "message": (
            f"{len(automatic_return_sync['failedOrderIds'])} terugbetaalde bestelling(en) konden niet "
            "automatisch op geretourneerd worden gezet. Probeer de Ecwid-verversing opnieuw."
            if automatic_return_sync["failedOrderIds"]
            else None
        ),
    }


def get_ecwid_headers(secret_token: str) -> Dict[str, str]:
    return {
        "Authorization": f"Bearer {secret_token}",
        "Content-Type": "application/json",
    }


def invalidate_ecwid_orders_cache() -> None:
    with ecwid_orders_cache_lock:
        ecwid_orders_cache["payload"] = None
        ecwid_orders_cache["cached_at"] = 0.0
        ecwid_orders_cache["config_fingerprint"] = None


def update_ecwid_order_fulfillment_status(order_id: str, fulfillment_status: str) -> bool:
    normalized_order_id = str(order_id or "").strip()
    if not normalized_order_id:
        raise ValueError("Bestelling ontbreekt.")
    normalized_fulfillment_status = str(fulfillment_status or "").strip().upper()
    if not normalized_fulfillment_status:
        raise ValueError("Afhandelstatus ontbreekt.")

    config = get_config()
    if not config["store_id"] or not config["secret_token"]:
        return False

    try:
        response = requests.put(
            f"{ECWID_API_BASE}/{config['store_id']}/orders/{normalized_order_id}",
            headers=get_ecwid_headers(config["secret_token"]),
            json={"fulfillmentStatus": normalized_fulfillment_status},
            timeout=20,
        )
        response.raise_for_status()
    except requests.RequestException as exc:
        raise RuntimeError("Ecwid-bestelling kon niet worden bijgewerkt.") from exc

    response_content = getattr(response, "content", b"")
    payload = response.json() if response_content else {}
    if isinstance(payload, dict) and int(payload.get("updateCount") or 0) not in {0, 1}:
        raise RuntimeError("Ecwid gaf een ongeldige reactie terug bij het bijwerken van de bestelling.")
    if isinstance(payload, dict) and "updateCount" in payload and int(payload.get("updateCount") or 0) != 1:
        raise RuntimeError("Ecwid heeft de bestelling niet bijgewerkt.")

    invalidate_ecwid_orders_cache()
    return True


def update_ecwid_order_to_processing(order_id: str) -> bool:
    try:
        return update_ecwid_order_fulfillment_status(order_id, ECWID_PROCESSING_FULFILLMENT_STATUS)
    except RuntimeError as exc:
        raise RuntimeError("Ecwid-bestelling kon niet op in verwerking worden gezet.") from exc


def update_ecwid_order_to_delivered(order_id: str) -> bool:
    try:
        return update_ecwid_order_fulfillment_status(order_id, ECWID_DELIVERED_FULFILLMENT_STATUS)
    except RuntimeError as exc:
        raise RuntimeError("Ecwid-bestelling kon niet op geleverd worden gezet.") from exc


def update_ecwid_order_to_returned(order_id: str) -> bool:
    try:
        return update_ecwid_order_fulfillment_status(order_id, ECWID_RETURNED_FULFILLMENT_STATUS)
    except RuntimeError as exc:
        raise RuntimeError("Ecwid-bestelling kon niet op geretourneerd worden gezet.") from exc


def sync_refunded_orders_to_returned(orders: List[Dict[str, Any]]) -> Dict[str, Any]:
    enabled = get_env_bool("ECWID_AUTO_RETURN_REFUNDED_ORDERS", True)
    matched_order_ids: List[str] = []
    synced_order_ids: List[str] = []
    failed_order_ids: List[str] = []
    orders_by_id: Dict[str, Dict[str, Any]] = {}

    for order in orders:
        order_id = str(order.get("id") or "").strip()
        payment_status = str(order.get("paymentStatus") or "").strip().upper()
        fulfillment_status = str(order.get("fulfillmentStatus") or "").strip().upper()
        if (
            not order_id
            or order_id in orders_by_id
            or payment_status != "REFUNDED"
            or fulfillment_status == ECWID_RETURNED_FULFILLMENT_STATUS
        ):
            continue
        orders_by_id[order_id] = order
        matched_order_ids.append(order_id)

    if enabled:
        for order_id in matched_order_ids:
            try:
                updated = update_ecwid_order_to_returned(order_id)
            except RuntimeError as exc:
                app.logger.warning(
                    "Automatisch retourneren van Ecwid-bestelling %s mislukt: %s",
                    order_id,
                    exc,
                )
                failed_order_ids.append(order_id)
                continue

            if not updated:
                failed_order_ids.append(order_id)
                continue

            orders_by_id[order_id]["fulfillmentStatus"] = ECWID_RETURNED_FULFILLMENT_STATUS
            synced_order_ids.append(order_id)

    return {
        "enabled": enabled,
        "matchedOrderIds": matched_order_ids,
        "syncedOrderIds": synced_order_ids,
        "failedOrderIds": failed_order_ids,
    }


def sync_emailed_registration_orders_to_ecwid(order_ids: Optional[List[str]] = None) -> Dict[str, Any]:
    normalized_order_ids = normalize_registration_email_status_order_ids(
        load_all_registration_emailed_order_ids() if order_ids is None else order_ids
    )
    synced_order_ids: List[str] = []
    failed_order_ids: List[str] = []

    for order_id in normalized_order_ids:
        try:
            updated = update_ecwid_order_to_processing(order_id)
        except RuntimeError:
            failed_order_ids.append(order_id)
            continue

        if updated:
            synced_order_ids.append(order_id)

    return {
        "orderIds": normalized_order_ids,
        "syncedOrderIds": synced_order_ids,
        "failedOrderIds": failed_order_ids,
    }


def sync_registration_event_orders_to_delivered(order_ids: List[str]) -> Dict[str, Any]:
    normalized_order_ids = normalize_registration_email_status_order_ids(order_ids)
    synced_order_ids: List[str] = []
    failed_order_ids: List[str] = []

    for order_id in normalized_order_ids:
        try:
            updated = update_ecwid_order_to_delivered(order_id)
        except RuntimeError:
            failed_order_ids.append(order_id)
            continue

        if updated:
            synced_order_ids.append(order_id)

    return {
        "orderIds": normalized_order_ids,
        "syncedOrderIds": synced_order_ids,
        "failedOrderIds": failed_order_ids,
    }


def sync_registration_event_orders_to_returned(order_ids: List[str]) -> Dict[str, Any]:
    normalized_order_ids = normalize_registration_email_status_order_ids(order_ids)
    synced_order_ids: List[str] = []
    failed_order_ids: List[str] = []

    for order_id in normalized_order_ids:
        try:
            updated = update_ecwid_order_to_returned(order_id)
        except RuntimeError:
            failed_order_ids.append(order_id)
            continue

        if updated:
            synced_order_ids.append(order_id)

    return {
        "orderIds": normalized_order_ids,
        "syncedOrderIds": synced_order_ids,
        "failedOrderIds": failed_order_ids,
    }


def get_moneybird_headers(token: str) -> Dict[str, str]:
    return {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
    }


def fetch_moneybird_administration(config: Dict[str, str]) -> Dict[str, Any]:
    token = config["moneybird_token"]
    administration_id = config["moneybird_administration_id"]
    if not token:
        return {}

    response = requests.get(
        f"{MONEYBIRD_API_BASE}/administrations.json",
        headers=get_moneybird_headers(token),
        timeout=20,
    )
    response.raise_for_status()
    administrations = response.json()
    if not isinstance(administrations, list) or not administrations:
        return {}

    if administration_id:
        for administration in administrations:
            if str(administration.get("id")) == str(administration_id):
                return administration

    return administrations[0]


def fetch_moneybird_ledger_account_types(token: str, administration_id: Any) -> Dict[str, str]:
    response = requests.get(
        f"{MONEYBIRD_API_BASE}/{administration_id}/ledger_accounts.json",
        headers=get_moneybird_headers(token),
        timeout=20,
    )
    response.raise_for_status()
    accounts = response.json()
    if not isinstance(accounts, list):
        return {}

    return {
        str(account.get("id")): str(account.get("account_type", "")).strip()
        for account in accounts
        if account.get("id")
    }


def find_moneybird_contact_by_name(token: str, administration_id: Any, query: str) -> Optional[Dict[str, Any]]:
    response = requests.get(
        f"{MONEYBIRD_API_BASE}/{administration_id}/contacts.json",
        headers=get_moneybird_headers(token),
        params={"query": query, "per_page": 20},
        timeout=20,
    )
    response.raise_for_status()
    contacts = response.json()
    if not isinstance(contacts, list):
        return None

    normalized_query = normalize_agenda_label(query).lower().replace(".", "")
    active_contacts = [contact for contact in contacts if not contact.get("archived")]
    for contact in active_contacts:
        company_name = normalize_agenda_label(contact.get("company_name")).lower().replace(".", "")
        if company_name == normalized_query:
            return contact
    return active_contacts[0] if active_contacts else None


def fetch_moneybird_sales_tax_rate_id(token: str, administration_id: Any, percentage: Decimal = Decimal("9")) -> str:
    response = requests.get(
        f"{MONEYBIRD_API_BASE}/{administration_id}/tax_rates.json",
        headers=get_moneybird_headers(token),
        timeout=20,
    )
    response.raise_for_status()
    tax_rates = response.json()
    if not isinstance(tax_rates, list):
        return ""

    for tax_rate in tax_rates:
        tax_percentage = decimal_from_value(tax_rate.get("percentage"))
        if (
            tax_rate.get("active")
            and str(tax_rate.get("tax_rate_type") or "").strip() == "sales_invoice"
            and tax_percentage == percentage
        ):
            return str(tax_rate.get("id") or "").strip()
    return ""


def fetch_moneybird_sales_no_tax_rate_id(token: str, administration_id: Any) -> str:
    response = requests.get(
        f"{MONEYBIRD_API_BASE}/{administration_id}/tax_rates.json",
        headers=get_moneybird_headers(token),
        timeout=20,
    )
    response.raise_for_status()
    tax_rates = response.json()
    if not isinstance(tax_rates, list):
        return ""

    for tax_rate in tax_rates:
        if (
            tax_rate.get("active")
            and str(tax_rate.get("tax_rate_type") or "").strip() == "sales_invoice"
            and str(tax_rate.get("name") or "").strip().lower() == "geen btw"
        ):
            return str(tax_rate.get("id") or "").strip()
    return ""


def format_invoice_decimal(value: Decimal) -> str:
    return f"{value.quantize(Decimal('0.01'))}"


def format_invoice_display(value: Any) -> str:
    decimal_value = decimal_from_value(value).quantize(Decimal("0.01"))
    return f"€ {str(decimal_value).replace('.', ',')}"


def format_invoice_amount_without_currency(value: Any) -> str:
    decimal_value = decimal_from_value(value).quantize(Decimal("0.01"))
    return str(decimal_value).replace(".", ",")


def get_month_bounds(process_date: date) -> Tuple[date, date]:
    month_start = process_date.replace(day=1)
    month_end = add_months(month_start, 1) - timedelta(days=1)
    return month_start, month_end


def get_previous_month_bounds(process_date: date) -> Tuple[date, date]:
    previous_month_start = add_months(process_date.replace(day=1), -1)
    previous_month_end = process_date.replace(day=1) - timedelta(days=1)
    return previous_month_start, previous_month_end


def build_automatic_invoice_season_label(setting: Dict[str, Any], process_date: date) -> str:
    period_start = parse_iso_date(setting.get("periodStart", ""))
    period_end = parse_iso_date(setting.get("periodEnd", ""))
    if period_start and period_end:
        return f"{period_start.year}/{period_end.year}"
    month_start = process_date.replace(day=1)
    season_start_year = month_start.year if month_start.month >= 7 else month_start.year - 1
    return f"{season_start_year}/{season_start_year + 1}"


def get_automatic_invoice_sequence_number(setting: Dict[str, Any], process_date: date) -> int:
    period_start = parse_iso_date(setting.get("periodStart", ""))
    invoice_month = process_date.replace(day=1)
    if not period_start:
        return 1
    period_month = period_start.replace(day=1)
    if invoice_month < period_month:
        return 1
    return ((invoice_month.year - period_month.year) * 12) + invoice_month.month - period_month.month + 1


def load_automatic_invoice_settings(active_only: bool = False) -> List[Dict[str, Any]]:
    query = """
        SELECT id, club_name, standard_amount, training_amount, invoice_day, repeat_enabled,
               period_start, period_end, active, created_at, updated_at
        FROM automatic_invoice_settings
    """
    if active_only:
        query += "\n        WHERE active = 1"
    query += "\n        ORDER BY club_name ASC, id ASC"

    with get_db_connection() as connection:
        rows = connection.execute(query).fetchall()

    settings = []
    for row in rows:
        standard_amount = decimal_from_value(row["standard_amount"])
        training_amount = decimal_from_value(row["training_amount"])
        settings.append(
            {
                "id": int(row["id"]),
                "clubName": str(row["club_name"] or "").strip(),
                "standardAmount": str(row["standard_amount"] or "").strip(),
                "standardAmountLabel": format_invoice_display(standard_amount),
                "trainingAmount": str(row["training_amount"] or "").strip(),
                "trainingAmountLabel": format_invoice_display(training_amount),
                "invoiceDay": int(row["invoice_day"] or 1),
                "repeatEnabled": bool(row["repeat_enabled"]),
                "periodStart": str(row["period_start"] or "").strip(),
                "periodEnd": str(row["period_end"] or "").strip(),
                "active": bool(row["active"]),
                "createdAt": str(row["created_at"] or "").strip(),
                "updatedAt": str(row["updated_at"] or "").strip(),
                "updatedAtLabel": format_datetime_display(str(row["updated_at"] or "").strip()),
            }
        )
    return settings


def load_automatic_invoice_setting(setting_id: int) -> Optional[Dict[str, Any]]:
    for setting in load_automatic_invoice_settings():
        if int(setting["id"]) == int(setting_id):
            return setting
    return None


def save_automatic_invoice_setting(form: Any) -> Tuple[bool, str]:
    club_name = normalize_agenda_club(form.get("club_name", ""))
    standard_amount = normalize_price_input(form.get("standard_amount", ""))
    training_amount = normalize_price_input(form.get("training_amount", ""))
    repeat_enabled = form.get("repeat_enabled") == "1"
    period_start = str(form.get("period_start", "") or "").strip()
    period_end = str(form.get("period_end", "") or "").strip()
    active = form.get("active", "1") == "1"

    try:
        invoice_day = int(str(form.get("invoice_day", "") or "").strip())
    except ValueError:
        invoice_day = 0

    if not club_name:
        return False, "Selecteer een samenwerkende amateurclub."
    if decimal_from_value(standard_amount) <= 0:
        return False, "Vul een standaard factuurbedrag groter dan 0 in."
    if decimal_from_value(training_amount) < 0:
        return False, "Vul een bedrag per training van 0 of hoger in."
    if invoice_day < 1 or invoice_day > 31:
        return False, "Vul een factuurdag tussen 1 en 31 in."
    if repeat_enabled:
        if parse_iso_date(period_start) is None or parse_iso_date(period_end) is None:
            return False, "Vul bij herhalen een geldige begin- en einddatum in."
        if parse_iso_date(period_end) < parse_iso_date(period_start):
            return False, "De einddatum moet na de begindatum liggen."
    else:
        period_start = ""
        period_end = ""

    now_value = datetime.now().isoformat(timespec="seconds")
    setting_id = form.get("setting_id", type=int)
    with get_db_connection() as connection:
        if setting_id:
            connection.execute(
                """
                UPDATE automatic_invoice_settings
                SET club_name = ?, standard_amount = ?, training_amount = ?, invoice_day = ?,
                    repeat_enabled = ?, period_start = ?, period_end = ?, active = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    club_name,
                    standard_amount,
                    training_amount,
                    invoice_day,
                    1 if repeat_enabled else 0,
                    period_start,
                    period_end,
                    1 if active else 0,
                    now_value,
                    setting_id,
                ),
            )
        else:
            connection.execute(
                """
                INSERT INTO automatic_invoice_settings (
                    club_name, standard_amount, training_amount, invoice_day, repeat_enabled,
                    period_start, period_end, active, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    club_name,
                    standard_amount,
                    training_amount,
                    invoice_day,
                    1 if repeat_enabled else 0,
                    period_start,
                    period_end,
                    1 if active else 0,
                    now_value,
                    now_value,
                ),
            )
    clear_local_data_cache()
    return True, "Automatische factuurinstelling opgeslagen."


def delete_automatic_invoice_setting(setting_id: int) -> None:
    with get_db_connection() as connection:
        connection.execute("DELETE FROM automatic_invoice_runs WHERE setting_id = ?", (setting_id,))
        connection.execute("DELETE FROM automatic_invoice_settings WHERE id = ?", (setting_id,))
    clear_local_data_cache()


def setting_is_due_for_invoice(setting: Dict[str, Any], process_date: date) -> bool:
    if not setting.get("active"):
        return False
    last_day = calendar.monthrange(process_date.year, process_date.month)[1]
    due_day = min(int(setting.get("invoiceDay") or 1), last_day)
    if process_date.day != due_day:
        return False
    if setting.get("repeatEnabled"):
        period_start = parse_iso_date(setting.get("periodStart", ""))
        period_end = parse_iso_date(setting.get("periodEnd", ""))
        return bool(period_start and period_end and period_start <= process_date <= period_end)
    return not automatic_invoice_run_exists(int(setting["id"]))


def automatic_invoice_run_exists(setting_id: int, invoice_month: Optional[str] = None) -> bool:
    with get_db_connection() as connection:
        if invoice_month:
            row = connection.execute(
                """
                SELECT id
                FROM automatic_invoice_runs
                WHERE setting_id = ? AND invoice_month = ? AND status = 'created'
                """,
                (setting_id, invoice_month),
            ).fetchone()
        else:
            row = connection.execute(
                """
                SELECT id
                FROM automatic_invoice_runs
                WHERE setting_id = ? AND status = 'created'
                """,
                (setting_id,),
            ).fetchone()
    return row is not None


def build_automatic_invoice_lines(setting: Dict[str, Any], process_date: date) -> Dict[str, Any]:
    month_start, month_end = get_month_bounds(process_date)
    previous_month_start, previous_month_end = get_previous_month_bounds(process_date)
    trainings = load_agenda_trainings(previous_month_start.isoformat(), previous_month_end.isoformat())
    cancelled_trainings = [
        training
        for training in trainings
        if training.get("location") == setting.get("clubName")
        and training.get("trainingType") == "samenwerkende_amateurclub"
        and training.get("status") == "geannuleerd"
    ]
    standard_amount = decimal_from_value(setting.get("standardAmount"))
    training_amount = decimal_from_value(setting.get("trainingAmount"))
    cancelled_total = training_amount * Decimal(len(cancelled_trainings))
    sequence_number = get_automatic_invoice_sequence_number(setting, process_date)
    season_label = build_automatic_invoice_season_label(setting, process_date)
    invoice_lines = [
        {
            "description": f"Factuurbedrag {sequence_number} seizoen {season_label}",
            "amount": "1",
            "price": format_invoice_decimal(standard_amount),
            "taxRateKind": "sales_9",
        }
    ]
    if cancelled_trainings:
        invoice_lines.append(
            {
                "description": (
                    f"Niet gegeven trainingen {build_full_month_label(previous_month_start)} "
                    f"({len(cancelled_trainings)} x {format_invoice_amount_without_currency(training_amount)})"
                ),
                "amount": "1",
                "price": format_invoice_decimal(-cancelled_total),
                "taxRateKind": "no_tax",
            }
        )
    total_amount = standard_amount - cancelled_total
    return {
        "invoiceMonth": month_start.strftime("%Y-%m"),
        "monthLabel": build_month_label(month_start),
        "deductionMonthLabel": build_full_month_label(previous_month_start),
        "sequenceNumber": sequence_number,
        "seasonLabel": season_label,
        "cancelledTrainings": cancelled_trainings,
        "cancelledCount": len(cancelled_trainings),
        "invoiceLines": invoice_lines,
        "totalAmount": total_amount,
        "totalAmountLabel": format_invoice_display(total_amount),
    }


def create_moneybird_concept_invoice_for_setting(setting: Dict[str, Any], process_date: date) -> Dict[str, Any]:
    config = get_config()
    token = config["moneybird_token"]
    if not token:
        raise ValueError("MoneyBird API-token ontbreekt.")

    administration = fetch_moneybird_administration(config)
    administration_id = administration.get("id")
    if not administration_id:
        raise ValueError("Geen MoneyBird-administratie gevonden.")

    invoice_payload = build_automatic_invoice_lines(setting, process_date)
    invoice_month = invoice_payload["invoiceMonth"]
    if automatic_invoice_run_exists(int(setting["id"]), invoice_month):
        return {"skipped": True, "message": "Voor deze maand bestaat al een conceptfactuur."}

    contact = find_moneybird_contact_by_name(token, administration_id, setting["clubName"])
    if not contact or not contact.get("id"):
        raise ValueError(f"Geen MoneyBird-contact gevonden voor {setting['clubName']}.")
    tax_rate_id = fetch_moneybird_sales_tax_rate_id(token, administration_id, Decimal("9"))
    if not tax_rate_id:
        raise ValueError("Geen actieve MoneyBird-btwcode van 9% gevonden voor verkoopfacturen.")
    no_tax_rate_id = fetch_moneybird_sales_no_tax_rate_id(token, administration_id)
    if not no_tax_rate_id:
        raise ValueError("Geen actieve MoneyBird-btwcode 'Geen btw' gevonden voor verkoopfacturen.")

    invoice_lines = []
    for line in invoice_payload["invoiceLines"]:
        line_payload = {key: value for key, value in line.items() if key != "taxRateKind"}
        line_payload["tax_rate_id"] = no_tax_rate_id if line.get("taxRateKind") == "no_tax" else tax_rate_id
        invoice_lines.append(line_payload)

    response = requests.post(
        f"{MONEYBIRD_API_BASE}/{administration_id}/sales_invoices.json",
        headers={**get_moneybird_headers(token), "Content-Type": "application/json"},
        json={
            "sales_invoice": {
                "contact_id": contact["id"],
                "invoice_date": process_date.isoformat(),
                "prices_are_incl_tax": False,
                "details_attributes": invoice_lines,
            }
        },
        timeout=20,
    )
    response.raise_for_status()
    invoice = response.json()
    now_value = datetime.now().isoformat(timespec="seconds")
    with get_db_connection() as connection:
        connection.execute(
            """
            INSERT OR REPLACE INTO automatic_invoice_runs (
                setting_id, invoice_month, moneybird_invoice_id, moneybird_draft_id,
                status, error_message, created_at
            )
            VALUES (?, ?, ?, ?, 'created', '', ?)
            """,
            (
                int(setting["id"]),
                invoice_month,
                str(invoice.get("id") or ""),
                str(invoice.get("draft_id") or ""),
                now_value,
            ),
        )
    clear_local_data_cache()
    return {
        "skipped": False,
        "invoice": invoice,
        **invoice_payload,
    }


def process_automatic_invoices(process_date: Optional[date] = None) -> Dict[str, Any]:
    target_date = process_date or date.today()
    results: List[Dict[str, Any]] = []
    for setting in load_automatic_invoice_settings(active_only=True):
        if not setting_is_due_for_invoice(setting, target_date):
            continue
        try:
            result = create_moneybird_concept_invoice_for_setting(setting, target_date)
            results.append({"setting": setting, "result": result, "error": ""})
        except (requests.RequestException, ValueError) as exc:
            month_start, _ = get_month_bounds(target_date)
            now_value = datetime.now().isoformat(timespec="seconds")
            with get_db_connection() as connection:
                connection.execute(
                    """
                    INSERT OR REPLACE INTO automatic_invoice_runs (
                        setting_id, invoice_month, moneybird_invoice_id, moneybird_draft_id,
                        status, error_message, created_at
                    )
                    VALUES (?, ?, '', '', 'failed', ?, ?)
                    """,
                    (int(setting["id"]), month_start.strftime("%Y-%m"), str(exc), now_value),
                )
            results.append({"setting": setting, "result": None, "error": str(exc)})
    return {"date": target_date.isoformat(), "processed": results}


def build_automatic_invoice_page_settings() -> List[Dict[str, Any]]:
    today = date.today()
    settings = load_automatic_invoice_settings()
    for setting in settings:
        preview = build_automatic_invoice_lines(setting, today)
        setting["preview"] = preview
        setting["statusLabel"] = "Actief" if setting["active"] else "Inactief"
        setting["repeatLabel"] = (
            f"Herhaalt van {setting['periodStart']} t/m {setting['periodEnd']}"
            if setting["repeatEnabled"]
            else "Eenmalig"
        )
    return settings


def fetch_moneybird_summary() -> Dict[str, Any]:
    config = get_config()
    token = config["moneybird_token"]
    if not token:
        return {
            "source": "missing",
            "invoiceCount": 0,
            "revenue_received": 0.0,
            "message": "Moneybird-koppeling staat nog niet aan. Voeg MONEYBIRD_API_TOKEN toe.",
        }

    try:
        administration = fetch_moneybird_administration(config)
        administration_id = administration.get("id")
        if not administration_id:
            return {
                "source": "missing",
                "invoiceCount": 0,
                "revenue_received": 0.0,
                "expenses_total": 0.0,
                "financialMutations": [],
                "message": "Geen Moneybird-administratie gevonden voor de huidige API-token.",
            }

        ledger_account_types = fetch_moneybird_ledger_account_types(token, administration_id)

        sync_url = f"{MONEYBIRD_API_BASE}/{administration_id}/sales_invoices/synchronization.json"
        response = requests.get(
            sync_url,
            headers=get_moneybird_headers(token),
            timeout=20,
        )
        response.raise_for_status()
        sync_items = response.json()
        if not isinstance(sync_items, list):
            sync_items = []

        invoice_ids = [item.get("id") for item in sync_items if item.get("id")]
        invoices: List[Dict[str, Any]] = []
        batch_size = 100

        for start_index in range(0, len(invoice_ids), batch_size):
            batch_ids = invoice_ids[start_index : start_index + batch_size]
            detail_response = requests.post(
                sync_url,
                headers={
                    **get_moneybird_headers(token),
                    "Content-Type": "application/json",
                },
                json={"ids": batch_ids},
                timeout=20,
            )
            detail_response.raise_for_status()
            batch = detail_response.json()
            if isinstance(batch, list):
                invoices.extend(batch)

        invoice_years = []
        for invoice in invoices:
            invoice_date = parse_iso_date(str(invoice.get("invoice_date", "")).strip())
            if invoice_date is not None:
                invoice_years.append(invoice_date.year)

        start_year = min(invoice_years, default=date.today().year)
        end_year = max(invoice_years, default=date.today().year) + 1

        mutations_sync_url = f"{MONEYBIRD_API_BASE}/{administration_id}/financial_mutations/synchronization.json"
        mutation_response = requests.get(
            mutations_sync_url,
            headers=get_moneybird_headers(token),
            params={"filter": f"period:{start_year}01..{end_year}12,state:all"},
            timeout=20,
        )
        mutation_response.raise_for_status()
        mutation_sync_items = mutation_response.json()
        if not isinstance(mutation_sync_items, list):
            mutation_sync_items = []

        mutation_ids = [item.get("id") for item in mutation_sync_items if item.get("id")]
        financial_mutations: List[Dict[str, Any]] = []

        for start_index in range(0, len(mutation_ids), batch_size):
            batch_ids = mutation_ids[start_index : start_index + batch_size]
            detail_response = requests.post(
                mutations_sync_url,
                headers={
                    **get_moneybird_headers(token),
                    "Content-Type": "application/json",
                },
                json={"ids": batch_ids},
                timeout=20,
            )
            detail_response.raise_for_status()
            batch = detail_response.json()
            if isinstance(batch, list):
                financial_mutations.extend(batch)

        invoiced_total = sum(decimal_from_value(invoice.get("total_price_incl_tax")) for invoice in invoices)
        received_total = sum(decimal_from_value(invoice.get("total_paid")) for invoice in invoices)
        expenses_total = sum(
            abs(decimal_from_value(mutation.get("amount")))
            for mutation in financial_mutations
            if is_cost_mutation(mutation, ledger_account_types)
        )
        outstanding_total = invoiced_total - received_total
        last_synced_at = max(
            (
                str(value).strip()
                for value in [
                    *(invoice.get("updated_at", "") for invoice in invoices),
                    *(mutation.get("updated_at", "") for mutation in financial_mutations),
                ]
                if value
            ),
            default="",
        )

        return {
            "source": "moneybird",
            "administrationId": str(administration_id),
            "administrationName": administration.get("name", ""),
            "invoiceCount": len(invoices),
            "revenue_total": round(float(invoiced_total), 2),
            "revenue_received": round(float(received_total), 2),
            "revenue_outstanding": round(float(outstanding_total), 2),
            "expenses_total": round(float(expenses_total), 2),
            "lastSyncedAt": last_synced_at,
            "invoices": invoices,
            "financialMutations": financial_mutations,
            "ledgerAccountTypes": ledger_account_types,
        }
    except requests.HTTPError as exc:
        status_code = exc.response.status_code if exc.response is not None else 0
        if status_code == 401:
            message = "Moneybird API-token is ongeldig of ingetrokken. Controleer MONEYBIRD_API_TOKEN."
        elif status_code == 403:
            message = "Moneybird-token heeft onvoldoende rechten voor deze administratie."
        elif status_code == 404:
            message = "Moneybird-administratie niet gevonden. Controleer MONEYBIRD_ADMINISTRATION_ID."
        else:
            message = "Moneybird reageerde met een fout. Controleer token, administratie en rechten."
        return {
            "source": "error",
            "administrationId": str(config.get("moneybird_administration_id") or ""),
            "invoiceCount": 0,
            "revenue_total": 0.0,
            "revenue_received": 0.0,
            "revenue_outstanding": 0.0,
            "expenses_total": 0.0,
            "lastSyncedAt": "",
            "invoices": [],
            "financialMutations": [],
            "ledgerAccountTypes": {},
            "message": message,
        }
    except requests.RequestException:
        return {
            "source": "error",
            "administrationId": str(config.get("moneybird_administration_id") or ""),
            "invoiceCount": 0,
            "revenue_total": 0.0,
            "revenue_received": 0.0,
            "revenue_outstanding": 0.0,
            "expenses_total": 0.0,
            "lastSyncedAt": "",
            "invoices": [],
            "financialMutations": [],
            "ledgerAccountTypes": {},
            "message": "Moneybird is tijdelijk niet bereikbaar. Probeer het zo opnieuw.",
        }


def fetch_dashboard_payload() -> Dict[str, Any]:
    with ThreadPoolExecutor(max_workers=2) as executor:
        ecwid_future = executor.submit(fetch_orders_from_ecwid)
        moneybird_future = executor.submit(fetch_moneybird_summary)
        ecwid_payload = ecwid_future.result()
        moneybird_payload = moneybird_future.result()

    ecwid_summary = ecwid_payload.get("summary", build_summary(ecwid_payload.get("items", [])))
    report_summary = build_report_summary(ecwid_summary, moneybird_payload)
    messages = [message for message in [ecwid_payload.get("message"), moneybird_payload.get("message")] if message]

    return {
        **ecwid_payload,
        "moneybird": moneybird_payload,
        "reportSummary": report_summary,
        "message": " ".join(messages) if messages else None,
    }


def fetch_ecwid_orders_payload() -> Dict[str, Any]:
    payload = fetch_orders_from_ecwid()
    items = payload.get("items", [])
    if "summary" not in payload:
        payload["summary"] = build_summary(items)
    return payload


def refresh_orders_cache() -> None:
    global refresh_in_progress
    try:
        payload = fetch_dashboard_payload()
        cached_at = time.time()
        config_fingerprint = get_external_cache_fingerprint(include_moneybird=True)
        with cache_lock:
            orders_cache["payload"] = payload
            orders_cache["cached_at"] = cached_at
            orders_cache["config_fingerprint"] = config_fingerprint
        moneybird_payload = payload.get("moneybird")
        if isinstance(moneybird_payload, dict):
            store_moneybird_cache_payload(moneybird_payload, cached_at, config_fingerprint)
    finally:
        refresh_in_progress = False


def start_background_refresh() -> None:
    global refresh_in_progress
    if refresh_in_progress:
        return

    refresh_in_progress = True
    threading.Thread(target=refresh_orders_cache, daemon=True).start()


def refresh_ecwid_orders_cache() -> None:
    global ecwid_refresh_in_progress
    try:
        payload = fetch_ecwid_orders_payload()
        with ecwid_orders_cache_lock:
            ecwid_orders_cache["payload"] = payload
            ecwid_orders_cache["cached_at"] = time.time()
            ecwid_orders_cache["config_fingerprint"] = get_external_cache_fingerprint()
    finally:
        ecwid_refresh_in_progress = False


def start_ecwid_orders_background_refresh() -> None:
    global ecwid_refresh_in_progress
    if ecwid_refresh_in_progress:
        return

    ecwid_refresh_in_progress = True
    threading.Thread(target=refresh_ecwid_orders_cache, daemon=True).start()


def get_empty_moneybird_payload(message: Optional[str] = None) -> Dict[str, Any]:
    return {
        "source": "pending",
        "invoiceCount": 0,
        "revenue_received": 0.0,
        "revenue_total": 0.0,
        "revenue_outstanding": 0.0,
        "expenses_total": 0.0,
        "lastSyncedAt": "",
        "invoices": [],
        "financialMutations": [],
        "ledgerAccountTypes": {},
        "message": message,
        "cachedAt": 0.0,
    }


def serialize_external_cache_fingerprint(config_fingerprint: Any) -> str:
    try:
        return json.dumps(config_fingerprint, ensure_ascii=True, sort_keys=True, default=str)
    except (TypeError, ValueError):
        return str(config_fingerprint)


def load_persistent_external_cache(cache_key: str, config_fingerprint: Any) -> Optional[Dict[str, Any]]:
    serialized_fingerprint = serialize_external_cache_fingerprint(config_fingerprint)
    try:
        with get_db_connection() as connection:
            row = connection.execute(
                """
                SELECT payload_json, cached_at
                FROM external_api_cache
                WHERE cache_key = ? AND config_fingerprint = ?
                """,
                (cache_key, serialized_fingerprint),
            ).fetchone()
    except sqlite3.Error:
        return None
    if row is None:
        return None
    try:
        payload = json.loads(str(row["payload_json"] or "{}"))
    except json.JSONDecodeError:
        return None
    if not isinstance(payload, dict):
        return None
    return {
        "payload": payload,
        "cached_at": float(row["cached_at"] or 0.0),
    }


def save_persistent_external_cache(cache_key: str, config_fingerprint: Any, payload: Dict[str, Any], cached_at: float) -> None:
    try:
        serialized_fingerprint = serialize_external_cache_fingerprint(config_fingerprint)
        payload_json = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
        with get_db_connection() as connection:
            connection.execute(
                """
                INSERT OR REPLACE INTO external_api_cache (cache_key, config_fingerprint, payload_json, cached_at)
                VALUES (?, ?, ?, ?)
                """,
                (cache_key, serialized_fingerprint, payload_json, cached_at),
            )
            connection.commit()
    except (sqlite3.Error, TypeError, ValueError) as exc:
        app.logger.warning("Externe API-cache opslaan mislukt voor %s: %s", cache_key, exc)
        return


def store_moneybird_cache_payload(payload: Dict[str, Any], cached_at: float, config_fingerprint: str) -> None:
    with moneybird_cache_lock:
        moneybird_cache["payload"] = payload
        moneybird_cache["cached_at"] = cached_at
        moneybird_cache["config_fingerprint"] = config_fingerprint
    save_persistent_external_cache("moneybird_summary", config_fingerprint, payload, cached_at)


def refresh_moneybird_cache() -> None:
    global moneybird_refresh_in_progress
    try:
        payload = fetch_moneybird_summary()
        store_moneybird_cache_payload(
            payload,
            time.time(),
            get_external_cache_fingerprint(include_moneybird=True),
        )
    finally:
        moneybird_refresh_in_progress = False


def start_moneybird_background_refresh() -> None:
    global moneybird_refresh_in_progress
    if moneybird_refresh_in_progress:
        return

    moneybird_refresh_in_progress = True
    threading.Thread(target=refresh_moneybird_cache, daemon=True).start()


def fetch_moneybird_non_blocking(force_refresh: bool = False) -> Dict[str, Any]:
    now = time.time()
    config_fingerprint = get_external_cache_fingerprint(include_moneybird=True)
    with moneybird_cache_lock:
        cached_payload = moneybird_cache.get("payload")
        cached_at = float(moneybird_cache.get("cached_at") or 0.0)
        cached_fingerprint = moneybird_cache.get("config_fingerprint")

    cache_matches_config = cached_fingerprint == config_fingerprint
    cache_is_fresh = cached_payload is not None and cache_matches_config and now - cached_at < CACHE_TTL_SECONDS

    if not force_refresh and cache_is_fresh:
        payload = dict(cached_payload)
        payload["cachedAt"] = cached_at
        return payload

    if not force_refresh and cached_payload is not None and cache_matches_config:
        payload = dict(cached_payload)
        payload["cachedAt"] = cached_at
        start_moneybird_background_refresh()
        return payload

    with cache_lock:
        dashboard_payload = orders_cache.get("payload")
        dashboard_cached_at = float(orders_cache.get("cached_at") or 0.0)
        dashboard_fingerprint = orders_cache.get("config_fingerprint")

    if not force_refresh and dashboard_fingerprint == config_fingerprint and dashboard_payload:
        moneybird_payload = dashboard_payload.get("moneybird") if isinstance(dashboard_payload, dict) else None
        if moneybird_payload:
            payload = dict(moneybird_payload)
            payload["cachedAt"] = dashboard_cached_at
            store_moneybird_cache_payload(moneybird_payload, dashboard_cached_at, config_fingerprint)
            if now - dashboard_cached_at >= CACHE_TTL_SECONDS:
                start_moneybird_background_refresh()
            return payload

    persistent_cache = load_persistent_external_cache("moneybird_summary", config_fingerprint)
    if not force_refresh and persistent_cache is not None:
        cached_at = float(persistent_cache.get("cached_at") or 0.0)
        payload = dict(persistent_cache.get("payload") or {})
        payload["cachedAt"] = cached_at
        with moneybird_cache_lock:
            moneybird_cache["payload"] = persistent_cache.get("payload")
            moneybird_cache["cached_at"] = cached_at
            moneybird_cache["config_fingerprint"] = config_fingerprint
        if now - cached_at >= CACHE_TTL_SECONDS:
            start_moneybird_background_refresh()
        return payload

    try:
        payload = fetch_moneybird_summary()
        cached_at = time.time()
        store_moneybird_cache_payload(payload, cached_at, config_fingerprint)
        payload_with_cache = dict(payload)
        payload_with_cache["cachedAt"] = cached_at
        return payload_with_cache
    except requests.RequestException:
        start_moneybird_background_refresh()

    start_moneybird_background_refresh()
    return get_empty_moneybird_payload(
        "Moneybird wordt op de achtergrond bijgewerkt. De nieuwste betalingen verschijnen zo automatisch."
    )


def fetch_orders(force_refresh: bool = False) -> Dict[str, Any]:
    now = time.time()
    config_fingerprint = get_external_cache_fingerprint(include_moneybird=True)
    with cache_lock:
        cached_payload = orders_cache.get("payload")
        cached_at = float(orders_cache.get("cached_at") or 0.0)
        cached_fingerprint = orders_cache.get("config_fingerprint")

    cache_matches_config = cached_fingerprint == config_fingerprint
    cache_is_fresh = cached_payload is not None and cache_matches_config and now - cached_at < CACHE_TTL_SECONDS

    if not force_refresh and cache_is_fresh:
        payload = dict(cached_payload)
        payload["cachedAt"] = cached_at
        return payload

    if not force_refresh and cached_payload is not None and cache_matches_config:
        payload = dict(cached_payload)
        payload["cachedAt"] = cached_at
        start_background_refresh()
        return payload

    try:
        payload = fetch_dashboard_payload()
    except requests.RequestException:
        if cached_payload is not None:
            payload = dict(cached_payload)
            payload["cachedAt"] = cached_at
            payload["message"] = (
                "Er wordt tijdelijk een recente cacheversie getoond omdat Ecwid niet direct reageerde."
            )
            return payload
        raise

    with cache_lock:
        orders_cache["payload"] = payload
        orders_cache["cached_at"] = now
        orders_cache["config_fingerprint"] = config_fingerprint

    moneybird_payload = payload.get("moneybird")
    if isinstance(moneybird_payload, dict):
        store_moneybird_cache_payload(moneybird_payload, now, config_fingerprint)

    payload_with_cache = dict(payload)
    payload_with_cache["cachedAt"] = now
    return payload_with_cache


def fetch_ecwid_orders(force_refresh: bool = False) -> Dict[str, Any]:
    now = time.time()
    config_fingerprint = get_external_cache_fingerprint()
    with ecwid_orders_cache_lock:
        cached_payload = ecwid_orders_cache.get("payload")
        cached_at = float(ecwid_orders_cache.get("cached_at") or 0.0)
        cached_fingerprint = ecwid_orders_cache.get("config_fingerprint")

    cache_matches_config = cached_fingerprint == config_fingerprint
    cache_is_fresh = cached_payload is not None and cache_matches_config and now - cached_at < CACHE_TTL_SECONDS

    if not force_refresh and cache_is_fresh:
        payload = dict(cached_payload)
        payload["cachedAt"] = cached_at
        return payload

    if not force_refresh and cached_payload is not None and cache_matches_config:
        payload = dict(cached_payload)
        payload["cachedAt"] = cached_at
        start_ecwid_orders_background_refresh()
        return payload

    try:
        payload = fetch_ecwid_orders_payload()
    except requests.RequestException:
        if cached_payload is not None:
            payload = dict(cached_payload)
            payload["cachedAt"] = cached_at
            payload["message"] = (
                "Er wordt tijdelijk een recente cacheversie getoond omdat Ecwid niet direct reageerde."
            )
            return payload
        raise

    with ecwid_orders_cache_lock:
        ecwid_orders_cache["payload"] = payload
        ecwid_orders_cache["cached_at"] = now
        ecwid_orders_cache["config_fingerprint"] = config_fingerprint

    payload_with_cache = dict(payload)
    payload_with_cache["cachedAt"] = now
    return payload_with_cache


def get_empty_dashboard_payload(message: Optional[str] = None) -> Dict[str, Any]:
    empty_summary = build_summary([])
    empty_moneybird = {
        "source": "pending",
        "invoiceCount": 0,
        "revenue_received": 0.0,
        "revenue_total": 0.0,
        "revenue_outstanding": 0.0,
        "expenses_total": 0.0,
        "lastSyncedAt": "",
        "invoices": [],
        "financialMutations": [],
        "ledgerAccountTypes": {},
    }
    return {
        "source": "pending",
        "items": [],
        "summary": empty_summary,
        "moneybird": empty_moneybird,
        "reportSummary": build_report_summary(empty_summary, empty_moneybird),
        "message": message,
        "cachedAt": 0.0,
    }


def fetch_orders_non_blocking() -> Dict[str, Any]:
    now = time.time()
    with cache_lock:
        cached_payload = orders_cache.get("payload")
        cached_at = float(orders_cache.get("cached_at") or 0.0)

    if cached_payload is not None:
        payload = dict(cached_payload)
        payload["cachedAt"] = cached_at
        if now - cached_at >= CACHE_TTL_SECONDS:
            start_background_refresh()
        return payload

    start_background_refresh()
    return get_empty_dashboard_payload()


def format_cache_timestamp(timestamp: float) -> str:
    if not timestamp:
        return datetime.now().strftime("%d-%m-%Y %H:%M")
    return datetime.fromtimestamp(timestamp).strftime("%d-%m-%Y %H:%M")


def build_dashboard_frontend_payload(payload: Dict[str, Any]) -> Dict[str, Any]:
    moneybird_payload = payload.get("moneybird", {})
    ledger_account_types = moneybird_payload.get("ledgerAccountTypes", {})
    monthly_revenue_series = build_monthly_revenue_series(
        payload.get("items", []),
        moneybird_payload.get("invoices", []),
        moneybird_payload.get("financialMutations", []),
        ledger_account_types,
    )
    return {
        "source": payload.get("source", "mock"),
        "summary": payload.get("summary", {}),
        "reportSummary": payload.get("reportSummary", {}),
        "productSummary": build_product_summary(payload.get("items", [])),
        "monthlyRevenueSeries": monthly_revenue_series[-12:],
        "moneybird": moneybird_payload,
        "message": payload.get("message"),
        "cachedAt": payload.get("cachedAt", 0.0),
        "lastUpdated": format_cache_timestamp(payload.get("cachedAt", 0.0)),
    }


def parse_iso_datetime(value: str) -> Optional[datetime]:
    if not value:
        return None

    try:
        return datetime.fromisoformat(value)
    except ValueError:
        return None


def invite_is_expired(user: Dict[str, Any]) -> bool:
    expires_at = parse_iso_datetime(user.get("inviteExpiresAt", ""))
    if expires_at is None:
        return False
    return expires_at < datetime.utcnow()


def parse_iso_date(value: str) -> Optional[date]:
    if not value:
        return None

    try:
        return date.fromisoformat(value)
    except ValueError:
        return None


def format_currency(value: float) -> str:
    return f"EUR {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def get_month_label(month_key: str) -> str:
    month_names = [
        "januari",
        "februari",
        "maart",
        "april",
        "mei",
        "juni",
        "juli",
        "augustus",
        "september",
        "oktober",
        "november",
        "december",
    ]
    month_date = datetime.strptime(f"{month_key}-01", "%Y-%m-%d")
    return f"{month_names[month_date.month - 1]} {month_date.year}"


def build_month_options(orders: List[Dict[str, Any]], moneybird_invoices: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    month_keys = set()

    for order in orders:
        created_at = parse_iso_datetime(order.get("createdAt", ""))
        if created_at is not None:
            month_keys.add(created_at.strftime("%Y-%m"))

    for invoice in moneybird_invoices:
        invoice_date = parse_iso_date(str(invoice.get("invoice_date", "")).strip())
        if invoice_date is not None:
            month_keys.add(invoice_date.strftime("%Y-%m"))

    sorted_months = sorted(month_keys, reverse=True)
    options = []
    for month_key in sorted_months:
        options.append(
            {
                "value": month_key,
                "label": get_month_label(month_key),
            }
        )

    return options


def build_profit_month_options(
    orders: List[Dict[str, Any]],
    moneybird_invoices: List[Dict[str, Any]],
    financial_mutations: List[Dict[str, Any]],
    ledger_account_types: Dict[str, str],
) -> List[Dict[str, str]]:
    month_keys = set()

    for order in orders:
        created_at = parse_iso_datetime(order.get("createdAt", ""))
        if created_at is not None:
            month_keys.add(created_at.strftime("%Y-%m"))

    for invoice in moneybird_invoices:
        for payment in invoice.get("payments") or []:
            payment_date = parse_iso_date(str(payment.get("payment_date", "")).strip())
            if payment_date is not None:
                month_keys.add(payment_date.strftime("%Y-%m"))

    for mutation in financial_mutations:
        mutation_date = parse_iso_date(str(mutation.get("date", "")).strip())
        if mutation_date is not None and is_cost_mutation(mutation, ledger_account_types):
            month_keys.add(mutation_date.strftime("%Y-%m"))

    return [
        {"value": month_key, "label": get_month_label(month_key)}
        for month_key in sorted(month_keys, reverse=True)
    ]


def get_football_season_label(start_year: int) -> str:
    return f"{start_year}/{start_year + 1}"


def build_football_season_options(start_year: int = 2022, reference_date: Optional[date] = None) -> List[Dict[str, str]]:
    current_date = reference_date or date.today()
    current_season_start_year = current_date.year if current_date.month >= 7 else current_date.year - 1
    latest_season_start_year = max(current_season_start_year, start_year)

    return [
        {
            "value": str(season_start_year),
            "label": get_football_season_label(season_start_year),
        }
        for season_start_year in range(latest_season_start_year, start_year - 1, -1)
    ]


def get_football_season_range(season_start_year: int) -> Dict[str, date]:
    return {
        "start": date(season_start_year, 7, 1),
        "end": date(season_start_year + 1, 6, 30),
    }


def get_season_start_year_for_date(value: date) -> int:
    return value.year if value.month >= 7 else value.year - 1


def get_season_months(season_start_year: int) -> List[Dict[str, Any]]:
    return [
        {
            "month": month,
            "year": season_start_year if month >= 7 else season_start_year + 1,
            "label": DUTCH_FULL_MONTH_NAMES[month - 1],
            "shortLabel": DUTCH_MONTH_NAMES[month - 1],
        }
        for month in (7, 8, 9, 10, 11, 12, 1, 2, 3, 4, 5, 6)
    ]


def get_season_quarters(season_start_year: int) -> List[Dict[str, Any]]:
    return [
        {
            "quarter": 1,
            "label": "Jul - Sep",
            "periodLabel": f"1 juli {season_start_year} t/m 30 september {season_start_year}",
            "months": {7, 8, 9},
            "yearsByMonth": {7: season_start_year, 8: season_start_year, 9: season_start_year},
        },
        {
            "quarter": 2,
            "label": "Okt - Dec",
            "periodLabel": f"1 oktober {season_start_year} t/m 31 december {season_start_year}",
            "months": {10, 11, 12},
            "yearsByMonth": {10: season_start_year, 11: season_start_year, 12: season_start_year},
        },
        {
            "quarter": 3,
            "label": "Jan - Mrt",
            "periodLabel": f"1 januari {season_start_year + 1} t/m 31 maart {season_start_year + 1}",
            "months": {1, 2, 3},
            "yearsByMonth": {1: season_start_year + 1, 2: season_start_year + 1, 3: season_start_year + 1},
        },
        {
            "quarter": 4,
            "label": "Apr - Jun",
            "periodLabel": f"1 april {season_start_year + 1} t/m 30 juni {season_start_year + 1}",
            "months": {4, 5, 6},
            "yearsByMonth": {4: season_start_year + 1, 5: season_start_year + 1, 6: season_start_year + 1},
        },
    ]


def build_moneybird_revenue_by_month(moneybird_invoices: List[Dict[str, Any]]) -> Dict[str, Decimal]:
    revenue_by_month: Dict[str, Decimal] = {}

    for invoice in moneybird_invoices:
        for payment in invoice.get("payments") or []:
            payment_date = parse_iso_date(str(payment.get("payment_date", "")).strip())
            if payment_date is None:
                continue

            month_key = payment_date.strftime("%Y-%m")
            revenue_by_month[month_key] = revenue_by_month.get(month_key, Decimal("0")) + decimal_from_value(
                payment.get("price")
            )

    return revenue_by_month


def get_quarter_for_month(month: int) -> int:
    return ((month - 1) // 3) + 1


def get_quarter_label(quarter: int) -> str:
    labels = {
        1: "JAN-FEB-MAA",
        2: "APR-MEI-JUN",
        3: "JUL-AUG-SEP",
        4: "OKT-NOV-DEC",
    }
    return labels.get(quarter, "")


def get_quarter_period_label(year: int, quarter: int) -> str:
    month_ranges = {
        1: "1 januari t/m 31 maart",
        2: "1 april t/m 30 juni",
        3: "1 juli t/m 30 september",
        4: "1 oktober t/m 31 december",
    }
    return f"{month_ranges.get(quarter, '')} {year}".strip()


def build_spaarpot_payment_account_label(invoice: Dict[str, Any], payment: Dict[str, Any]) -> str:
    payment_account_values = [
        payment.get("account_name"),
        payment.get("bank_account_name"),
        payment.get("counterparty_name"),
        payment.get("counterparty_account"),
        payment.get("bank_account"),
    ]
    for value in payment_account_values:
        label = str(value or "").strip()
        if label:
            return label

    contact = invoice.get("contact") if isinstance(invoice.get("contact"), dict) else {}
    contact_name = (
        str(contact.get("company_name") or "").strip()
        or " ".join(
            part
            for part in [
                str(contact.get("firstname") or "").strip(),
                str(contact.get("lastname") or "").strip(),
            ]
            if part
        )
        or str(contact.get("email") or "").strip()
    )
    contact_bank_account = str(contact.get("bank_account") or "").strip()
    if contact_name and contact_bank_account:
        return f"{contact_name} - {contact_bank_account}"
    if contact_name:
        return contact_name
    if contact_bank_account:
        return contact_bank_account
    return "Onbekende rekening"


def build_spaarpot_stripe_mutation_label(mutation: Dict[str, Any]) -> str:
    label_values = [
        mutation.get("contra_account_name"),
        mutation.get("account_name"),
        mutation.get("message"),
        mutation.get("description"),
    ]
    for value in label_values:
        label = str(value or "").strip()
        if label:
            return label
    return "Stripe"


def build_spaarpot_payment_entries(
    moneybird_invoices: List[Dict[str, Any]],
    financial_mutations: Optional[List[Dict[str, Any]]] = None,
) -> List[Dict[str, Any]]:
    entries: List[Dict[str, Any]] = []

    for invoice in moneybird_invoices:
        for payment in invoice.get("payments") or []:
            payment_date = parse_iso_date(str(payment.get("payment_date", "")).strip())
            if payment_date is None:
                continue

            amount = decimal_from_value(payment.get("price"))
            if amount <= 0:
                continue

            quarter = get_quarter_for_month(payment_date.month)
            reserve = amount * VAT_SAVINGS_RATE
            entries.append(
                {
                    "source": "payment",
                    "date": payment_date.isoformat(),
                    "dateLabel": payment_date.strftime("%d-%m-%Y"),
                    "year": payment_date.year,
                    "quarter": quarter,
                    "quarterLabel": get_quarter_label(quarter),
                    "invoiceId": str(invoice.get("invoice_id") or invoice.get("id") or "").strip(),
                    "contactName": str((invoice.get("contact") or {}).get("company_name") or "").strip(),
                    "accountLabel": build_spaarpot_payment_account_label(invoice, payment),
                    "amount": round(float(amount), 2),
                    "reserve": round(float(reserve), 2),
                }
            )

    for mutation in financial_mutations or []:
        mutation_date = parse_iso_date(str(mutation.get("date", "")).strip())
        if mutation_date is None or not is_spaarpot_stripe_income_mutation(mutation):
            continue

        amount = get_spaarpot_stripe_mutation_amount(mutation)
        quarter = get_quarter_for_month(mutation_date.month)
        reserve = amount * VAT_SAVINGS_RATE
        mutation_reference = str(
            mutation.get("code") or mutation.get("id") or mutation.get("financial_account_id") or ""
        ).strip()
        entries.append(
            {
                "source": "stripe",
                "date": mutation_date.isoformat(),
                "dateLabel": mutation_date.strftime("%d-%m-%Y"),
                "year": mutation_date.year,
                "quarter": quarter,
                "quarterLabel": get_quarter_label(quarter),
                "invoiceId": f"Stripe {mutation_reference}".strip(),
                "contactName": "Stripe",
                "accountLabel": build_spaarpot_stripe_mutation_label(mutation),
                "amount": round(float(amount), 2),
                "reserve": round(float(reserve), 2),
            }
        )

    return sorted(entries, key=lambda item: (item["date"], item["invoiceId"]), reverse=True)


def normalize_spaarpot_manual_entry(row: sqlite3.Row) -> Dict[str, Any]:
    amount = decimal_from_value(row["amount"])
    created_at = str(row["created_at"] or "").strip()
    created_date = parse_iso_date(created_at[:10]) if created_at else None
    date_label = created_date.strftime("%d-%m-%Y") if created_date is not None else ""
    quarter = int(row["quarter"])
    return {
        "source": "manual",
        "id": int(row["id"]),
        "date": created_at[:10],
        "dateLabel": date_label,
        "year": int(row["year"]),
        "quarter": quarter,
        "quarterLabel": get_quarter_label(quarter),
        "invoiceId": "Handmatig",
        "contactName": "",
        "accountLabel": str(row["description"] or "").strip(),
        "amount": 0.0,
        "reserve": round(float(amount), 2),
    }


def load_spaarpot_manual_entries() -> List[Dict[str, Any]]:
    def loader() -> List[Dict[str, Any]]:
        with get_db_connection() as connection:
            rows = connection.execute(
                """
                SELECT id, year, quarter, description, amount, created_at
                FROM spaarpot_manual_entries
                ORDER BY year DESC, quarter DESC, created_at DESC, id DESC
                """
            ).fetchall()
        return [normalize_spaarpot_manual_entry(row) for row in rows]

    return get_cached_local_data("spaarpot_manual_entries", (), loader)


def create_spaarpot_manual_entry(year: int, quarter: int, description: str, amount: Decimal) -> int:
    created_at = datetime.now().isoformat(timespec="seconds")
    with get_db_connection() as connection:
        cursor = connection.execute(
            """
            INSERT INTO spaarpot_manual_entries (year, quarter, description, amount, created_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (year, quarter, description.strip(), round(float(amount), 2), created_at),
        )
        entry_id = int(cursor.lastrowid)
    clear_local_data_cache()
    return entry_id


def delete_spaarpot_manual_entry(entry_id: int) -> None:
    with get_db_connection() as connection:
        connection.execute("DELETE FROM spaarpot_manual_entries WHERE id = ?", (entry_id,))
    clear_local_data_cache()


def format_euro_amount(amount: Decimal) -> str:
    quantized = amount.quantize(Decimal("0.01"))
    return f"EUR {quantized:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def build_spaarpot_weekly_reminder(
    payment_entries: List[Dict[str, Any]],
    monday: date,
) -> Dict[str, Any]:
    week_start = monday - timedelta(days=7)
    week_end = monday
    previous_balance = Decimal("0")
    weekly_added = Decimal("0")
    weekly_payment_count = 0

    for entry in payment_entries:
        entry_date = parse_iso_date(str(entry.get("date", "")).strip())
        if entry_date is None:
            continue

        reserve = decimal_from_value(entry.get("reserve"))
        if entry_date < week_start:
            previous_balance += reserve
        elif week_start <= entry_date < week_end:
            weekly_added += reserve
            if entry.get("source") != "manual":
                weekly_payment_count += 1

    current_balance = previous_balance + weekly_added
    return {
        "monday": monday.isoformat(),
        "weekStart": week_start.isoformat(),
        "weekEnd": week_end.isoformat(),
        "weekLabel": f"{week_start.strftime('%d-%m-%Y')} t/m {(week_end - timedelta(days=1)).strftime('%d-%m-%Y')}",
        "previousBalance": round(float(previous_balance), 2),
        "weeklyAdded": round(float(weekly_added), 2),
        "topUpAmount": round(float(weekly_added), 2),
        "currentBalance": round(float(current_balance), 2),
        "paymentCount": weekly_payment_count,
        "previousBalanceLabel": format_euro_amount(previous_balance),
        "weeklyAddedLabel": format_euro_amount(weekly_added),
        "topUpAmountLabel": format_euro_amount(weekly_added),
        "currentBalanceLabel": format_euro_amount(current_balance),
    }


def build_current_spaarpot_weekly_reminder(target_date: Optional[date] = None) -> Dict[str, Any]:
    monday = target_date or date.today()
    monday = monday - timedelta(days=monday.weekday())
    payload = fetch_orders(force_refresh=True)
    moneybird = payload.get("moneybird", {})
    payment_entries = build_spaarpot_payment_entries(
        moneybird.get("invoices", []),
        moneybird.get("financialMutations", []),
    )
    payment_entries.extend(load_spaarpot_manual_entries())
    reminder = build_spaarpot_weekly_reminder(payment_entries, monday)
    reminder["message"] = payload.get("message") or ""
    reminder["lastUpdated"] = format_cache_timestamp(payload.get("cachedAt", 0.0))
    return reminder


def get_web_push_vapid_public_key() -> str:
    return get_env("WEB_PUSH_VAPID_PUBLIC_KEY")


def get_web_push_vapid_private_key() -> str:
    return get_env("WEB_PUSH_VAPID_PRIVATE_KEY")


def get_web_push_vapid_subject() -> str:
    return get_env("WEB_PUSH_VAPID_SUBJECT") or f"mailto:{get_env('ADMIN_EMAIL') or 'info@hwsvoetbalschool.nl'}"


def is_web_push_configured() -> bool:
    return bool(webpush and get_web_push_vapid_public_key() and get_web_push_vapid_private_key())


def normalize_push_subscription(subscription: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    endpoint = str(subscription.get("endpoint") or "").strip()
    keys = subscription.get("keys") if isinstance(subscription.get("keys"), dict) else {}
    p256dh = str(keys.get("p256dh") or "").strip()
    auth = str(keys.get("auth") or "").strip()
    if not endpoint or not p256dh or not auth:
        return None
    return {"endpoint": endpoint, "keys": {"p256dh": p256dh, "auth": auth}}


def save_web_push_subscription(user_id: str, subscription: Dict[str, Any], user_agent: str = "") -> None:
    normalized = normalize_push_subscription(subscription)
    if normalized is None:
        raise ValueError("Ongeldige push-subscription.")

    now_value = datetime.now().isoformat(timespec="seconds")
    with get_db_connection() as connection:
        connection.execute(
            """
            INSERT INTO web_push_subscriptions (
                user_id, endpoint, subscription_json, user_agent, created_at, updated_at, last_error
            )
            VALUES (?, ?, ?, ?, ?, ?, NULL)
            ON CONFLICT(endpoint) DO UPDATE SET
                user_id = excluded.user_id,
                subscription_json = excluded.subscription_json,
                user_agent = excluded.user_agent,
                updated_at = excluded.updated_at,
                last_error = NULL
            """,
            (
                user_id,
                normalized["endpoint"],
                json.dumps(normalized, separators=(",", ":")),
                user_agent[:500],
                now_value,
                now_value,
            ),
        )


def delete_web_push_subscription(endpoint: str) -> None:
    normalized_endpoint = str(endpoint or "").strip()
    if not normalized_endpoint:
        return
    with get_db_connection() as connection:
        connection.execute("DELETE FROM web_push_subscriptions WHERE endpoint = ?", (normalized_endpoint,))


def load_web_push_subscriptions() -> List[Dict[str, Any]]:
    with get_db_connection() as connection:
        rows = connection.execute(
            """
            SELECT id, user_id, endpoint, subscription_json
            FROM web_push_subscriptions
            ORDER BY updated_at DESC, id DESC
            """
        ).fetchall()
    return [dict(row) for row in rows]


def mark_web_push_subscription_sent(subscription_id: int) -> None:
    with get_db_connection() as connection:
        connection.execute(
            """
            UPDATE web_push_subscriptions
            SET last_sent_at = ?, last_error = NULL
            WHERE id = ?
            """,
            (datetime.now().isoformat(timespec="seconds"), subscription_id),
        )


def mark_web_push_subscription_error(subscription_id: int, error: str) -> None:
    with get_db_connection() as connection:
        connection.execute(
            "UPDATE web_push_subscriptions SET last_error = ? WHERE id = ?",
            (str(error)[:500], subscription_id),
        )


def build_spaarpot_push_payload(reminder: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "title": "Spaarpot bijstorten",
        "body": (
            f"Bijstorten: {reminder['topUpAmountLabel']}. "
            f"Vorige maandag stond er {reminder['previousBalanceLabel']} in de spaarpot; "
            f"deze week kwam er {reminder['weeklyAddedLabel']} bij."
        ),
        "url": "/spaarpot",
        "tag": f"spaarpot-week-{reminder['monday']}",
        "data": reminder,
    }


def send_spaarpot_push_reminders(target_date: Optional[date] = None) -> Dict[str, Any]:
    if not is_web_push_configured():
        return {
            "sent": 0,
            "failed": 0,
            "deleted": 0,
            "error": "Web Push is nog niet geconfigureerd. Zet WEB_PUSH_VAPID_PUBLIC_KEY en WEB_PUSH_VAPID_PRIVATE_KEY.",
        }

    reminder = build_current_spaarpot_weekly_reminder(target_date)
    notification_payload = build_spaarpot_push_payload(reminder)
    subscriptions = load_web_push_subscriptions()
    sent = 0
    failed = 0
    deleted = 0

    for subscription in subscriptions:
        try:
            webpush(
                subscription_info=json.loads(subscription["subscription_json"]),
                data=json.dumps(notification_payload),
                vapid_private_key=get_web_push_vapid_private_key(),
                vapid_claims={"sub": get_web_push_vapid_subject()},
            )
            mark_web_push_subscription_sent(int(subscription["id"]))
            sent += 1
        except Exception as exc:
            status_code = getattr(getattr(exc, "response", None), "status_code", 0)
            if status_code in {404, 410}:
                delete_web_push_subscription(subscription["endpoint"])
                deleted += 1
            else:
                mark_web_push_subscription_error(int(subscription["id"]), str(exc))
                failed += 1

    return {
        "sent": sent,
        "failed": failed,
        "deleted": deleted,
        "reminder": reminder,
        "subscriptionCount": len(subscriptions),
    }


def build_spaarpot_season_options(payment_entries: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    season_years = {
        get_season_start_year_for_date(entry_date)
        for entry in payment_entries
        for entry_date in [get_spaarpot_entry_period_date(entry)]
        if entry_date is not None
    }
    current_season_start_year = get_season_start_year_for_date(date.today())
    season_years.add(current_season_start_year)
    season_years.add(current_season_start_year + 1)
    return [
        {"value": str(year), "label": get_football_season_label(year)}
        for year in sorted(season_years, reverse=True)
    ]


def get_default_spaarpot_season(payment_entries: List[Dict[str, Any]]) -> str:
    current_season_start_year = get_season_start_year_for_date(date.today())
    entry_seasons = {
        get_season_start_year_for_date(entry_date)
        for entry in payment_entries
        for entry_date in [get_spaarpot_entry_period_date(entry)]
        if entry_date is not None
    }
    if current_season_start_year in entry_seasons:
        return str(current_season_start_year)
    if entry_seasons:
        return str(max(entry_seasons))
    return str(current_season_start_year)


def get_spaarpot_entry_season_quarter(entry: Dict[str, Any]) -> Optional[int]:
    entry_date = get_spaarpot_entry_period_date(entry)
    if entry_date is None:
        return None
    if entry_date.month in {7, 8, 9}:
        return 1
    if entry_date.month in {10, 11, 12}:
        return 2
    if entry_date.month in {1, 2, 3}:
        return 3
    return 4


def get_spaarpot_entry_period_date(entry: Dict[str, Any]) -> Optional[date]:
    if entry.get("source") == "manual":
        try:
            year = int(entry.get("year"))
            quarter = int(entry.get("quarter"))
        except (TypeError, ValueError):
            return None
        quarter_start_month = {1: 1, 2: 4, 3: 7, 4: 10}.get(quarter)
        return date(year, quarter_start_month, 1) if quarter_start_month else None
    return parse_iso_date(str(entry.get("date") or ""))


def get_calendar_period_for_season_quarter(season_start_year: int, season_quarter: int) -> Tuple[int, int]:
    if season_quarter == 1:
        return season_start_year, 3
    if season_quarter == 2:
        return season_start_year, 4
    if season_quarter == 3:
        return season_start_year + 1, 1
    return season_start_year + 1, 2


def build_spaarpot_quarter_summary(
    payment_entries: List[Dict[str, Any]],
    season_start_year: int,
) -> Dict[str, Any]:
    quarters = []
    season_range = get_football_season_range(season_start_year)
    season_income = Decimal("0")
    season_reserve = Decimal("0")
    season_payment_count = 0

    for quarter_meta in get_season_quarters(season_start_year):
        quarter = int(quarter_meta["quarter"])
        quarter_entries = [
            entry
            for entry in payment_entries
            for entry_date in [get_spaarpot_entry_period_date(entry)]
            if entry_date is not None
            and season_range["start"] <= entry_date <= season_range["end"]
            and get_spaarpot_entry_season_quarter(entry) == quarter
        ]
        income = sum(decimal_from_value(entry["amount"]) for entry in quarter_entries)
        reserve = sum(decimal_from_value(entry["reserve"]) for entry in quarter_entries)
        payment_count = sum(1 for entry in quarter_entries if entry.get("source") != "manual")
        manual_count = sum(1 for entry in quarter_entries if entry.get("source") == "manual")
        season_income += income
        season_reserve += reserve
        season_payment_count += payment_count
        quarters.append(
            {
                "quarter": quarter,
                "label": str(quarter_meta["label"]),
                "periodLabel": str(quarter_meta["periodLabel"]),
                "income": round(float(income), 2),
                "reserve": round(float(reserve), 2),
                "paymentCount": payment_count,
                "manualCount": manual_count,
                "entryCount": len(quarter_entries),
                "payments": quarter_entries,
            }
        )

    return {
        "selectedSeason": str(season_start_year),
        "selectedSeasonLabel": get_football_season_label(season_start_year),
        "quarters": quarters,
        "income": round(float(season_income), 2),
        "reserve": round(float(season_reserve), 2),
        "paymentCount": season_payment_count,
        "ratePercentage": round(float(VAT_SAVINGS_RATE * Decimal("100")), 2),
    }


def build_spaarpot_year_totals(payment_entries: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    totals_by_year: Dict[int, Dict[str, Decimal]] = {}
    counts_by_year: Dict[int, int] = {}

    for entry in payment_entries:
        year = int(entry["year"])
        totals = totals_by_year.setdefault(year, {"income": Decimal("0"), "reserve": Decimal("0")})
        totals["income"] += decimal_from_value(entry["amount"])
        totals["reserve"] += decimal_from_value(entry["reserve"])
        counts_by_year[year] = counts_by_year.get(year, 0) + 1

    return [
        {
            "year": str(year),
            "income": round(float(totals_by_year[year]["income"]), 2),
            "reserve": round(float(totals_by_year[year]["reserve"]), 2),
            "paymentCount": counts_by_year.get(year, 0),
        }
        for year in sorted(totals_by_year, reverse=True)
    ]


def build_moneybird_expenses_by_month(
    financial_mutations: List[Dict[str, Any]],
    ledger_account_types: Dict[str, str],
) -> Dict[str, Decimal]:
    expenses_by_month: Dict[str, Decimal] = {}

    for mutation in financial_mutations:
        mutation_date = parse_iso_date(str(mutation.get("date", "")).strip())
        if mutation_date is None or not is_cost_mutation(mutation, ledger_account_types):
            continue

        month_key = mutation_date.strftime("%Y-%m")
        expenses_by_month[month_key] = expenses_by_month.get(month_key, Decimal("0")) + abs(
            decimal_from_value(mutation.get("amount"))
        )

    return expenses_by_month


def build_period_revenue_summary(
    orders: List[Dict[str, Any]],
    moneybird_invoices: List[Dict[str, Any]],
    financial_mutations: List[Dict[str, Any]],
    ledger_account_types: Dict[str, str],
    period_start: date,
    period_end: date,
    period_label: str,
    period_value: str,
) -> Dict[str, Any]:
    ecwid_revenue = Decimal("0")
    ecwid_order_count = 0

    for order in orders:
        created_at = parse_iso_datetime(order.get("createdAt", ""))
        if created_at is None or order.get("paymentStatus") == "REFUNDED":
            continue

        created_date = created_at.date()
        if created_date < period_start or created_date > period_end:
            continue

        ecwid_revenue += decimal_from_value(order.get("total"))
        ecwid_order_count += 1

    moneybird_revenue = Decimal("0")
    moneybird_payment_count = 0

    for invoice in moneybird_invoices:
        for payment in invoice.get("payments") or []:
            payment_date = parse_iso_date(str(payment.get("payment_date", "")).strip())
            if payment_date is None or payment_date < period_start or payment_date > period_end:
                continue

            moneybird_revenue += decimal_from_value(payment.get("price"))
            moneybird_payment_count += 1

    expenses = Decimal("0")
    for mutation in financial_mutations:
        mutation_date = parse_iso_date(str(mutation.get("date", "")).strip())
        if (
            mutation_date is None
            or mutation_date < period_start
            or mutation_date > period_end
            or not is_cost_mutation(mutation, ledger_account_types)
        ):
            continue

        expenses += abs(decimal_from_value(mutation.get("amount")))

    combined_revenue = ecwid_revenue + moneybird_revenue
    profit = combined_revenue - expenses
    return {
        "selectedPeriod": period_value,
        "selectedPeriodLabel": period_label,
        "periodStart": period_start.isoformat(),
        "periodEnd": period_end.isoformat(),
        "ecwidRevenue": round(float(ecwid_revenue), 2),
        "ecwidOrderCount": ecwid_order_count,
        "moneybirdRevenue": round(float(moneybird_revenue), 2),
        "moneybirdPaymentCount": moneybird_payment_count,
        "expenses": round(float(expenses), 2),
        "combinedRevenue": round(float(combined_revenue), 2),
        "profit": round(float(profit), 2),
        "profitMarginPercentage": calculate_margin_percentage(combined_revenue, profit),
    }


def build_football_season_summary(
    orders: List[Dict[str, Any]],
    moneybird_invoices: List[Dict[str, Any]],
    financial_mutations: List[Dict[str, Any]],
    ledger_account_types: Dict[str, str],
    season_start_year: int,
) -> Dict[str, Any]:
    season_range = get_football_season_range(season_start_year)
    return build_period_revenue_summary(
        orders,
        moneybird_invoices,
        financial_mutations,
        ledger_account_types,
        season_range["start"],
        season_range["end"],
        get_football_season_label(season_start_year),
        str(season_start_year),
    )


def build_monthly_revenue_summary(
    orders: List[Dict[str, Any]],
    moneybird_invoices: List[Dict[str, Any]],
    financial_mutations: List[Dict[str, Any]],
    ledger_account_types: Dict[str, str],
    selected_month: str,
) -> Dict[str, Any]:
    month_start = datetime.strptime(f"{selected_month}-01", "%Y-%m-%d").date()
    next_month = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)
    month_end = next_month - timedelta(days=1)
    summary = build_period_revenue_summary(
        orders,
        moneybird_invoices,
        financial_mutations,
        ledger_account_types,
        month_start,
        month_end,
        get_month_label(selected_month),
        selected_month,
    )
    summary["selectedMonth"] = selected_month
    summary["selectedMonthLabel"] = summary["selectedPeriodLabel"]
    return summary


def build_monthly_revenue_series(
    orders: List[Dict[str, Any]],
    moneybird_invoices: List[Dict[str, Any]],
    financial_mutations: List[Dict[str, Any]],
    ledger_account_types: Dict[str, str],
) -> List[Dict[str, Any]]:
    month_keys: Set[str] = set()
    ecwid_by_month: Dict[str, Decimal] = {}
    moneybird_by_month = build_moneybird_revenue_by_month(moneybird_invoices)
    expenses_by_month = build_moneybird_expenses_by_month(financial_mutations, ledger_account_types)

    for order in orders:
        created_at = parse_iso_datetime(order.get("createdAt", ""))
        if created_at is None or order.get("paymentStatus") == "REFUNDED":
            continue
        month_key = created_at.strftime("%Y-%m")
        month_keys.add(month_key)
        ecwid_by_month[month_key] = ecwid_by_month.get(month_key, Decimal("0")) + decimal_from_value(order.get("total"))

    for month_key in moneybird_by_month:
        month_keys.add(month_key)

    for month_key in expenses_by_month:
        month_keys.add(month_key)

    series = []
    for month_key in sorted(month_keys):
        ecwid_revenue = ecwid_by_month.get(month_key, Decimal("0"))
        moneybird_revenue = moneybird_by_month.get(month_key, Decimal("0"))
        expenses = expenses_by_month.get(month_key, Decimal("0"))
        combined_revenue = ecwid_revenue + moneybird_revenue
        profit = combined_revenue - expenses
        series.append(
            {
                "month": month_key,
                "label": get_month_label(month_key),
                "ecwidRevenue": round(float(ecwid_revenue), 2),
                "moneybirdRevenue": round(float(moneybird_revenue), 2),
                "combinedRevenue": round(float(combined_revenue), 2),
                "expenses": round(float(expenses), 2),
                "profit": round(float(profit), 2),
                "profitMarginPercentage": calculate_margin_percentage(combined_revenue, profit),
            }
        )

    return series


def build_profit_totals(
    orders: List[Dict[str, Any]],
    moneybird_invoices: List[Dict[str, Any]],
    financial_mutations: List[Dict[str, Any]],
    ledger_account_types: Dict[str, str],
) -> Dict[str, Any]:
    monthly_series = build_monthly_revenue_series(orders, moneybird_invoices, financial_mutations, ledger_account_types)
    combined_revenue = sum(decimal_from_value(item.get("combinedRevenue")) for item in monthly_series)
    expenses = sum(decimal_from_value(item.get("expenses")) for item in monthly_series)
    profit = combined_revenue - expenses

    return {
        "combinedRevenue": round(float(combined_revenue), 2),
        "expenses": round(float(expenses), 2),
        "profit": round(float(profit), 2),
        "profitMarginPercentage": calculate_margin_percentage(combined_revenue, profit),
    }


def mock_orders() -> List[Dict[str, Any]]:
    return [
        {
            "id": "WEB-1001",
            "orderNumber": "WEB-1001",
            "createdAt": "2026-04-04T14:12:00+02:00",
            "status": "PAID",
            "paymentStatus": "PAID",
            "fulfillmentStatus": "AWAITING_PROCESSING",
            "total": 89.95,
            "email": "anne@example.com",
            "customerName": "Anne de Vries",
            "paymentMethod": "iDEAL",
            "shippingMethod": "PostNL pakket",
            "itemCount": 3,
            "items": [
                {"name": "Linnen blouse", "quantity": 1, "price": 49.95, "sku": "BL-01"},
                {"name": "Canvas tas", "quantity": 2, "price": 20.00, "sku": "TS-02"},
            ],
        },
        {
            "id": "WEB-1002",
            "orderNumber": "WEB-1002",
            "createdAt": "2026-04-03T09:05:00+02:00",
            "status": "PROCESSING",
            "paymentStatus": "AWAITING_PAYMENT",
            "fulfillmentStatus": "AWAITING_PROCESSING",
            "total": 129.00,
            "email": "milan@example.com",
            "customerName": "Milan Jansen",
            "paymentMethod": "Bankoverschrijving",
            "shippingMethod": "Afhalen",
            "itemCount": 1,
            "items": [
                {"name": "Leren portefeuille", "quantity": 1, "price": 129.00, "sku": "PF-09"},
            ],
        },
        {
            "id": "WEB-1003",
            "orderNumber": "WEB-1003",
            "createdAt": "2026-04-01T16:45:00+02:00",
            "status": "SHIPPED",
            "paymentStatus": "PAID",
            "fulfillmentStatus": "SHIPPED",
            "total": 62.50,
            "email": "noor@example.com",
            "customerName": "Noor Bakker",
            "paymentMethod": "Creditcard",
            "shippingMethod": "DHL",
            "itemCount": 2,
            "items": [
                {"name": "Keramische mok", "quantity": 2, "price": 17.50, "sku": "MK-11"},
                {"name": "Theeblik", "quantity": 1, "price": 27.50, "sku": "TB-03"},
            ],
        },
    ]


run_storage_migrations()


@app.route("/login", methods=["GET", "POST"])
def login_page() -> str:
    existing_user = get_current_user()
    if existing_user is not None and request.method == "GET":
        return redirect(get_default_post_login_path(existing_user))

    login_error = ""
    next_path = request.args.get("next", "").strip() or request.form.get("next", "").strip()

    if request.method == "POST":
        login_value = request.form.get("email", "").strip()
        password = request.form.get("password", "")
        user = authenticate_user(login_value, password)
        if user is None:
            login_error = GENERIC_AUTH_ERROR_MESSAGE
        else:
            rotate_authenticated_session(user["id"])
            return redirect(get_default_post_login_path(user))

    if not next_path:
        fallback_user = existing_user or {"isAdmin": True}
        next_path = get_default_post_login_path(fallback_user)

    return render_template("login.html", login_error=login_error, next_path=next_path)


@app.route("/uitnodiging/<invite_token>", methods=["GET", "POST"])
def invite_accept_page(invite_token: str) -> str:
    invited_user = get_user_by_invite_token(invite_token)
    if invited_user is None:
        return render_template(
            "invite_accept.html",
            invited_user=None,
            invite_error="Deze aanmeldlink is niet geldig of is al gebruikt.",
            invite_success="",
        )

    if invited_user.get("passwordHash"):
        return redirect(url_for("login_page"))

    if invite_is_expired(invited_user):
        return render_template(
            "invite_accept.html",
            invited_user=invited_user,
            invite_error="Deze aanmeldlink is verlopen. Maak een nieuwe uitnodiging aan voor dit teamlid.",
            invite_success="",
        )

    invite_error = ""
    invite_success = ""

    if request.method == "POST":
        password = request.form.get("password", "")
        password_confirm = request.form.get("password_confirm", "")

        if len(password) < 12:
            invite_error = "Kies een wachtwoord van minimaal 12 tekens."
        elif password != password_confirm:
            invite_error = "De wachtwoorden komen niet overeen."
        else:
            accept_trainer_invite(invited_user["id"], password)
            refreshed_user = get_user_by_id(invited_user["id"])
            if refreshed_user is not None:
                rotate_authenticated_session(refreshed_user["id"])
            invite_success = "Wachtwoord opgeslagen. Je account is geactiveerd."
            if refreshed_user is not None:
                return redirect(get_default_post_login_path(refreshed_user))

    return render_template(
        "invite_accept.html",
        invited_user=invited_user,
        invite_error=invite_error,
        invite_success=invite_success,
    )


@app.post("/logout")
def logout_page():
    session.clear()
    return redirect(url_for("login_page"))


@app.get("/")
def index() -> str:
    access_redirect = require_page_access("dashboard")
    if access_redirect is not None:
        return access_redirect

    user = get_current_user()
    payload = fetch_orders_non_blocking()
    dashboard_payload = build_dashboard_frontend_payload(payload)
    return render_template(
        "index.html",
        active_page="dashboard",
        is_trainer_dashboard_user=is_trainer_user(user),
        trainer_week_schedule=build_trainer_dashboard_week_schedule(user),
        dashboard_weather=load_dashboard_weather_settings(),
        source=dashboard_payload["source"],
        summary=dashboard_payload["summary"],
        report_summary=dashboard_payload["reportSummary"],
        product_summary=dashboard_payload["productSummary"],
        last_updated=dashboard_payload["lastUpdated"],
        message=dashboard_payload["message"],
    )


@app.get("/bestellingen")
def orders_page() -> str:
    access_redirect = require_page_access("orders")
    if access_redirect is not None:
        return access_redirect

    page = max(request.args.get("page", default=1, type=int), 1)
    search_query = request.args.get("q", "").strip()
    selected_status = request.args.get("status", "").strip()
    selected_payment_status = request.args.get("payment_status", "").strip()
    selected_month = request.args.get("month", "").strip()
    per_page = 20
    payload = fetch_ecwid_orders()
    all_orders = sort_orders_desc(payload.get("items", []))
    filter_options = build_orders_filter_options(all_orders)
    filtered_orders = filter_orders(
        all_orders,
        search_query=search_query,
        status=selected_status,
        payment_status=selected_payment_status,
        month=selected_month,
    )
    total_orders = len(filtered_orders)
    total_pages = max(ceil(total_orders / per_page), 1)
    current_page = min(page, total_pages)
    start_index = (current_page - 1) * per_page
    end_index = start_index + per_page
    page_orders = decorate_orders_for_list(filtered_orders[start_index:end_index])

    pagination_links = []
    for page_number in range(1, total_pages + 1):
        pagination_links.append(
            {
                "page": page_number,
                "url": build_orders_page_url(
                    page=page_number,
                    search_query=search_query,
                    status=selected_status,
                    payment_status=selected_payment_status,
                    month=selected_month,
                ),
            }
        )

    return render_template(
        "orders.html",
        active_page="orders",
        source=payload.get("source", "mock"),
        summary=payload.get("summary", build_summary(payload.get("items", []))),
        orders=page_orders,
        current_page=current_page,
        total_pages=total_pages,
        total_orders=total_orders,
        start_number=start_index + 1 if total_orders else 0,
        end_number=min(end_index, total_orders),
        total_unfiltered_orders=len(all_orders),
        last_updated=format_cache_timestamp(payload.get("cachedAt", 0.0)),
        message=payload.get("message"),
        search_query=search_query,
        selected_status=selected_status,
        selected_payment_status=selected_payment_status,
        selected_month=selected_month,
        filter_options=filter_options,
        has_active_filters=bool(search_query or selected_status or selected_payment_status or selected_month),
        refresh_url=build_orders_page_url(
            page=current_page,
            search_query=search_query,
            status=selected_status,
            payment_status=selected_payment_status,
            month=selected_month,
        ),
        reset_filters_url=url_for("orders_page"),
        prev_page_url=(
            build_orders_page_url(
                page=current_page - 1,
                search_query=search_query,
                status=selected_status,
                payment_status=selected_payment_status,
                month=selected_month,
            )
            if current_page > 1
            else ""
        ),
        next_page_url=(
            build_orders_page_url(
                page=current_page + 1,
                search_query=search_query,
                status=selected_status,
                payment_status=selected_payment_status,
                month=selected_month,
            )
            if current_page < total_pages
            else ""
        ),
        pagination_links=pagination_links,
    )


@app.get("/aanmeldingen")
def registrations_page() -> str:
    access_redirect = require_page_access("orders")
    if access_redirect is not None:
        return access_redirect

    selected_product_key = request.args.get("product", "").strip()
    if selected_product_key:
        return redirect(build_registration_detail_url(selected_product_key))

    products_payload = fetch_catalog_products()
    product_message = products_payload.get("message")
    registrations = build_registrations_overview_entries(products_payload.get("items", []))

    return render_template(
        "registrations.html",
        active_page="registrations",
        products=registrations,
        total_products=len(registrations),
        refresh_url=build_registrations_page_url(),
        last_updated=format_cache_timestamp(products_payload.get("cachedAt", 0.0)),
        message=product_message or None,
    )


@app.get("/leads")
def leads_page() -> str:
    access_redirect = require_page_access("leads")
    if access_redirect is not None:
        return access_redirect

    products_payload = fetch_catalog_products()
    orders_payload = fetch_ecwid_orders()
    product_summaries = build_product_registration_summary(
        products_payload.get("items", []),
        orders_payload.get("items", []),
    )

    leads_products = [
        {
            "productKey": product["productKey"],
            "productId": product["productId"],
            "name": product["name"],
            "sku": product["sku"],
            "orderCount": product["orderCount"],
            "participantCount": product["participantCount"],
            "emailCount": product["emailCount"],
            "emails": product["emails"],
            "searchText": product["searchText"],
        }
        for product in product_summaries
    ]

    product_message = products_payload.get("message")
    order_message = orders_payload.get("message")
    message_parts = []
    for message in (product_message, order_message):
        if message and message not in message_parts:
            message_parts.append(message)

    return render_template(
        "leads.html",
        active_page="leads",
        products=leads_products,
        blocked_emails_value=load_blocked_lead_emails(),
        total_products=len(leads_products),
        refresh_url=build_leads_page_url(),
        last_updated=format_cache_timestamp(orders_payload.get("cachedAt", 0.0)),
        message=" ".join(message_parts) if message_parts else None,
    )


@app.post("/api/leads/blocked-emails")
def api_save_leads_blocked_emails():
    access_redirect = require_page_access("leads")
    if access_redirect is not None:
        return access_redirect

    payload = request.get_json(silent=True) or {}
    normalized_value = save_blocked_lead_emails(payload.get("blockedEmails", ""))
    blocked_count = len(normalized_value.splitlines()) if normalized_value else 0
    return jsonify(
        {
            "ok": True,
            "blockedEmails": normalized_value,
            "blockedCount": blocked_count,
        }
    )


def normalize_planning_program(program: Any) -> List[Dict[str, str]]:
    if not isinstance(program, list):
        return []

    normalized_rows: List[Dict[str, str]] = []
    for item in program[:100]:
        if not isinstance(item, dict):
            continue
        start_time = str(item.get("startTime") or "").strip()[:5]
        end_time = str(item.get("endTime") or "").strip()[:5]
        if start_time and not re.fullmatch(r"(?:[01]\d|2[0-3]):[0-5]\d", start_time):
            start_time = ""
        if end_time and not re.fullmatch(r"(?:[01]\d|2[0-3]):[0-5]\d", end_time):
            end_time = ""
        activity = str(item.get("activity") or "").strip()[:240]
        details = str(item.get("details") or "").strip()[:500]
        if not (start_time or end_time or activity or details):
            continue
        normalized_rows.append(
            {
                "startTime": start_time,
                "endTime": end_time,
                "activity": activity,
                "details": details,
                "icon": infer_football_activity_icon(activity),
            }
        )
    return normalized_rows


def normalize_planning_date(value: Any) -> str:
    date_value = str(value or "").strip()[:10]
    return date_value if re.fullmatch(r"\d{4}-\d{2}-\d{2}", date_value) else ""


def normalize_planning_title(value: Any, fallback: str = "Planning") -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()[:160] or fallback


def normalize_planning_days(
    days: Any,
    fallback_date: Any = "",
    fallback_program: Any = None,
    fallback_title: Any = "Planning",
) -> List[Dict[str, Any]]:
    default_title = normalize_planning_title(fallback_title)
    normalized_days: List[Dict[str, Any]] = []
    if isinstance(days, list):
        for day in days[:31]:
            if not isinstance(day, dict):
                continue
            normalized_days.append(
                {
                    "title": normalize_planning_title(day.get("title"), default_title),
                    "date": normalize_planning_date(day.get("date") or day.get("planningDate")),
                    "program": normalize_planning_program(day.get("program")),
                }
            )
    if not normalized_days:
        normalized_days.append(
            {
                "title": default_title,
                "date": normalize_planning_date(fallback_date),
                "program": normalize_planning_program(fallback_program),
            }
        )
    return normalized_days


def normalize_planning_document(row: Optional[sqlite3.Row]) -> Dict[str, Any]:
    if row is None:
        blank_program = [{"startTime": "", "endTime": "", "activity": "", "details": "", "icon": "clock"}]
        return {
            "id": 0,
            "title": "Nieuwe planning",
            "planningDate": "",
            "location": "",
            "includeIcons": True,
            "program": blank_program,
            "days": [{"title": "Nieuwe planning", "date": "", "program": blank_program}],
            "programCount": 0,
            "createdAt": "",
            "updatedAt": "",
        }
    try:
        stored_program = json.loads(str(row["program_json"] or "[]"))
    except (TypeError, json.JSONDecodeError):
        stored_program = []
    stored_days = stored_program.get("days") if isinstance(stored_program, dict) else None
    legacy_program = stored_program if isinstance(stored_program, list) else []
    days = normalize_planning_days(stored_days, row["planning_date"], legacy_program, row["title"])
    first_day = days[0]
    return {
        "id": int(row["id"]),
        "title": normalize_planning_title(row["title"]),
        "planningDate": first_day["date"],
        "location": str(row["location"] or "").strip(),
        "includeIcons": bool(row["include_icons"]),
        "program": first_day["program"],
        "days": days,
        "programCount": sum(len(day["program"]) for day in days),
        "createdAt": str(row["created_at"] or "").strip(),
        "updatedAt": str(row["updated_at"] or "").strip(),
    }


def load_planning_documents() -> List[Dict[str, Any]]:
    with get_db_connection() as connection:
        rows = connection.execute(
            """
            SELECT id, title, planning_date, location, include_icons, program_json, created_at, updated_at
            FROM planning_documents
            ORDER BY COALESCE(NULLIF(planning_date, ''), updated_at) DESC, id DESC
            """
        ).fetchall()
    return [normalize_planning_document(row) for row in rows]


def load_planning_document(planning_id: int) -> Optional[Dict[str, Any]]:
    with get_db_connection() as connection:
        row = connection.execute(
            """
            SELECT id, title, planning_date, location, include_icons, program_json, created_at, updated_at
            FROM planning_documents
            WHERE id = ?
            """,
            (planning_id,),
        ).fetchone()
    return normalize_planning_document(row) if row is not None else None


def build_planning_document_from_form() -> Dict[str, Any]:
    starts = request.form.getlist("program_start")
    ends = request.form.getlist("program_end")
    activities = request.form.getlist("program_activity")
    details = request.form.getlist("program_details")
    row_count = min(100, max(len(starts), len(ends), len(activities), len(details)))
    legacy_program = normalize_planning_program(
        [
            {
                "startTime": starts[index] if index < len(starts) else "",
                "endTime": ends[index] if index < len(ends) else "",
                "activity": activities[index] if index < len(activities) else "",
                "details": details[index] if index < len(details) else "",
            }
            for index in range(row_count)
        ]
    )
    submitted_days: Any = None
    days_json = str(request.form.get("days_json", "") or "").strip()
    if days_json:
        try:
            submitted_days = json.loads(days_json)
        except json.JSONDecodeError:
            submitted_days = None
    title = normalize_planning_title(request.form.get("title"), "Nieuwe planning")
    days = normalize_planning_days(
        submitted_days,
        request.form.get("planning_date", ""),
        legacy_program,
        title,
    )
    return {
        "title": title,
        "planningDate": days[0]["date"],
        "location": str(request.form.get("location", "") or "").strip()[:160],
        "includeIcons": "1" in request.form.getlist("include_icons"),
        "program": days[0]["program"],
        "days": days,
    }


def save_planning_document(planning: Dict[str, Any], planning_id: Optional[int] = None) -> int:
    now = utcnow_iso()
    title = normalize_planning_title(planning.get("title"), "Nieuwe planning")
    days = normalize_planning_days(
        planning.get("days"),
        planning.get("planningDate"),
        planning.get("program"),
        title,
    )
    payload = (
        title,
        days[0]["date"],
        str(planning.get("location") or "").strip()[:160],
        1 if planning.get("includeIcons", True) else 0,
        json.dumps({"days": days}, ensure_ascii=False),
    )
    with get_db_connection() as connection:
        if planning_id is not None:
            connection.execute(
                """
                UPDATE planning_documents
                SET title = ?, planning_date = ?, location = ?, include_icons = ?, program_json = ?, updated_at = ?
                WHERE id = ?
                """,
                (*payload, now, planning_id),
            )
            return planning_id
        cursor = connection.execute(
            """
            INSERT INTO planning_documents (
                title, planning_date, location, include_icons, program_json, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (*payload, now, now),
        )
        return int(cursor.lastrowid)


def normalize_planning_export_payload(payload: Dict[str, Any]) -> Dict[str, Any]:
    title = normalize_planning_title(payload.get("title"))
    days = normalize_planning_days(
        payload.get("days"),
        payload.get("planningDate"),
        payload.get("program"),
        title,
    )
    return {
        "title": title,
        "planningDate": days[0]["date"],
        "location": str(payload.get("location") or "").strip()[:160],
        "includeIcons": bool(payload.get("includeIcons", True)),
        "program": days[0]["program"],
        "days": days,
    }


def planning_pdf_filename(planning: Dict[str, Any]) -> str:
    title = unicodedata.normalize("NFKD", str(planning.get("title") or "planning")).encode("ascii", "ignore").decode("ascii")
    safe_title = re.sub(r"[^a-zA-Z0-9]+", "-", title).strip("-").lower() or "planning"
    date_suffix = str(planning.get("planningDate") or "").strip()
    return f"{safe_title}{f'-{date_suffix}' if date_suffix else ''}.pdf"


def planning_png_filename(planning: Dict[str, Any]) -> str:
    return f"{planning_pdf_filename(planning)[:-4]}.png"


def create_planning_pdf(planning: Dict[str, Any]) -> bytes:
    try:
        from reportlab.lib import colors
        from reportlab.lib.utils import ImageReader
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.pdfmetrics import stringWidth
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.pdfgen import canvas
    except ImportError as exc:
        raise RuntimeError("De PDF-library ontbreekt. Installeer de packages uit requirements.txt.") from exc

    font_root = os.path.join(os.path.dirname(__file__), "static", "assets", "fonts")
    font_names = {
        "regular": "PoppinsPDF",
        "bold": "PoppinsPDF-Bold",
        "black": "PoppinsPDF-Black",
    }
    font_files = {
        "regular": "Poppins-Regular.ttf",
        "bold": "Poppins-Bold.ttf",
        "black": "Poppins-Black.ttf",
    }
    try:
        registered_fonts = set(pdfmetrics.getRegisteredFontNames())
        for key, font_name in font_names.items():
            if font_name not in registered_fonts:
                pdfmetrics.registerFont(TTFont(font_name, os.path.join(font_root, font_files[key])))
    except Exception as exc:
        raise RuntimeError("De Poppins-fontbestanden ontbreken of kunnen niet worden geladen.") from exc

    regular_font = font_names["regular"]
    bold_font = font_names["bold"]
    black_font = font_names["black"]
    pdf_white = colors.HexColor("#FFFFFF")
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=(FOOTBALL_DAYS_PDF_WIDTH, FOOTBALL_DAYS_PDF_HEIGHT))
    backgrounds = football_days_background_paths()

    def draw_background(page_index: int, shade_alpha: float = 0.34) -> None:
        background_path = backgrounds[page_index % len(backgrounds)] if backgrounds else ""
        if background_path and os.path.exists(background_path):
            pdf.drawImage(ImageReader(background_path), 0, 0, FOOTBALL_DAYS_PDF_WIDTH, FOOTBALL_DAYS_PDF_HEIGHT)
        else:
            pdf.setFillColor(colors.HexColor("#161616"))
            pdf.rect(0, 0, FOOTBALL_DAYS_PDF_WIDTH, FOOTBALL_DAYS_PDF_HEIGHT, fill=1, stroke=0)
        pdf.saveState()
        pdf.setFillColor(colors.Color(0, 0, 0, alpha=shade_alpha))
        pdf.rect(0, 0, FOOTBALL_DAYS_PDF_WIDTH, FOOTBALL_DAYS_PDF_HEIGHT, fill=1, stroke=0)
        pdf.restoreState()

    def trim_text(value: Any, max_width: float, font_name: str, font_size: float) -> str:
        text_value = str(value or "").strip()
        if stringWidth(text_value, font_name, font_size) <= max_width:
            return text_value
        suffix = "..."
        while text_value and stringWidth(f"{text_value}{suffix}", font_name, font_size) > max_width:
            text_value = text_value[:-1].rstrip()
        return f"{text_value}{suffix}" if text_value else suffix

    def split_lines(value: Any, max_width: float, font_name: str, font_size: float, max_lines: int = 2) -> List[str]:
        words = str(value or "").strip().split()
        if not words:
            return [""]
        lines: List[str] = []
        current = ""
        for word in words:
            candidate = f"{current} {word}".strip()
            if not current or stringWidth(candidate, font_name, font_size) <= max_width:
                current = candidate
            else:
                lines.append(current)
                current = word
        if current:
            lines.append(current)
        if len(lines) > max_lines:
            lines = lines[:max_lines]
            lines[-1] = trim_text(lines[-1], max_width, font_name, font_size)
        return [trim_text(line, max_width, font_name, font_size) for line in lines]

    def draw_icon(icon_key: str, cx: float, cy: float, size: float = 18) -> None:
        key = icon_key if icon_key in FOOTBALL_ACTIVITY_ICON_KEYS else "clock"
        scale = size / 21
        pdf.saveState()
        pdf.setStrokeColor(colors.Color(1, 1, 1, alpha=0.95))
        pdf.setFillColor(colors.Color(1, 1, 1, alpha=0.08))
        pdf.setLineWidth(1.5)
        pdf.circle(cx, cy, size * 0.72, stroke=1, fill=1)
        pdf.translate(cx, cy)
        pdf.scale(scale, scale)
        pdf.setLineWidth(1.9)
        if key == "football":
            pdf.circle(0, 0, 7.6, stroke=1, fill=0)
            pdf.circle(0, 0, 2.2, stroke=1, fill=0)
            for dx, dy in ((0, 7.3), (6.9, 2.3), (4.4, -6.1), (-4.4, -6.1), (-6.9, 2.3)):
                pdf.line(0, 0, dx, dy)
        elif key == "trophy":
            pdf.roundRect(-7, 1, 14, 10, 2, stroke=1, fill=0)
            pdf.line(0, 1, 0, -8)
            pdf.line(-8, -9, 8, -9)
        elif key == "clipboard":
            pdf.roundRect(-8, -10, 16, 20, 2, stroke=1, fill=0)
            pdf.line(-4, 3, 5, 3)
            pdf.line(-4, -3, 5, -3)
        elif key == "flame":
            path = pdf.beginPath()
            path.moveTo(0, -10)
            path.curveTo(-14, -4, -7, 7, -1, 12)
            path.curveTo(-1, 5, 6, 5, 3, 12)
            path.curveTo(12, 4, 13, -7, 0, -10)
            pdf.drawPath(path, stroke=1, fill=0)
        elif key == "camera":
            pdf.roundRect(-11, -7, 22, 16, 2, stroke=1, fill=0)
            pdf.circle(0, 1, 4.2, stroke=1, fill=0)
        elif key == "medical":
            pdf.rect(-3, -10, 6, 20, stroke=1, fill=0)
            pdf.rect(-10, -3, 20, 6, stroke=1, fill=0)
        elif key == "cones":
            pdf.line(-8, -10, -2, 10)
            pdf.line(8, -10, 2, 10)
            pdf.line(-11, -10, 11, -10)
            pdf.line(-5, 1, 5, 1)
        elif key == "utensils":
            pdf.line(-8, 10, -8, -9)
            pdf.line(-4, 10, -4, -9)
            pdf.line(7, 10, 7, -10)
        else:
            pdf.circle(0, 0, 7.6, stroke=1, fill=0)
            pdf.line(0, 0, 0, 6)
            pdf.line(0, 0, 6, -4)
        pdf.restoreState()

    def draw_sheet_title(title: Any, x: float, y: float, max_width: float, max_size: float = 48) -> None:
        title_text = str(title or "PLANNING").strip().upper()
        font_size = max_size
        while font_size > 22 and stringWidth(title_text, black_font, font_size) > max_width:
            font_size -= 1
        title_text = trim_text(title_text, max_width, black_font, font_size)
        pdf.setFillColor(pdf_white)
        pdf.setFont(black_font, font_size)
        pdf.drawString(x, y, title_text)

    days = normalize_planning_days(
        planning.get("days"),
        planning.get("planningDate"),
        planning.get("program"),
        planning.get("title"),
    )
    sheets: List[Dict[str, Any]] = []
    for day_index, day in enumerate(days, start=1):
        program = day["program"] or [
            {
                "startTime": "",
                "endTime": "",
                "activity": "Nog geen programmaonderdelen toegevoegd",
                "details": "",
                "icon": "clock",
            }
        ]
        sheets.append(
            {
                "dayIndex": day_index,
                "title": day["title"],
                "date": day["date"],
                "rows": program,
            }
        )

    include_icons = bool(planning.get("includeIcons", True))
    for page_index, sheet in enumerate(sheets, start=1):
        rows = sheet["rows"]
        draw_background(page_index)
        draw_sheet_title(sheet["title"], 99, 454, 816, 42)
        date_label = format_football_days_date(sheet["date"])
        pdf.setFont(bold_font, 14)
        pdf.setFillColor(pdf_white)
        pdf.drawCentredString(FOOTBALL_DAYS_PDF_WIDTH / 2, 425, trim_text(date_label, 850, bold_font, 14))
        location_label = str(planning.get("location") or "").strip()
        if location_label:
            pdf.setFont(regular_font, 8.5)
            pdf.setFillColor(colors.Color(1, 1, 1, alpha=0.9))
            pdf.drawCentredString(FOOTBALL_DAYS_PDF_WIDTH / 2, 408, trim_text(location_label, 850, regular_font, 8.5))

        table_x = 55
        table_top = 382
        table_width = 850
        header_height = 28
        icon_width = 48 if include_icons else 0
        time_width = 125
        activity_width = 280 if include_icons else 300
        details_width = table_width - icon_width - time_width - activity_width
        columns = ([icon_width] if include_icons else []) + [time_width, activity_width, details_width]
        headers = ([""] if include_icons else []) + ["Tijd", "Onderdeel", "Toelichting"]
        pdf.saveState()
        pdf.setFillColor(colors.Color(0, 0, 0, alpha=0.76))
        pdf.roundRect(table_x, table_top - header_height, table_width, header_height, 5, fill=1, stroke=0)
        cursor_x = table_x
        pdf.setFillColor(pdf_white)
        pdf.setFont(bold_font, 9)
        for column_index, header in enumerate(headers):
            if header:
                pdf.drawString(cursor_x + 10, table_top - 19, header.upper())
            cursor_x += columns[column_index]
        pdf.restoreState()

        available_row_height = table_top - header_height - 24
        row_height = min(38, available_row_height / max(1, len(rows)))
        row_scale = min(1.0, max(0.24, row_height / 38))
        row_gap = min(3, max(0.5, row_height * 0.08))
        horizontal_padding = max(4, 10 * row_scale)
        max_lines = 2 if row_height >= 28 else 1
        time_font_size = max(2.6, 10.5 * row_scale)
        activity_font_size = max(2.6, 10 * row_scale)
        details_font_size = max(2.5, 9 * row_scale)
        current_top = table_top - header_height - 4
        for row_index, item in enumerate(rows):
            pdf.saveState()
            pdf.setFillColor(colors.Color(0, 0, 0, alpha=0.58 if row_index % 2 == 0 else 0.47))
            pdf.setStrokeColor(colors.Color(1, 1, 1, alpha=0.16))
            pdf.roundRect(table_x, current_top - row_height + row_gap, table_width, row_height - row_gap, min(4, row_height / 4), fill=1, stroke=1)
            pdf.restoreState()
            cursor_x = table_x
            if include_icons:
                draw_icon(str(item.get("icon") or "clock"), cursor_x + (icon_width / 2), current_top - (row_height / 2), max(4, 15 * row_scale))
                cursor_x += icon_width
            time_label = " - ".join(
                part for part in (str(item.get("startTime") or "").strip(), str(item.get("endTime") or "").strip()) if part
            ) or "--:--"
            pdf.setFillColor(pdf_white)
            pdf.setFont(bold_font, time_font_size)
            pdf.drawString(cursor_x + horizontal_padding, current_top - (row_height / 2) - (time_font_size * 0.34), trim_text(time_label, time_width - (horizontal_padding * 2), bold_font, time_font_size))
            cursor_x += time_width
            activity_lines = split_lines(item.get("activity") or "-", activity_width - (horizontal_padding * 2), bold_font, activity_font_size, max_lines)
            pdf.setFont(bold_font, activity_font_size)
            activity_leading = activity_font_size * 1.18
            activity_y = current_top - (row_height / 2) + ((activity_leading / 2) if len(activity_lines) > 1 else -(activity_font_size * 0.34))
            for line in activity_lines:
                pdf.drawString(cursor_x + horizontal_padding, activity_y, line)
                activity_y -= activity_leading
            cursor_x += activity_width
            detail_lines = split_lines(item.get("details") or "-", details_width - (horizontal_padding * 2), regular_font, details_font_size, max_lines)
            pdf.setFont(regular_font, details_font_size)
            pdf.setFillColor(colors.Color(1, 1, 1, alpha=0.84))
            detail_leading = details_font_size * 1.18
            detail_y = current_top - (row_height / 2) + ((detail_leading / 2) if len(detail_lines) > 1 else -(details_font_size * 0.34))
            for line in detail_lines:
                pdf.drawString(cursor_x + horizontal_padding, detail_y, line)
                detail_y -= detail_leading
            current_top -= row_height

        pdf.showPage()

    pdf.save()
    buffer.seek(0)
    return buffer.read()


def create_planning_png(planning: Dict[str, Any]) -> bytes:
    try:
        from PIL import Image, ImageDraw, ImageFont
    except ImportError as exc:
        raise RuntimeError("De afbeeldingsbibliotheek ontbreekt. Installeer de packages uit requirements.txt.") from exc

    scale = 2
    page_width = FOOTBALL_DAYS_PDF_WIDTH * scale
    page_height = FOOTBALL_DAYS_PDF_HEIGHT * scale
    font_root = os.path.join(os.path.dirname(__file__), "static", "assets", "fonts")
    font_files = {
        "regular": "Poppins-Regular.ttf",
        "bold": "Poppins-Bold.ttf",
        "black": "Poppins-Black.ttf",
    }
    font_cache: Dict[Tuple[str, int], Any] = {}

    def px(value: float) -> int:
        return round(value * scale)

    def get_font(weight: str, size: float) -> Any:
        cache_key = (weight, px(size))
        if cache_key not in font_cache:
            try:
                font_cache[cache_key] = ImageFont.truetype(
                    os.path.join(font_root, font_files[weight]),
                    px(size),
                )
            except (OSError, KeyError) as exc:
                raise RuntimeError("De Poppins-fontbestanden ontbreken of kunnen niet worden geladen.") from exc
        return font_cache[cache_key]

    def text_width(draw: Any, value: Any, font: Any) -> float:
        return float(draw.textlength(str(value or ""), font=font))

    def trim_text(draw: Any, value: Any, max_width: float, font: Any) -> str:
        text_value = str(value or "").strip()
        if text_width(draw, text_value, font) <= max_width:
            return text_value
        suffix = "..."
        while text_value and text_width(draw, f"{text_value}{suffix}", font) > max_width:
            text_value = text_value[:-1].rstrip()
        return f"{text_value}{suffix}" if text_value else suffix

    def split_lines(draw: Any, value: Any, max_width: float, font: Any, max_lines: int = 2) -> List[str]:
        words = str(value or "").strip().split()
        if not words:
            return [""]
        lines: List[str] = []
        current = ""
        for word in words:
            candidate = f"{current} {word}".strip()
            if not current or text_width(draw, candidate, font) <= max_width:
                current = candidate
            else:
                lines.append(current)
                current = word
        if current:
            lines.append(current)
        lines = lines[:max_lines]
        return [trim_text(draw, line, max_width, font) for line in lines]

    def draw_centered_text(draw: Any, value: Any, center_x: float, top_y: float, font: Any, fill: Any) -> None:
        text_value = str(value or "")
        draw.text((px(center_x) - (text_width(draw, text_value, font) / 2), px(top_y)), text_value, font=font, fill=fill)

    def draw_vertical_centered_text(draw: Any, value: Any, x: float, center_y: float, font: Any, fill: Any) -> None:
        text_value = str(value or "")
        bounds = draw.textbbox((0, 0), text_value, font=font)
        text_height = bounds[3] - bounds[1]
        draw.text((px(x), px(center_y) - (text_height / 2) - bounds[1]), text_value, font=font, fill=fill)

    def draw_icon(draw: Any, icon_key: str, center_x: float, center_y: float, size: float = 17) -> None:
        key = icon_key if icon_key in FOOTBALL_ACTIVITY_ICON_KEYS else "clock"
        cx = px(center_x)
        cy = px(center_y)
        radius = px(size * 0.72)
        line_width = max(2, px(1.5))
        stroke = (255, 255, 255, 242)
        draw.ellipse((cx - radius, cy - radius, cx + radius, cy + radius), outline=stroke, width=line_width)
        unit = scale * (size / 21)

        def point(x_value: float, y_value: float) -> Tuple[int, int]:
            return (round(cx + (x_value * unit)), round(cy - (y_value * unit)))

        icon_width = max(2, round(1.8 * unit))
        if key == "football":
            icon_radius = round(7.6 * unit)
            inner_radius = round(2.2 * unit)
            draw.ellipse((cx - icon_radius, cy - icon_radius, cx + icon_radius, cy + icon_radius), outline=stroke, width=icon_width)
            draw.ellipse((cx - inner_radius, cy - inner_radius, cx + inner_radius, cy + inner_radius), outline=stroke, width=icon_width)
            for end_x, end_y in ((0, 7.3), (6.9, 2.3), (4.4, -6.1), (-4.4, -6.1), (-6.9, 2.3)):
                draw.line((point(0, 0), point(end_x, end_y)), fill=stroke, width=icon_width)
        elif key == "trophy":
            draw.rounded_rectangle((*point(-7, 11), *point(7, 1)), radius=max(1, round(2 * unit)), outline=stroke, width=icon_width)
            draw.line((point(0, 1), point(0, -8)), fill=stroke, width=icon_width)
            draw.line((point(-8, -9), point(8, -9)), fill=stroke, width=icon_width)
        elif key == "clipboard":
            draw.rounded_rectangle((*point(-8, 10), *point(8, -10)), radius=max(1, round(2 * unit)), outline=stroke, width=icon_width)
            draw.line((point(-4, 3), point(5, 3)), fill=stroke, width=icon_width)
            draw.line((point(-4, -3), point(5, -3)), fill=stroke, width=icon_width)
        elif key == "flame":
            draw.line(
                [point(0, -10), point(-7, -5), point(-6, 4), point(-1, 12), point(0, 4), point(3, 12), point(9, 3), point(7, -6), point(0, -10)],
                fill=stroke,
                width=icon_width,
                joint="curve",
            )
        elif key == "camera":
            draw.rounded_rectangle((*point(-11, 9), *point(11, -7)), radius=max(1, round(2 * unit)), outline=stroke, width=icon_width)
            camera_radius = round(4.2 * unit)
            camera_center_y = point(0, 1)[1]
            draw.ellipse((cx - camera_radius, camera_center_y - camera_radius, cx + camera_radius, camera_center_y + camera_radius), outline=stroke, width=icon_width)
        elif key == "medical":
            draw.rectangle((*point(-3, 10), *point(3, -10)), outline=stroke, width=icon_width)
            draw.rectangle((*point(-10, 3), *point(10, -3)), outline=stroke, width=icon_width)
        elif key == "cones":
            draw.line((point(-8, -10), point(-2, 10)), fill=stroke, width=icon_width)
            draw.line((point(8, -10), point(2, 10)), fill=stroke, width=icon_width)
            draw.line((point(-11, -10), point(11, -10)), fill=stroke, width=icon_width)
            draw.line((point(-5, 1), point(5, 1)), fill=stroke, width=icon_width)
        elif key == "utensils":
            draw.line((point(-8, 10), point(-8, -9)), fill=stroke, width=icon_width)
            draw.line((point(-4, 10), point(-4, -9)), fill=stroke, width=icon_width)
            draw.line((point(7, 10), point(7, -10)), fill=stroke, width=icon_width)
        else:
            icon_radius = round(7.6 * unit)
            draw.ellipse((cx - icon_radius, cy - icon_radius, cx + icon_radius, cy + icon_radius), outline=stroke, width=icon_width)
            draw.line((point(0, 0), point(0, 6)), fill=stroke, width=icon_width)
            draw.line((point(0, 0), point(6, -4)), fill=stroke, width=icon_width)

    backgrounds = football_days_background_paths()
    resampling = getattr(Image, "Resampling", Image).LANCZOS
    logo_path = os.path.join(os.path.dirname(__file__), "static", "assets", "hws-logo.png")
    logo_image = None
    if os.path.exists(logo_path):
        with Image.open(logo_path) as source_logo:
            logo_image = source_logo.convert("RGBA")
        logo_bounds = logo_image.getchannel("A").getbbox()
        if logo_bounds:
            logo_image = logo_image.crop(logo_bounds)
            logo_height = px(68)
            logo_width = round(logo_image.width * (logo_height / logo_image.height))
            logo_image = logo_image.resize((logo_width, logo_height), resampling)
        else:
            logo_image = None

    def create_background(page_index: int) -> Any:
        background_path = backgrounds[page_index % len(backgrounds)] if backgrounds else ""
        if background_path and os.path.exists(background_path):
            with Image.open(background_path) as source_image:
                page = source_image.convert("RGBA").resize((page_width, page_height), resampling)
        else:
            page = Image.new("RGBA", (page_width, page_height), (22, 22, 22, 255))
        shade = Image.new("RGBA", page.size, (0, 0, 0, 87))
        return Image.alpha_composite(page, shade)

    def draw_translucent_rounded_rectangle(
        page: Any,
        box: Tuple[int, int, int, int],
        radius: int,
        fill: Tuple[int, int, int, int],
        outline: Optional[Tuple[int, int, int, int]] = None,
        width: int = 1,
    ) -> None:
        overlay = Image.new("RGBA", page.size, (0, 0, 0, 0))
        overlay_draw = ImageDraw.Draw(overlay)
        overlay_draw.rounded_rectangle(box, radius=radius, fill=fill, outline=outline, width=width)
        page.alpha_composite(overlay)

    days = normalize_planning_days(
        planning.get("days"),
        planning.get("planningDate"),
        planning.get("program"),
        planning.get("title"),
    )
    sheets: List[Dict[str, Any]] = []
    for day_index, day in enumerate(days, start=1):
        program = day["program"] or [
            {
                "startTime": "",
                "endTime": "",
                "activity": "Nog geen programmaonderdelen toegevoegd",
                "details": "",
                "icon": "clock",
            }
        ]
        sheets.append(
            {
                "dayIndex": day_index,
                "title": day["title"],
                "date": day["date"],
                "rows": program,
            }
        )

    include_icons = bool(planning.get("includeIcons", True))
    rendered_pages: List[Any] = []

    for page_index, sheet in enumerate(sheets, start=1):
        rows = sheet["rows"]
        page = create_background(page_index)
        if logo_image is not None:
            # The background artwork already contains a subdued logo. Redrawing the
            # source logo after the dark overlay keeps it crisp in the PNG export.
            page.alpha_composite(logo_image, (px(28), px(20)))
        draw = ImageDraw.Draw(page, "RGBA")
        title = str(sheet["title"] or planning.get("title") or "Planning").strip().upper()
        title_size = 42
        title_font = get_font("black", title_size)
        title_left = 99
        title_right = 915
        title_max_width = px(title_right - title_left)
        while title_size > 22 and text_width(draw, title, title_font) > title_max_width:
            title_size -= 1
            title_font = get_font("black", title_size)
        title = trim_text(draw, title, title_max_width, title_font)
        draw_vertical_centered_text(draw, title, title_left, 54, title_font, (255, 255, 255, 255))

        date_label = format_football_days_date(sheet["date"])
        date_font = get_font("bold", 14)
        date_label = trim_text(draw, date_label, px(850), date_font)
        draw_centered_text(draw, date_label, FOOTBALL_DAYS_PDF_WIDTH / 2, 101, date_font, (255, 255, 255, 255))
        location_label = str(planning.get("location") or "").strip()
        if location_label:
            location_font = get_font("regular", 8.5)
            location_label = trim_text(draw, location_label, px(850), location_font)
            draw_centered_text(draw, location_label, FOOTBALL_DAYS_PDF_WIDTH / 2, 126, location_font, (255, 255, 255, 230))

        table_x = 55
        table_top = 150
        table_width = 850
        header_height = 28
        icon_width = 48 if include_icons else 0
        time_width = 125
        activity_width = 280 if include_icons else 300
        details_width = table_width - icon_width - time_width - activity_width
        columns = ([icon_width] if include_icons else []) + [time_width, activity_width, details_width]
        headers = ([""] if include_icons else []) + ["Tijd", "Onderdeel", "Toelichting"]
        draw_translucent_rounded_rectangle(
            page,
            (px(table_x), px(table_top), px(table_x + table_width), px(table_top + header_height)),
            radius=px(5),
            fill=(0, 0, 0, 194),
        )
        draw = ImageDraw.Draw(page, "RGBA")
        cursor_x = table_x
        header_font = get_font("bold", 9)
        for column_index, header in enumerate(headers):
            if header:
                draw_vertical_centered_text(draw, header.upper(), cursor_x + 10, table_top + (header_height / 2), header_font, (255, 255, 255, 255))
            cursor_x += columns[column_index]

        available_row_height = 522 - table_top - header_height - 4
        row_height = min(38, available_row_height / max(1, len(rows)))
        row_scale = min(1.0, max(0.24, row_height / 38))
        row_gap = min(3, max(0.5, row_height * 0.08))
        horizontal_padding = max(4, 10 * row_scale)
        max_lines = 2 if row_height >= 28 else 1
        time_font_size = max(2.6, 10.5 * row_scale)
        activity_font_size = max(2.6, 10 * row_scale)
        details_font_size = max(2.5, 9 * row_scale)
        current_top = table_top + header_height + 4
        for row_index, item in enumerate(rows):
            row_bottom = current_top + row_height - row_gap
            draw_translucent_rounded_rectangle(
                page,
                (px(table_x), px(current_top), px(table_x + table_width), px(row_bottom)),
                radius=px(min(4, row_height / 4)),
                fill=(0, 0, 0, 148 if row_index % 2 == 0 else 120),
                outline=(255, 255, 255, 42),
                width=max(1, px(0.7)),
            )
            draw = ImageDraw.Draw(page, "RGBA")
            center_y = current_top + ((row_height - row_gap) / 2)
            cursor_x = table_x
            if include_icons:
                draw_icon(draw, str(item.get("icon") or infer_football_activity_icon(str(item.get("activity") or ""))), cursor_x + (icon_width / 2), center_y, max(4, 15 * row_scale))
                cursor_x += icon_width

            time_label = " - ".join(
                part for part in (str(item.get("startTime") or "").strip(), str(item.get("endTime") or "").strip()) if part
            ) or "--:--"
            time_font = get_font("bold", time_font_size)
            time_label = trim_text(draw, time_label, px(time_width - (horizontal_padding * 2)), time_font)
            draw_vertical_centered_text(draw, time_label, cursor_x + horizontal_padding, center_y, time_font, (255, 255, 255, 255))
            cursor_x += time_width

            activity_font = get_font("bold", activity_font_size)
            activity_lines = split_lines(draw, item.get("activity") or "-", px(activity_width - (horizontal_padding * 2)), activity_font, max_lines)
            activity_leading = activity_font_size * 1.18
            activity_start = center_y - ((activity_leading / 2) * (len(activity_lines) - 1))
            for line_index, line in enumerate(activity_lines):
                draw_vertical_centered_text(draw, line, cursor_x + horizontal_padding, activity_start + (line_index * activity_leading), activity_font, (255, 255, 255, 255))
            cursor_x += activity_width

            details_font = get_font("regular", details_font_size)
            details_lines = split_lines(draw, item.get("details") or "-", px(details_width - (horizontal_padding * 2)), details_font, max_lines)
            details_leading = details_font_size * 1.18
            details_start = center_y - ((details_leading / 2) * (len(details_lines) - 1))
            for line_index, line in enumerate(details_lines):
                draw_vertical_centered_text(draw, line, cursor_x + horizontal_padding, details_start + (line_index * details_leading), details_font, (255, 255, 255, 218))
            current_top += row_height

        rendered_pages.append(page.convert("RGB"))

    output_image = Image.new("RGB", (page_width, page_height * len(rendered_pages)), (22, 22, 22))
    for page_index, page in enumerate(rendered_pages):
        output_image.paste(page, (0, page_index * page_height))
    buffer = BytesIO()
    output_image.save(buffer, format="PNG", optimize=True)
    buffer.seek(0)
    return buffer.read()


@app.route("/planning", methods=["GET", "POST"])
def planning_page() -> str:
    access_redirect = require_page_access("planning")
    if access_redirect is not None:
        return access_redirect

    if request.method == "POST":
        planning_id = save_planning_document(build_planning_document_from_form())
        return redirect(f"/planning/{planning_id}?success=Planning opgeslagen.")

    return render_template(
        "planning.html",
        active_page="planning",
        page_mode="overview",
        plannings=load_planning_documents(),
        planning=normalize_planning_document(None),
        success=request.args.get("success", "").strip(),
    )


@app.route("/planning/<int:planning_id>", methods=["GET", "POST"])
def planning_edit_page(planning_id: int) -> str:
    access_redirect = require_page_access("planning")
    if access_redirect is not None:
        return access_redirect

    planning = load_planning_document(planning_id)
    if planning is None:
        return redirect("/planning")
    if request.method == "POST":
        save_planning_document(build_planning_document_from_form(), planning_id)
        return redirect(f"/planning/{planning_id}?success=Planning opgeslagen.")
    for day in planning["days"]:
        if not day["program"]:
            day["program"] = [{"startTime": "", "endTime": "", "activity": "", "details": "", "icon": "clock"}]
    planning["program"] = planning["days"][0]["program"]
    return render_template(
        "planning.html",
        active_page="planning",
        page_mode="edit",
        plannings=[],
        planning=planning,
        success=request.args.get("success", "").strip(),
    )


@app.post("/api/planning/export-pdf")
def api_planning_export_pdf():
    access_redirect = require_page_access("planning")
    if access_redirect is not None:
        return access_redirect

    planning = normalize_planning_export_payload(request.get_json(silent=True) or {})
    try:
        pdf_bytes = create_planning_pdf(planning)
    except RuntimeError as exc:
        return jsonify({"error": str(exc)}), 500
    return (
        pdf_bytes,
        200,
        {
            "Content-Type": "application/pdf",
            "Content-Disposition": f'attachment; filename="{planning_pdf_filename(planning)}"',
            "Cache-Control": "no-store",
        },
    )


@app.post("/api/planning/export-png")
def api_planning_export_png():
    access_redirect = require_page_access("planning")
    if access_redirect is not None:
        return access_redirect

    planning = normalize_planning_export_payload(request.get_json(silent=True) or {})
    try:
        png_bytes = create_planning_png(planning)
    except RuntimeError as exc:
        return jsonify({"error": str(exc)}), 500
    return (
        png_bytes,
        200,
        {
            "Content-Type": "image/png",
            "Content-Disposition": f'attachment; filename="{planning_png_filename(planning)}"',
            "Cache-Control": "no-store",
        },
    )


@app.get("/management")
def management_page() -> str:
    access_redirect = require_page_access("management")
    if access_redirect is not None:
        return access_redirect

    return render_template("management.html", active_page="management")


def get_current_request_base_url() -> str:
    request_url = str(request.url or "").strip()
    request_path = str(request.path or "").strip()
    if request_url and request_path and request_path in request_url:
        return request_url.split(request_path, 1)[0].rstrip("/")
    scheme = "https" if is_request_secure() else "http"
    host = str(request.headers.get("Host", "") or "").strip()
    if not host:
        return ""
    return f"{scheme}://{host}"


@app.route("/management/api", methods=["GET", "POST"])
def api_management_page() -> str:
    access_redirect = require_page_access("api")
    if access_redirect is not None:
        return access_redirect

    if request.method == "POST":
        action = str(request.form.get("action", "") or "").strip()
        if action != "rotate":
            return "Ongeldige API-actie.", 400
        rotate_agenda_api_credential()
        return redirect(
            url_for(
                "api_management_page",
                success="De API-sleutel is vernieuwd. Werk de koppeling in het andere project bij.",
            )
        )

    credential = get_agenda_api_credential() or rotate_agenda_api_credential()
    api_token = credential["token"]
    base_url = get_current_request_base_url()
    events_url = f"{base_url}/api/v1/agenda/events"
    calendar_feed_url = f"{base_url}/api/v1/agenda/calendar.ics?token={api_token}"
    integration_environment = "\n".join(
        [
            f"HWS_AGENDA_API_URL={events_url}",
            f"HWS_AGENDA_API_TOKEN={api_token}",
            f"HWS_AGENDA_ICS_URL={calendar_feed_url}",
        ]
    )
    python_example = "\n".join(
        [
            "import os",
            "import requests",
            "",
            'response = requests.get(',
            '    os.environ["HWS_AGENDA_API_URL"],',
            '    headers={"Authorization": f\'Bearer {os.environ["HWS_AGENDA_API_TOKEN"]}\'},',
            '    params={"start": "2026-01-01", "include_day_plans": "1"},',
            "    timeout=20,",
            ")",
            "response.raise_for_status()",
            'appointments = response.json()["events"]',
        ]
    )
    javascript_example = "\n".join(
        [
            "const url = new URL(process.env.HWS_AGENDA_API_URL);",
            'url.searchParams.set("start", "2026-01-01");',
            "",
            "const response = await fetch(url, {",
            "  headers: {",
            "    Authorization: `Bearer ${process.env.HWS_AGENDA_API_TOKEN}`,",
            "  },",
            "});",
            "",
            "if (!response.ok) throw new Error(`HWS API: ${response.status}`);",
            "const { events } = await response.json();",
        ]
    )
    return render_template(
        "api.html",
        active_page="api",
        api_token=api_token,
        events_url=events_url,
        calendar_feed_url=calendar_feed_url,
        integration_environment=integration_environment,
        python_example=python_example,
        javascript_example=javascript_example,
        api_created_at=credential.get("createdAt", ""),
        api_last_used_at=credential.get("lastUsedAt", ""),
        api_allowed_origins=sorted(get_agenda_api_allowed_origins()),
        success=request.args.get("success", "").strip(),
    )


@app.route("/management/begroting", methods=["GET", "POST"])
@app.route("/begroting", methods=["GET", "POST"])
def budget_page() -> str:
    access_redirect = require_page_access("begroting")
    if access_redirect is not None:
        return access_redirect

    all_trainings = load_agenda_trainings()
    season_options = build_trainer_fee_season_options(all_trainings)
    available_seasons = {option["value"] for option in season_options}
    season_source = request.form if request.method == "POST" else request.args
    default_budget_season = season_options[0]["value"] if season_options else str(get_season_start_year_for_date(date.today()))
    selected_season = season_source.get("season", "").strip() or default_budget_season
    if selected_season not in available_seasons:
        selected_season = season_options[0]["value"] if season_options else str(get_season_start_year_for_date(date.today()))

    try:
        season_start_year = int(selected_season)
    except ValueError:
        season_start_year = get_season_start_year_for_date(date.today())

    if request.method == "POST":
        rows = parse_budget_lines_from_form(request.form, season_start_year)
        save_budget_lines(season_start_year, rows)
        action = request.form.get("action", "save").strip()
        if action == "save_and_forward":
            selected_indexes = {
                parse_non_negative_int(value)
                for value in request.form.getlist("forward_line")
                if str(value).strip().isdigit()
            }
            added, skipped = forward_budget_rows_to_trainer_profiles(rows, selected_indexes, season_start_year)
            message = f"Begroting opgeslagen en {added} planningsregel{'s' if added != 1 else ''} doorgestuurd."
            if skipped:
                message += f" {skipped} geselecteerde regel{'s' if skipped != 1 else ''} zonder trainer overgeslagen."
            return redirect(url_for("budget_page", season=season_start_year, success=message))
        return redirect(url_for("budget_page", season=season_start_year, success="Begroting opgeslagen."))

    return render_template(
        "begroting.html",
        active_page="begroting",
        season_options=season_options,
        budget_summary=build_budget_summary(season_start_year),
        trainer_options=build_agenda_trainer_options(),
        success=request.args.get("success", "").strip(),
    )


@app.route("/materialen", methods=["GET", "POST"])
def materialen_page() -> str:
    access_redirect = require_page_access("materialen")
    if access_redirect is not None:
        return access_redirect

    saved = False
    if request.method == "POST":
        save_materials_inventory(build_materials_inventory_from_form())
        saved = True

    inventory = load_materials_inventory()
    if not inventory["materials"]:
        inventory["materials"] = [
            {"id": "", "key": "material-new-1", "name": "Ballen", "totalCount": 0, "allocatedCount": 0, "availableCount": 0},
            {"id": "", "key": "material-new-2", "name": "Hoedjes", "totalCount": 0, "allocatedCount": 0, "availableCount": 0},
            {"id": "", "key": "material-new-3", "name": "Hesjes", "totalCount": 0, "allocatedCount": 0, "availableCount": 0},
            {"id": "", "key": "material-new-4", "name": "Pionnen", "totalCount": 0, "allocatedCount": 0, "availableCount": 0},
        ]

    return render_template(
        "materialen.html",
        active_page="materialen",
        inventory=inventory,
        saved=saved,
    )


@app.get("/materialen/clubs/<int:club_id>/export-pdf")
def materialen_club_export_pdf(club_id: int):
    access_redirect = require_page_access("materialen")
    if access_redirect is not None:
        return access_redirect

    inventory = load_materials_inventory()
    club = next((item for item in inventory["clubs"] if item["id"] == club_id), None)
    if club is None:
        return redirect(url_for("materialen_page"))

    try:
        pdf_bytes = create_materials_club_pdf(club, inventory["materials"])
    except RuntimeError as exc:
        return jsonify({"error": str(exc)}), 500

    club_slug = slugify_value(club.get("name") or "club")
    return (
        pdf_bytes,
        200,
        {
            "Content-Type": "application/pdf",
            "Content-Disposition": f'inline; filename="materialenkrat-{club_slug}.pdf"',
            "Cache-Control": "no-store",
        },
    )


@app.get("/materialen/clubs/export-pdf")
def materialen_all_clubs_export_pdf():
    access_redirect = require_page_access("materialen")
    if access_redirect is not None:
        return access_redirect

    inventory = load_materials_inventory()
    if not inventory["clubs"]:
        return redirect(url_for("materialen_page"))

    requested_club_ids = request.args.getlist("club_id")
    selected_clubs = inventory["clubs"]
    if requested_club_ids:
        selected_club_ids = {
            int(club_id)
            for club_id in requested_club_ids
            if str(club_id).strip().isdigit() and int(club_id) > 0
        }
        selected_clubs = [club for club in inventory["clubs"] if club["id"] in selected_club_ids]
        if not selected_clubs:
            return redirect(url_for("materialen_page"))

    try:
        pdf_bytes = create_materials_all_clubs_pdf(selected_clubs, inventory["materials"])
    except RuntimeError as exc:
        return jsonify({"error": str(exc)}), 500

    return (
        pdf_bytes,
        200,
        {
            "Content-Type": "application/pdf",
            "Content-Disposition": (
                'attachment; filename="materialenkratten-geselecteerde-clubs.pdf"'
                if requested_club_ids
                else 'attachment; filename="materialenkratten-alle-clubs.pdf"'
            ),
            "Cache-Control": "no-store",
        },
    )


@app.get("/aanmeldingen/<path:product_key>")
def registrations_detail_page(product_key: str) -> str:
    access_redirect = require_page_access("orders")
    if access_redirect is not None:
        return access_redirect

    normalized_product_key = str(product_key or "").strip()
    if not normalized_product_key:
        return redirect(build_registrations_page_url())

    products_payload = fetch_catalog_products()
    orders_payload = fetch_ecwid_orders()
    selected_product = build_registration_product_detail(
        products_payload.get("items", []),
        orders_payload.get("items", []),
        normalized_product_key,
    )

    if selected_product is None:
        abort(404)

    product_message = products_payload.get("message")
    order_message = orders_payload.get("message")
    message_parts = []
    for message in (product_message, order_message):
        if message and message not in message_parts:
            message_parts.append(message)

    return render_template(
        "registration_detail.html",
        active_page="registrations",
        selected_product=selected_product,
        email_templates=load_registration_event_email_templates(normalized_product_key),
        refresh_url=build_registration_detail_url(normalized_product_key),
        team_assignment_export_url=url_for("export_registration_team_assignment", product_key=normalized_product_key),
        back_url=build_registrations_page_url(),
        last_updated=format_cache_timestamp(orders_payload.get("cachedAt", 0.0)),
        message=" ".join(message_parts) if message_parts else None,
    )


@app.post("/aanmeldingen/<path:product_key>/teamindeling-export")
def export_registration_team_assignment(product_key: str):
    access_redirect = require_page_access("orders")
    if access_redirect is not None:
        return access_redirect

    normalized_product_key = str(product_key or "").strip()
    if not normalized_product_key:
        return redirect(build_registrations_page_url())

    try:
        group_count = int(str(request.form.get("group_count", "") or "0").strip())
    except ValueError:
        group_count = 0
    if group_count < 1:
        group_count = 1
    group_count = min(group_count, 100)

    products_payload = fetch_catalog_products()
    orders_payload = fetch_ecwid_orders()
    selected_product = build_registration_product_detail(
        products_payload.get("items", []),
        orders_payload.get("items", []),
        normalized_product_key,
    )
    if selected_product is None:
        abort(404)

    participants = build_registration_participant_rows(selected_product)
    if not participants:
        return redirect(build_registration_detail_url(normalized_product_key))

    workbook_buffer = build_registration_team_assignment_workbook(selected_product, group_count)
    safe_product_name = re.sub(r"[^a-zA-Z0-9_-]+", "-", str(selected_product.get("name", "") or "aanmeldingen")).strip("-")
    filename = f"teamindeling-{safe_product_name or 'aanmeldingen'}-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
    return (
        workbook_buffer.getvalue(),
        200,
        {
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


@app.post("/api/registrations/email-status")
def api_update_registration_email_status():
    access_redirect = require_page_access("orders")
    if access_redirect is not None:
        return access_redirect

    payload = request.get_json(silent=True) or {}
    product_key = str(payload.get("productKey", "") or "").strip()
    order_ids = normalize_registration_email_status_order_ids(payload.get("orderIds", []))
    emailed = payload.get("emailed")

    if not product_key:
        return jsonify({"error": "Product ontbreekt."}), 400
    if not order_ids:
        return jsonify({"error": "Geen bestellingen geselecteerd."}), 400
    if len(order_ids) > 500:
        return jsonify({"error": "Te veel bestellingen in één verzoek."}), 400
    if not isinstance(emailed, bool):
        return jsonify({"error": "Ongeldige e-mailstatus."}), 400

    ecwid_updated_order_ids: List[str] = []
    if emailed:
        try:
            for order_id in order_ids:
                if update_ecwid_order_to_processing(order_id):
                    ecwid_updated_order_ids.append(order_id)
        except RuntimeError as exc:
            return jsonify({"error": str(exc)}), 502

    updated_order_ids = set_registration_orders_emailed(product_key, order_ids, emailed)
    return jsonify(
        {
            "ok": True,
            "productKey": product_key,
            "orderIds": updated_order_ids,
            "emailed": emailed,
            "ecwidUpdatedOrderIds": ecwid_updated_order_ids,
        }
    )


@app.post("/api/registrations/event-email-settings")
def api_save_registration_event_email_settings():
    access_redirect = require_page_access("orders")
    if access_redirect is not None:
        return access_redirect

    payload = request.get_json(silent=True) or {}
    product_key = str(payload.get("productKey", "") or "").strip()
    product_name = str(payload.get("productName", "") or "").strip()
    if not product_key:
        return jsonify({"error": "Product ontbreekt."}), 400

    try:
        settings_row = save_registration_event_email_settings(
            product_key=product_key,
            product_name=product_name,
            event_date=payload.get("eventDate", ""),
            event_date_2=payload.get("eventDate2", "") if payload.get("useSecondEventDate") else "",
            email_subject=payload.get("emailSubject", ""),
            email_body=payload.get("emailBody", ""),
        )
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    return jsonify(
        {
            "ok": True,
            "settings": settings_row,
            "templates": load_registration_event_email_templates(product_key),
            "message": "Mailinstellingen opgeslagen.",
        }
    )


@app.post("/api/registrations/send-event-email")
def api_send_registration_event_email():
    access_redirect = require_page_access("orders")
    if access_redirect is not None:
        return access_redirect

    if not registration_auto_email_is_configured():
        return jsonify({"error": "Automatische inschrijvingsmail is nog niet volledig geconfigureerd."}), 400

    payload = request.get_json(silent=True) or {}
    product_key = str(payload.get("productKey", "") or "").strip()
    if not product_key:
        return jsonify({"error": "Product ontbreekt."}), 400

    products_payload = fetch_catalog_products()
    orders_payload = fetch_ecwid_orders(force_refresh=True)
    selected_product = build_registration_product_detail(
        products_payload.get("items", []),
        orders_payload.get("items", []),
        product_key,
    )
    if selected_product is None:
        return jsonify({"error": "Product niet gevonden."}), 404

    order_ids = normalize_registration_email_status_order_ids(
        [order.get("id", "") for order in selected_product.get("orders", [])]
    )
    if len(order_ids) > 500:
        return jsonify({"error": "Te veel bestellingen in één verzoek."}), 400

    send_result = send_registration_product_emails(product_key, orders_payload.get("items", []))
    sent_count = len(send_result["sentOrderIds"])
    failed_count = len(send_result["failedOrderIds"])
    skipped_count = len(send_result["skippedOrderIds"])
    if failed_count:
        message = f"{sent_count} mails verstuurd, {failed_count} mislukt en {skipped_count} overgeslagen."
    elif sent_count:
        message = f"{sent_count} standaardmail(s) verstuurd en op gemaild gezet."
    else:
        message = "Geen mails verstuurd. Alle passende bestellingen waren al gemaild, onbetaald of hadden geen e-mailadres."

    return jsonify(
        {
            "ok": failed_count == 0,
            "productKey": product_key,
            "sentOrderIds": send_result["sentOrderIds"],
            "skippedOrderIds": send_result["skippedOrderIds"],
            "failedOrderIds": send_result["failedOrderIds"],
            "ecwidUpdatedOrderIds": send_result["ecwidUpdatedOrderIds"],
            "message": message,
        }
    )


@app.post("/api/registrations/sync-emailed-orders")
def api_sync_emailed_registration_orders():
    access_redirect = require_page_access("orders")
    if access_redirect is not None:
        return access_redirect

    config = get_config()
    if not config["store_id"] or not config["secret_token"]:
        return jsonify({"error": "Live Ecwid-koppeling staat nog niet aan."}), 400

    order_ids = load_all_registration_emailed_order_ids()
    if not order_ids:
        return jsonify(
            {
                "ok": True,
                "orderIds": [],
                "syncedOrderIds": [],
                "failedOrderIds": [],
                "message": "Er staan nog geen bestellingen op gemaild om te synchroniseren.",
            }
        )

    sync_result = sync_emailed_registration_orders_to_ecwid(order_ids)
    status_code = 200 if not sync_result["failedOrderIds"] else 502
    message = (
        f"{len(sync_result['syncedOrderIds'])} gemailde bestellingen zijn naar Ecwid gesynchroniseerd."
        if not sync_result["failedOrderIds"]
        else (
            f"{len(sync_result['syncedOrderIds'])} bestellingen gesynchroniseerd, "
            f"{len(sync_result['failedOrderIds'])} niet bijgewerkt."
        )
    )
    return (
        jsonify(
            {
                "ok": not sync_result["failedOrderIds"],
                "orderIds": sync_result["orderIds"],
                "syncedOrderIds": sync_result["syncedOrderIds"],
                "failedOrderIds": sync_result["failedOrderIds"],
                "message": message,
            }
        ),
        status_code,
    )


@app.post("/api/registrations/event-completed")
def api_complete_registration_event():
    access_redirect = require_page_access("orders")
    if access_redirect is not None:
        return access_redirect

    config = get_config()
    if not config["store_id"] or not config["secret_token"]:
        return jsonify({"error": "Live Ecwid-koppeling staat nog niet aan."}), 400

    payload = request.get_json(silent=True) or {}
    product_key = str(payload.get("productKey", "") or "").strip()
    if not product_key:
        return jsonify({"error": "Product ontbreekt."}), 400

    products_payload = fetch_catalog_products()
    orders_payload = fetch_ecwid_orders(force_refresh=True)
    selected_product = build_registration_product_detail(
        products_payload.get("items", []),
        orders_payload.get("items", []),
        product_key,
    )
    if selected_product is None:
        return jsonify({"error": "Product niet gevonden."}), 404

    order_ids = normalize_registration_email_status_order_ids(
        [order.get("id", "") for order in selected_product.get("orders", [])]
    )
    if len(order_ids) > 500:
        return jsonify({"error": "Te veel bestellingen in één verzoek."}), 400

    sync_result = sync_registration_event_orders_to_delivered(order_ids)
    if sync_result["failedOrderIds"]:
        message = (
            f"{len(sync_result['syncedOrderIds'])} bestellingen op geleverd gezet, "
            f"{len(sync_result['failedOrderIds'])} niet bijgewerkt."
        )
        return (
            jsonify(
                {
                    "ok": False,
                    "productKey": product_key,
                    "orderIds": sync_result["orderIds"],
                    "syncedOrderIds": sync_result["syncedOrderIds"],
                    "failedOrderIds": sync_result["failedOrderIds"],
                    "message": message,
                }
            ),
            502,
        )

    set_registration_event_completed(product_key)
    message = (
        "Event afgerond. Er waren geen Ecwid-bestellingen om bij te werken."
        if not sync_result["orderIds"]
        else f"Event afgerond. {len(sync_result['syncedOrderIds'])} bestellingen zijn op geleverd gezet."
    )
    return jsonify(
        {
            "ok": True,
            "productKey": product_key,
            "orderIds": sync_result["orderIds"],
            "syncedOrderIds": sync_result["syncedOrderIds"],
            "failedOrderIds": [],
            "eventCompleted": True,
            "eventCanceled": False,
            "message": message,
        }
    )


@app.post("/api/registrations/event-canceled")
def api_cancel_registration_event():
    access_redirect = require_page_access("orders")
    if access_redirect is not None:
        return access_redirect

    config = get_config()
    if not config["store_id"] or not config["secret_token"]:
        return jsonify({"error": "Live Ecwid-koppeling staat nog niet aan."}), 400

    payload = request.get_json(silent=True) or {}
    product_key = str(payload.get("productKey", "") or "").strip()
    if not product_key:
        return jsonify({"error": "Product ontbreekt."}), 400

    products_payload = fetch_catalog_products()
    orders_payload = fetch_ecwid_orders(force_refresh=True)
    selected_product = build_registration_product_detail(
        products_payload.get("items", []),
        orders_payload.get("items", []),
        product_key,
    )
    if selected_product is None:
        return jsonify({"error": "Product niet gevonden."}), 404

    order_ids = normalize_registration_email_status_order_ids(
        [order.get("id", "") for order in selected_product.get("orders", [])]
    )
    if len(order_ids) > 500:
        return jsonify({"error": "Te veel bestellingen in één verzoek."}), 400

    sync_result = sync_registration_event_orders_to_returned(order_ids)
    if sync_result["failedOrderIds"]:
        message = (
            f"{len(sync_result['syncedOrderIds'])} bestellingen op geretourneerd gezet, "
            f"{len(sync_result['failedOrderIds'])} niet bijgewerkt."
        )
        return (
            jsonify(
                {
                    "ok": False,
                    "productKey": product_key,
                    "orderIds": sync_result["orderIds"],
                    "syncedOrderIds": sync_result["syncedOrderIds"],
                    "failedOrderIds": sync_result["failedOrderIds"],
                    "message": message,
                }
            ),
            502,
        )

    set_registration_event_canceled(product_key)
    message = (
        "Event geannuleerd. Er waren geen Ecwid-bestellingen om bij te werken."
        if not sync_result["orderIds"]
        else f"Event geannuleerd. {len(sync_result['syncedOrderIds'])} bestellingen zijn op geretourneerd gezet."
    )
    return jsonify(
        {
            "ok": True,
            "productKey": product_key,
            "orderIds": sync_result["orderIds"],
            "syncedOrderIds": sync_result["syncedOrderIds"],
            "failedOrderIds": [],
            "eventCompleted": False,
            "eventCanceled": True,
            "message": message,
        }
    )


@app.post("/bestellingen/teamindeling-export")
def export_orders_team_assignment():
    access_redirect = require_page_access("orders")
    if access_redirect is not None:
        return access_redirect

    selected_ids = parse_selected_order_ids(request.form.getlist("selected_order_ids"))
    if not selected_ids:
        return redirect(url_for("orders_page"))

    payload = fetch_ecwid_orders()
    orders_by_id = {
        str(order.get("id", "") or order.get("orderNumber", "")): order
        for order in payload.get("items", [])
    }
    selected_orders = [orders_by_id[order_id] for order_id in selected_ids if order_id in orders_by_id]

    if not selected_orders:
        return redirect(url_for("orders_page"))

    workbook_buffer = build_team_assignment_workbook(selected_orders)
    filename = f"teamindeling-bestellingen-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
    return (
        workbook_buffer.getvalue(),
        200,
        {
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


@app.get("/omzet")
def revenue_home_page() -> str:
    access_redirect = require_page_access("revenue")
    if access_redirect is not None:
        return access_redirect

    payload = fetch_orders_non_blocking()
    return render_template(
        "revenue_home.html",
        active_page="revenue",
        last_updated=format_cache_timestamp(payload.get("cachedAt", 0.0)),
        message=payload.get("message"),
    )


@app.get("/financien")
def financien_page() -> str:
    access_redirect = require_page_access("financien")
    if access_redirect is not None:
        return access_redirect

    return render_template("financien.html", active_page="financien")


@app.route("/financien/automatisch-facturen", methods=["GET", "POST"])
def automatic_invoices_page() -> str:
    access_redirect = require_page_access("financien")
    if access_redirect is not None:
        return access_redirect

    success = request.args.get("success", "").strip()
    error = request.args.get("error", "").strip()

    if request.method == "POST":
        action = request.form.get("action", "").strip()
        if action == "save_setting":
            saved, message = save_automatic_invoice_setting(request.form)
            return redirect(
                url_for(
                    "automatic_invoices_page",
                    success=message if saved else "",
                    error="" if saved else message,
                )
            )
        if action == "delete_setting":
            setting_id = request.form.get("setting_id", type=int)
            if setting_id:
                delete_automatic_invoice_setting(setting_id)
                return redirect(url_for("automatic_invoices_page", success="Automatische factuurinstelling verwijderd."))
            return redirect(url_for("automatic_invoices_page", error="Deze instelling kon niet worden verwijderd."))
        if action == "process_today":
            result = process_automatic_invoices(date.today())
            processed_count = len(result.get("processed", []))
            failed_count = sum(1 for item in result.get("processed", []) if item.get("error"))
            if failed_count:
                return redirect(
                    url_for(
                        "automatic_invoices_page",
                        error=f"{failed_count} automatische factuur{' is' if failed_count == 1 else 'en zijn'} mislukt.",
                    )
                )
            return redirect(
                url_for(
                    "automatic_invoices_page",
                    success=f"{processed_count} automatische factuur{' is' if processed_count == 1 else 'en zijn'} verwerkt.",
                )
            )

    return render_template(
        "automatic_invoices.html",
        active_page="financien",
        agenda_club_options=AGENDA_CLUB_OPTIONS,
        automatic_invoice_settings=build_automatic_invoice_page_settings(),
        today=date.today().isoformat(),
        success=success,
        error=error,
    )


@app.get("/omzet/totaal")
def revenue_total_page() -> str:
    access_redirect = require_page_access("revenue")
    if access_redirect is not None:
        return access_redirect

    payload = fetch_orders_non_blocking()
    orders = payload.get("items", [])
    moneybird = payload.get("moneybird", {})
    ledger_account_types = moneybird.get("ledgerAccountTypes", {})
    monthly_revenue_series = build_monthly_revenue_series(
        orders,
        moneybird.get("invoices", []),
        moneybird.get("financialMutations", []),
        ledger_account_types,
    )

    return render_template(
        "revenue_total.html",
        active_page="revenue",
        report_summary=payload.get("reportSummary", build_report_summary(payload.get("summary", {}), {})),
        monthly_revenue_series=monthly_revenue_series,
        last_updated=format_cache_timestamp(payload.get("cachedAt", 0.0)),
        message=payload.get("message"),
    )


@app.get("/omzet/per-maand")
def revenue_monthly_page() -> str:
    access_redirect = require_page_access("revenue")
    if access_redirect is not None:
        return access_redirect

    payload = fetch_orders_non_blocking()
    orders = payload.get("items", [])
    moneybird = payload.get("moneybird", {})
    moneybird_invoices = moneybird.get("invoices", [])
    financial_mutations = moneybird.get("financialMutations", [])
    ledger_account_types = moneybird.get("ledgerAccountTypes", {})
    month_options = build_profit_month_options(orders, moneybird_invoices, financial_mutations, ledger_account_types)

    selected_month = request.args.get("month", "").strip()
    available_months = {option["value"] for option in month_options}
    if selected_month not in available_months:
        selected_month = month_options[0]["value"] if month_options else datetime.now().strftime("%Y-%m")

    monthly_summary = build_monthly_revenue_summary(
        orders,
        moneybird_invoices,
        financial_mutations,
        ledger_account_types,
        selected_month,
    )

    return render_template(
        "revenue_monthly.html",
        active_page="revenue",
        month_options=month_options,
        monthly_summary=monthly_summary,
        last_updated=format_cache_timestamp(payload.get("cachedAt", 0.0)),
        message=payload.get("message"),
    )


@app.get("/omzet/winst")
def revenue_profit_page() -> str:
    access_redirect = require_page_access("revenue")
    if access_redirect is not None:
        return access_redirect

    payload = fetch_orders_non_blocking()
    orders = payload.get("items", [])
    moneybird = payload.get("moneybird", {})
    moneybird_invoices = moneybird.get("invoices", [])
    financial_mutations = moneybird.get("financialMutations", [])
    ledger_account_types = moneybird.get("ledgerAccountTypes", {})
    month_options = build_profit_month_options(orders, moneybird_invoices, financial_mutations, ledger_account_types)

    selected_month = request.args.get("month", "").strip()
    available_months = {option["value"] for option in month_options}
    if selected_month not in available_months:
        selected_month = month_options[0]["value"] if month_options else datetime.now().strftime("%Y-%m")

    monthly_summary = build_monthly_revenue_summary(
        orders,
        moneybird_invoices,
        financial_mutations,
        ledger_account_types,
        selected_month,
    )
    total_summary = build_profit_totals(orders, moneybird_invoices, financial_mutations, ledger_account_types)

    return render_template(
        "revenue_profit.html",
        active_page="revenue",
        month_options=month_options,
        total_summary=total_summary,
        monthly_summary=monthly_summary,
        last_updated=format_cache_timestamp(payload.get("cachedAt", 0.0)),
        message=payload.get("message"),
    )


@app.get("/omzet/per-seizoen")
def revenue_season_page() -> str:
    access_redirect = require_page_access("revenue")
    if access_redirect is not None:
        return access_redirect

    payload = fetch_orders_non_blocking()
    orders = payload.get("items", [])
    moneybird = payload.get("moneybird", {})
    moneybird_invoices = moneybird.get("invoices", [])
    financial_mutations = moneybird.get("financialMutations", [])
    ledger_account_types = moneybird.get("ledgerAccountTypes", {})
    season_options = build_football_season_options(start_year=2022)

    selected_season = request.args.get("season", "").strip()
    available_seasons = {option["value"] for option in season_options}
    if selected_season not in available_seasons:
        selected_season = season_options[0]["value"] if season_options else "2022"

    season_summary = build_football_season_summary(
        orders,
        moneybird_invoices,
        financial_mutations,
        ledger_account_types,
        int(selected_season),
    )

    return render_template(
        "revenue_season.html",
        active_page="revenue",
        season_options=season_options,
        season_summary=season_summary,
        last_updated=format_cache_timestamp(payload.get("cachedAt", 0.0)),
        message=payload.get("message"),
    )


@app.route("/spaarpot", methods=["GET", "POST"])
def spaarpot_page() -> str:
    access_redirect = require_page_access("spaarpot")
    if access_redirect is not None:
        return access_redirect

    if request.method == "POST":
        action = request.form.get("action", "").strip()
        selected_season = request.form.get("season", request.form.get("year", "")).strip() or str(get_season_start_year_for_date(date.today()))
        selected_quarter = request.form.get("quarter", "").strip() or "1"
        try:
            season_start_year = int(selected_season)
            quarter_number = int(selected_quarter)
        except ValueError:
            return redirect(url_for("spaarpot_page"))
        if quarter_number not in {1, 2, 3, 4}:
            return redirect(url_for("spaarpot_page", season=season_start_year))

        if action == "add_manual_entry":
            description = request.form.get("description", "").strip()
            amount = parse_decimal_amount(request.form.get("amount", ""))
            if description and amount > 0:
                calendar_year, calendar_quarter = get_calendar_period_for_season_quarter(season_start_year, quarter_number)
                create_spaarpot_manual_entry(calendar_year, calendar_quarter, description, amount)
            return redirect(url_for("spaarpot_page", season=season_start_year, quarter=quarter_number))

        if action == "delete_manual_entry":
            try:
                entry_id = int(request.form.get("entry_id", "0"))
            except ValueError:
                entry_id = 0
            if entry_id > 0:
                delete_spaarpot_manual_entry(entry_id)
            return redirect(url_for("spaarpot_page", season=season_start_year, quarter=quarter_number))

        return redirect(url_for("spaarpot_page", season=season_start_year, quarter=quarter_number))

    moneybird = fetch_moneybird_non_blocking()
    payment_entries = build_spaarpot_payment_entries(
        moneybird.get("invoices", []),
        moneybird.get("financialMutations", []),
    )
    payment_entries.extend(load_spaarpot_manual_entries())
    payment_entries = sorted(payment_entries, key=lambda item: (item["date"], item["invoiceId"]), reverse=True)
    season_options = build_spaarpot_season_options(payment_entries)

    selected_season = request.args.get("season", request.args.get("year", "")).strip()
    available_seasons = {option["value"] for option in season_options}
    if selected_season not in available_seasons:
        default_season = get_default_spaarpot_season(payment_entries)
        selected_season = default_season if default_season in available_seasons else (
            season_options[0]["value"] if season_options else str(get_season_start_year_for_date(date.today()))
        )

    spaarpot_summary = build_spaarpot_quarter_summary(payment_entries, int(selected_season))
    if not any(quarter["entryCount"] > 0 for quarter in spaarpot_summary["quarters"]):
        default_season = get_default_spaarpot_season(payment_entries)
        if default_season != selected_season and default_season in available_seasons:
            selected_season = default_season
            spaarpot_summary = build_spaarpot_quarter_summary(payment_entries, int(selected_season))

    selected_quarter = request.args.get("quarter", "").strip()
    available_quarters = {"1", "2", "3", "4"}
    if selected_quarter not in available_quarters:
        selected_quarter = next(
            (
                str(quarter["quarter"])
                for quarter in spaarpot_summary["quarters"]
                if quarter["entryCount"] > 0
            ),
            "1",
        )
    selected_quarter_number = int(selected_quarter)
    selected_quarter_summary = next(
        quarter
        for quarter in spaarpot_summary["quarters"]
        if quarter["quarter"] == selected_quarter_number
    )

    return render_template(
        "spaarpot.html",
        active_page="spaarpot",
        season_options=season_options,
        spaarpot_summary=spaarpot_summary,
        selected_quarter=selected_quarter_number,
        selected_quarter_summary=selected_quarter_summary,
        last_updated=format_cache_timestamp(moneybird.get("cachedAt", 0.0)),
        message=moneybird.get("message"),
    )


@app.get("/api/push/status")
def api_push_status():
    user = get_current_user()
    if not user:
        return jsonify({"error": "Je bent niet ingelogd."}), 401

    return jsonify(
        {
            "enabled": is_web_push_configured(),
            "publicKey": get_web_push_vapid_public_key(),
            "message": ""
            if is_web_push_configured()
            else "Pushmeldingen zijn nog niet volledig geconfigureerd op de server.",
        }
    )


@app.post("/api/push/subscribe")
def api_push_subscribe():
    user = get_current_user()
    if not user:
        return jsonify({"error": "Je bent niet ingelogd."}), 401
    if not is_web_push_configured():
        return jsonify({"error": "Pushmeldingen zijn nog niet volledig geconfigureerd op de server."}), 503

    payload = request.get_json(silent=True) or {}
    subscription = payload.get("subscription") if isinstance(payload.get("subscription"), dict) else payload
    try:
        save_web_push_subscription(str(user["id"]), subscription, request.headers.get("User-Agent", ""))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    return jsonify({"ok": True})


@app.post("/api/push/unsubscribe")
def api_push_unsubscribe():
    user = get_current_user()
    if not user:
        return jsonify({"error": "Je bent niet ingelogd."}), 401

    payload = request.get_json(silent=True) or {}
    delete_web_push_subscription(str(payload.get("endpoint") or "").strip())
    return jsonify({"ok": True})


@app.route("/trainersvergoedingen", methods=["GET", "POST"])
def trainer_fees_home_page() -> str:
    access_redirect = require_page_access("trainer-fees")
    if access_redirect is not None:
        return access_redirect

    if request.method == "POST":
        action = request.form.get("action", "").strip()
        selected_season = request.form.get("season", "").strip() or str(get_season_start_year_for_date(date.today()))
        selected_month = request.form.get("month", "").strip()
        try:
            season_start_year = int(selected_season)
            month_number = int(selected_month)
        except ValueError:
            return redirect(url_for("trainer_fees_home_page"))
        if action == "set_payment_status":
            save_trainer_fee_payment_status(
                request.form.get("trainer_id", "").strip(),
                season_start_year,
                month_number,
                request.form.get("paid", "").strip() == "1",
            )
        redirect_kwargs: Dict[str, Any] = {"season": season_start_year}
        selected_month_filter = request.form.get("selected_month", "").strip()
        if selected_month_filter:
            redirect_kwargs["month"] = selected_month_filter
        return redirect(url_for("trainer_fees_home_page", **redirect_kwargs))

    auto_mark_completed_agenda_trainings()
    all_trainings = load_agenda_trainings()
    season_options = build_trainer_fee_season_options(all_trainings)
    available_seasons = {option["value"] for option in season_options}
    selected_season = request.args.get("season", request.args.get("year", "")).strip() or str(get_season_start_year_for_date(date.today()))
    if selected_season not in available_seasons:
        selected_season = season_options[0]["value"] if season_options else str(get_season_start_year_for_date(date.today()))
    try:
        selected_month = int(request.args.get("month", "").strip())
    except ValueError:
        selected_month = None
    if selected_month is not None and (selected_month < 1 or selected_month > 12):
        selected_month = None
    fee_summary = build_trainer_fee_monthly_summary(int(selected_season), selected_month)

    return render_template(
        "trainer_fees_home.html",
        active_page="trainer-fees",
        season_options=season_options,
        fee_summary=fee_summary,
    )


@app.route("/profiel", methods=["GET", "POST"])
def personal_profile_page() -> str:
    user = get_current_user()
    if user is None:
        return redirect(url_for("login_page"))

    form_error = request.args.get("error", "").strip()
    form_success = request.args.get("success", "").strip()

    if request.method == "POST":
        if request.form.get("action", "").strip() == "save_planning" and is_trainer_user(user):
            update_trainer_fee_rows(user["id"], parse_trainer_fee_rows_from_form(request.form))
            return redirect(url_for("personal_profile_page", onderdeel="planning", success="Planning opgeslagen."))

        first_name = request.form.get("first_name", "").strip()
        last_name = request.form.get("last_name", "").strip()
        full_name = " ".join(part for part in [first_name, last_name] if part).strip()
        email = request.form.get("email", "").strip()
        submitted_role = request.form.get("system_role", "").strip()
        current_role = normalize_system_role(str(user.get("systemRole") or user.get("role") or ""))
        system_role = normalize_system_role(submitted_role) if user.get("isAdmin") else current_role
        is_admin = role_grants_admin_access(system_role)
        member_type = derive_member_type_from_role(system_role)
        knvb_license = request.form.get("knvb_license", "").strip()
        education = request.form.get("education", "").strip()
        phone = request.form.get("phone", "").strip()
        address = request.form.get("address", "").strip()
        city = request.form.get("city", "").strip()
        postal_code = request.form.get("postal_code", "").strip()
        bank_account_number = request.form.get("bank_account_number", "").strip()
        bank_account_name = request.form.get("bank_account_name", "").strip()
        notes = request.form.get("notes", "").strip()
        availability_days = request.form.getlist("availability_days")

        if not full_name or not email or not system_role or not address or not city or not postal_code or not bank_account_number or not bank_account_name:
            return redirect(url_for("personal_profile_page", error="Vul alle verplichte velden in."))
        if not is_valid_email_address(email):
            return redirect(url_for("personal_profile_page", error="Vul een geldig e-mailadres in."))
        if user.get("isAdmin") and not is_allowed_system_role(system_role):
            return redirect(url_for("personal_profile_page", error="Kies een geldige rol."))
        if trainer_email_exists(email, exclude_profile_id=user["id"]):
            return redirect(url_for("personal_profile_page", error="Dit e-mailadres is al gekoppeld aan een ander account."))

        update_trainer_profile(
            user["id"],
            full_name,
            email,
            build_internal_username(full_name, email, exclude_profile_id=user["id"]),
            member_type,
            system_role,
            knvb_license,
            education,
            phone,
            address,
            city,
            postal_code,
            bank_account_number,
            bank_account_name,
            notes,
            availability_days,
            is_admin,
        )
        return redirect(url_for("personal_profile_page", success="Profiel opgeslagen."))

    profile = dict(user)
    name_parts = profile.get("fullName", "").split()
    profile["firstName"] = name_parts[0] if name_parts else ""
    profile["lastName"] = " ".join(name_parts[1:]) if len(name_parts) > 1 else ""
    profile["initials"] = "".join(part[:1] for part in name_parts[:2]).upper() or "PP"
    is_trainer_profile = is_trainer_user(user)
    profile_sections = {"persoonlijke-informatie", "kleding-en-sleutels", "planning"}
    active_profile_section = request.args.get("onderdeel", "persoonlijke-informatie").strip()
    if not is_trainer_profile or active_profile_section not in profile_sections:
        active_profile_section = "persoonlijke-informatie"

    return render_template(
        "personal_profile.html",
        active_page="profile",
        profile=profile,
        form_error=form_error,
        form_success=form_success,
        can_edit_role=bool(user.get("isAdmin")),
        is_trainer_profile=is_trainer_profile,
        active_profile_section=active_profile_section,
        trainer_fee_agenda_activity_options=build_trainer_fee_agenda_activity_options(),
        trainer_fee_club_options_by_type=build_trainer_fee_club_options_by_type(),
    )


@app.route("/trainers", methods=["GET", "POST"])
def trainers_page() -> str:
    admin_redirect = require_admin_user()
    if admin_redirect is not None:
        return admin_redirect

    form_error = request.args.get("error", "").strip()
    form_success = request.args.get("success", "").strip()
    invite_link = str(session.pop("latest_invite_link", "") or "").strip()

    if request.method == "POST":
        action = request.form.get("action", "").strip()
        if action == "update":
            profile_id = request.form.get("profile_id", "").strip()
            first_name = request.form.get("first_name", "").strip()
            last_name = request.form.get("last_name", "").strip()
            full_name = " ".join(part for part in [first_name, last_name] if part).strip()
            email = request.form.get("email", "").strip()
            system_role = normalize_system_role(request.form.get("system_role", "").strip())
            is_admin = role_grants_admin_access(system_role)
            member_type = derive_member_type_from_role(system_role)
            knvb_license = request.form.get("knvb_license", "").strip()
            education = request.form.get("education", "").strip()
            availability_days = request.form.getlist("availability_days")
            phone = request.form.get("phone", "").strip()
            address = request.form.get("address", "").strip()
            city = request.form.get("city", "").strip()
            postal_code = request.form.get("postal_code", "").strip()
            bank_account_number = request.form.get("bank_account_number", "").strip()
            bank_account_name = request.form.get("bank_account_name", "").strip()
            notes = request.form.get("notes", "").strip()
            trainer_fees = parse_trainer_fee_rows_from_form(request.form)

            if not profile_id or not full_name or not email or not system_role:
                return redirect(url_for("trainers_page", error="Vul alle verplichte velden in."))
            if not is_valid_email_address(email):
                return redirect(url_for("trainers_page", error="Vul een geldig e-mailadres in."))
            if not is_allowed_system_role(system_role):
                return redirect(url_for("trainers_page", error="Kies een geldige rol."))

            existing_profile = next((item for item in load_trainer_profiles() if item.get("id") == profile_id), None)
            if existing_profile is None:
                return redirect(url_for("trainers_page", error="Dit teamlid bestaat niet meer."))

            if trainer_email_exists(email, exclude_profile_id=profile_id):
                return redirect(url_for("trainers_page", error="Dit e-mailadres bestaat al."))

            update_trainer_profile(
                profile_id,
                full_name,
                email,
                build_internal_username(full_name, email, exclude_profile_id=profile_id),
                member_type,
                system_role,
                knvb_license,
                education,
                phone,
                address,
                city,
                postal_code,
                bank_account_number,
                bank_account_name,
                notes,
                availability_days,
                is_admin,
                trainer_fees,
            )
            return redirect(url_for("trainers_page", success="Teamlid opgeslagen."))
        if action == "delete":
            profile_id = request.form.get("profile_id", "").strip()
            if not profile_id:
                return redirect(url_for("trainers_page", error="Teamlid kon niet worden verwijderd."))

            current_user = get_current_user()
            existing_profile = next((item for item in load_trainer_profiles() if item.get("id") == profile_id), None)
            if existing_profile is None:
                return redirect(url_for("trainers_page", error="Dit teamlid bestaat niet meer."))
            if current_user is not None and current_user.get("id") == profile_id:
                return redirect(url_for("trainers_page", error="Je kunt je eigen account niet verwijderen."))
            if existing_profile.get("isAdmin"):
                admin_count = sum(1 for item in load_trainer_profiles() if item.get("isAdmin"))
                if admin_count <= 1:
                    return redirect(url_for("trainers_page", error="De laatste admin kan niet worden verwijderd."))

            delete_trainer_profile(profile_id)
            return redirect(url_for("trainers_page", success="Teamlid verwijderd."))

        first_name = request.form.get("first_name", "").strip()
        last_name = request.form.get("last_name", "").strip()
        full_name = " ".join(part for part in [first_name, last_name] if part).strip()
        email = request.form.get("email", "").strip()
        system_role = normalize_system_role(request.form.get("system_role", "").strip())
        is_admin = role_grants_admin_access(system_role)
        member_type = derive_member_type_from_role(system_role)
        role = system_role or request.form.get("role", "").strip()
        knvb_license = request.form.get("knvb_license", "").strip()
        education = request.form.get("education", "").strip()
        availability_days = request.form.getlist("availability_days")
        phone = request.form.get("phone", "").strip()
        address = request.form.get("address", "").strip()
        city = request.form.get("city", "").strip()
        postal_code = request.form.get("postal_code", "").strip()
        bank_account_number = request.form.get("bank_account_number", "").strip()
        bank_account_name = request.form.get("bank_account_name", "").strip()
        notes = request.form.get("notes", "").strip()

        if not full_name or not email or not system_role or not address or not city or not postal_code or not bank_account_number or not bank_account_name:
            return redirect(url_for("trainers_page", error="Vul alle verplichte velden in."))
        if not is_valid_email_address(email):
            return redirect(url_for("trainers_page", error="Vul een geldig e-mailadres in."))
        if not is_allowed_system_role(system_role):
            return redirect(url_for("trainers_page", error="Kies een geldige rol."))

        if trainer_email_exists(email):
            return redirect(url_for("trainers_page", error="Dit e-mailadres bestaat al."))

        try:
            invite = create_trainer_invite_profile(
                full_name,
                email,
                role,
                member_type,
                system_role,
                knvb_license,
                education,
                availability_days,
                phone,
                address,
                city,
                postal_code,
                bank_account_number,
                bank_account_name,
                notes,
                is_admin=is_admin,
            )
        except sqlite3.IntegrityError:
            return redirect(url_for("trainers_page", error="Dit account kon niet worden opgeslagen. Controleer of het e-mailadres uniek is."))
        session["latest_invite_link"] = url_for("invite_accept_page", invite_token=invite["inviteToken"], _external=True)
        return redirect(url_for("trainers_page", success="Teamlid opgeslagen. De aanmeldlink is klaar om te delen."))

    profiles = load_trainer_profiles()
    for profile in profiles:
        created_at = parse_iso_datetime(profile.get("createdAt", ""))
        profile["createdAtDisplay"] = created_at.strftime("%d-%m-%Y %H:%M") if created_at else "-"
        initials = "".join(part[:1] for part in profile.get("fullName", "").split()[:2]).upper() or "TM"
        profile["initials"] = initials
        profile["availabilityLabel"] = ", ".join(profile.get("availabilityDays", [])) or "Niet ingesteld"
        profile["inviteLink"] = (
            url_for("invite_accept_page", invite_token=profile["inviteToken"], _external=True)
            if profile.get("inviteToken")
            else ""
        )

    return render_template(
        "trainers.html",
        active_page="trainers",
        trainer_profiles=profiles,
        account_debug=build_admin_account_debug_summary(),
        form_error=form_error,
        form_success=form_success,
        invite_link=invite_link,
        agenda_club_options=AGENDA_CLUB_OPTIONS,
        trainer_fee_activity_options=TRAINER_FEE_ACTIVITY_OPTIONS,
        trainer_fee_agenda_activity_options=build_trainer_fee_agenda_activity_options(),
        trainer_fee_club_options_by_type=build_trainer_fee_club_options_by_type(),
    )


@app.route("/agenda", methods=["GET", "POST"])
def agenda_page() -> str:
    access_redirect = require_page_access("agenda")
    if access_redirect is not None:
        return access_redirect

    user = get_current_user()
    is_trainer_agenda_user = is_trainer_user(user)
    can_manage_agenda = bool(
        user
        and role_grants_admin_access(str(user.get("systemRole") or user.get("role") or ""))
    )
    if request.method == "POST" and not can_manage_agenda:
        return "Trainers kunnen de agenda alleen bekijken.", 403

    auto_mark_completed_agenda_trainings()
    view_mode = normalize_agenda_label(request.args.get("view", "week")).lower() or "week"
    if view_mode not in {"week", "month"}:
        view_mode = "week"
    summary_filter = normalize_agenda_summary_filter(request.args.get("summary_filter", "total"))
    week_offset = request.args.get("week", default=0, type=int)
    month_offset = request.args.get("month", default=0, type=int)
    redirect_week = request.form.get("week", "").strip()
    redirect_month = request.form.get("month", "").strip()
    redirect_view = normalize_agenda_label(request.form.get("view", "")).lower()
    redirect_summary_filter = request.form.get("summary_filter", "").strip()
    if redirect_view in {"week", "month"}:
        view_mode = redirect_view
    if redirect_summary_filter:
        summary_filter = normalize_agenda_summary_filter(redirect_summary_filter)
    if redirect_week:
        try:
            week_offset = int(redirect_week)
        except ValueError:
            week_offset = week_offset
    if redirect_month:
        try:
            month_offset = int(redirect_month)
        except ValueError:
            month_offset = month_offset

    if request.method == "POST":
        action = request.form.get("action", "").strip()
        if action == "save_day_plans":
            raw_day_plans = request.form.get("day_plans", "").strip()
            raw_visible_dates = request.form.get("visible_dates", "").strip() or request.form.get("week_dates", "").strip()
            day_plans_payload = {}
            visible_dates: List[str] = []
            if raw_day_plans:
                try:
                    parsed_payload = json.loads(raw_day_plans)
                except json.JSONDecodeError:
                    parsed_payload = {}
                if isinstance(parsed_payload, dict):
                    day_plans_payload = {
                        str(key or "").strip(): str(value or "").strip()
                        for key, value in parsed_payload.items()
                    }
            if raw_visible_dates:
                try:
                    parsed_visible_dates = json.loads(raw_visible_dates)
                except json.JSONDecodeError:
                    parsed_visible_dates = []
                if isinstance(parsed_visible_dates, list):
                    visible_dates = [str(value or "").strip() for value in parsed_visible_dates if str(value or "").strip()]

            try:
                save_agenda_day_plans(day_plans_payload, replace_dates=visible_dates)
                return redirect(
                    url_for(
                        "agenda_page",
                        view=view_mode,
                        summary_filter=summary_filter,
                        week=week_offset,
                        month=month_offset,
                        success="Dagplanning opgeslagen.",
                    )
                )
            except ValueError as exc:
                return redirect(
                    url_for(
                        "agenda_page",
                        view=view_mode,
                        summary_filter=summary_filter,
                        week=week_offset,
                        month=month_offset,
                        error=str(exc),
                    )
                )

        if action == "delete_training":
            training_id = request.form.get("training_id", "").strip()
            delete_scope = request.form.get("update_scope", "").strip()
            original_signature = request.form.get("original_signature", "").strip()
            deleted_count = delete_agenda_training(training_id, delete_scope, original_signature)

            if deleted_count:
                return redirect(
                    url_for(
                        "agenda_page",
                        view=view_mode,
                        summary_filter=summary_filter,
                        week=week_offset,
                        month=month_offset,
                        success=f"{deleted_count} afspraak{' is' if deleted_count == 1 else 'en zijn'} verwijderd.",
                    )
                )
            return redirect(
                url_for(
                    "agenda_page",
                    view=view_mode,
                    summary_filter=summary_filter,
                    week=week_offset,
                    month=month_offset,
                    error="De afspraak kon niet worden verwijderd.",
                )
            )

        if action == "update_training":
            title = request.form.get("title", "").strip()
            training_id = request.form.get("training_id", "").strip()
            update_scope = request.form.get("update_scope", "").strip()
            original_signature = request.form.get("original_signature", "").strip()
            date_value = request.form.get("date", "").strip()
            time_value = request.form.get("time", "").strip()
            end_time_value = request.form.get("end_time", "").strip()
            location = normalize_agenda_club(request.form.get("location", ""))
            training_type = normalize_agenda_training_type(request.form.get("training_type", ""))
            status = normalize_agenda_training_status(request.form.get("status", ""))
            trainers = normalize_agenda_trainers(request.form.getlist("trainer_ids"))
            notes = request.form.get("notes", "").strip()
            resolved_end_time = end_time_value or (compute_default_end_time(time_value) if time_value else "")

            updated_count = update_agenda_training(
                training_id,
                update_scope,
                original_signature,
                title,
                date_value,
                time_value,
                resolved_end_time,
                location,
                training_type,
                status,
                trainers,
                notes,
            )
            if updated_count:
                return redirect(
                    url_for(
                        "agenda_page",
                        view=view_mode,
                        summary_filter=summary_filter,
                        week=week_offset,
                        month=month_offset,
                        success=f"{updated_count} afspraak{' is' if updated_count == 1 else 'en zijn'} bijgewerkt.",
                    )
                )
            return redirect(
                url_for(
                    "agenda_page",
                    view=view_mode,
                    summary_filter=summary_filter,
                    week=week_offset,
                    month=month_offset,
                    error="De afspraak kon niet worden bijgewerkt. Controleer titel, datum, type, club en starttijd.",
                )
            )

        if action == "bulk_add_trainings":
            title = request.form.get("title", "").strip()
            bulk_dates = request.form.getlist("bulk_dates")
            time_value = request.form.get("time", "").strip()
            end_time_value = request.form.get("end_time", "").strip()
            location = normalize_agenda_club(request.form.get("location", ""))
            training_type = normalize_agenda_training_type(request.form.get("training_type", ""))
            trainers = normalize_agenda_trainers(request.form.getlist("trainer_ids"))
            notes = request.form.get("notes", "").strip()
            resolved_end_time = end_time_value or (compute_default_end_time(time_value) if time_value else "")

            created_count = add_agenda_trainings_bulk(
                title,
                bulk_dates,
                time_value,
                resolved_end_time,
                location,
                training_type,
                trainers,
                notes,
            )
            if created_count:
                return redirect(
                    url_for(
                        "agenda_page",
                        view=view_mode,
                        summary_filter=summary_filter,
                        week=week_offset,
                        month=month_offset,
                        success=f"{created_count} trainingen toegevoegd.",
                    )
                )
            return redirect(
                url_for(
                    "agenda_page",
                    view=view_mode,
                    summary_filter=summary_filter,
                    week=week_offset,
                    month=month_offset,
                    error="Kies minimaal een datum en vul titel, type, club en starttijd in.",
                )
            )

        title = request.form.get("title", "").strip()
        date_value = request.form.get("date", "").strip()
        time_value = request.form.get("time", "").strip()
        end_time_value = request.form.get("end_time", "").strip()
        location = normalize_agenda_club(request.form.get("location", ""))
        training_type = normalize_agenda_training_type(request.form.get("training_type", ""))
        trainers = normalize_agenda_trainers(request.form.getlist("trainer_ids"))
        notes = request.form.get("notes", "").strip()

        if title and date_value and time_value and location and training_type:
            add_agenda_training(
                title,
                date_value,
                time_value,
                end_time_value or compute_default_end_time(time_value),
                location,
                training_type,
                trainers,
                notes,
            )
            return redirect(
                url_for(
                    "agenda_page",
                    view=view_mode,
                    summary_filter=summary_filter,
                    week=week_offset,
                    month=month_offset,
                    success="Training toegevoegd.",
                )
            )

        return redirect(
            url_for(
                "agenda_page",
                view=view_mode,
                summary_filter=summary_filter,
                week=week_offset,
                month=month_offset,
                error="Vul titel, datum, type, club en starttijd in.",
            )
        )

    today = date.today()
    week_start = today - timedelta(days=today.weekday()) + timedelta(days=week_offset * 7)
    month_start = add_months(today.replace(day=1), month_offset)
    week_days = get_week_days(week_start)
    month_weeks = build_agenda_month_days(month_start)
    month_days = [day for week in month_weeks for day in week]
    visible_days = week_days if view_mode == "week" else month_days
    visible_day_keys = [day["key"] for day in visible_days]
    day_plans = load_agenda_day_plans(visible_day_keys)
    all_day_plans = load_all_agenda_day_plans()
    filtered_summary_day_plans = filter_agenda_day_plans_for_summary(all_day_plans, summary_filter)
    selected_summary_filter = get_agenda_summary_filter_option(summary_filter)
    selected_summary_start = selected_summary_filter.get("start")
    selected_summary_end = selected_summary_filter.get("end")
    if isinstance(selected_summary_start, date) and isinstance(selected_summary_end, date):
        summary_trainings = filter_agenda_trainings_for_user(
            load_agenda_trainings(
                selected_summary_start.isoformat(),
                selected_summary_end.isoformat(),
            ),
            user,
        )
    else:
        summary_trainings = filter_agenda_trainings_for_user(load_agenda_trainings(), user)
    for day in visible_days:
        day["planType"] = day_plans.get(day["key"], "")
    agenda_day_plan_summary = build_agenda_day_plan_summary(
        add_football_day_only_no_activity_days(filtered_summary_day_plans, summary_trainings)
    )
    week_end = week_start + timedelta(days=6)
    month_visible_start = month_days[0]["date"] if month_days else month_start
    month_visible_end = month_days[-1]["date"] if month_days else month_start
    agenda_external_labels: Dict[str, List[str]] = {day["key"]: [] for day in visible_days}
    try:
        agenda_external_labels = build_agenda_external_labels(
            visible_day_keys,
            AGENDA_SCHOOL_REGION,
        )
    except requests.RequestException:
        agenda_external_labels = {day["key"]: [] for day in visible_days}
    trainings = filter_agenda_trainings_for_user(
        load_agenda_trainings(week_start.isoformat(), week_end.isoformat()),
        user,
    )
    month_trainings = filter_agenda_trainings_for_user(
        load_agenda_trainings(month_visible_start.isoformat(), month_visible_end.isoformat()),
        user,
    )
    calendar_events = build_agenda_week_events(trainings, week_start)
    month_events = build_agenda_month_events(month_trainings, set(visible_day_keys))
    time_slots = [f"{hour:02d}" for hour in range(24)]
    month_day_names = ["Zo", "Ma", "Di", "Wo", "Do", "Vr", "Za"]

    return render_template(
        "agenda.html",
        active_page="agenda",
        trainings=trainings,
        agenda_view=view_mode,
        agenda_summary_filter=summary_filter,
        agenda_summary_filter_options=AGENDA_SUMMARY_FILTER_OPTIONS,
        agenda_summary_filter_label=selected_summary_filter.get("label", "Totaal"),
        agenda_summary_filter_description=selected_summary_filter.get("description", ""),
        week_days=week_days,
        week_offset=week_offset,
        week_label=build_week_label(week_start),
        month_offset=month_offset,
        month_label=build_month_label(month_start),
        month_weeks=month_weeks,
        month_day_names=month_day_names,
        month_events=month_events,
        calendar_events=calendar_events,
        time_slots=time_slots,
        agenda_visible_dates=visible_day_keys,
        today_week_offset=0,
        agenda_day_plan_options=AGENDA_DAY_PLAN_OPTIONS,
        agenda_club_options=AGENDA_CLUB_OPTIONS,
        agenda_club_options_by_training_type=build_agenda_club_options_by_training_type(),
        agenda_training_type_options=AGENDA_TRAINING_TYPE_OPTIONS,
        agenda_training_status_options=AGENDA_TRAINING_STATUS_OPTIONS,
        agenda_trainer_options=build_agenda_trainer_options(),
        agenda_day_plan_summary=agenda_day_plan_summary,
        agenda_external_labels=agenda_external_labels,
        agenda_school_region=AGENDA_SCHOOL_REGION,
        can_manage_agenda=can_manage_agenda,
        success=request.args.get("success", "").strip(),
        error=request.args.get("error", "").strip(),
    )


@app.route("/oefeningen-bibliotheek", methods=["GET", "POST"])
def oefeningen_bibliotheek_page() -> str:
    access_redirect = require_page_access("oefeningen-bibliotheek")
    if access_redirect is not None:
        return access_redirect

    success = request.args.get("success", "").strip()
    error = request.args.get("error", "").strip()

    if request.method == "POST":
        action = request.form.get("action", "preview").strip() or "preview"
        if request.form.get("import_one_index") is not None:
            action = "import_one"
        if action == "preview":
            upload = request.files.get("exercise_file") or request.files.get("pptx_file")
            if upload is None or not upload.filename:
                return redirect(url_for("oefeningen_bibliotheek_page", error="Kies eerst een PowerPoint- of Word-bestand."))
            filename = upload.filename.lower()
            if not filename.endswith((".pptx", ".docx")):
                return redirect(url_for("oefeningen_bibliotheek_page", error="Upload een .pptx- of .docx-bestand."))

            try:
                file_bytes = upload.read()
                if filename.endswith(".docx"):
                    preview_exercises = parse_exercises_from_docx(file_bytes)
                    source_label = "Word"
                else:
                    preview_exercises = parse_exercises_from_pptx(file_bytes)
                    source_label = "PowerPoint"
                preview_exercises, skipped_count = filter_importable_exercises(preview_exercises)
                preview_id = save_exercise_import_preview(preview_exercises)
            except (zipfile.BadZipFile, XmlElementTree.ParseError, KeyError, ValueError, OSError):
                app.logger.exception("Importpreview kon niet worden gemaakt")
                return redirect(url_for("oefeningen_bibliotheek_page", error="Dit bestand kon niet worden gelezen."))

            if not preview_exercises:
                clear_exercise_import_preview(preview_id)
                if skipped_count:
                    return redirect(url_for("oefeningen_bibliotheek_page", error=f"Geen nieuwe oefeningen gevonden. {skipped_count} dubbele oefeningen overgeslagen."))
                return redirect(url_for("oefeningen_bibliotheek_page", error=f"Geen oefeningen gevonden in deze {source_label}."))

            exercises = load_exercises()
            skipped_message = f" {skipped_count} dubbele oefeningen overgeslagen." if skipped_count else ""
            return render_template(
                "oefeningen_bibliotheek.html",
                active_page="oefeningen-bibliotheek",
                exercises=add_exercise_field_svgs(exercises),
                categories=list(EXERCISE_CATEGORY_OPTIONS),
                age_groups=list(EXERCISE_AGE_GROUP_OPTIONS),
                import_preview=add_exercise_field_svgs(preview_exercises),
                import_preview_id=preview_id,
                success=f"{len(preview_exercises)} nieuwe oefeningen gevonden.{skipped_message} Controleer de preview en upload daarna wat je wilt bewaren.",
                error="",
            )

        preview_id = request.form.get("preview_id", "").strip()
        try:
            preview_exercises = load_exercise_import_preview(preview_id)
        except (OSError, ValueError, json.JSONDecodeError):
            return redirect(url_for("oefeningen_bibliotheek_page", error="De importpreview is verlopen. Upload het bestand opnieuw."))

        if action == "import_all":
            apply_submitted_exercise_import_edits(preview_exercises)
            imported_count = insert_exercises(preview_exercises)
            clear_exercise_import_preview(preview_id)
            return redirect(url_for("oefeningen_bibliotheek_page", success=f"{imported_count} oefeningen geupload."))

        if action == "import_one":
            exercise_index = request.form.get("import_one_index", type=int)
            if exercise_index is None:
                exercise_index = request.form.get("exercise_index", type=int)
            if exercise_index is None or exercise_index < 0 or exercise_index >= len(preview_exercises):
                return redirect(url_for("oefeningen_bibliotheek_page", error="Deze oefening kon niet worden gevonden in de preview."))
            apply_submitted_exercise_import_edits(preview_exercises)
            imported_count = insert_exercises([preview_exercises[exercise_index]])
            remaining_preview = [
                item
                for index, item in enumerate(preview_exercises)
                if index != exercise_index
            ]
            if remaining_preview:
                save_existing_exercise_import_preview(preview_id, remaining_preview)
            else:
                clear_exercise_import_preview(preview_id)
            exercises = load_exercises()
            return render_template(
                "oefeningen_bibliotheek.html",
                active_page="oefeningen-bibliotheek",
                exercises=add_exercise_field_svgs(exercises),
                categories=list(EXERCISE_CATEGORY_OPTIONS),
                age_groups=list(EXERCISE_AGE_GROUP_OPTIONS),
                import_preview=add_exercise_field_svgs(remaining_preview),
                import_preview_id=preview_id if remaining_preview else "",
                success="Oefening geupload." if imported_count else "Geen oefening geupload.",
                error="",
            )

        return redirect(url_for("oefeningen_bibliotheek_page", error="Onbekende importactie."))

    exercises = load_exercises()
    return render_template(
        "oefeningen_bibliotheek.html",
        active_page="oefeningen-bibliotheek",
        exercises=add_exercise_field_svgs(exercises),
        categories=list(EXERCISE_CATEGORY_OPTIONS),
        age_groups=list(EXERCISE_AGE_GROUP_OPTIONS),
        import_preview=[],
        import_preview_id="",
        success=success,
        error=error,
    )


@app.get("/oefenstof")
def oefenstof_page() -> str:
    access_redirect = require_page_access("oefenstof")
    if access_redirect is not None:
        return access_redirect

    return render_template("oefenstof.html", active_page="oefenstof")


@app.route("/oefeningen-videos", methods=["GET", "POST"])
def exercise_videos_page() -> str:
    access_redirect = require_page_access("exercise-videos")
    if access_redirect is not None:
        return access_redirect
    user = get_current_user()
    if not user or not user.get("isAdmin"):
        return redirect(url_for("oefeningen_bibliotheek_page", error="Alleen Admins mogen video's koppelen."))

    success = request.args.get("success", "").strip()
    error = request.args.get("error", "").strip()
    if request.method == "POST":
        action = request.form.get("action", "upload_video").strip() or "upload_video"
        exercise_id = request.form.get("exercise_id")
        if action == "delete_video":
            if delete_exercise_video(exercise_id):
                return redirect(url_for("exercise_videos_page", success="Video verwijderd."))
            return redirect(url_for("exercise_videos_page", error="Video kon niet worden verwijderd."))

        updated_exercise, upload_error = upload_exercise_video(exercise_id, request.files.get("exercise_video"))
        if updated_exercise is None:
            return redirect(url_for("exercise_videos_page", error=upload_error or "Video uploaden mislukt."))
        return redirect(url_for("exercise_videos_page", success=f"Video gekoppeld aan {updated_exercise['title']}."))

    exercises = load_exercises()
    return render_template(
        "exercise_videos.html",
        active_page="exercise-videos",
        exercises=exercises,
        video_storage=get_exercise_video_storage_config(),
        success=success,
        error=error,
    )


@app.post("/api/oefeningen-bibliotheek/category")
def api_update_exercise_category():
    access_redirect = require_page_access("oefeningen-bibliotheek")
    if access_redirect is not None:
        return access_redirect
    user = get_current_user()
    if not user or not user.get("isAdmin"):
        return jsonify({"error": "Alleen Admins mogen oefeningen bewerken."}), 403

    payload = request.get_json(silent=True) or {}
    exercise_id = payload.get("id")
    category = payload.get("category")
    normalized_category = normalize_exercise_category(category)
    if normalized_category not in EXERCISE_CATEGORY_OPTIONS:
        return jsonify({"error": "Kies een geldige categorie."}), 400
    if not update_exercise_category(exercise_id, normalized_category):
        return jsonify({"error": "Oefening niet gevonden."}), 404
    return jsonify({"ok": True, "category": normalized_category})


@app.post("/api/oefeningen-bibliotheek/update")
def api_update_exercise():
    access_redirect = require_page_access("oefeningen-bibliotheek")
    if access_redirect is not None:
        return access_redirect
    user = get_current_user()
    if not user or not user.get("isAdmin"):
        return jsonify({"error": "Alleen Admins mogen oefeningen bewerken."}), 403

    payload = request.get_json(silent=True) or {}
    updated_exercise = update_exercise(payload.get("id"), payload)
    if updated_exercise is None:
        return jsonify({"error": "Controleer titel en categorie."}), 400
    return jsonify({"ok": True, "exercise": updated_exercise})


@app.post("/api/oefeningen-bibliotheek/field-image")
def api_update_exercise_field_image():
    access_redirect = require_page_access("oefeningen-bibliotheek")
    if access_redirect is not None:
        return access_redirect
    user = get_current_user()
    if not user or not user.get("isAdmin"):
        return jsonify({"error": "Alleen Admins mogen oefeningen bewerken."}), 403

    field, error = normalize_exercise_field_image_upload(request.files.get("field_image"))
    if field is None:
        return jsonify({"error": error or "Afbeelding uploaden mislukt."}), 400

    updated_exercise = update_exercise_field_image(request.form.get("id"), field)
    if updated_exercise is None:
        return jsonify({"error": "Oefening niet gevonden."}), 404
    return jsonify({"ok": True, "exercise": updated_exercise})


@app.post("/api/oefeningen-bibliotheek/field-overlay")
def api_update_exercise_field_overlay():
    access_redirect = require_page_access("oefeningen-bibliotheek")
    if access_redirect is not None:
        return access_redirect
    user = get_current_user()
    if not user or not user.get("isAdmin"):
        return jsonify({"error": "Alleen Admins mogen veldtekeningen bewerken."}), 403

    payload = request.get_json(silent=True) or {}
    updated_exercise = update_exercise_field_overlay(payload.get("id"), payload)
    if updated_exercise is None:
        return jsonify({"error": "Veldtekening kon niet worden opgeslagen."}), 400
    return jsonify({"ok": True, "exercise": updated_exercise})


@app.post("/api/oefeningen-bibliotheek/video")
def api_update_exercise_video():
    access_redirect = require_page_access("exercise-videos")
    if access_redirect is not None:
        return access_redirect
    user = get_current_user()
    if not user or not user.get("isAdmin"):
        return jsonify({"error": "Alleen Admins mogen video's koppelen."}), 403

    updated_exercise, error = upload_exercise_video(request.form.get("id"), request.files.get("exercise_video"))
    if updated_exercise is None:
        return jsonify({"error": error or "Video uploaden mislukt."}), 400
    return jsonify({"ok": True, "exercise": updated_exercise})


@app.post("/api/oefeningen-bibliotheek/video/delete")
def api_delete_exercise_video():
    access_redirect = require_page_access("exercise-videos")
    if access_redirect is not None:
        return access_redirect
    user = get_current_user()
    if not user or not user.get("isAdmin"):
        return jsonify({"error": "Alleen Admins mogen video's verwijderen."}), 403

    payload = request.get_json(silent=True) or {}
    if not delete_exercise_video(payload.get("id")):
        return jsonify({"error": "Video niet gevonden."}), 404
    return jsonify({"ok": True})


@app.post("/api/oefeningen-bibliotheek/delete")
def api_delete_exercise():
    access_redirect = require_page_access("oefeningen-bibliotheek")
    if access_redirect is not None:
        return access_redirect
    user = get_current_user()
    if not user or not user.get("isAdmin"):
        return jsonify({"error": "Alleen Admins mogen oefeningen verwijderen."}), 403

    payload = request.get_json(silent=True) or {}
    if not delete_exercise(payload.get("id")):
        return jsonify({"error": "Oefening niet gevonden."}), 404
    return jsonify({"ok": True})


@app.get("/trainingen")
def trainingen_page() -> str:
    access_redirect = require_page_access("trainingen")
    if access_redirect is not None:
        return access_redirect

    return render_trainingen_page("home")


@app.get("/trainingen/opgeslagen")
def trainingen_saved_page() -> str:
    access_redirect = require_page_access("trainingen")
    if access_redirect is not None:
        return access_redirect

    return render_trainingen_page("saved")


@app.get("/trainingen/maker")
def trainingen_maker_page() -> str:
    access_redirect = require_page_access("trainingen")
    if access_redirect is not None:
        return access_redirect

    return render_trainingen_page("maker")


def render_trainingen_page(training_mode: str) -> str:
    return render_template(
        "trainingen.html",
        active_page="trainingen",
        training_mode=training_mode,
        exercises=add_exercise_field_svgs(load_exercises()),
        trainings=load_training_sessions(),
    )


@app.post("/api/trainingen")
def api_save_training():
    access_redirect = require_page_access("trainingen")
    if access_redirect is not None:
        return access_redirect

    payload = request.get_json(silent=True) or {}
    training = save_training_session(payload)
    if training is None:
        return jsonify({"error": "Geef de training een titel en voeg minimaal één oefening toe."}), 400
    return jsonify({"ok": True, "training": training, "trainings": load_training_sessions()})


@app.get("/voetbaldagen")
def football_days_page() -> str:
    return render_football_playbook_overview("voetbaldagen")


@app.get("/draaiboeken")
def draaiboeken_page() -> str:
    access_redirect = require_page_access("draaiboeken")
    if access_redirect is not None:
        return access_redirect

    return render_template("draaiboeken.html", active_page="draaiboeken")


@app.get("/samenwerkende-amateurclubs")
def amateur_clubs_page() -> str:
    return render_football_playbook_overview("samenwerkende-amateurclubs")


def render_football_playbook_overview(playbook_type: str) -> str:
    context = get_football_playbook_context(playbook_type)
    access_redirect = require_page_access(context["pageKey"])
    if access_redirect is not None:
        return access_redirect

    return render_template(
        "voetbaldagen.html",
        active_page=context["pageKey"],
        page_context=context,
        playbooks=attach_football_days_registration_counts(load_football_days_playbooks(context["playbookType"]), cached_only=True),
        success=request.args.get("success", "").strip(),
    )


@app.route("/voetbaldagen/nieuw", methods=["GET", "POST"])
def football_days_new_page() -> str:
    return render_football_playbook_new_page("voetbaldagen")


@app.route("/samenwerkende-amateurclubs/nieuw", methods=["GET", "POST"])
def amateur_clubs_new_page() -> str:
    return render_football_playbook_new_page("samenwerkende-amateurclubs")


@app.post("/samenwerkende-amateurclubs/<int:playbook_id>/dupliceren")
def amateur_clubs_duplicate_page(playbook_id: int) -> str:
    context = get_football_playbook_context("samenwerkende-amateurclubs")
    access_redirect = require_page_access(context["pageKey"])
    if access_redirect is not None:
        return access_redirect

    duplicate_id = duplicate_football_days_playbook(playbook_id, context["playbookType"])
    if duplicate_id is None:
        return redirect(context["overviewPath"])
    return redirect(f"{context['editPathPrefix']}/{duplicate_id}?success=Draaiboek gedupliceerd.")


def render_football_playbook_new_page(playbook_type: str) -> str:
    context = get_football_playbook_context(playbook_type)
    access_redirect = require_page_access(context["pageKey"])
    if access_redirect is not None:
        return access_redirect

    if request.method == "POST":
        playbook_id = save_football_days_playbook(
            build_football_days_playbook_from_form(context["playbookType"]),
            playbook_type=context["playbookType"],
        )
        return redirect(f"{context['editPathPrefix']}/{playbook_id}?success=Draaiboek opgeslagen.")

    playbook = create_empty_football_days_playbook(context["playbookType"])

    return render_template(
        "voetbaldagen_form.html",
        active_page=context["pageKey"],
        page_context=context,
        playbook=playbook,
        exercises=add_exercise_field_svgs(load_exercises()),
        previous_playbooks=attach_football_days_registration_counts(load_football_days_playbooks(context["playbookType"]), cached_only=True),
        form_action=context["newPath"],
        page_mode="new",
        success=request.args.get("success", "").strip(),
    )


@app.route("/voetbaldagen/<int:playbook_id>", methods=["GET", "POST"])
def football_days_edit_page(playbook_id: int) -> str:
    return render_football_playbook_edit_page("voetbaldagen", playbook_id)


@app.route("/samenwerkende-amateurclubs/<int:playbook_id>", methods=["GET", "POST"])
def amateur_clubs_edit_page(playbook_id: int) -> str:
    return render_football_playbook_edit_page("samenwerkende-amateurclubs", playbook_id)


def render_football_playbook_edit_page(playbook_type: str, playbook_id: int) -> str:
    context = get_football_playbook_context(playbook_type)
    access_redirect = require_page_access(context["pageKey"])
    if access_redirect is not None:
        return access_redirect

    playbook = load_football_days_playbook(playbook_id, context["playbookType"])
    if playbook is None:
        return redirect(context["overviewPath"])

    if request.method == "POST":
        save_football_days_playbook(
            build_football_days_playbook_from_form(context["playbookType"]),
            playbook_id=playbook_id,
            playbook_type=context["playbookType"],
        )
        return redirect(f"{context['editPathPrefix']}/{playbook_id}?success=Draaiboek opgeslagen.")

    if not playbook["staff"]:
        playbook["staff"] = [{"name": "", "role": "", "setupTask": ""}]
    if not playbook["program"]:
        playbook["program"] = [{"startTime": "", "endTime": "", "activity": "", "icon": "clock"}]
    attach_football_days_registration_counts([playbook])

    return render_template(
        "voetbaldagen_form.html",
        active_page=context["pageKey"],
        page_context=context,
        playbook=playbook,
        exercises=add_exercise_field_svgs(load_exercises()),
        previous_playbooks=attach_football_days_registration_counts(
            [item for item in load_football_days_playbooks(context["playbookType"]) if item["id"] != playbook_id],
            cached_only=True,
        ),
        form_action=f"{context['editPathPrefix']}/{playbook_id}",
        page_mode="edit",
        success=request.args.get("success", "").strip(),
    )


@app.route("/voorstellen-maker", methods=["GET", "POST"])
def voorstellen_maker_page() -> str:
    access_redirect = require_page_access("voorstellen-maker")
    if access_redirect is not None:
        return access_redirect

    form_state = build_proposal_form_state()

    if request.method == "POST":
        action = request.form.get("action", "").strip()
        if action == "create_proposal":
            club_name = request.form.get("club_name", "").strip()
            proposal_type = request.form.get("proposal_type", "").strip()
            season_start_year = request.form.get("season_start_year", "").strip()
            price_per_training = request.form.get("price_per_training", "").strip()
            lines = parse_proposal_lines_from_form(request.form)

            form_state = build_proposal_form_state(
                club_name=club_name,
                proposal_type=proposal_type,
                season_start_year=season_start_year,
                price_per_training=price_per_training,
                lines=lines,
            )
            validated_payload, error_message = validate_proposal_input(
                club_name,
                proposal_type,
                season_start_year,
                price_per_training,
                lines,
            )
            if error_message:
                return render_template(
                    "voorstellen_maker.html",
                    active_page="voorstellen-maker",
                    proposal_form=form_state,
                    proposal_type_options=PROPOSAL_TYPE_OPTIONS,
                    proposal_weekday_options=PROPOSAL_WEEKDAY_OPTIONS,
                    proposal_training_kind_options=PROPOSAL_TRAINING_KIND_OPTIONS,
                    proposal_season_options=build_football_season_options(
                        start_year=PROPOSAL_MIN_SEASON_START_YEAR
                    ),
                    proposals=load_proposals(),
                    error=error_message,
                    success="",
                )

            proposal_id = create_proposal(
                validated_payload["clubName"],
                validated_payload["proposalType"],
                validated_payload["seasonStartYear"],
                validated_payload["pricePerTraining"],
                validated_payload["lines"],
            )
            return redirect(
                url_for(
                    "voorstellen_maker_detail_page",
                    proposal_id=proposal_id,
                    success="Voorstel opgeslagen.",
                )
            )
        elif action == "delete_proposal":
            proposal_id = request.form.get("proposal_id", type=int)
            if proposal_id:
                delete_proposal(proposal_id)
                return redirect(url_for("voorstellen_maker_page", success="Voorstel verwijderd."))

    return render_template(
        "voorstellen_maker.html",
        active_page="voorstellen-maker",
        proposal_form=form_state,
        proposal_type_options=PROPOSAL_TYPE_OPTIONS,
        proposal_weekday_options=PROPOSAL_WEEKDAY_OPTIONS,
        proposal_training_kind_options=PROPOSAL_TRAINING_KIND_OPTIONS,
        proposal_season_options=build_football_season_options(start_year=PROPOSAL_MIN_SEASON_START_YEAR),
        proposals=load_proposals(),
        success=request.args.get("success", "").strip(),
        error=request.args.get("error", "").strip(),
    )


@app.route("/voorstellen-maker/<int:proposal_id>", methods=["GET", "POST"])
def voorstellen_maker_detail_page(proposal_id: int) -> str:
    access_redirect = require_page_access("voorstellen-maker")
    if access_redirect is not None:
        return access_redirect

    if request.method == "POST":
        action = request.form.get("action", "").strip()
        if action == "delete_proposal":
            delete_proposal(proposal_id)
            return redirect(url_for("voorstellen_maker_page", success="Voorstel verwijderd."))

    proposal = load_proposal_by_id(proposal_id)
    if proposal is None:
        return redirect(url_for("voorstellen_maker_page", error="Dit voorstel bestaat niet."))

    return render_template(
        "voorstellen_maker_detail.html",
        active_page="voorstellen-maker",
        proposal=proposal,
        success=request.args.get("success", "").strip(),
        error=request.args.get("error", "").strip(),
    )


@app.get("/api/voorstellen-maker/training-counts")
def voorstellen_maker_training_counts_api():
    access_redirect = require_page_access("voorstellen-maker")
    if access_redirect is not None:
        return jsonify({"error": "Je hebt geen toegang tot deze pagina."}), 403

    season_start_year_raw = str(request.args.get("season_start_year", "") or "").strip()
    try:
        season_start_year = int(season_start_year_raw)
    except ValueError:
        return jsonify({"error": "Kies een geldig seizoen."}), 400

    available_seasons = {
        int(option["value"])
        for option in build_football_season_options(start_year=PROPOSAL_MIN_SEASON_START_YEAR)
        if str(option.get("value", "")).isdigit()
    }
    if season_start_year not in available_seasons:
        return jsonify({"error": "Kies een seizoen uit de lijst."}), 400

    proposal_type = normalize_proposal_type(request.args.get("proposal_type", ""))
    proposal_type_option = get_proposal_type_option(proposal_type) if proposal_type else None
    weekday_counts = build_proposal_weekday_counts(
        season_start_year,
        proposal_type_option["agenda_plan_type"] if proposal_type_option else None,
    )
    return jsonify(
        {
            "weekdayCounts": weekday_counts,
            "totalTrainings": sum(int(value or 0) for value in weekday_counts.values()),
        }
    )


@app.get("/overeenkomsten")
def overeenkomsten_page() -> str:
    access_redirect = require_page_access("overeenkomsten")
    if access_redirect is not None:
        return access_redirect

    return render_template(
        "overeenkomsten.html",
        active_page="overeenkomsten",
        contracts=load_contracts(),
        success=request.args.get("success", "").strip(),
    )


@app.route("/overeenkomsten/nieuw", methods=["GET", "POST"])
def overeenkomsten_new_page() -> str:
    access_redirect = require_page_access("overeenkomsten")
    if access_redirect is not None:
        return access_redirect

    if request.method == "POST":
        contract = build_contract_from_form()
        if not contract["clubName"]:
            return render_template(
                "overeenkomsten_form.html",
                active_page="overeenkomsten",
                contract=contract,
                previous_contracts=load_contracts(),
                agenda_attachment_options=build_contract_agenda_attachment_options(contract),
                form_action=url_for("overeenkomsten_new_page"),
                page_mode="new",
                success="",
                error="Vul minimaal een clubnaam in.",
            )
        contract_id = save_contract(contract)
        return redirect(url_for("overeenkomsten_edit_page", contract_id=contract_id, success="Overeenkomst opgeslagen."))

    return render_template(
        "overeenkomsten_form.html",
        active_page="overeenkomsten",
        contract=normalize_contract(None),
        previous_contracts=load_contracts(),
        agenda_attachment_options=build_contract_agenda_attachment_options(normalize_contract(None)),
        form_action=url_for("overeenkomsten_new_page"),
        page_mode="new",
        success=request.args.get("success", "").strip(),
        error=request.args.get("error", "").strip(),
    )


@app.route("/overeenkomsten/<int:contract_id>", methods=["GET", "POST"])
def overeenkomsten_edit_page(contract_id: int) -> str:
    access_redirect = require_page_access("overeenkomsten")
    if access_redirect is not None:
        return access_redirect

    contract = load_contract(contract_id)
    if contract is None:
        return redirect(url_for("overeenkomsten_page"))

    if request.method == "POST":
        action = request.form.get("action", "save").strip() or "save"
        if action == "delete":
            with get_db_connection() as connection:
                connection.execute("DELETE FROM contracts WHERE id = ?", (contract_id,))
            clear_local_data_cache()
            return redirect(url_for("overeenkomsten_page", success="Overeenkomst verwijderd."))

        updated_contract = build_contract_from_form()
        if not updated_contract["clubName"]:
            contract.update(updated_contract)
            return render_template(
                "overeenkomsten_form.html",
                active_page="overeenkomsten",
                contract=contract,
                previous_contracts=[item for item in load_contracts() if item["id"] != contract_id],
                agenda_attachment_options=build_contract_agenda_attachment_options(contract),
                form_action=url_for("overeenkomsten_edit_page", contract_id=contract_id),
                page_mode="edit",
                success="",
                error="Vul minimaal een clubnaam in.",
            )
        save_contract(updated_contract, contract_id=contract_id)
        return redirect(url_for("overeenkomsten_edit_page", contract_id=contract_id, success="Overeenkomst opgeslagen."))

    return render_template(
        "overeenkomsten_form.html",
        active_page="overeenkomsten",
        contract=contract,
        previous_contracts=[item for item in load_contracts() if item["id"] != contract_id],
        agenda_attachment_options=build_contract_agenda_attachment_options(contract),
        form_action=url_for("overeenkomsten_edit_page", contract_id=contract_id),
        page_mode="edit",
        success=request.args.get("success", "").strip(),
        error=request.args.get("error", "").strip(),
    )


@app.get("/overeenkomsten/<int:contract_id>/export-pdf")
def overeenkomsten_export_pdf(contract_id: int):
    access_redirect = require_page_access("overeenkomsten")
    if access_redirect is not None:
        return access_redirect

    contract = load_contract(contract_id)
    if contract is None:
        return redirect(url_for("overeenkomsten_page"))
    try:
        pdf_bytes = create_contract_pdf(contract)
    except RuntimeError as exc:
        return jsonify({"error": str(exc)}), 500
    return (
        pdf_bytes,
        200,
        {
            "Content-Type": "application/pdf",
            "Content-Disposition": f'attachment; filename="{contract_filename(contract, "pdf")}"',
            "Cache-Control": "no-store",
        },
    )


@app.get("/overeenkomsten/<int:contract_id>/export-docx")
def overeenkomsten_export_docx(contract_id: int):
    access_redirect = require_page_access("overeenkomsten")
    if access_redirect is not None:
        return access_redirect

    contract = load_contract(contract_id)
    if contract is None:
        return redirect(url_for("overeenkomsten_page"))
    try:
        docx_bytes = create_contract_docx(contract)
    except RuntimeError as exc:
        return jsonify({"error": str(exc)}), 500
    return (
        docx_bytes,
        200,
        {
            "Content-Type": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "Content-Disposition": f'attachment; filename="{contract_filename(contract, "docx")}"',
            "Cache-Control": "no-store",
        },
    )


@app.route("/social-media", methods=["GET", "POST"])
def social_media_page() -> str:
    access_redirect = require_page_access("social-media")
    if access_redirect is not None:
        return access_redirect

    redirect_week = request.form.get("week_offset", default=0, type=int)

    if request.method == "POST":
        action = request.form.get("action", "").strip()
        if action == "create_idea":
            title = request.form.get("title", "").strip()
            platforms = parse_social_media_platforms(request.form.getlist("platform") or request.form.get("platform", ""))
            content_type = request.form.get("content_type", "").strip()
            priority = request.form.get("priority", "").strip() or "Midden"
            notes = request.form.get("notes", "").strip()
            if title and platforms and content_type:
                add_social_media_idea(title, platforms, content_type, priority, notes)
                return redirect(url_for("social_media_page", week=redirect_week, success="Contentidee opgeslagen."))
        elif action == "update_idea":
            idea_id = request.form.get("idea_id", type=int)
            title = request.form.get("title", "").strip()
            platforms = parse_social_media_platforms(request.form.getlist("platform") or request.form.get("platform", ""))
            content_type = request.form.get("content_type", "").strip()
            priority = request.form.get("priority", "").strip() or "Midden"
            notes = request.form.get("notes", "").strip()
            if idea_id and title and platforms and content_type:
                update_social_media_idea(idea_id, title, platforms, content_type, priority, notes)
                return redirect(url_for("social_media_page", week=redirect_week, success="Contentidee opgeslagen."))
            return redirect(url_for("social_media_page", week=redirect_week, error="Vul alle velden van het contentidee in."))
        elif action == "toggle_idea_scheduled":
            idea_id = request.form.get("idea_id", type=int)
            is_scheduled = request.form.get("is_scheduled") == "1"
            if idea_id:
                set_social_media_idea_scheduled(idea_id, is_scheduled)
                return redirect(url_for("social_media_page", week=redirect_week, success="Contentidee bijgewerkt."))
        elif action == "delete_idea":
            idea_id = request.form.get("idea_id", type=int)
            if idea_id:
                delete_social_media_idea(idea_id)
                return redirect(url_for("social_media_page", week=redirect_week, success="Contentidee verwijderd."))
        elif action == "create_plan":
            title = request.form.get("title", "").strip()
            platform = request.form.get("platform", "").strip()
            publish_date = request.form.get("publish_date", "").strip()
            publish_time = request.form.get("publish_time", "").strip()
            status = request.form.get("status", "").strip() or "Gepland"
            notes = request.form.get("notes", "").strip()
            idea_id = request.form.get("idea_id", type=int)
            if title and platform and publish_date and publish_time:
                add_social_media_schedule_item(title, platform, publish_date, publish_time, status, notes)
                if idea_id:
                    set_social_media_idea_scheduled(idea_id, True)
                return redirect(url_for("social_media_page", week=redirect_week, success="Contentplanning opgeslagen."))
            return redirect(url_for("social_media_page", week=redirect_week, error="Vul alle velden van de planning in."))
        elif action == "update_plan":
            plan_id = request.form.get("plan_id", type=int)
            title = request.form.get("title", "").strip()
            platform = request.form.get("platform", "").strip()
            publish_date = request.form.get("publish_date", "").strip()
            publish_time = request.form.get("publish_time", "").strip()
            status = request.form.get("status", "").strip() or "Gepland"
            notes = request.form.get("notes", "").strip()
            if plan_id and title and platform and publish_date and publish_time:
                update_social_media_schedule_item(plan_id, title, platform, publish_date, publish_time, status, notes)
                return redirect(url_for("social_media_page", week=redirect_week, success="Afspraak bijgewerkt."))
            return redirect(url_for("social_media_page", week=redirect_week, error="Vul alle velden van de afspraak in."))
        elif action == "delete_plan":
            plan_id = request.form.get("plan_id", type=int)
            if plan_id:
                delete_social_media_schedule_item(plan_id)
                return redirect(url_for("social_media_page", week=redirect_week, success="Geplande post verwijderd."))

    week_offset = request.args.get("week", default=0, type=int)
    today = date.today()
    week_start = today - timedelta(days=today.weekday()) + timedelta(days=week_offset * 7)
    week_days = get_week_days(week_start)
    schedule_items = load_social_media_schedule()
    ideas = load_social_media_ideas()
    calendar_events = build_social_media_week_events(schedule_items, week_start)
    time_slots = [f"{hour:02d}" for hour in range(24)]
    return render_template(
        "social_media.html",
        active_page="social-media",
        ideas=ideas,
        schedule_items=schedule_items,
        week_offset=week_offset,
        week_days=week_days,
        week_label=build_week_label(week_start),
        calendar_events=calendar_events,
        time_slots=time_slots,
        error=request.args.get("error", "").strip(),
        success=request.args.get("success", "").strip(),
    )


@app.get("/marketing")
def marketing_page() -> str:
    access_redirect = require_page_access("marketing")
    if access_redirect is not None:
        return access_redirect

    return render_template("marketing.html", active_page="marketing")


@app.route("/content", methods=["GET", "POST"])
def content_page() -> str:
    access_redirect = require_page_access("content")
    if access_redirect is not None:
        return access_redirect

    user = get_current_user()
    if request.method == "POST":
        if not can_manage_content(user):
            return redirect(url_for("content_page", error="Je hebt geen rechten om content te beheren."))

        action = request.form.get("action", "").strip()
        if action == "create_album":
            existing_album_id = request.form.get("album_id", default=0, type=int)
            new_album_title = request.form.get("album_title", "").strip()
            album_id = resolve_content_album_id(existing_album_id, new_album_title)
            if album_id is None:
                error_message = "Kies een bestaand album of vul een nieuwe albumtitel in."
                if request_prefers_json():
                    return jsonify({"ok": False, "error": error_message}), 400
                return redirect(url_for("content_page", error=error_message))

            album_url = url_for("content_album_page", album_id=album_id)
            if request_prefers_json():
                return jsonify({"ok": True, "albumId": album_id, "albumUrl": album_url})
            return redirect(album_url)

        if action == "upload_album_photos":
            existing_album_id = request.form.get("album_id", default=0, type=int)
            new_album_title = request.form.get("album_title", "").strip()
            uploaded_files = collect_content_upload_files()
            album_id = resolve_content_album_id(existing_album_id, new_album_title)
            if album_id is None:
                return redirect(url_for("content_page", error="Kies een bestaand album of vul een nieuwe albumtitel in."))

            created_new_album = not existing_album_id and bool(new_album_title)
            try:
                uploaded_count = upload_files_to_content_album(album_id, uploaded_files)
            except ValueError as exc:
                if created_new_album:
                    delete_empty_content_album(album_id)
                return redirect(url_for("content_page", error=str(exc)))
            except requests.RequestException:
                if created_new_album:
                    delete_empty_content_album(album_id)
                return redirect(url_for("content_page", error="Upload mislukt. Probeer het opnieuw."))

            return redirect(
                url_for(
                    "content_album_page",
                    album_id=album_id,
                    success=f"{uploaded_count} foto{'s' if uploaded_count != 1 else ''} geupload.",
                )
            )
        if action == "delete_album":
            album_id = request.form.get("album_id", type=int)
            if not album_id:
                return redirect(url_for("content_page", error="Geen album geselecteerd om te verwijderen."))
            try:
                deleted = delete_content_album(album_id)
            except requests.RequestException:
                return redirect(url_for("content_page", error="Album verwijderen mislukt. Probeer het opnieuw."))
            if not deleted:
                return redirect(url_for("content_page", error="Het gekozen album kon niet worden gevonden."))
            return redirect(url_for("content_page", success="Fotoalbum verwijderd."))

    repaired_albums = ensure_content_album_records_exist()
    albums = load_content_album_summaries()
    return render_template(
        "content.html",
        active_page="content",
        albums=albums,
        content_storage=build_content_storage_status(),
        content_debug=(
            build_admin_content_debug_summary(repaired_albums=repaired_albums)
            if user and user.get("isAdmin")
            else None
        ),
        can_manage_content=can_manage_content(user),
        success=request.args.get("success", "").strip(),
        error=request.args.get("error", "").strip(),
    )


@app.route("/content/<int:album_id>", methods=["GET", "POST"])
def content_album_page(album_id: int) -> str:
    access_redirect = require_page_access("content")
    if access_redirect is not None:
        return access_redirect

    user = get_current_user()
    album = load_content_album(album_id)
    if album is None:
        return redirect(url_for("content_page", error="Dit fotoalbum bestaat niet."))

    if request.method == "POST":
        if not can_manage_content(user):
            return redirect(url_for("content_album_page", album_id=album_id, error="Je hebt geen rechten voor deze actie."))

        action = request.form.get("action", "").strip()
        if action == "upload_album_photos":
            uploaded_files = collect_content_upload_files()
            try:
                uploaded_count = upload_files_to_content_album(album_id, uploaded_files)
            except ValueError as exc:
                return redirect(url_for("content_album_page", album_id=album_id, error=str(exc)))
            except requests.RequestException:
                return redirect(url_for("content_album_page", album_id=album_id, error="Upload mislukt. Probeer het opnieuw."))
            return redirect(
                url_for(
                    "content_album_page",
                    album_id=album_id,
                    success=f"{uploaded_count} foto{'s' if uploaded_count != 1 else ''} geupload.",
                )
            )
        if action == "delete_photo":
            photo_id = request.form.get("photo_id", type=int)
            if not photo_id:
                return redirect(url_for("content_album_page", album_id=album_id, error="Geen foto geselecteerd om te verwijderen."))
            try:
                deleted = delete_content_photo(photo_id, album_id)
            except requests.RequestException:
                return redirect(url_for("content_album_page", album_id=album_id, error="Foto verwijderen mislukt. Probeer het opnieuw."))
            if not deleted:
                return redirect(url_for("content_album_page", album_id=album_id, error="De gekozen foto kon niet worden gevonden."))
            return redirect(url_for("content_album_page", album_id=album_id, success="Foto verwijderd."))
        if action == "delete_album":
            try:
                deleted = delete_content_album(album_id)
            except requests.RequestException:
                return redirect(url_for("content_album_page", album_id=album_id, error="Album verwijderen mislukt. Probeer het opnieuw."))
            if not deleted:
                return redirect(url_for("content_page", error="Het gekozen album kon niet worden gevonden."))
            return redirect(url_for("content_page", success="Fotoalbum verwijderd."))

    photos = load_content_album_photos(album_id)
    return render_template(
        "content_album.html",
        active_page="content",
        album=album,
        photos=photos,
        content_storage=build_content_storage_status(),
        can_manage_content=can_manage_content(user),
        success=request.args.get("success", "").strip(),
        error=request.args.get("error", "").strip(),
    )


@app.get("/api/orders")
def api_orders():
    access_redirect = require_page_access("orders")
    if access_redirect is not None:
        return access_redirect

    try:
        force_refresh = request.args.get("refresh") == "1"
        return jsonify(fetch_ecwid_orders(force_refresh=force_refresh))
    except requests.HTTPError as exc:
        status_code = exc.response.status_code if exc.response is not None else 502
        return jsonify({"error": "Ecwid API request mislukt"}), status_code
    except requests.RequestException:
        return jsonify({"error": "Netwerkfout bij Ecwid"}), 502


@app.route("/api/v1/agenda/events", methods=["GET", "OPTIONS"])
def api_agenda_events():
    if request.method == "OPTIONS":
        return agenda_api_preflight_response()
    authentication_error = validate_agenda_api_request()
    if authentication_error is not None:
        return authentication_error

    start_date = str(request.args.get("start", "") or "").strip()
    end_date = str(request.args.get("end", "") or "").strip()
    if start_date and parse_iso_date(start_date) is None:
        return agenda_api_json_response(
            {"error": {"code": "invalid_start", "message": "Gebruik voor start het formaat JJJJ-MM-DD."}},
            400,
        )
    if end_date and parse_iso_date(end_date) is None:
        return agenda_api_json_response(
            {"error": {"code": "invalid_end", "message": "Gebruik voor end het formaat JJJJ-MM-DD."}},
            400,
        )
    if start_date and end_date and start_date > end_date:
        return agenda_api_json_response(
            {"error": {"code": "invalid_range", "message": "De startdatum moet voor de einddatum liggen."}},
            400,
        )

    include_cancelled = parse_agenda_api_bool(request.args.get("include_cancelled", ""), default=True)
    include_day_plans = parse_agenda_api_bool(request.args.get("include_day_plans", ""), default=False)
    events = build_agenda_api_events(
        start_date=start_date or None,
        end_date=end_date or None,
        include_cancelled=include_cancelled,
        include_day_plans=include_day_plans,
    )
    return agenda_api_json_response(
        {
            "apiVersion": "1",
            "timezone": str(settings.TIME_ZONE or "Europe/Amsterdam"),
            "generatedAt": datetime.now(tz=ZoneInfo("UTC")).isoformat(timespec="seconds").replace("+00:00", "Z"),
            "count": len(events),
            "filters": {
                "start": start_date or None,
                "end": end_date or None,
                "includeCancelled": include_cancelled,
                "includeDayPlans": include_day_plans,
            },
            "events": events,
        }
    )


@app.route("/api/v1/agenda/calendar.ics", methods=["GET", "OPTIONS"])
def api_agenda_calendar():
    if request.method == "OPTIONS":
        return agenda_api_preflight_response()
    authentication_error = validate_agenda_api_request(allow_query_parameter=True)
    if authentication_error is not None:
        return authentication_error

    include_cancelled = parse_agenda_api_bool(request.args.get("include_cancelled", ""), default=True)
    include_day_plans = parse_agenda_api_bool(request.args.get("include_day_plans", ""), default=False)
    events = build_agenda_api_events(
        include_cancelled=include_cancelled,
        include_day_plans=include_day_plans,
    )
    calendar_text = build_agenda_icalendar(events)
    response_headers = {
        "Content-Type": "text/calendar; charset=utf-8",
        "Content-Disposition": 'inline; filename="hws-agenda.ics"',
        "Cache-Control": "no-store",
        "X-Robots-Tag": "noindex, nofollow",
    }
    origin = str(request.headers.get("Origin", "") or "").strip().rstrip("/")
    if origin and origin in get_agenda_api_allowed_origins():
        response_headers.update(
            {
                "Access-Control-Allow-Origin": origin,
                "Access-Control-Allow-Methods": "GET, OPTIONS",
                "Access-Control-Allow-Headers": "Authorization, Content-Type",
                "Vary": "Origin",
            }
        )
    return calendar_text.encode("utf-8"), 200, response_headers


@app.get("/api/dashboard-summary")
def api_dashboard_summary():
    access_redirect = require_page_access("dashboard")
    if access_redirect is not None:
        return access_redirect

    try:
        force_refresh = request.args.get("refresh") == "1"
        payload = fetch_orders(force_refresh=force_refresh)
        frontend_payload = build_dashboard_frontend_payload(payload)
        user = get_current_user()
        if user is not None and not user.get("isAdmin"):
            frontend_payload["summary"] = {}
            frontend_payload["reportSummary"] = {}
            frontend_payload["monthlyRevenueSeries"] = []
            frontend_payload["moneybird"] = {}
        return jsonify(frontend_payload)
    except requests.HTTPError as exc:
        status_code = exc.response.status_code if exc.response is not None else 502
        return jsonify({"error": "Dashboardgegevens ophalen mislukt"}), status_code
    except requests.RequestException:
        return jsonify({"error": "Netwerkfout bij dashboardgegevens"}), 502


@app.get("/api/products/search")
def api_product_search():
    user = get_current_user()
    if user is None:
        return redirect(url_for("login_page", next=request.path))
    if not (user_can_access_page(user, "dashboard") or user_can_access_page(user, "voetbaldagen")):
        return redirect(url_for("personal_profile_page"))

    query = request.args.get("q", "").strip()
    try:
        return jsonify({"items": search_catalog_products(query)})
    except requests.HTTPError as exc:
        status_code = exc.response.status_code if exc.response is not None else 502
        return jsonify({"error": "Productzoekopdracht mislukt"}), status_code
    except requests.RequestException:
        return jsonify({"error": "Netwerkfout bij productzoekopdracht"}), 502


@app.get("/api/products/registration-count")
def api_product_registration_count():
    access_redirect = require_page_access("voetbaldagen")
    if access_redirect is not None:
        return access_redirect

    product_id = request.args.get("product_id", "").strip()
    product_name = request.args.get("product_name", "").strip()
    product_sku = request.args.get("product_sku", "").strip()
    if not product_id and not product_name and not product_sku:
        return jsonify({"productId": "", "registrationCount": 0})

    try:
        orders_payload = fetch_ecwid_orders()
        return jsonify(
            {
                "productId": product_id,
                "registrationCount": count_ecwid_product_registrations(
                    orders_payload.get("items", []),
                    product_id,
                    product_name,
                    product_sku,
                ),
            }
        )
    except requests.HTTPError as exc:
        status_code = exc.response.status_code if exc.response is not None else 502
        return jsonify({"error": "Aanmeldingen ophalen mislukt"}), status_code
    except requests.RequestException:
        return jsonify({"error": "Netwerkfout bij aanmeldingen"}), 502


@app.get("/api/voetbaldagen/registration-counts")
def api_football_days_registration_counts():
    return football_playbook_registration_counts_api("voetbaldagen")


@app.get("/api/samenwerkende-amateurclubs/registration-counts")
def api_amateur_clubs_registration_counts():
    return football_playbook_registration_counts_api("samenwerkende-amateurclubs")


def football_playbook_registration_counts_api(playbook_type: str):
    context = get_football_playbook_context(playbook_type)
    access_redirect = require_page_access(context["pageKey"])
    if access_redirect is not None:
        return access_redirect

    requested_ids = {
        int(item)
        for item in re.split(r"[, ]+", request.args.get("playbook_ids", "").strip())
        if item.isdigit()
    }
    playbooks = load_football_days_playbooks(context["playbookType"])
    if requested_ids:
        playbooks = [playbook for playbook in playbooks if int(playbook.get("id") or 0) in requested_ids]

    try:
        orders_payload = fetch_ecwid_orders()
        return jsonify(
            {
                "counts": build_football_days_registration_counts(playbooks, orders_payload.get("items", [])),
                "cachedAt": orders_payload.get("cachedAt", 0.0),
            }
        )
    except requests.HTTPError as exc:
        status_code = exc.response.status_code if exc.response is not None else 502
        return jsonify({"error": "Aanmeldingen ophalen mislukt"}), status_code
    except requests.RequestException:
        return jsonify({"error": "Netwerkfout bij aanmeldingen"}), 502


@app.get("/api/dashboard-events")
def api_dashboard_events():
    access_redirect = require_page_access("dashboard")
    if access_redirect is not None:
        return access_redirect

    return jsonify({"items": load_dashboard_events_config()})


@app.post("/api/dashboard-events")
def api_save_dashboard_events():
    access_redirect = require_page_access("dashboard")
    if access_redirect is not None:
        return access_redirect

    payload = request.get_json(silent=True) or {}
    items = payload.get("items", [])
    if not isinstance(items, list):
        return jsonify({"error": "Ongeldige payload"}), 400
    if len(items) > 50:
        return jsonify({"error": "Te veel items in één verzoek."}), 400

    sanitized_items = []
    for item in items:
        if not isinstance(item, dict):
            continue
        label = str(item.get("label", "")).strip()[:120]
        if not label:
            continue
        product_id = item.get("productId")
        sanitized_items.append(
            {
                "productId": product_id,
                "label": label,
                "matchTerms": [
                    str(term).strip()[:120]
                    for term in item.get("matchTerms", [label])
                    if str(term).strip()
                ][:10] or [label],
            }
        )

    save_dashboard_events_config(sanitized_items)
    return jsonify({"ok": True, "items": sanitized_items})


@app.get("/api/dashboard-weather")
def api_dashboard_weather():
    access_redirect = require_page_access("dashboard")
    if access_redirect is not None:
        return access_redirect

    settings = load_dashboard_weather_settings()
    lat = request.args.get("lat", type=float)
    lon = request.args.get("lon", type=float)
    location_name = request.args.get("name", "").strip()

    if lat is None or lon is None:
        try:
            lat = float(settings.get("weather_lat", "52.25"))
            lon = float(settings.get("weather_lon", "6.16"))
        except ValueError:
            lat, lon = 52.25, 6.16

    if not location_name:
        location_name = settings.get("weather_name", "Deventer")

    try:
        weather_response = requests.get(
            "https://api.open-meteo.com/v1/forecast",
            params={
                "latitude": lat,
                "longitude": lon,
                "current_weather": "true",
                "timezone": "auto",
            },
            timeout=8,
        )
        weather_response.raise_for_status()
        weather_payload = weather_response.json()
        current_weather = weather_payload.get("current_weather") or {}
        weather_code = int(current_weather.get("weathercode", -1))
        weather_meta = get_weather_description(weather_code)
        temperature = float(current_weather.get("temperature", 0))
        windspeed = float(current_weather.get("windspeed", 0))
        return jsonify(
            {
                "location": location_name,
                "temperature": round(temperature),
                "windspeed": round(windspeed, 1),
                "weatherCode": weather_code,
                "condition": weather_meta["label"],
                "icon": weather_meta["icon"],
                "isWarning": weather_code >= 61,
            }
        )
    except (requests.RequestException, TypeError, ValueError):
        return jsonify({"error": "Weergegevens ophalen mislukt"}), 502


@app.get("/api/agenda-school-holidays")
def api_agenda_school_holidays():
    access_redirect = require_page_access("agenda")
    if access_redirect is not None:
        return access_redirect

    raw_school_years = request.args.get("schoolYears", "").strip()
    school_years = []
    if raw_school_years:
        school_years = [normalize_agenda_label(value) for value in raw_school_years.split(",") if normalize_agenda_label(value)]
    if not school_years:
        current_year = date.today().year
        school_years = [f"{current_year}-{current_year + 1}"]

    region = normalize_agenda_region(request.args.get("region", "")) or "all"
    items: List[Dict[str, Any]] = []
    latest_cached_at = 0.0

    try:
        for school_year in school_years:
            payload = fetch_school_holidays_for_schoolyear(school_year, region)
            items.extend(payload.get("items", []))
            latest_cached_at = max(latest_cached_at, float(payload.get("cachedAt") or 0.0))
        return jsonify(
            {
                "items": items,
                "region": region,
                "cachedAt": latest_cached_at,
            }
        )
    except requests.HTTPError as exc:
        status_code = exc.response.status_code if exc.response is not None else 502
        return jsonify({"error": "Schoolvakanties ophalen mislukt"}), status_code
    except requests.RequestException:
        return jsonify({"error": "Netwerkfout bij schoolvakanties"}), 502


@app.get("/api/agenda-public-holidays")
def api_agenda_public_holidays():
    access_redirect = require_page_access("agenda")
    if access_redirect is not None:
        return access_redirect

    raw_years = request.args.get("years", "").strip()
    years: List[int] = []
    if raw_years:
        for value in raw_years.split(","):
            normalized_value = normalize_agenda_label(value)
            if not normalized_value:
                continue
            try:
                years.append(int(normalized_value))
            except ValueError:
                continue
    if not years:
        current_year = date.today().year
        years = [current_year, current_year + 1]

    items: List[Dict[str, Any]] = []
    latest_cached_at = 0.0

    try:
        for year in years:
            payload = fetch_public_holidays_for_year(year)
            items.extend(payload.get("items", []))
            latest_cached_at = max(latest_cached_at, float(payload.get("cachedAt") or 0.0))
        return jsonify(
            {
                "items": items,
                "cachedAt": latest_cached_at,
            }
        )
    except requests.HTTPError as exc:
        status_code = exc.response.status_code if exc.response is not None else 502
        return jsonify({"error": "Feestdagen ophalen mislukt"}), status_code
    except requests.RequestException:
        return jsonify({"error": "Netwerkfout bij feestdagen"}), 502


@app.get("/service-worker.js")
def service_worker():
    response = send_from_directory(os.path.join(os.path.dirname(__file__), "static"), "service-worker.js")
    response.headers["Service-Worker-Allowed"] = "/"
    response.cache_control.public = True
    response.cache_control.max_age = 60
    return response


@app.get("/manifest.webmanifest")
def web_manifest():
    response = send_from_directory(os.path.join(os.path.dirname(__file__), "static"), "manifest.webmanifest")
    response.headers["Content-Type"] = "application/manifest+json"
    response.cache_control.public = True
    response.cache_control.max_age = 3600
    return response


@app.after_request
def set_response_headers(response):
    if request.method == "GET" and response.status_code == 200:
        response.add_etag()
        response.make_conditional(request)

    if request.path in {"/manifest.webmanifest", "/service-worker.js"}:
        response.cache_control.public = True
        response.cache_control.max_age = 60 if request.path == "/service-worker.js" else 3600
    elif request.path.startswith("/static/"):
        response.cache_control.public = True
        response.cache_control.max_age = 31536000
        response.cache_control.immutable = True
    elif request.path.startswith("/api/dashboard-weather"):
        response.cache_control.private = True
        response.cache_control.max_age = 300
        response.cache_control.must_revalidate = True
    elif request.path.startswith("/api/"):
        response.cache_control.private = True
        response.cache_control.max_age = 60
        response.cache_control.must_revalidate = True
    else:
        response.cache_control.private = True
        response.cache_control.no_cache = True
        response.cache_control.must_revalidate = True
    response.headers["Vary"] = "Cookie, Accept-Encoding"
    return apply_security_headers(response)


@app.errorhandler(413)
def handle_request_entity_too_large(_exc):
    message = "Upload te groot. Verklein het bestand of upload minder bestanden tegelijk."
    if request.path.startswith("/api/") or request_prefers_json():
        return jsonify({"error": message}), 413
    return redirect(url_for("content_page", error=message))


@app.errorhandler(Exception)
def handle_unexpected_exception(exc):
    if isinstance(exc, HTTPException):
        return exc
    app.logger.exception("Onverwachte fout tijdens request", exc_info=exc)
    if request.path.startswith("/api/"):
        return jsonify({"error": "Interne serverfout"}), 500
    return "Er is een interne fout opgetreden.", 500


if __name__ == "__main__":
    debug_mode = get_env("FLASK_DEBUG") != "0"
    port = int(get_env("PORT") or "5001")
    app.run(debug=debug_mode, use_reloader=debug_mode, port=port)
